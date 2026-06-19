import sqlite3
import shutil
import uuid
import os
import json
import logging
import io
import csv
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, Form, Request, Response, Cookie, HTTPException, Depends
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import cloudinary
import cloudinary.uploader
import hashlib, time, secrets
from collections import defaultdict

# --- WHATSAPP ALERT ---
import threading

def send_whatsapp_alert(payload: dict):
    try:
        from twilio.rest import Client
        account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
        auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
        from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")
        recipients_env = os.environ.get("WHATSAPP_RECIPIENTS", "+2349160420100,+2349039587686,+2349072707396,+2348051383900,+2348089377590")
        to_numbers = [f"whatsapp:{n.strip()}" for n in recipients_env.split(",")]
        if not account_sid or not auth_token:
            logger.warning("Twilio credentials not set — WhatsApp alert skipped.")
            return
        import json as _json
        votes = payload.get("votes", {})
        accord_votes = votes.get("ACCORD", 0)
        top_rivals = sorted(
            [(p, v) for p, v in votes.items() if p != "ACCORD" and v > 0],
            key=lambda x: -x[1]
        )[:3]
        rival_str = ", ".join([f"{p}: {v}" for p, v in top_rivals]) or "None"
        msg = (
            f"🗳 *NEW PU SUBMISSION*\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📍 *PU:* {payload.get('location', 'N/A')}\n"
            f"🏛 *Ward:* {payload.get('ward', 'N/A')} | *LGA:* {payload.get('lg', 'N/A')}\n"
            f"🔑 *PU Code:* {payload.get('pu_code', 'N/A')}\n"
            f"👤 *Officer:* {payload.get('officer_id', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"✅ *ACCORD:* {accord_votes}\n"
            f"🔴 *Rivals:* {rival_str}\n"
            f"📊 *Total Cast:* {payload.get('total_cast', 0)} | *Accredited:* {payload.get('total_accredited', 0)}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🕐 {payload.get('timestamp', '')}\n"
            f"_Powered by Popson Geospatial Services_"
        )
        client = Client(account_sid, auth_token)
        for to_number in to_numbers:
            try:
                client.messages.create(from_=f"whatsapp:{from_number}", to=to_number, body=msg)
                logger.info(f"✅ WhatsApp alert sent to {to_number} for PU: {payload.get('pu_code')}")
            except Exception as sms_err:
                logger.error(f"Failed to send to {to_number}: {sms_err}")
    except Exception as e:
        logger.error(f"WhatsApp alert failed: {type(e).__name__}: {e}", exc_info=True)


# --- INCIDENT WHATSAPP ALERT ---
def send_incident_alert(payload: dict):
    try:
        from twilio.rest import Client
        account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
        auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
        from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")
        recipients_env = os.environ.get("WHATSAPP_RECIPIENTS", "+2349160420100,+2349039587686,+2349072707396,+2348051383900,+2348089377590")
        to_numbers = [f"whatsapp:{n.strip()}" for n in recipients_env.split(",")]
        if not account_sid or not auth_token:
            logger.warning("Twilio credentials not set — incident alert skipped.")
            return

        severity = payload.get("severity", "Unknown").upper()
        severity_icon = {"CRITICAL": "🔴", "MEDIUM": "🟡", "LOW": "🟢"}.get(severity, "⚪")

        msg = (
            f"🚨 *INCIDENT REPORT ALERT*\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"{severity_icon} *Severity:* {severity}\n"
            f"⚠️ *Type:* {payload.get('incident_type', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📍 *PU:* {payload.get('location', 'N/A')}\n"
            f"🏛 *Ward:* {payload.get('ward', 'N/A')} | *LGA:* {payload.get('lg', 'N/A')}\n"
            f"🔑 *PU Code:* {payload.get('pu_code', 'N/A')}\n"
            f"👤 *Officer:* {payload.get('officer_id', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📝 *Description:*\n{payload.get('description', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🕐 {payload.get('timestamp', '')}\n"
            f"_Powered by Popson Geospatial Services_"
        )

        client = Client(account_sid, auth_token)
        for to_number in to_numbers:
            try:
                client.messages.create(from_=f"whatsapp:{from_number}", to=to_number, body=msg)
                logger.info(f"✅ Incident alert sent to {to_number}")
            except Exception as sms_err:
                logger.error(f"Failed to send incident alert to {to_number}: {sms_err}")
    except Exception as e:
        logger.error(f"Incident alert failed: {type(e).__name__}: {e}", exc_info=True)

# --- CONFIGURATION ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

# ── CORS — restricted to same origin only ─────────────────────────────────────
from fastapi.middleware.cors import CORSMiddleware
_ALLOWED_ORIGINS = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "").split(",") if o.strip()]
if not _ALLOWED_ORIGINS:
    _ALLOWED_ORIGINS = ["*"]   # fallback for local dev; set ALLOWED_ORIGINS in prod
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type", "X-Requested-With"],
)
# ─────────────────────────────────────────────────────────────────────────────

# ── Security headers middleware ───────────────────────────────────────────────
from starlette.middleware.base import BaseHTTPMiddleware
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "SAMEORIGIN"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "geolocation=(self), camera=(self)"
        response.headers["Cache-Control"] = "no-store"
        return response
app.add_middleware(SecurityHeadersMiddleware)
# ─────────────────────────────────────────────────────────────────────────────

# ── Dashboard key (set DASHBOARD_KEY env var — keep it secret!) ───────────────
_DASHBOARD_KEY_HASH = hashlib.sha256(
    os.environ.get("DASHBOARD_KEY", "changeme-set-in-env").encode()
).hexdigest()
_SESSION_TOKENS: dict = {}   # token -> expiry timestamp
_SESSION_TTL = 8 * 3600      # 8 hours

def _make_session_token() -> str:
    token = secrets.token_hex(32)
    _SESSION_TOKENS[token] = time.time() + _SESSION_TTL
    return token

def _is_valid_token(token) -> bool:  # accepts str or None
    if not token:
        return False
    expiry = _SESSION_TOKENS.get(token)
    if not expiry:
        return False
    if time.time() > expiry:
        _SESSION_TOKENS.pop(token, None)
        return False
    return True

def _require_dashboard(request: Request):
    token = request.cookies.get("ds_session")
    if not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Dashboard access denied")

def _require_admin(request: Request):
    """Accepts EITHER the admin Bearer key OR a valid ds_session cookie —
    matches the pattern used by the existing /api/admin/* routes, so the
    same admin-portal login (Bearer token) works for Supabase-backed routes too."""
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    bearer_ok = auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )
    if not bearer_ok and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")

# ── OTP store ────────────────────────────────────────────────────────────────
# Structure: { officer_id: { otp, expiry, phone_hint, used, attempts, locked_until } }
_OTP_STORE: dict = {}
_OTP_TTL        = 5 * 60        # 5 minutes
_OTP_MAX_TRIES  = 3             # wrong attempts before lockout
_OTP_LOCKOUT    = 15 * 60       # 15-minute lockout after max attempts
_SUBMIT_TOKENS: dict = {}       # token -> { officer_id, expiry }
_SUBMIT_TOKEN_TTL = 30 * 60     # submission token valid 30 min

def _generate_otp() -> str:
    """Cryptographically random 6-digit OTP."""
    import random as _rnd
    return str(_rnd.SystemRandom().randint(100000, 999999))

def _mask_phone(phone: str) -> str:
    """Return +234***4567 style hint."""
    if len(phone) < 6:
        return "****"
    return phone[:4] + "***" + phone[-4:]

def _clean_phone(p: str) -> str:
    """Normalize Nigerian phone numbers to E.164 format (+234XXXXXXXXXX).
    Strips ALL non-digit chars first, then rebuilds E.164 — handles stray
    plus signs mid-string, spaces, dashes, and any other garbage.
    """
    import re as _re
    digits = _re.sub(r"[^0-9]", "", str(p).strip())
    if not digits:
        return ""
    if digits.startswith("234") and len(digits) == 13:
        return "+" + digits          # 2348012345678  -> +2348012345678
    if digits.startswith("0") and len(digits) == 11:
        return "+234" + digits[1:]   # 08012345678    -> +2348012345678
    if len(digits) == 10:
        return "+234" + digits       # 8012345678     -> +2348012345678
    if digits.startswith("234"):
        return "+" + digits          # any other 234X -> +234X
    return "+" + digits              # best-effort

def _make_submit_token(officer_id: str) -> str:
    """Short-lived signed token that authorises one submission."""
    import hmac as _hmac
    token = secrets.token_hex(24)
    expiry = time.time() + _SUBMIT_TOKEN_TTL
    # Sign: HMAC of token+officer_id with DASHBOARD_KEY_HASH as key material
    sig = _hmac.new(
        _DASHBOARD_KEY_HASH.encode(),
        (token + officer_id).encode(),
        "sha256"
    ).hexdigest()[:16]
    full = f"{token}.{sig}"
    _SUBMIT_TOKENS[full] = {"officer_id": officer_id, "expiry": expiry}
    return full

def _verify_submit_token(token: str, officer_id: str) -> bool:
    """Verify token is valid, not expired, and belongs to this officer."""
    entry = _SUBMIT_TOKENS.get(token)
    if not entry:
        return False
    if time.time() > entry["expiry"]:
        _SUBMIT_TOKENS.pop(token, None)
        return False
    if entry["officer_id"] != officer_id:
        return False
    # Single-use: remove after first successful check
    _SUBMIT_TOKENS.pop(token, None)
    return True
# ─────────────────────────────────────────────────────────────────────────────

# ── Rate limiter for validate_officer (10 attempts / 60s per IP) ──────────────
_rl_store: dict = defaultdict(list)
_RL_MAX   = 10
_RL_WINDOW = 60

def _check_rate_limit(ip: str):
    now = time.time()
    calls = [t for t in _rl_store[ip] if now - t < _RL_WINDOW]
    calls.append(now)
    _rl_store[ip] = calls
    if len(calls) > _RL_MAX:
        raise HTTPException(status_code=429, detail="Too many attempts. Try again in 60 seconds.")

# ── File upload guard ─────────────────────────────────────────────────────────
_MAX_UPLOAD_BYTES = 5 * 1024 * 1024   # 5 MB
_ALLOWED_MIME_PREFIXES = ("image/jpeg", "image/png", "image/webp", "image/gif")
_ALLOWED_MAGIC = {
    b"\xff\xd8\xff": "image/jpeg",
    b"\x89PNG":        "image/png",
    b"RIFF":           "image/webp",
    b"GIF8":           "image/gif",
}

async def _safe_read_image(upload: UploadFile, max_bytes: int = _MAX_UPLOAD_BYTES) -> bytes:
    data = await upload.read(max_bytes + 1)
    if len(data) > max_bytes:
        raise HTTPException(status_code=413, detail="Image too large (max 5MB)")
    magic = data[:4]
    ok = any(magic.startswith(sig) for sig in _ALLOWED_MAGIC)
    if not ok:
        raise HTTPException(status_code=415, detail="Unsupported file type. JPEG/PNG/WebP only.")
    return data
# ─────────────────────────────────────────────────────────────────────────────

# ── Cloudinary Configuration ──────────────────────────────────────────────────
cloudinary.config(
    cloud_name = os.environ.get("CLOUDINARY_CLOUD_NAME", ""),
    api_key    = os.environ.get("CLOUDINARY_API_KEY", ""),
    api_secret = os.environ.get("CLOUDINARY_API_SECRET", "")
)
# ─────────────────────────────────────────────────────────────────────────────

# Render-safe Pathing
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "static", "logos")
os.makedirs(LOGO_PATH, exist_ok=True)
app.mount("/logos", StaticFiles(directory=LOGO_PATH), name="logos")

STATIC_PATH = os.path.join(BASE_DIR, "static")
EC8E_PATH = os.path.join(BASE_DIR, "static", "ec8e")
os.makedirs(EC8E_PATH, exist_ok=True)
os.makedirs(STATIC_PATH, exist_ok=True)
app.mount("/static", StaticFiles(directory=STATIC_PATH), name="static")

# --- DATABASE CONNECTION ---
DB_RELEASE_URL = os.environ.get(
    "DB_RELEASE_URL",
    "https://github.com/Popson0205/Election_Dashboard03/releases/download/v1/election_v3.db"
)
DB_PATH = os.path.join(BASE_DIR, "election_v3.db")

def _ensure_db():
    """Download the SQLite DB from GitHub Releases if not already present."""
    if not os.path.exists(DB_PATH):
        import urllib.request
        logger.info(f"Downloading database from {DB_RELEASE_URL} ...")
        urllib.request.urlretrieve(DB_RELEASE_URL, DB_PATH)
        logger.info(f"Database saved to {DB_PATH}")

_ensure_db()

class _DictRow(sqlite3.Row):
    """Make sqlite3.Row behave like a dict (supports row['key'] and row.get())."""
    def get(self, key, default=None):
        try:
            return self[key]
        except (IndexError, KeyError):
            return default

class _FakeCursor:
    """Wraps sqlite3.Cursor so it supports 'with conn.cursor() as cur:' syntax."""
    def __init__(self, cur):
        self._cur = cur
    def execute(self, sql, params=()):
        self._cur.execute(sql, params)
    def fetchone(self):
        return self._cur.fetchone()
    def fetchall(self):
        return self._cur.fetchall()
    def close(self):
        self._cur.close()
    def __enter__(self):
        return self
    def __exit__(self, *args):
        self._cur.close()

class _FakeConn:
    """Wraps sqlite3 connection so it works as a context manager like psycopg2."""
    def __init__(self, conn):
        self._conn = conn
    def cursor(self):
        return _FakeCursor(self._conn.cursor())
    def commit(self):
        self._conn.commit()
    def close(self):
        self._conn.close()
    def __enter__(self):
        return self
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self._conn.commit()
        self._conn.close()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _DictRow
    return _FakeConn(conn)

# --- SUPABASE POSTGRES CONNECTION (for field_submissions / incidents) ---
# polling_units stays on SQLite above (static reference data, safe to be
# ephemeral — it's just re-downloaded from GitHub on every boot).
# field_submissions and incidents are USER-GENERATED data that must survive
# redeploys, so they live in Supabase's actual Postgres database instead.
#
# Get this connection string from: Supabase → Settings → Database →
# Connection string → "Transaction pooler" (or "Session pooler"), then set
# it as SUPABASE_DB_URL in Render's environment variables.
import psycopg2
import psycopg2.extras

SUPABASE_DB_URL = os.environ.get("SUPABASE_DB_URL", "")
if not SUPABASE_DB_URL:
    raise RuntimeError(
        "SUPABASE_DB_URL environment variable is not set. "
        "Set it to your Supabase Postgres connection string "
        "(Supabase → Settings → Database → Connection string)."
    )
if SUPABASE_DB_URL.startswith("postgres://"):
    SUPABASE_DB_URL = SUPABASE_DB_URL.replace("postgres://", "postgresql://", 1)


class _PgDictRow(dict):
    """Dict-like row, matching the sqlite3.Row-based _DictRow interface above."""
    def get(self, key, default=None):
        return super().get(key, default)


class _PgFakeCursor:
    """Wraps a psycopg2 cursor so 'with conn.cursor() as cur:' call sites keep working."""
    def __init__(self, cur):
        self._cur = cur
    def execute(self, sql, params=()):
        self._cur.execute(sql, params)
    def fetchone(self):
        row = self._cur.fetchone()
        return _PgDictRow(row) if row is not None else None
    def fetchall(self):
        return [_PgDictRow(r) for r in self._cur.fetchall()]
    def close(self):
        self._cur.close()
    def __enter__(self):
        return self
    def __exit__(self, *args):
        self._cur.close()


class _PgFakeConn:
    """Wraps a psycopg2 connection so 'with get_pg() as conn:' call sites keep working."""
    def __init__(self, conn):
        self._conn = conn
    def cursor(self):
        return _PgFakeCursor(self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor))
    def commit(self):
        self._conn.commit()
    def close(self):
        self._conn.close()
    def __enter__(self):
        return self
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._conn.close()


def get_pg():
    conn = psycopg2.connect(SUPABASE_DB_URL)
    return _PgFakeConn(conn)

# --- DATABASE INITIALIZATION ---
def init_db():
    # ── SQLite: polling_units reference data only (officer_phone migration) ──
    try:
        conn = get_db()
        cur = conn.cursor()
        try:
            cur.execute("ALTER TABLE polling_units ADD COLUMN officer_phone TEXT")
        except Exception:
            pass  # Column already exists — fine
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ SQLite (polling_units) INIT ERROR: {e}")

    # ── Supabase Postgres: field_submissions / incidents / result_audit_log ──
    try:
        conn = get_pg()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS field_submissions (
                id SERIAL PRIMARY KEY,
                officer_id TEXT,
                state TEXT,
                lg TEXT,
                ward TEXT,
                ward_code TEXT,
                pu_code TEXT,
                location TEXT,
                reg_voters INTEGER,
                total_accredited INTEGER,
                valid_votes INTEGER,
                rejected_votes INTEGER,
                total_cast INTEGER,
                lat REAL,
                lon REAL,
                timestamp TEXT,
                votes_json TEXT,
                ec8e_image TEXT,
                reviewed INTEGER DEFAULT 0,
                reviewed_by TEXT,
                reviewed_at TEXT,
                edited_votes_json TEXT,
                edit_note TEXT,
                UNIQUE(pu_code)
            )
        """)

        # Audit log table — records every edit made in the results portal
        cur.execute("""
            CREATE TABLE IF NOT EXISTS result_audit_log (
                id          SERIAL PRIMARY KEY,
                submission_id INTEGER,
                action      TEXT,        -- 'edit' | 'approve' | 'unapprove'
                field       TEXT,        -- field that changed (or 'bulk')
                old_value   TEXT,
                new_value   TEXT,
                changed_by  TEXT,        -- 'admin' for now
                changed_at  TEXT
            )
        """)

        # ── Incidents table ────────────────────────────────────────────────────
        cur.execute("""
            CREATE TABLE IF NOT EXISTS incidents (
                id SERIAL PRIMARY KEY,
                officer_id TEXT,
                pu_code TEXT,
                ward TEXT,
                ward_code TEXT,
                lg TEXT,
                state TEXT,
                location TEXT,
                incident_type TEXT,
                severity TEXT,
                description TEXT,
                evidence_url TEXT,
                lat REAL,
                lon REAL,
                timestamp TEXT,
                status TEXT DEFAULT 'open'
            )
        """)

        conn.commit()
        conn.close()
        print("✅ Supabase Postgres tables (field_submissions/incidents) ready")
    except Exception as e:
        print(f"❌ Supabase Postgres INIT ERROR: {e}")

init_db()

# =============================================================================
# ── SUPABASE OFFICER MANAGEMENT ───────────────────────────────────────────────
# Officers uploaded via /api/upload-officers are stored permanently in Supabase.
# This survives all Render redeployments. Add SUPABASE_URL + SUPABASE_KEY env
# vars on Render, and create the officers table in your Supabase SQL editor:
#
#   CREATE TABLE officers (
#     officer_id   TEXT PRIMARY KEY,
#     lga          TEXT,
#     ward         TEXT,
#     polling_unit TEXT,
#     phone        TEXT
#   );
# =============================================================================

def get_supabase():
    """Return a Supabase client, or None if env vars are not set."""
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_KEY", "").strip()
    if not url or not key:
        logger.warning("SUPABASE_URL / SUPABASE_KEY not set — Supabase disabled.")
        return None
    try:
        from supabase import create_client
        return create_client(url, key)
    except Exception as e:
        logger.error(f"Supabase client error: {e}")
        return None

def _ensure_officers_table():
    """Check the officers table exists at startup and log a clear error if not."""
    sb = get_supabase()
    if not sb:
        return
    try:
        sb.table("officers").select("officer_id").limit(1).execute()
        logger.info("✅ Supabase 'officers' table found.")
    except Exception as e:
        logger.error(
            "❌ Supabase 'officers' table missing. Create it in Supabase SQL editor:\n"
            "  CREATE TABLE officers (\n"
            "    officer_id   TEXT PRIMARY KEY,\n"
            "    lga          TEXT,\n"
            "    ward         TEXT,\n"
            "    polling_unit TEXT,\n"
            "    phone        TEXT\n"
            "  );\n"
            f"Error: {e}"
        )

_ensure_officers_table()

def _get_supabase_officer(officer_id: str, lg: str = "") -> dict | None:
    """Look up one officer in Supabase. Returns row dict or None."""
    sb = get_supabase()
    if not sb:
        return None
    try:
        query = sb.table("officers").select("*").eq("officer_id", officer_id)
        if lg:
            query = query.ilike("lga", lg)
        res = query.limit(1).execute()
        return res.data[0] if res.data else None
    except Exception as e:
        logger.error(f"Supabase officer lookup error: {e}")
        return None


@app.post("/api/upload-officers")
async def upload_officers(request: Request, file: UploadFile = File(...)):
    """
    Admin-only. Upload a CSV of officers → saved permanently to Supabase.
    Required columns: officer_id, lga, ward, polling_unit, phone
    Existing officers are updated (upsert). Safe to re-upload an expanded list.
    """
    _require_admin(request)

    sb = get_supabase()
    if not sb:
        raise HTTPException(
            status_code=503,
            detail="Supabase not configured. Set SUPABASE_URL and SUPABASE_KEY env vars on Render."
        )

    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")   # strips Excel BOM automatically
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    required = {"officer_id", "lga", "ward", "polling_unit", "phone"}
    if not reader.fieldnames or not required.issubset({c.strip().lower() for c in reader.fieldnames}):
        raise HTTPException(
            status_code=400,
            detail=f"CSV must contain columns: {', '.join(sorted(required))}"
        )

    rows, skipped = [], []
    for i, row in enumerate(reader, start=2):
        r     = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
        oid   = r.get("officer_id", "")
        phone = _clean_phone(r.get("phone", ""))
        if not oid:
            skipped.append(f"Row {i}: missing officer_id")
            continue
        rows.append({
            "officer_id":   oid,
            "lga":          r.get("lga", ""),
            "ward":         r.get("ward", ""),
            "polling_unit": r.get("polling_unit", ""),
            "phone":        phone,
        })

    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows found in CSV.")

    total_saved = 0
    for start in range(0, len(rows), 500):
        batch = rows[start:start + 500]
        try:
            sb.table("officers").upsert(batch, on_conflict="officer_id").execute()
            total_saved += len(batch)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Database error: {e}")

    logger.info(f"✅ {total_saved} officers upserted to Supabase.")
    return {
        "status":          "success",
        "saved":           total_saved,
        "skipped":         len(skipped),
        "skipped_details": skipped[:20],
        "message":         f"{total_saved} officers saved permanently to Supabase.",
    }


@app.get("/api/officers")
async def list_officers(request: Request, lga: str = "", search: str = "", page: int = 1, filter: str = "all", page_size: int = 50):
    """Admin-only: list officers stored in Supabase, paginated and filterable."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        q = sb.table("officers").select("*").order("officer_id")
        if lga:
            q = q.ilike("lga", f"%{lga}%")
        if search:
            q = q.ilike("officer_id", f"%{search}%")
        res = q.execute()
        rows = res.data or []
        if filter == "registered":
            rows = [r for r in rows if (r.get("phone") or "").strip()]
        elif filter == "unregistered":
            rows = [r for r in rows if not (r.get("phone") or "").strip()]
        total = len(rows)
        pages = max(1, (total + page_size - 1) // page_size)
        page = max(1, min(page, pages))
        start = (page - 1) * page_size
        page_rows = rows[start:start + page_size]
        return {"status": "success", "officers": page_rows, "total": total, "page": page, "pages": pages, "count": total}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/officers/stats")
async def officer_stats(request: Request):
    """Admin-only: registration counts for the dashboard stat cards (Supabase-backed)."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        res = sb.table("officers").select("phone").execute()
        rows = res.data or []
        total = len(rows)
        with_phone = sum(1 for r in rows if (r.get("phone") or "").strip())
        return {"total": total, "with_phone": with_phone, "without_phone": total - with_phone}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/officers/update")
async def update_officer(request: Request):
    """Admin-only: update a single officer's phone number in Supabase."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        body  = await request.json()
        oid   = str(body.get("officer_id", "")).strip()
        if not oid:
            raise HTTPException(status_code=400, detail="officer_id required.")
        phone = _clean_phone(str(body.get("phone", "")))
        sb.table("officers").update({"phone": phone}).eq("officer_id", oid).execute()
        return {"status": "success", "message": f"Officer {oid} updated."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/officers/delete")
async def delete_officer(request: Request):
    """Admin-only: remove a single officer by officer_id."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        body = await request.json()
        oid  = str(body.get("officer_id", "")).strip()
        if not oid:
            raise HTTPException(status_code=400, detail="officer_id required.")
        sb.table("officers").delete().eq("officer_id", oid).execute()
        return {"status": "success", "message": f"Officer {oid} deleted."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# END SUPABASE BLOCK
# =============================================================================

# --- API ENDPOINTS ---
@app.get("/api/validate_officer/{officer_id}")
def validate_officer(officer_id: str, request: Request, lg: str = ""):
    _check_rate_limit(request.client.host)
    officer_id = officer_id[:60]
    import re as _re
    officer_id = _re.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)
    lg = lg.strip()[:60]
    if not lg:
        return {"valid": False, "message": "Please select your LGA first."}

    # ── 1. Check Supabase officers table first ────────────────────────────────
    sb_officer = _get_supabase_officer(officer_id, lg)
    if sb_officer:
        parts = officer_id.split("-", 1)
        pu_data = {}
        if len(parts) == 2:
            ward_code, pu_code = parts[0].strip(), parts[1].strip()
            try:
                with get_db() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """SELECT ward, lg, location, pu_code, ward_code, state
                               FROM polling_units
                               WHERE ward_code = ? AND pu_code = ?
                               AND LOWER(state) = 'osun' AND LOWER(lg) = LOWER(?)
                               LIMIT 1""",
                            (ward_code, pu_code, lg)
                        )
                        row = cur.fetchone()
                        if row:
                            pu_data = {
                                "state":     row["state"] or "osun",
                                "ward":      row["ward"],
                                "lg":        row["lg"],
                                "location":  row["location"],
                                "pu_code":   row["pu_code"],
                                "ward_code": row["ward_code"],
                            }
            except Exception:
                pass
        if not pu_data:
            pu_data = {
                "state":     "osun",
                "ward":      sb_officer.get("ward", ""),
                "lg":        sb_officer.get("lga", lg),
                "location":  sb_officer.get("polling_unit", ""),
                "pu_code":   officer_id.split("-", 1)[1] if "-" in officer_id else officer_id,
                "ward_code": officer_id.split("-", 1)[0] if "-" in officer_id else "",
            }
        return {
            "valid":   True,
            "message": f"Access Granted: {pu_data.get('location', sb_officer.get('polling_unit', ''))}",
            **pu_data,
        }

    # ── 2. Fallback: check polling_units SQLite table ─────────────────────────
    try:
        parts = officer_id.split("-", 1)
        if len(parts) != 2:
            return {"valid": False, "message": "Invalid ID format. Expected: WARDCODE-PUCODE"}
        ward_code, pu_code = parts[0].strip(), parts[1].strip()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT ward, lg, location, pu_code, ward_code, state
                       FROM polling_units
                       WHERE ward_code = ? AND pu_code = ?
                       AND LOWER(state) = 'osun'
                       AND LOWER(lg) = LOWER(?)
                       LIMIT 1""",
                    (ward_code, pu_code, lg)
                )
                row = cur.fetchone()
                if row:
                    return {
                        "valid":     True,
                        "message":   f"Access Granted: {row['location']}",
                        "state":     row["state"] or "osun",
                        "ward":      row["ward"],
                        "lg":        row["lg"],
                        "location":  row["location"],
                        "pu_code":   row["pu_code"],
                        "ward_code": row["ward_code"],
                    }
                else:
                    return {"valid": False, "message": "Officer ID not found in selected LGA. Check your LGA and ID."}
    except Exception as e:
        return {"valid": False, "message": f"Validation error: {str(e)}"}


@app.get("/api/states")
def get_states():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT state FROM polling_units WHERE state = 'osun' ORDER BY state")
            rows = cur.fetchall()
            return [r["state"] for r in rows]

@app.get("/api/lgas/{state}")
def get_lgas(state: str):
    with get_db() as conn:
        with conn.cursor() as cur:
            # BUG FIX #6: Use the actual state param (lowercased) instead of hardcoded 'osun'
            cur.execute("SELECT DISTINCT lg FROM polling_units WHERE LOWER(state) = LOWER(?) ORDER BY lg", (state,))
            rows = cur.fetchall()
            return [r["lg"] for r in rows]

@app.get("/api/wards/{state}/{lg}")
def get_wards(state: str, lg: str):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT ward, ward_code FROM polling_units WHERE LOWER(state) = LOWER(?) AND lg = ? ORDER BY ward", (state, lg))
            rows = cur.fetchall()
            return [{"name": r["ward"], "code": r["ward_code"]} for r in rows]

@app.get("/api/pus/{state}/{lg}/{ward}")
def get_pus(state: str, lg: str, ward: str):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT location, pu_code FROM polling_units WHERE LOWER(state) = LOWER(?) AND lg = ? AND ward = ?", (state, lg, ward))
            rows = cur.fetchall()
            return [{"location": r["location"], "pu_code": r["pu_code"]} for r in rows]

@app.post("/submit")
async def submit(
    data: str = Form(...),
    ec8e_image: UploadFile = File(None)
):
    try:
        payload = json.loads(data)
        # Verify submission token — proves officer completed OTP auth
        submit_token = payload.get("submit_token", "")
        officer_id   = payload.get("officer_id", "")
        if not _verify_submit_token(submit_token, officer_id):
            return {"status": "error", "message": "Session expired or invalid. Please log in again."}
        votes_json = json.dumps(payload.get("votes", {}))
        ec8e_filename = None
        if ec8e_image and ec8e_image.filename:
            safe_pu = str(payload.get("pu_code", "unk")).replace("/", "_").replace(" ", "_")
            public_id = f"ec8e_forms/{safe_pu}_{uuid.uuid4().hex[:8]}"
            try:
                img_bytes = await _safe_read_image(ec8e_image)
                upload_result = cloudinary.uploader.upload(
                    img_bytes,
                    public_id=public_id,
                    resource_type="image",
                    overwrite=True
                )
                ec8e_filename = upload_result["secure_url"]
                logger.info(f"EC8E uploaded to Cloudinary: {ec8e_filename}")
            except Exception as cloud_err:
                logger.error(f"Cloudinary upload failed: {cloud_err} — saving locally")
                ext = os.path.splitext(ec8e_image.filename)[1].lower()
                local_name = f"{safe_pu}_{uuid.uuid4().hex[:8]}{ext}"
                with open(os.path.join(EC8E_PATH, local_name), "wb") as img_f:
                    shutil.copyfileobj(ec8e_image.file, img_f)
                ec8e_filename = f"/ec8e/{local_name}"
        with get_pg() as conn:
            with conn.cursor() as cur:
                cur.execute("""INSERT INTO field_submissions (
                    officer_id, state, lg, ward, ward_code, pu_code, location,
                    reg_voters, total_accredited, valid_votes, rejected_votes, total_cast,
                    lat, lon, timestamp, votes_json, ec8e_image
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
                    payload.get("officer_id"),
                    (payload.get("state") or "osun").lower(),  # always store lowercase
                    payload.get("lg"),
                    payload.get("ward"), payload.get("ward_code"), payload.get("pu_code"),
                    payload.get("location"), payload.get("reg_voters"), payload.get("total_accredited"),
                    payload.get("valid_votes"), payload.get("rejected_votes"), payload.get("total_cast"),
                    payload.get("lat"), payload.get("lon"),
                    datetime.now().isoformat(), votes_json, ec8e_filename
                ))
                conn.commit()
        alert_payload = {**payload, "timestamp": datetime.now().strftime("%d %b %Y %H:%M")}
        threading.Thread(target=send_whatsapp_alert, args=(alert_payload,)).start()
        return {"status": "success", "message": "Result Uploaded Successfully"}
    except psycopg2.IntegrityError:
        return {"status": "error", "message": "REJECTED: A submission for this Polling Unit already exists."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/ai_interpret")
async def ai_interpret(data: dict):
    STATS = {
        "lgas": 30,
        "wards": 332,
        "pus": 3763,
        "urban_hubs": ["OSOGBO", "OLORUNDA", "ILESA EAST", "IFE CENTRAL", "IWO"]
    }
    OSUN_PARTIES_AI = ["ACCORD", "AA", "AAC", "ADC", "ADP", "APGA", "APC", "APM", "APP", "BP", "NNPP", "PRP", "YPP", "ZLP"]
    party_votes = {p: data.get(p, 0) for p in OSUN_PARTIES_AI}
    acc = party_votes["ACCORD"]
    rivals = {p: v for p, v in party_votes.items() if p != "ACCORD"}

    ta = data.get('total_accredited', 0)
    rv = data.get('reg_voters', 0)
    current_lg = str(data.get('lg', "")).upper()

    total_votes = sum(party_votes.values())
    if total_votes == 0:
        return {"analysis": "SYSTEM READY: Awaiting live feed from 3,763 Polling Units across Osun State."}

    share = (acc / total_votes) * 100
    top_rival = max(rivals, key=rivals.get)
    margin = acc - rivals[top_rival]
    turnout = (ta / rv * 100) if rv > 0 else 0

    is_urban = current_lg in STATS["urban_hubs"]

    if share > 55:
        trend = "LANDSLIDE"
    elif share > 40:
        trend = "STRONG LEAD"
    else:
        trend = "BATTLEGROUND"

    location_tag = " [URBAN HUB]" if is_urban else " [RURAL SECTOR]"

    analysis = (
        f"OSUN STATISTICAL AUDIT ({current_lg}{location_tag}): "
        f"Accord is in a {trend} position with {share:.1f}% of the current tally. "
        f"Lead Margin over {top_rival}: **{margin:+,}** votes. "
    )

    if turnout > 0:
        analysis += f"Voter Productivity is at **{turnout:.1f}%**. "
        if turnout > 65:
            analysis += "⚠️ ALERT: Unusually high turnout detected; verify PU logs. "

    if is_urban and share < 45:
        analysis += "STRATEGY: Increase urban mobilization; Osogbo/Iwo volume is critical."
    elif not is_urban and share > 50:
        analysis += "STRATEGY: Rural stronghold confirmed. Protect the lead during collation."

    return {
        "analysis": analysis,
        "is_alert": turnout > 70 or margin < 100,
        "stats": {
            "turnout_gap": 100 - turnout,
            "osun_progress": f"Active in {STATS['lgas']} LGAs"
        }
    }

@app.get("/api/dashboard_filters")
def get_dash_filters(request: Request):
    _require_dashboard(request)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""SELECT DISTINCT
                               LOWER(state) as state,
                               lg,
                               ward
                           FROM polling_units
                           WHERE LOWER(state) = 'osun'
                           ORDER BY lg, ward""")
            return [dict(r) for r in cur.fetchall()]

@app.get("/export/csv")
async def export_csv(request: Request):
    _require_dashboard(request)
    try:
        import openpyxl
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=500, content={"error": "openpyxl not installed. Add openpyxl to requirements.txt and redeploy."})
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response as _Resp
    import io as _io
    from datetime import datetime as _dt

    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM field_submissions ORDER BY timestamp DESC")
            rows = cur.fetchall()

    PARTIES = ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Election Results"

    GREEN="00008751"; GOLD="00FFC107"; DARK="00121212"; WHITE="00FFFFFF"
    LIGHT_GREY="00F5F5F5"; ACCORD_LIGHT="00E8F5E9"
    hdr_fill=PatternFill("solid",fgColor=GREEN)
    hdr_font=Font(bold=True,color=WHITE,size=10,name="Calibri")
    party_fill=PatternFill("solid",fgColor=GOLD)
    party_font=Font(bold=True,color=DARK,size=10,name="Calibri")
    accord_fill=PatternFill("solid",fgColor=ACCORD_LIGHT)
    alt_fill=PatternFill("solid",fgColor=LIGHT_GREY)
    center=Alignment(horizontal="center",vertical="center",wrap_text=True)
    left_align=Alignment(horizontal="left",vertical="center")
    ts=Side(style="thin",color="CCCCCC")
    tb=Border(left=ts,right=ts,top=ts,bottom=ts)

    TOTAL_COLS=17+len(PARTIES)+1
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=TOTAL_COLS)
    tc=ws.cell(row=1,column=1,value="ACCORD PARTY - OSUN 2026 GOVERNORSHIP ELECTION RESULTS")
    tc.fill=PatternFill("solid",fgColor=GREEN); tc.font=Font(bold=True,color=GOLD,size=14,name="Calibri")
    tc.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30

    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=TOTAL_COLS)
    sc=ws.cell(row=2,column=1,value=f"Exported: {_dt.now().strftime('%d %B %Y  %H:%M')}  |  Total PUs: {len(rows)}")
    sc.fill=PatternFill("solid",fgColor=GOLD); sc.font=Font(bold=True,color=DARK,size=10,name="Calibri")
    sc.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=18

    headers=["#","Officer ID","State","LGA","Ward","Ward Code","PU Code","Polling Unit",
             "Reg. Voters","Accredited","Valid Votes","Rejected","Total Cast",
             "Latitude","Longitude","Timestamp","EC8E",*PARTIES,"ACCORD TOTAL"]
    for col,h in enumerate(headers,1):
        cell=ws.cell(row=3,column=col,value=h)
        ip=h in PARTIES or h=="ACCORD TOTAL"
        cell.fill=party_fill if ip else hdr_fill
        cell.font=party_font if ip else hdr_font
        cell.alignment=center; cell.border=tb
    ws.row_dimensions[3].height=22

    for ri,r in enumerate(rows,4):
        v=json.loads(r["votes_json"]) if isinstance(r["votes_json"],str) else (r["votes_json"] or {})
        av=v.get("ACCORD",0); is_alt=(ri%2==0)
        rf=accord_fill if av>0 else (alt_fill if is_alt else None)
        row_data=[r["id"],r["officer_id"],(r["state"] or "").upper(),(r["lg"] or "").upper(),
                  (r["ward"] or "").upper(),r["ward_code"],r["pu_code"],r["location"],
                  r["reg_voters"],r["total_accredited"],r["valid_votes"],r["rejected_votes"],
                  r["total_cast"],r["lat"],r["lon"],r["timestamp"],
                  "YES" if r.get("ec8e_image") else "NO",
                  *[v.get(p,0) for p in PARTIES],av]
        for col,val in enumerate(row_data,1):
            cell=ws.cell(row=ri,column=col,value=val)
            cell.border=tb; cell.alignment=left_align if col==8 else center
            if rf: cell.fill=rf
            if col==len(headers): cell.font=Font(bold=True,color="00008751",size=10,name="Calibri")
        ws.row_dimensions[ri].height=16

    col_widths=[4,12,8,14,16,11,10,28,10,11,10,9,10,10,10,22,6]+[8]*len(PARTIES)+[12]
    for i,w in enumerate(col_widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A4"
    ws.auto_filter.ref=f"A3:{get_column_letter(TOTAL_COLS)}3"

    ws2=wb.create_sheet("Party Summary")
    ws2.merge_cells("A1:C1")
    s2t=ws2.cell(row=1,column=1,value="PARTY VOTE SUMMARY")
    s2t.fill=PatternFill("solid",fgColor=GREEN); s2t.font=Font(bold=True,color=GOLD,size=12,name="Calibri")
    s2t.alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=24
    for col,h in enumerate(["Party","Total Votes","% Share"],1):
        c=ws2.cell(row=2,column=col,value=h); c.fill=hdr_fill; c.font=hdr_font; c.alignment=center; c.border=tb
    pt={p:sum((json.loads(r["votes_json"]) if isinstance(r["votes_json"],str) else (r["votes_json"] or {})).get(p,0) for r in rows) for p in PARTIES}
    gt=sum(pt.values()) or 1
    for ri2,(party,total) in enumerate(sorted(pt.items(),key=lambda x:-x[1]),3):
        pct=round((total/gt)*100,2); ia=party=="ACCORD"
        for col,val in enumerate([party,total,f"{pct}%"],1):
            c=ws2.cell(row=ri2,column=col,value=val); c.border=tb; c.alignment=center
            if ia: c.fill=PatternFill("solid",fgColor=ACCORD_LIGHT); c.font=Font(bold=True,color="00008751",size=10,name="Calibri")
    ws2.column_dimensions["A"].width=12; ws2.column_dimensions["B"].width=14; ws2.column_dimensions["C"].width=10

    buf=_io.BytesIO(); wb.save(buf); buf.seek(0)
    return _Resp(content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=Accord_Osun2026_Results.xlsx"})

# ── Dashboard authentication endpoint ────────────────────────────────────────
@app.post("/api/verify-dashboard")
async def verify_dashboard(request: Request, response: Response):
    try:
        body = await request.json()
        key = str(body.get("key", ""))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request")
    given_hash = hashlib.sha256(key.encode()).hexdigest()
    if not secrets.compare_digest(given_hash, _DASHBOARD_KEY_HASH):
        raise HTTPException(status_code=401, detail="Invalid dashboard key")
    token = _make_session_token()
    # secure flag: True only when request came in over HTTPS
    # (Render/cloud proxies forward as HTTP internally — check X-Forwarded-Proto)
    is_https = request.headers.get("x-forwarded-proto", "http") == "https"
    response.set_cookie(
        key="ds_session",
        value=token,
        httponly=True,
        samesite="lax",
        secure=is_https,
        max_age=_SESSION_TTL,
        path="/"
    )
    return {"status": "ok"}

@app.post("/api/logout-dashboard")
async def logout_dashboard(request: Request, response: Response):
    token = request.cookies.get("ds_session")
    if token:
        _SESSION_TOKENS.pop(token, None)
    response.delete_cookie("ds_session", path="/")
    return {"status": "ok"}
# ─────────────────────────────────────────────────────────────────────────────

# ── OTP endpoints ────────────────────────────────────────────────────────────
@app.post("/api/request-otp")
async def request_otp(request: Request):
    """
    Step 1 of officer auth.
    Body: { "officer_id": "WARDCODE-PUCODE", "lg": "Osogbo" }
    Looks up phone from Supabase first, then falls back to polling_units SQLite.
    Returns: { "status": "sent", "phone_hint": "+234***4567" }
             or { "status": "no_phone", "message": "..." }
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    lg         = str(body.get("lg", "")).strip()[:60]
    import re as _re_otp
    officer_id = _re_otp.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)

    if not lg:
        raise HTTPException(status_code=400, detail="LGA is required.")

    parts = officer_id.split("-", 1)
    if len(parts) != 2:
        raise HTTPException(status_code=400, detail="Invalid officer ID format")
    ward_code, pu_code = parts[0].strip(), parts[1].strip()

    # Check lockout — key includes LGA to prevent cross-LGA lockout bleed
    otp_key      = f"{officer_id}|{lg.lower()}"
    entry        = _OTP_STORE.get(otp_key, {})
    locked_until = entry.get("locked_until", 0)
    if time.time() < locked_until:
        remaining = int(locked_until - time.time())
        raise HTTPException(
            status_code=429,
            detail=f"Too many failed attempts. Try again in {remaining // 60}m {remaining % 60}s."
        )

    # ── Resolve phone + PU data: Supabase first, SQLite fallback ─────────────
    phone   = None
    pu_data = {}

    sb_officer = _get_supabase_officer(officer_id, lg)
    if sb_officer:
        phone = _clean_phone(sb_officer.get("phone", "") or "")
        # Enrich pu_data from SQLite polling_units if available
        try:
            with get_db() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT ward, lg, location, pu_code, ward_code, state
                           FROM polling_units
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun' AND LOWER(lg) = LOWER(?)
                           LIMIT 1""",
                        (ward_code, pu_code, lg)
                    )
                    row = cur.fetchone()
                    if row:
                        pu_data = {
                            "state":     row["state"] or "osun",
                            "ward":      row["ward"],
                            "lg":        row["lg"],
                            "location":  row["location"],
                            "pu_code":   row["pu_code"],
                            "ward_code": row["ward_code"],
                        }
        except Exception:
            pass
        if not pu_data:
            pu_data = {
                "state":     "osun",
                "ward":      sb_officer.get("ward", ""),
                "lg":        sb_officer.get("lga", lg),
                "location":  sb_officer.get("polling_unit", ""),
                "pu_code":   pu_code,
                "ward_code": ward_code,
            }
    else:
        # Fallback: check polling_units SQLite table
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT ward, lg, location, pu_code, ward_code, state, officer_phone
                       FROM polling_units
                       WHERE ward_code = ? AND pu_code = ?
                       AND LOWER(state) = 'osun'
                       AND LOWER(lg) = LOWER(?)
                       LIMIT 1""",
                    (ward_code, pu_code, lg)
                )
                row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Officer ID not found. Access Denied.")

        phone = _clean_phone(row["officer_phone"] or "") if row["officer_phone"] else None
        pu_data = {
            "state":     row["state"] or "osun",
            "ward":      row["ward"],
            "lg":        row["lg"],
            "location":  row["location"],
            "pu_code":   row["pu_code"],
            "ward_code": row["ward_code"],
        }

    if not phone or len(phone) < 10:
        return {
            "status":  "no_phone",
            "message": "No phone number registered for this officer ID. Contact your supervisor."
        }

    # Generate OTP and store
    otp = _generate_otp()
    _OTP_STORE[otp_key] = {
        "otp":          otp,
        "expiry":       time.time() + _OTP_TTL,
        "phone_hint":   _mask_phone(phone),
        "phone":        phone,
        "used":         False,
        "attempts":     0,
        "locked_until": 0,
        "pu_data":      pu_data,
    }

    # Send via Twilio WhatsApp
    account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
    auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
    from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")

    if not account_sid or not auth_token:
        dev_mode = os.environ.get("OTP_DEV_MODE", "").lower() in ("1", "true", "yes")
        if dev_mode:
            logger.warning("Twilio not configured — OTP not sent (dev mode)")
            logger.info(f"[DEV OTP] {officer_id}: {otp}")
        else:
            logger.error("Twilio credentials missing — cannot send OTP")
            raise HTTPException(
                status_code=503,
                detail="OTP service is not configured. Contact your administrator."
            )
    else:
        try:
            from twilio.rest import Client as _TC
            client = _TC(account_sid, auth_token)
            msg = (
                f"🗳 *ACCORD FIELD COLLATION*\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"Your One-Time Password:\n\n"
                f"*{otp}*\n\n"
                f"Valid for *5 minutes*. Do NOT share this code.\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"_Powered by Popson Geospatial Services_"
            )
            client.messages.create(
                from_=f"whatsapp:{from_number}",
                to=f"whatsapp:{phone}",
                body=msg
            )
            logger.info(f"✅ OTP sent to {_mask_phone(phone)} for officer {officer_id}")
        except Exception as e:
            twilio_msg = str(e)
            logger.error(f"Twilio OTP send failed for {officer_id}: {twilio_msg}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Failed to send OTP: {twilio_msg}"
            )

    return {
        "status":     "sent",
        "phone_hint": _mask_phone(phone)
    }


@app.post("/api/verify-otp")
async def verify_otp(request: Request):
    """
    Step 2 of officer auth.
    Body: { "officer_id": "WARDCODE-PUCODE", "otp": "123456" }
    Returns: { "status": "ok", "token": "...", "pu_data": {...} }
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    lg         = str(body.get("lg", "")).strip()[:60]
    import re as _re3
    officer_id = _re3.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)
    given_otp  = str(body.get("otp", "")).strip()[:6]

    otp_key = f"{officer_id}|{lg.lower()}"
    entry = _OTP_STORE.get(otp_key)

    # No OTP requested
    if not entry:
        raise HTTPException(status_code=400, detail="No OTP requested for this officer. Start again.")

    # Lockout check
    if time.time() < entry.get("locked_until", 0):
        remaining = int(entry["locked_until"] - time.time())
        raise HTTPException(
            status_code=429,
            detail=f"Account locked. Try again in {remaining // 60}m {remaining % 60}s."
        )

    # Already used
    if entry.get("used"):
        raise HTTPException(status_code=400, detail="OTP already used. Request a new one.")

    # Expired
    if time.time() > entry["expiry"]:
        _OTP_STORE.pop(otp_key, None)
        raise HTTPException(status_code=400, detail="OTP expired. Request a new one.")

    # Wrong OTP — increment attempts
    if not secrets.compare_digest(given_otp, entry["otp"]):
        entry["attempts"] = entry.get("attempts", 0) + 1
        remaining_tries = _OTP_MAX_TRIES - entry["attempts"]
        if entry["attempts"] >= _OTP_MAX_TRIES:
            entry["locked_until"] = time.time() + _OTP_LOCKOUT
            entry["used"] = True  # invalidate
            raise HTTPException(
                status_code=429,
                detail=f"Too many wrong attempts. Account locked for {_OTP_LOCKOUT // 60} minutes."
            )
        raise HTTPException(
            status_code=401,
            detail=f"Incorrect OTP. {remaining_tries} attempt{'s' if remaining_tries != 1 else ''} remaining."
        )

    # ✅ Correct OTP — mark used, issue submission token
    entry["used"] = True
    token = _make_submit_token(officer_id)
    pu_data = entry["pu_data"]

    return {
        "status": "ok",
        "token": token,
        "officer_id": officer_id,
        "state":     pu_data["state"],
        "ward":      pu_data["ward"],
        "lg":        pu_data["lg"],
        "location":  pu_data["location"],
        "pu_code":   pu_data["pu_code"],
        "ward_code": pu_data["ward_code"],
    }
# ─────────────────────────────────────────────────────────────────────────────

# ── Admin: officer stats ─────────────────────────────────────────────────────
@app.get("/api/admin/officer-stats")
async def officer_stats(request: Request):
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing Authorization")
    given_hash = hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest()
    if not secrets.compare_digest(given_hash, _DASHBOARD_KEY_HASH):
        raise HTTPException(status_code=403, detail="Invalid key")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as total FROM polling_units WHERE LOWER(state)='osun'")
            total = cur.fetchone()["total"]
            cur.execute("SELECT COUNT(*) as c FROM polling_units WHERE LOWER(state)='osun' AND officer_phone IS NOT NULL AND officer_phone != ''")
            with_phone = cur.fetchone()["c"]
    return {"total": total, "with_phone": with_phone, "without_phone": total - with_phone}
# ─────────────────────────────────────────────────────────────────────────────

# ── Admin: set officer phone numbers ─────────────────────────────────────────
@app.post("/api/admin/set-officer-phone")
async def set_officer_phone(request: Request):
    """
    Protected by DASHBOARD_KEY (same secret as dashboard).
    Accepts JSON:
      Single:  {"officer_id": "WARDCODE-PUCODE", "phone": "+2348012345678"}
      Bulk:    {"officers": [{"officer_id": "...", "phone": "..."}, ...]}
    """
    # Verify admin key from Authorization header
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    given = auth.split(" ", 1)[1].strip()
    given_hash = hashlib.sha256(given.encode()).hexdigest()
    if not secrets.compare_digest(given_hash, _DASHBOARD_KEY_HASH):
        raise HTTPException(status_code=403, detail="Invalid admin key")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    records = body.get("officers") or ([body] if "officer_id" in body else [])
    if not records:
        raise HTTPException(status_code=400, detail="Provide officer_id+phone or officers array")

    updated, skipped, no_match = 0, 0, []
    with get_db() as conn:
        with conn.cursor() as cur:
            for rec in records:
                oid   = str(rec.get("officer_id", "")).strip()[:60]
                phone = _clean_phone(rec.get("phone", ""))
                lga   = str(rec.get("lga", "")).strip()[:80]

                # Skip rows with no phone (blank = not yet assigned, not an error)
                if not phone or len(phone) < 10:
                    skipped += 1
                    continue
                if not oid:
                    skipped += 1
                    continue
                parts = oid.split("-", 1)
                if len(parts) != 2:
                    skipped += 1
                    continue
                ward_code, pu_code = parts[0].strip(), parts[1].strip()

                if lga:
                    # Preferred: match on ward_code + pu_code + LGA (unique)
                    cur.execute(
                        """UPDATE polling_units SET officer_phone = ?
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun'
                           AND LOWER(lg) = LOWER(?)""",
                        (phone, ward_code, pu_code, lga)
                    )
                else:
                    # Fallback: no LGA provided — only safe if combo is unique
                    cur.execute(
                        """UPDATE polling_units SET officer_phone = ?
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun'""",
                        (phone, ward_code, pu_code)
                    )

                if conn._conn.total_changes > 0:
                    updated += 1
                else:
                    no_match.append(f"{oid} ({lga})")
                    skipped += 1
        conn.commit()

    result = {"status": "ok", "updated": updated, "skipped": skipped}
    if no_match:
        result["not_found"] = no_match[:20]  # return first 20 unmatched for debugging
    return result
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/submissions")
async def get_dashboard_data(request: Request):
    _require_dashboard(request)
    require_review = os.environ.get("REQUIRE_REVIEW", "0").lower() in ("1", "true", "yes")
    where = "WHERE LOWER(state) = 'osun'"
    if require_review:
        where += " AND reviewed = 1"
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT * FROM field_submissions {where} ORDER BY timestamp DESC")
            rows = cur.fetchall()
            data = []
            for r in rows:
                # Use edited votes if admin has overridden them
                votes_src = r.get('edited_votes_json') or r.get('votes_json')
                v = json.loads(votes_src) if isinstance(votes_src, str) else (votes_src or {})
                raw = r.get('ec8e_image')
                if raw:
                    ec8e_url = raw if raw.startswith('http') else f"/ec8e/{raw}"
                else:
                    ec8e_url = None
                entry = {
                    "pu_name": r['location'], "state": r['state'], "lga": r['lg'], "ward": r['ward'],
                    "latitude": r['lat'], "longitude": r['lon'],
                    "ec8e_image": ec8e_url,
                    "reg_voters": r.get('reg_voters') or 0,
                    "total_accredited": r.get('total_accredited') or 0,
                    "total_cast": r.get('total_cast') or 0,
                    "officer_id": r.get('officer_id') or '',
                    "timestamp": str(r.get('timestamp') or ''),
                    "pu_code": r.get('pu_code') or '',
                    "reviewed": bool(r.get('reviewed')),
                }
                for p in ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"]:
                    entry[f"votes_party_{p}"] = v.get(p, 0)
                data.append(entry)
            return data


@app.get("/ec8e/{filename}")
async def serve_ec8e(filename: str):
    """Legacy fallback for locally-stored EC8E images. New uploads use Cloudinary CDN."""
    import mimetypes
    filepath = os.path.join(EC8E_PATH, filename)
    if not os.path.exists(filepath):
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Image not found")
    mime = mimetypes.guess_type(filepath)[0] or "image/jpeg"
    with open(filepath, "rb") as f:
        content = f.read()
    from fastapi.responses import Response
    return Response(content=content, media_type=mime, headers={
        "Cache-Control": "public, max-age=86400",
        "Access-Control-Allow-Origin": "*"
    })



RESULTS_PORTAL_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ACCORD — Results Review Portal</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{--green:#008751;--green-light:#00b368;--gold:#ffc107;--dark:#020c06;--panel:rgba(255,255,255,0.04);--border:rgba(255,255,255,0.1)}
body{font-family:'Inter',sans-serif;background:var(--dark);color:#fff;min-height:100vh}
body::before{content:'';position:fixed;inset:0;z-index:0;background:radial-gradient(ellipse 120% 80% at 50% -10%,rgba(0,135,81,0.15) 0%,transparent 60%),linear-gradient(160deg,#020c06 0%,#041508 40%,#020c06 100%)}
.page{position:relative;z-index:1;max-width:1100px;margin:0 auto;padding:28px 20px 60px}
#loginGate{min-height:100vh;display:flex;align-items:center;justify-content:center;position:relative;z-index:1}
.login-card{background:var(--panel);border:1px solid rgba(0,135,81,0.25);border-radius:20px;padding:48px 40px;width:100%;max-width:400px;text-align:center}
.lock-icon{width:60px;height:60px;border-radius:50%;background:rgba(0,135,81,0.15);border:2px solid rgba(0,135,81,0.3);display:flex;align-items:center;justify-content:center;font-size:1.6rem;margin:0 auto 20px}
h1{font-size:1.5rem;font-weight:800;margin-bottom:6px}
.sub{color:rgba(255,255,255,0.4);font-size:0.82rem;margin-bottom:28px}
.form-control{width:100%;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:10px;color:#fff;padding:12px 16px;font-size:0.9rem;outline:none;margin-bottom:12px}
.form-control:focus{border-color:var(--green)}
.btn-primary{width:100%;background:linear-gradient(135deg,var(--green),var(--green-light));border:none;border-radius:10px;color:#fff;font-weight:700;font-size:0.9rem;padding:13px;cursor:pointer;margin-top:4px}
.btn-primary:hover{opacity:0.9}
.err-msg{color:#ff6b6b;font-size:0.78rem;margin-top:8px;display:none}
/* Portal layout */
#portal{display:none}
.portal-header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:24px;flex-wrap:wrap;gap:16px}
.badge-tag{font-size:0.62rem;font-weight:700;letter-spacing:0.15em;color:var(--gold);background:rgba(255,193,7,0.1);border:1px solid rgba(255,193,7,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:8px;text-transform:uppercase}
.portal-header h1{font-size:1.6rem;font-weight:900;margin-bottom:4px}
.portal-header p{color:rgba(255,255,255,0.45);font-size:0.8rem}
.btn-secondary{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:rgba(255,255,255,0.7);font-size:0.78rem;font-weight:600;padding:8px 16px;cursor:pointer}
.btn-secondary:hover{background:rgba(255,255,255,0.1)}
/* Filters */
.filters{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:18px;align-items:center}
.filter-select{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:#fff;padding:7px 12px;font-size:0.78rem;outline:none;cursor:pointer}
.filter-select option{background:#041508}
.search-input{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:#fff;padding:7px 14px;font-size:0.78rem;outline:none;width:220px}
.search-input:focus{border-color:var(--green)}
/* Stats pills */
.stats-row{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:22px}
.stat-pill{background:var(--panel);border:1px solid var(--border);border-radius:12px;padding:12px 18px;display:flex;flex-direction:column;align-items:center;min-width:110px}
.stat-pill .val{font-size:1.4rem;font-weight:800;line-height:1}
.stat-pill .lbl{font-size:0.65rem;color:rgba(255,255,255,0.4);margin-top:4px;text-transform:uppercase;letter-spacing:0.06em}
.val-green{color:#00cc66}.val-gold{color:#ffc107}.val-red{color:#ff6b6b}.val-blue{color:#4fc3f7}
/* Table */
.tbl-wrap{overflow-x:auto;border-radius:12px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;font-size:0.78rem}
thead th{background:rgba(255,255,255,0.04);padding:10px 10px;text-align:left;font-size:0.65rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;white-space:nowrap;border-bottom:1px solid rgba(255,255,255,0.08)}
tbody tr{border-bottom:1px solid rgba(255,255,255,0.05);transition:background 0.1s}
tbody tr:hover{background:rgba(255,255,255,0.03)}
td{padding:9px 10px;vertical-align:middle}
.badge{display:inline-block;padding:2px 8px;border-radius:20px;font-size:0.62rem;font-weight:700;letter-spacing:0.05em}
.badge-ok{background:rgba(0,204,102,0.15);color:#00cc66;border:1px solid rgba(0,204,102,0.3)}
.badge-pending{background:rgba(255,193,7,0.12);color:#ffc107;border:1px solid rgba(255,193,7,0.25)}
.badge-edited{background:rgba(79,195,247,0.12);color:#4fc3f7;border:1px solid rgba(79,195,247,0.25)}
/* Action buttons */
.btn-a{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:6px;color:#fff;cursor:pointer;font-size:0.68rem;padding:3px 8px;margin-right:3px;white-space:nowrap}
.btn-a:hover{background:rgba(255,255,255,0.12)}
.btn-approve{border-color:rgba(0,204,102,0.4);color:#00cc66}
.btn-unapprove{border-color:rgba(255,193,7,0.4);color:#ffc107}
.btn-edit-r{border-color:rgba(79,195,247,0.4);color:#4fc3f7}
/* Edit modal */
.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,0.75);z-index:1000;display:flex;align-items:center;justify-content:center;padding:20px}
.modal{background:#041508;border:1px solid rgba(0,135,81,0.3);border-radius:16px;width:100%;max-width:560px;padding:28px;max-height:90vh;overflow-y:auto}
.modal h2{font-size:1.1rem;font-weight:800;margin-bottom:4px}
.modal .sub{margin-bottom:20px}
.field-row{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}
.field-group label{font-size:0.68rem;color:rgba(255,255,255,0.4);display:block;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.06em}
.field-group input{width:100%;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.15);border-radius:8px;color:#fff;padding:8px 10px;font-size:0.82rem;outline:none}
.field-group input:focus{border-color:var(--green)}
.parties-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(100px,1fr));gap:8px;margin-bottom:16px}
.modal-actions{display:flex;gap:10px;justify-content:flex-end;margin-top:16px}
.btn-save-modal{background:linear-gradient(135deg,var(--green),var(--green-light));border:none;border-radius:8px;color:#fff;font-weight:700;font-size:0.82rem;padding:10px 22px;cursor:pointer}
.btn-cancel-modal{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.15);border-radius:8px;color:rgba(255,255,255,0.7);font-size:0.82rem;padding:10px 18px;cursor:pointer}
.note-input{width:100%;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.15);border-radius:8px;color:#fff;padding:8px 10px;font-size:0.8rem;resize:vertical;min-height:60px;margin-bottom:12px;outline:none}
/* Pagination */
.pagination{display:flex;gap:8px;align-items:center;justify-content:flex-end;margin-top:14px;font-size:0.78rem;color:rgba(255,255,255,0.4)}
.page-btn{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:6px;color:#fff;cursor:pointer;padding:4px 10px;font-size:0.75rem}
.page-btn:disabled{opacity:0.3;cursor:default}
/* Pending badge on dashboard link */
.pending-badge{display:inline-block;background:#ff6b6b;color:#fff;border-radius:20px;font-size:0.6rem;font-weight:800;padding:1px 6px;margin-left:6px;vertical-align:middle}
</style>
</head>
<body>

<!-- Login gate -->
<div id="loginGate">
  <div class="login-card">
    <div class="lock-icon">🔐</div>
    <h1>Results Portal</h1>
    <p class="sub">ACCORD Field Collation — Review & Approve</p>
    <input type="password" id="rKey" class="form-control" placeholder="Admin key" onkeydown="if(event.key==='Enter')rLogin()">
    <button class="btn-primary" id="rLoginBtn" onclick="rLogin()">Access Results Portal →</button>
    <div class="err-msg" id="rLoginErr">Incorrect key. Try again.</div>
  </div>
</div>

<!-- Portal -->
<div id="portal" class="page">
  <div class="portal-header">
    <div>
      <div class="badge-tag">Admin — Results Review Portal</div>
      <h1>📊 Submitted Results</h1>
      <p>Review, edit and approve PU submissions before they appear on the dashboard.</p>
    </div>
    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
      <a href="/admin" target="_blank" class="btn-secondary">⚙️ Officer Admin</a>
      <button class="btn-secondary" onclick="rLogout()">🔒 Lock</button>
    </div>
  </div>

  <!-- Stats -->
  <div class="stats-row">
    <div class="stat-pill"><span class="val val-gold" id="rStatTotal">—</span><span class="lbl">Total</span></div>
    <div class="stat-pill"><span class="val val-green" id="rStatApproved">—</span><span class="lbl">Approved</span></div>
    <div class="stat-pill"><span class="val val-red" id="rStatPending">—</span><span class="lbl">Pending</span></div>
    <div class="stat-pill"><span class="val val-blue" id="rStatEdited">—</span><span class="lbl">Edited</span></div>
  </div>

  <!-- Filters -->
  <div class="filters">
    <select class="filter-select" id="rFilterLga" onchange="loadResults(1)">
      <option value="">All LGAs</option>
    </select>
    <select class="filter-select" id="rFilterStatus" onchange="loadResults(1)">
      <option value="">All Status</option>
      <option value="pending">Pending Review</option>
      <option value="reviewed">Approved</option>
    </select>
    <input type="text" class="search-input" id="rSearch" placeholder="Search PU, ward, officer..." oninput="filterTable(this.value)">
    <button class="btn-secondary" onclick="approveAllVisible()" style="margin-left:auto;">✅ Approve All Visible</button>
  </div>

  <!-- Table -->
  <div class="tbl-wrap">
    <table>
      <thead>
        <tr>
          <th>#</th>
          <th>PU Code</th>
          <th>Location</th>
          <th>LGA</th>
          <th>Ward</th>
          <th>Officer</th>
          <th>ACCORD</th>
          <th>Total Cast</th>
          <th>Accredited</th>
          <th>Timestamp</th>
          <th>Status</th>
          <th>Actions</th>
        </tr>
      </thead>
      <tbody id="rTableBody">
        <tr><td colspan="12" style="text-align:center;padding:30px;color:rgba(255,255,255,0.3);">Login to view results.</td></tr>
      </tbody>
    </table>
  </div>
  <div class="pagination">
    <span id="rPageInfo"></span>
    <button class="page-btn" id="rPrevBtn" onclick="loadResults(_rPage-1)" disabled>◀ Prev</button>
    <button class="page-btn" id="rNextBtn" onclick="loadResults(_rPage+1)" disabled>Next ▶</button>
  </div>
</div>

<!-- Edit Modal -->
<div class="modal-overlay" id="editModal" style="display:none;" onclick="if(event.target===this)closeModal()">
  <div class="modal">
    <h2>✏️ Edit Submission</h2>
    <p class="sub" id="modalSub"></p>
    <input type="hidden" id="modalId">
    <div class="field-row">
      <div class="field-group"><label>Total Accredited</label><input type="number" id="mAccredited" min="0"></div>
      <div class="field-group"><label>Total Cast</label><input type="number" id="mCast" min="0"></div>
      <div class="field-group"><label>Valid Votes</label><input type="number" id="mValid" min="0"></div>
      <div class="field-group"><label>Rejected Votes</label><input type="number" id="mRejected" min="0"></div>
    </div>
    <div style="font-size:0.68rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;margin-bottom:8px;">Party Votes</div>
    <div class="parties-grid" id="modalParties"></div>
    <div style="font-size:0.68rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;margin-bottom:6px;">Edit Note</div>
    <textarea class="note-input" id="mNote" placeholder="Reason for edit (optional)..."></textarea>
    <div class="modal-actions">
      <button class="btn-cancel-modal" onclick="closeModal()">Cancel</button>
      <button class="btn-save-modal" onclick="saveEdit()">💾 Save Changes</button>
    </div>
  </div>
</div>

<script>
const PARTIES = ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"];
let _rKey = '', _rPage = 1, _rData = [], _rTotal = 0, _rPages = 1;

async function rLogin() {
    const key = document.getElementById('rKey').value.trim();
    if (!key) return;
    const btn = document.getElementById('rLoginBtn');
    const err = document.getElementById('rLoginErr');
    btn.disabled = true; btn.textContent = 'Verifying...';
    err.style.display = 'none';
    try {
        const res = await fetch('/api/admin/officer-stats', {
            headers: { 'Authorization': 'Bearer ' + key }
        });
        if (!res.ok) {
            err.style.display = 'block';
            btn.disabled = false; btn.textContent = 'Access Results Portal →';
            return;
        }
        _rKey = key;
        document.getElementById('loginGate').style.display = 'none';
        document.getElementById('portal').style.display = 'block';
        loadResults(1);
        loadLgas();
        loadSummaryStats();
    } catch(e) {
        err.textContent = 'Server error.';
        err.style.display = 'block';
        btn.disabled = false; btn.textContent = 'Access Results Portal →';
    }
}

function rLogout() {
    _rKey = '';
    document.getElementById('portal').style.display = 'none';
    document.getElementById('loginGate').style.display = 'flex';
    document.getElementById('rKey').value = '';
}

async function loadSummaryStats() {
    try {
        const res = await fetch('/api/admin/pending-review-count', {
            headers: { 'Authorization': 'Bearer ' + _rKey },
            credentials: 'include'
        });
        if (!res.ok) return;
        const d = await res.json();
        document.getElementById('rStatTotal').textContent    = d.total.toLocaleString();
        document.getElementById('rStatPending').textContent  = d.pending.toLocaleString();
        document.getElementById('rStatApproved').textContent = (d.total - d.pending).toLocaleString();
    } catch(e) {}
}

async function loadLgas() {
    try {
        // /api/lgas/osun requires no auth — returns plain string array
        const res = await fetch('/api/lgas/osun');
        if (!res.ok) return;
        const d = await res.json();
        const sel = document.getElementById('rFilterLga');
        const items = Array.isArray(d) ? d : (d.lgas || []);
        items.forEach(l => {
            const o = document.createElement('option');
            o.value = typeof l === 'string' ? l : l.lg;
            o.textContent = typeof l === 'string' ? l : l.lg;
            sel.appendChild(o);
        });
    } catch(e) {}
}

async function loadResults(page) {
    _rPage = page || 1;
    const lga    = document.getElementById('rFilterLga').value;
    const status = document.getElementById('rFilterStatus').value;
    const tbody  = document.getElementById('rTableBody');
    tbody.innerHTML = '<tr><td colspan="12" style="text-align:center;padding:24px;color:rgba(255,255,255,0.3);">Loading...</td></tr>';
    try {
        const res = await fetch(`/api/admin/results?page=${_rPage}&lga=${encodeURIComponent(lga)}&status=${status}`, {
            headers: { 'Authorization': 'Bearer ' + _rKey }
        });
        const d = await res.json();
        _rData   = d.results || [];
        _rTotal  = d.total || 0;
        _rPages  = d.pages || 1;
        renderTable(_rData);
        document.getElementById('rPageInfo').textContent = `Page ${_rPage} of ${_rPages} · ${_rTotal.toLocaleString()} results`;
        document.getElementById('rPrevBtn').disabled = _rPage <= 1;
        document.getElementById('rNextBtn').disabled = _rPage >= _rPages;
        // Count edited
        const editedCount = _rData.filter(r => r.edit_note || false).length;
        document.getElementById('rStatEdited').textContent = editedCount;
        loadSummaryStats();
    } catch(e) {
        tbody.innerHTML = '<tr><td colspan="12" style="color:#ff6b6b;text-align:center;padding:20px;">Failed to load results.</td></tr>';
    }
}

function renderTable(data) {
    const tbody = document.getElementById('rTableBody');
    if (!data.length) {
        tbody.innerHTML = '<tr><td colspan="12" style="text-align:center;padding:30px;color:rgba(255,255,255,0.3);">No results found.</td></tr>';
        return;
    }
    const offset = (_rPage - 1) * 50;
    tbody.innerHTML = data.map((r, i) => {
        const statusBadge = r.reviewed
            ? '<span class="badge badge-ok">✅ Approved</span>'
            : '<span class="badge badge-pending">⏳ Pending</span>';
        const editedBadge = r.edit_note ? ' <span class="badge badge-edited">✏️ Edited</span>' : '';
        const accord = r.votes ? (r.votes.ACCORD || 0) : 0;
        const ts = r.timestamp ? r.timestamp.substring(0,16).replace('T',' ') : '—';
        return `<tr id="rrow-${r.id}">
            <td style="color:rgba(255,255,255,0.3);">${offset + i + 1}</td>
            <td style="font-weight:700;color:#ffc107;white-space:nowrap;">${r.pu_code || '—'}</td>
            <td style="max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${r.location||''}">${r.location || '—'}</td>
            <td style="font-size:0.72rem;color:rgba(255,255,255,0.6);">${r.lg || '—'}</td>
            <td style="font-size:0.7rem;color:rgba(255,255,255,0.5);">${r.ward || '—'}</td>
            <td style="font-size:0.72rem;">${r.officer_id || '—'}</td>
            <td style="font-weight:700;color:#00cc66;">${accord.toLocaleString()}</td>
            <td>${(r.total_cast||0).toLocaleString()}</td>
            <td>${(r.total_accredited||0).toLocaleString()}</td>
            <td style="font-size:0.68rem;color:rgba(255,255,255,0.4);white-space:nowrap;">${ts}</td>
            <td>${statusBadge}${editedBadge}</td>
            <td style="white-space:nowrap;">
                <button class="btn-a btn-edit-r" onclick="openEditModal(${r.id})">✏️ Edit</button>
                ${r.reviewed
                    ? `<button class="btn-a btn-unapprove" onclick="toggleApprove(${r.id}, false)">↩ Undo</button>`
                    : `<button class="btn-a btn-approve"   onclick="toggleApprove(${r.id}, true)">✅ Approve</button>`}
            </td>
        </tr>`;
    }).join('');
}

function filterTable(q) {
    if (!q) { renderTable(_rData); return; }
    const lq = q.toLowerCase();
    renderTable(_rData.filter(r =>
        (r.pu_code||'').toLowerCase().includes(lq) ||
        (r.location||'').toLowerCase().includes(lq) ||
        (r.ward||'').toLowerCase().includes(lq) ||
        (r.officer_id||'').toLowerCase().includes(lq)
    ));
}

async function toggleApprove(id, approve) {
    try {
        const res = await fetch(`/api/admin/results/${id}/approve`, {
            method: 'POST',
            headers: { 'Authorization': 'Bearer ' + _rKey, 'Content-Type': 'application/json' },
            body: JSON.stringify({ approve })
        });
        if (!res.ok) { alert('Failed to update approval'); return; }
        loadResults(_rPage);
    } catch(e) { alert('Server error'); }
}

async function approveAllVisible() {
    const pending = _rData.filter(r => !r.reviewed);
    if (!pending.length) { alert('No pending results on this page.'); return; }
    if (!confirm(`Approve all ${pending.length} pending results on this page?`)) return;
    for (const r of pending) {
        await fetch(`/api/admin/results/${r.id}/approve`, {
            method: 'POST',
            headers: { 'Authorization': 'Bearer ' + _rKey, 'Content-Type': 'application/json' },
            body: JSON.stringify({ approve: true })
        });
    }
    loadResults(_rPage);
}

function openEditModal(id) {
    const r = _rData.find(x => x.id === id);
    if (!r) return;
    document.getElementById('modalId').value = id;
    document.getElementById('modalSub').textContent = `${r.pu_code} · ${r.location} · ${r.lg}`;
    document.getElementById('mAccredited').value = r.total_accredited || 0;
    document.getElementById('mCast').value       = r.total_cast || 0;
    document.getElementById('mValid').value      = r.valid_votes || 0;
    document.getElementById('mRejected').value   = r.rejected_votes || 0;
    document.getElementById('mNote').value       = r.edit_note || '';
    const grid = document.getElementById('modalParties');
    grid.innerHTML = PARTIES.map(p => `
        <div class="field-group">
            <label>${p}</label>
            <input type="number" id="mp-${p}" min="0" value="${(r.votes && r.votes[p]) || 0}">
        </div>`).join('');
    document.getElementById('editModal').style.display = 'flex';
}

function closeModal() {
    document.getElementById('editModal').style.display = 'none';
}

async function saveEdit() {
    const id   = parseInt(document.getElementById('modalId').value);
    const body = {
        total_accredited: parseInt(document.getElementById('mAccredited').value) || 0,
        total_cast:       parseInt(document.getElementById('mCast').value)       || 0,
        valid_votes:      parseInt(document.getElementById('mValid').value)      || 0,
        rejected_votes:   parseInt(document.getElementById('mRejected').value)   || 0,
        edit_note:        document.getElementById('mNote').value.trim(),
        votes: {}
    };
    PARTIES.forEach(p => {
        body.votes[p] = parseInt(document.getElementById(`mp-${p}`).value) || 0;
    });
    try {
        const res = await fetch(`/api/admin/results/${id}`, {
            method: 'PUT',
            headers: { 'Authorization': 'Bearer ' + _rKey, 'Content-Type': 'application/json' },
            body: JSON.stringify(body)
        });
        if (!res.ok) { const e = await res.json(); alert(e.detail || 'Save failed'); return; }
        closeModal();
        loadResults(_rPage);
    } catch(e) { alert('Server error'); }
}
</script>
</body>
</html>
"""

ADMIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ACCORD — Admin Portal</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
        :root {
            --green: #008751; --green-light: #00b368; --gold: #ffc107;
            --dark: #020c06; --panel: rgba(255,255,255,0.04); --border: rgba(255,255,255,0.1);
        }
        body { font-family: 'Inter', sans-serif; background: var(--dark); color: #fff; min-height: 100vh; }
        body::before {
            content: ''; position: fixed; inset: 0; z-index: 0;
            background: radial-gradient(ellipse 120% 80% at 50% -10%, rgba(0,135,81,0.15) 0%, transparent 60%),
                        linear-gradient(160deg, #020c06 0%, #041508 40%, #020c06 100%);
        }
        .grid-bg {
            position: fixed; inset: 0; z-index: 0;
            background-image: linear-gradient(rgba(0,135,81,0.05) 1px, transparent 1px),
                              linear-gradient(90deg, rgba(0,135,81,0.05) 1px, transparent 1px);
            background-size: 60px 60px; pointer-events: none;
        }

        /* ── Layout ── */
        .page { position: relative; z-index: 1; max-width: 860px; margin: 0 auto; padding: 32px 20px 60px; }

        /* ── Login gate ── */
        #loginGate {
            min-height: 100vh; display: flex; align-items: center; justify-content: center;
        }
        .login-card {
            background: var(--panel); border: 1px solid rgba(0,135,81,0.25); border-radius: 20px;
            padding: 48px 40px; width: 100%; max-width: 400px; text-align: center;
            box-shadow: 0 0 60px rgba(0,135,81,0.1);
        }
        .lock-icon {
            width: 60px; height: 60px; border-radius: 50%;
            background: rgba(0,135,81,0.12); border: 1px solid rgba(0,135,81,0.3);
            display: flex; align-items: center; justify-content: center;
            font-size: 1.5rem; margin: 0 auto 20px;
        }
        .badge-tag {
            display: inline-block; font-size: 0.58rem; font-weight: 700;
            letter-spacing: 0.18em; text-transform: uppercase; color: var(--gold);
            background: rgba(255,193,7,0.1); border: 1px solid rgba(255,193,7,0.2);
            border-radius: 20px; padding: 3px 12px; margin-bottom: 14px;
        }
        .login-card h2 { font-size: 1.3rem; font-weight: 900; margin-bottom: 6px; }
        .login-card p { font-size: 0.78rem; color: rgba(255,255,255,0.4); margin-bottom: 28px; }

        /* ── Inputs ── */
        input[type="password"], input[type="text"], input[type="file"] {
            width: 100%; background: rgba(255,255,255,0.05);
            border: 1px solid rgba(0,135,81,0.3); border-radius: 10px;
            color: #fff; font-size: 0.9rem; padding: 12px 16px; outline: none;
            transition: border-color 0.2s; font-family: 'Inter', sans-serif;
        }
        input:focus { border-color: rgba(0,135,81,0.7); }
        input[type="file"] { cursor: pointer; padding: 10px 14px; }
        textarea {
            width: 100%; background: rgba(255,255,255,0.04);
            border: 1px solid rgba(0,135,81,0.25); border-radius: 10px;
            color: #fff; font-size: 0.82rem; padding: 12px 14px; outline: none;
            font-family: 'Inter', monospace; resize: vertical; min-height: 120px;
            transition: border-color 0.2s;
        }
        textarea:focus { border-color: rgba(0,135,81,0.6); }

        /* ── Buttons ── */
        .btn-primary {
            background: linear-gradient(135deg, #008751, #00b368); border: none;
            border-radius: 10px; color: #fff; font-size: 0.9rem; font-weight: 700;
            padding: 12px 24px; cursor: pointer; width: 100%; transition: opacity 0.2s;
        }
        .btn-primary:hover { opacity: 0.88; }
        .btn-primary:disabled { opacity: 0.45; cursor: not-allowed; }
        .btn-secondary {
            background: rgba(255,255,255,0.06); border: 1px solid var(--border);
            border-radius: 10px; color: rgba(255,255,255,0.7); font-size: 0.82rem;
            font-weight: 600; padding: 10px 20px; cursor: pointer; transition: background 0.2s;
        }
        .btn-secondary:hover { background: rgba(255,255,255,0.1); }
        .btn-gold {
            background: rgba(255,193,7,0.12); border: 1px solid rgba(255,193,7,0.3);
            border-radius: 10px; color: var(--gold); font-size: 0.82rem; font-weight: 700;
            padding: 10px 20px; cursor: pointer; text-decoration: none; display: inline-block;
            transition: background 0.2s;
        }
        .btn-gold:hover { background: rgba(255,193,7,0.22); }

        /* ── Admin portal ── */
        #adminPortal { display: none; }
        .portal-header { margin-bottom: 32px; }
        .portal-header h1 { font-size: 1.6rem; font-weight: 900; margin-bottom: 6px; }
        .portal-header p { font-size: 0.82rem; color: rgba(255,255,255,0.4); }
        .header-row { display: flex; align-items: flex-start; justify-content: space-between; gap: 16px; flex-wrap: wrap; }

        /* ── Cards ── */
        .card {
            background: var(--panel); border: 1px solid var(--border);
            border-radius: 16px; padding: 28px; margin-bottom: 20px;
        }
        .card-title {
            font-size: 0.65rem; font-weight: 700; letter-spacing: 0.15em;
            text-transform: uppercase; color: var(--gold);
            border-left: 3px solid var(--gold); padding-left: 10px;
            margin-bottom: 20px; display: block;
        }

        /* ── Upload zone ── */
        .upload-zone {
            border: 2px dashed rgba(0,135,81,0.3); border-radius: 12px;
            padding: 36px 20px; text-align: center; cursor: pointer;
            transition: all 0.2s; background: rgba(0,135,81,0.03); position: relative;
        }
        .upload-zone:hover, .upload-zone.drag-over {
            border-color: rgba(0,135,81,0.7); background: rgba(0,135,81,0.07);
        }
        .upload-zone input[type="file"] {
            position: absolute; inset: 0; opacity: 0; cursor: pointer;
            width: 100%; height: 100%; border: none; padding: 0;
        }
        .upload-icon { font-size: 2rem; margin-bottom: 10px; }
        .upload-label { font-size: 0.88rem; font-weight: 600; color: rgba(255,255,255,0.7); }
        .upload-hint { font-size: 0.72rem; color: rgba(255,255,255,0.3); margin-top: 4px; }
        .file-chosen { font-size: 0.78rem; color: var(--green-light); margin-top: 8px; font-weight: 600; }

        /* ── Preview table ── */
        .preview-wrap { overflow-x: auto; margin-top: 16px; border-radius: 8px; overflow: hidden; }
        table { width: 100%; border-collapse: collapse; font-size: 0.78rem; }
        thead tr { background: rgba(0,135,81,0.15); }
        th { padding: 8px 12px; text-align: left; color: var(--gold); font-weight: 700;
             font-size: 0.65rem; letter-spacing: 0.08em; text-transform: uppercase; }
        tbody tr { border-bottom: 1px solid rgba(255,255,255,0.05); }
        tbody tr:hover { background: rgba(255,255,255,0.03); }
        .btn-action { background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.12); border-radius: 6px; color: #fff; cursor: pointer; font-size: 0.7rem; padding: 3px 8px; margin-right: 4px; transition: background 0.15s; }
        .btn-action:hover { background: rgba(255,255,255,0.12); }
        .btn-action:disabled { opacity: 0.4; cursor: default; }
        .btn-edit { border-color: rgba(0,179,104,0.4); color: #00cc66; }
        .btn-del  { border-color: rgba(255,107,107,0.4); color: #ff6b6b; }
        .btn-save { border-color: rgba(0,179,104,0.6); color: #00cc66; background: rgba(0,135,81,0.15); }
        .btn-cancel { border-color: rgba(255,193,7,0.4); color: #ffc107; }
        .btn-filter { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.1); border-radius: 20px; color: rgba(255,255,255,0.5); cursor: pointer; font-size: 0.7rem; padding: 4px 12px; transition: all 0.15s; }
        .btn-filter:hover { background: rgba(255,255,255,0.08); color: #fff; }
        .btn-filter.active { background: rgba(0,135,81,0.2); border-color: rgba(0,135,81,0.5); color: #00cc66; font-weight: 700; }
        .phone-missing { color: rgba(255,255,255,0.25); font-style: italic; font-size: 0.72rem; }
        td { padding: 7px 12px; color: rgba(255,255,255,0.75); }
        td.valid { color: #00cc66; }
        td.invalid { color: #ff6b6b; }
        .row-num { color: rgba(255,255,255,0.25); font-size: 0.68rem; }

        /* ── Result banner ── */
        .result-banner {
            border-radius: 10px; padding: 14px 18px; margin-top: 16px;
            font-size: 0.82rem; font-weight: 600; display: none;
        }
        .result-banner.success { background: rgba(0,204,102,0.12); border: 1px solid rgba(0,204,102,0.3); color: #00cc66; }
        .result-banner.error   { background: rgba(255,107,107,0.12); border: 1px solid rgba(255,107,107,0.3); color: #ff6b6b; }
        .result-banner.info    { background: rgba(255,193,7,0.1); border: 1px solid rgba(255,193,7,0.25); color: var(--gold); }

        /* ── Stats row ── */
        .stats-row { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 20px; }
        .stat-pill {
            background: rgba(255,255,255,0.04); border: 1px solid var(--border);
            border-radius: 10px; padding: 10px 18px; text-align: center; flex: 1; min-width: 80px;
        }
        .stat-pill .val { font-size: 1.4rem; font-weight: 900; display: block; }
        .stat-pill .lbl { font-size: 0.6rem; color: rgba(255,255,255,0.35); text-transform: uppercase; letter-spacing: 0.08em; }
        .val-green { color: #00cc66; }
        .val-red   { color: #ff6b6b; }
        .val-gold  { color: var(--gold); }

        /* ── Error alert ── */
        .err-box { background: rgba(220,53,69,0.12); border: 1px solid rgba(220,53,69,0.3);
                   border-radius: 8px; color: #ff6b6b; font-size: 0.78rem; padding: 10px 14px;
                   margin-bottom: 12px; display: none; }
        .divider { border: none; border-top: 1px solid var(--border); margin: 24px 0; }
        label.field-label { display: block; font-size: 0.72rem; font-weight: 600;
                            color: rgba(255,255,255,0.5); margin-bottom: 6px; }
    </style>
</head>
<body>
<div class="grid-bg"></div>

<!-- ── Login gate ── -->
<div id="loginGate">
    <div class="login-card">
        <div class="lock-icon">🔐</div>
        <div class="badge-tag">Admin Portal</div>
        <h2>Officer Management</h2>
        <p>Enter your admin key to manage officer phone numbers.</p>
        <div class="err-box" id="loginErr"></div>
        <div style="position:relative; margin-bottom:12px;">
            <input type="password" id="adminKey" placeholder="Enter admin key"
                   autocomplete="off" onkeydown="if(event.key==='Enter')adminLogin()"
                   style="padding-right:44px;">
            <button onclick="togglePwd()" style="position:absolute;right:12px;top:50%;transform:translateY(-50%);
                background:none;border:none;color:rgba(255,255,255,0.3);cursor:pointer;font-size:1rem;">👁</button>
        </div>
        <button class="btn-primary" id="loginBtn" onclick="adminLogin()">Access Admin Portal →</button>
        <div style="margin-top:16px;"><a href="/" style="font-size:0.72rem;color:rgba(255,255,255,0.25);text-decoration:none;">← Back to home</a></div>
    </div>
</div>

<!-- ── Admin portal ── -->
<div id="adminPortal">
<div class="page">

    <div class="portal-header">
        <div class="header-row">
            <div>
                <div class="badge-tag">Admin Portal — Officer Management</div>
                <h1>📋 Officer Phone Numbers</h1>
                <p>Upload a CSV or paste officer IDs and phone numbers to register them for WhatsApp OTP authentication.</p>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;">
                <a href="/admin/results" target="_blank" class="btn-gold" style="background:rgba(0,179,104,0.15);border-color:rgba(0,179,104,0.4);">📊 Results Portal</a>
                <a id="templateDownloadBtn" href="#" onclick="downloadTemplate(event)" class="btn-gold">⬇ Download CSV Template</a>
                <button class="btn-secondary" onclick="adminLogout()">🔒 Lock</button>
            </div>
        </div>
    </div>

    <!-- ── Stats ── -->
    <div class="stats-row">
        <div class="stat-pill"><span class="val val-gold" id="statTotal">—</span><span class="lbl">Total Officers</span></div>
        <div class="stat-pill"><span class="val val-green" id="statWithPhone">—</span><span class="lbl">With Phone</span></div>
        <div class="stat-pill"><span class="val val-red" id="statNoPhone">—</span><span class="lbl">Missing Phone</span></div>
    </div>

    <!-- ── CSV Upload ── -->
    <div class="card">
        <span class="card-title">📁 Upload CSV File</span>
        <div class="upload-zone" id="uploadZone">
            <input type="file" id="csvFile" accept=".csv,text/csv" onchange="handleFileSelect(this)">
            <div class="upload-icon">📂</div>
            <div class="upload-label">Drop your CSV here or click to browse</div>
            <div class="upload-hint">Columns: <code>officer_id</code>, <code>lga</code>, <code>ward</code>, <code>polling_unit</code>, <code>phone</code> &nbsp;·&nbsp; Use the template from this portal</div>
            <div class="file-chosen d-none" id="fileChosen"></div>
        </div>

        <div id="previewSection" style="display:none;">
            <hr class="divider">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                <span style="font-size:0.78rem;color:rgba(255,255,255,0.5);" id="previewLabel"></span>
                <button class="btn-secondary" onclick="clearFile()" style="padding:6px 14px;font-size:0.72rem;">✕ Clear</button>
            </div>
            <div class="preview-wrap"><table id="previewTable"><thead></thead><tbody></tbody></table></div>
        </div>

        <div class="result-banner" id="uploadResult"></div>
        <div style="margin-top:16px;">
            <button class="btn-primary" id="uploadBtn" onclick="submitCSV()" disabled>Upload & Register Officers</button>
        </div>
    </div>

    <!-- ── Manual paste ── -->
    <div class="card">
        <span class="card-title">✏️ Manual Entry (Paste or Type)</span>
        <p style="font-size:0.78rem;color:rgba(255,255,255,0.4);margin-bottom:14px;">
            One officer per line. Format: <code style="color:#00cc66;">officer_id,lga,phone</code><br>
            <span style="font-size:0.72rem;color:rgba(255,255,255,0.25);">LGA is required to uniquely identify the polling unit across all 30 LGAs.</span>
        </p>
        <label class="field-label">Officer entries (officer_id, lga, phone)</label>
        <textarea id="manualInput" placeholder="10-001,Osogbo,+2348012345678&#10;06-001,Ife Central,+2348023456789&#10;04-001,Atakumosa East,+2348034567890"></textarea>
        <div class="result-banner" id="manualResult"></div>
        <div style="margin-top:14px;">
            <button class="btn-primary" id="manualBtn" onclick="submitManual()">Register These Officers</button>
        </div>
    </div>

</div>

    <!-- ── All Officers Table ── -->
    <div class="card" id="officerTableCard" style="margin-top:24px;">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px;flex-wrap:wrap;gap:10px;">
            <span class="card-title" style="margin-bottom:0;">📋 Officer Management</span>
            <input type="text" id="officerSearch" placeholder="Search ID, LGA, ward, phone..." onkeyup="loadOfficerTable(1, undefined, this.value)"
                   style="background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:#fff;padding:6px 12px;font-size:0.78rem;width:240px;">
        </div>
        <div style="display:flex;gap:8px;margin-bottom:12px;">
            <button class="btn-filter active" id="fAll"          onclick="setOfficerFilter('all')">All</button>
            <button class="btn-filter"        id="fRegistered"   onclick="setOfficerFilter('registered')">✅ Registered</button>
            <button class="btn-filter"        id="fUnregistered" onclick="setOfficerFilter('unregistered')">⚠️ No Phone</button>
        </div>
        <div class="preview-wrap" style="max-height:460px;overflow-y:auto;">
            <table style="width:100%;border-collapse:collapse;">
                <thead>
                    <tr style="font-size:0.68rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;border-bottom:1px solid rgba(255,255,255,0.08);">
                        <th style="padding:8px 6px;text-align:left;">Officer ID</th>
                        <th style="padding:8px 6px;text-align:left;">LGA</th>
                        <th style="padding:8px 6px;text-align:left;">Ward</th>
                        <th style="padding:8px 6px;text-align:left;">Polling Unit</th>
                        <th style="padding:8px 6px;text-align:left;">Phone</th>
                        <th style="padding:8px 6px;text-align:left;">Actions</th>
                    </tr>
                </thead>
                <tbody id="officerTableBody">
                    <tr><td colspan="6" style="text-align:center;padding:20px;color:rgba(255,255,255,0.3);">Login to manage officers.</td></tr>
                </tbody>
            </table>
        </div>
        <div id="officerTableInfo" style="margin-top:10px;font-size:0.72rem;color:rgba(255,255,255,0.4);display:flex;align-items:center;gap:10px;"></div>
    </div>

</div>

<script>
    let _adminKey = '';

    function togglePwd() {
        const i = document.getElementById('adminKey');
        i.type = i.type === 'password' ? 'text' : 'password';
    }

    async function adminLogin() {
        const key = document.getElementById('adminKey').value.trim();
        if (!key) return;
        const btn = document.getElementById('loginBtn');
        const err = document.getElementById('loginErr');
        btn.disabled = true; btn.textContent = 'Verifying...';
        err.style.display = 'none';
        // Test key against a protected endpoint
        try {
            const res = await fetch('/api/admin/set-officer-phone', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + key, 'Content-Type': 'application/json' },
                body: JSON.stringify({ officers: [] })
            });
            if (res.status === 401 || res.status === 403) {
                err.textContent = 'Invalid admin key.';
                err.style.display = 'block';
                btn.disabled = false; btn.textContent = 'Access Admin Portal →';
                return;
            }
            _adminKey = key;
            document.getElementById('loginGate').style.display = 'none';
            document.getElementById('adminPortal').style.display = 'block';
            loadStats();
        } catch(e) {
            err.textContent = 'Server error. Try again.';
            err.style.display = 'block';
            btn.disabled = false; btn.textContent = 'Access Admin Portal →';
        }
    }

    function adminLogout() {
        _adminKey = '';
        document.getElementById('adminPortal').style.display = 'none';
        document.getElementById('loginGate').style.display = 'flex';
        document.getElementById('adminKey').value = '';
        document.getElementById('loginBtn').disabled = false;
        document.getElementById('loginBtn').textContent = 'Access Admin Portal →';
    }

    async function _loadStatsCore() {
        try {
            const res = await fetch('/api/officers/stats', {
                headers: { 'Authorization': 'Bearer ' + _adminKey }
            });
            if (!res.ok) return;
            const d = await res.json();
            document.getElementById('statTotal').textContent     = d.total.toLocaleString();
            document.getElementById('statWithPhone').textContent = d.with_phone.toLocaleString();
            document.getElementById('statNoPhone').textContent   = d.without_phone.toLocaleString();
        } catch(e) {}
    }

    // ── CSV file handling ────────────────────────────────────────────────────
    let _parsedRows = [];

    function handleFileSelect(input) {
        const file = input.files[0];
        if (!file) return;
        document.getElementById('fileChosen').textContent = '📄 ' + file.name;
        document.getElementById('fileChosen').classList.remove('d-none');
        const reader = new FileReader();
        reader.onload = e => parseCSV(e.target.result);
        reader.readAsText(file);
    }

    // Drag-and-drop
    const zone = document.getElementById('uploadZone');
    zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag-over'); });
    zone.addEventListener('dragleave', () => zone.classList.remove('drag-over'));
    zone.addEventListener('drop', e => {
        e.preventDefault(); zone.classList.remove('drag-over');
        const file = e.dataTransfer.files[0];
        if (file) { document.getElementById('csvFile').files = e.dataTransfer.files; handleFileSelect({ files: [file] }); }
    });

    function parseCSV(text) {
        const lines = text.trim().split(/\\r?\\n/).filter(l => l.trim());
        if (!lines.length) return;
        const first = lines[0].toLowerCase();
        const hasHeader = first.includes('officer_id') || first.includes('phone') || first.includes('lga');
        const dataLines = hasHeader ? lines.slice(1) : lines;

        // Detect column positions from header
        let colOfficer = 0, colLga = 1, colWard = 2, colPu = 3, colPhone = 4;
        if (hasHeader) {
            const hCols = _splitCSVLine(lines[0].toLowerCase());
            hCols.forEach((h, i) => {
                if (h.includes('officer_id') || h === 'officer id') colOfficer = i;
                else if (h === 'lga')                               colLga     = i;
                else if (h === 'ward')                              colWard    = i;
                else if (h.includes('polling') || h.includes('unit')) colPu    = i;
                else if (h === 'phone')                             colPhone   = i;
            });
        }

        _parsedRows = [];
        dataLines.forEach((line, i) => {
            const parts      = _splitCSVLine(line);
            const officer_id = (parts[colOfficer] || '').trim();
            const lga        = (parts[colLga]     || '').trim();
            const ward       = (parts[colWard]    || '').trim();
            const pu         = (parts[colPu]      || '').trim();
            const phone      = (parts[colPhone]   || '').trim();
            const hasPhone   = phone.length >= 10;
            const valid      = officer_id.includes('-') && hasPhone;
            const noPhone    = officer_id.includes('-') && !hasPhone;
            _parsedRows.push({ officer_id, lga, ward, pu, phone, valid, noPhone, row: i + (hasHeader ? 2 : 1) });
        });
        renderPreview();
        document.getElementById('uploadBtn').disabled = _parsedRows.filter(r => r.valid).length === 0;
    }

    // CSV-aware splitter — handles quoted fields containing commas
    function _splitCSVLine(line) {
        const result = [];
        let cur = '', inQuote = false;
        for (let i = 0; i < line.length; i++) {
            const ch = line[i];
            if (ch === '"') { inQuote = !inQuote; }
            else if (ch === ',' && !inQuote) { result.push(cur); cur = ''; }
            else { cur += ch; }
        }
        result.push(cur);
        return result.map(s => s.replace(/^"|"$/g, '').trim());
    }

    function renderPreview() {
        const section  = document.getElementById('previewSection');
        const label    = document.getElementById('previewLabel');
        const valid    = _parsedRows.filter(r => r.valid).length;
        const noPhone  = _parsedRows.filter(r => r.noPhone).length;
        const invalid  = _parsedRows.filter(r => !r.valid && !r.noPhone).length;
        label.innerHTML =
            `<span style="color:#00cc66;">${valid} ready to upload</span>` +
            (noPhone  ? ` &nbsp;·&nbsp; <span style="color:#ffc107;">${noPhone} missing phone (will skip)</span>` : '') +
            (invalid  ? ` &nbsp;·&nbsp; <span style="color:#ff6b6b;">${invalid} invalid rows</span>` : '') +
            ` &nbsp;·&nbsp; ${_parsedRows.length} total`;
        section.style.display = 'block';

        const thead = document.querySelector('#previewTable thead');
        const tbody = document.querySelector('#previewTable tbody');
        thead.innerHTML = '<tr><th>#</th><th>Officer ID</th><th>LGA</th><th>Ward</th><th>Polling Unit</th><th>Phone</th><th>Status</th></tr>';
        tbody.innerHTML = _parsedRows.slice(0, 100).map(r => {
            const statusIcon = r.valid ? '✅' : r.noPhone ? '⏭ skip' : '⚠️';
            const rowClass   = r.valid ? 'valid' : r.noPhone ? '' : 'invalid';
            return `<tr>
                <td class="row-num">${r.row}</td>
                <td class="${rowClass}" style="font-weight:700;">${r.officer_id || '<em>—</em>'}</td>
                <td style="color:rgba(255,255,255,0.6);font-size:0.72rem;">${r.lga || '—'}</td>
                <td style="color:rgba(255,255,255,0.5);font-size:0.7rem;">${r.ward || '—'}</td>
                <td style="color:rgba(255,255,255,0.45);font-size:0.68rem;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${r.pu}">${r.pu || '—'}</td>
                <td class="${r.valid ? 'valid' : r.noPhone ? '' : 'invalid'}">${r.phone || '<em style="color:#555;">blank</em>'}</td>
                <td>${statusIcon}</td>
            </tr>`;
        }).join('') + (_parsedRows.length > 100 ? `<tr><td colspan="7" style="color:rgba(255,255,255,0.3);text-align:center;padding:10px;">... and ${_parsedRows.length - 100} more rows</td></tr>` : '');
    }

    function clearFile() {
        _parsedRows = [];
        document.getElementById('csvFile').value = '';
        document.getElementById('fileChosen').classList.add('d-none');
        document.getElementById('previewSection').style.display = 'none';
        document.getElementById('uploadBtn').disabled = true;
        document.getElementById('uploadResult').style.display = 'none';
    }

    async function submitCSV() {
        const valid = _parsedRows.filter(r => r.valid);
        if (!valid.length) return;
        const btn = document.getElementById('uploadBtn');
        const res_el = document.getElementById('uploadResult');
        btn.disabled = true; btn.textContent = `Uploading ${valid.length} officers...`;
        res_el.style.display = 'none';
        await _submitOfficers(valid, res_el);
        btn.disabled = false; btn.textContent = 'Upload & Register Officers';
        loadStats();
    }

    // ── Manual entry ─────────────────────────────────────────────────────────
    async function submitManual() {
        const raw = document.getElementById('manualInput').value.trim();
        if (!raw) return;
        const btn    = document.getElementById('manualBtn');
        const res_el = document.getElementById('manualResult');
        const lines  = raw.split(/\\n/).filter(l => l.trim());
        const officers = lines.map(l => {
            const p = _splitCSVLine(l);
            if (p.length >= 3) {
                return { officer_id: (p[0]||'').trim(), lga: (p[1]||'').trim(), ward: '', pu: '', phone: (p[2]||'').trim() };
            } else {
                return { officer_id: (p[0]||'').trim(), lga: '', ward: '', pu: '', phone: (p[1]||'').trim() };
            }
        }).filter(o => o.officer_id && o.phone);
        if (!officers.length) {
            showBanner(res_el, 'No valid entries. Format: officer_id,lga,phone  e.g.  10-001,Osogbo,+2348012345678', 'error');
            return;
        }
        btn.disabled = true; btn.textContent = `Registering ${officers.length} officers...`;
        res_el.style.display = 'none';
        await _submitOfficers(officers, res_el);
        btn.disabled = false; btn.textContent = 'Register These Officers';
        loadStats();
    }

    // ── Shared submit: builds a CSV in memory and posts it to the Supabase-backed
    //    /api/upload-officers endpoint, which is what actually persists officers.
    async function _submitOfficers(officers, res_el) {
        try {
            const header = 'officer_id,lga,ward,polling_unit,phone';
            const lines  = officers.map(o =>
                [o.officer_id, o.lga || '', o.ward || '', o.pu || '', o.phone]
                    .map(v => `"${String(v).replace(/"/g, '""')}"`).join(',')
            );
            const csvText = [header, ...lines].join('\\n');
            const blob = new Blob([csvText], { type: 'text/csv' });

            const formData = new FormData();
            formData.append('file', blob, 'officers_upload.csv');

            const res = await fetch('/api/upload-officers', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + _adminKey },
                body: formData
            });
            const out = await res.json();
            if (!res.ok) {
                showBanner(res_el, out.detail || 'Upload failed.', 'error');
                return;
            }
            let msg = `✅ ${out.saved} officer${out.saved !== 1 ? 's' : ''} saved permanently to Supabase`;
            if (out.skipped) {
                msg += ` · ${out.skipped} skipped`;
                if (out.skipped_details && out.skipped_details.length) {
                    msg += `<br><span style="color:#ffc107;font-size:0.78rem;">⚠️ ${out.skipped_details.slice(0,5).join('; ')}${out.skipped_details.length > 5 ? ' ...' : ''}</span>`;
                }
            }
            showBanner(res_el, msg, out.saved > 0 ? 'success' : 'info');
        } catch(e) {
            showBanner(res_el, 'Server error. Try again.', 'error');
        }
    }

    function showBanner(el, msg, type) {
        el.textContent = msg;
        el.className = 'result-banner ' + type;
        el.style.display = 'block';
    }

    async function downloadTemplate(e) {
        e.preventDefault();
        const res = await fetch('/admin/template.csv', {
            headers: { 'Authorization': 'Bearer ' + _adminKey }
        });
        if (!res.ok) { alert('Could not download template'); return; }
        const blob = await res.blob();
        const url  = URL.createObjectURL(blob);
        const a    = document.createElement('a');
        a.href = url; a.download = 'officer_phones_template.csv'; a.click();
        URL.revokeObjectURL(url);
    }

    document.addEventListener('DOMContentLoaded', () => {
        document.getElementById('adminKey').focus();
    });

    // ── Officer Management Table ─────────────────────────────────────────────
    let _officerPage = 1, _officerQ = '', _officerFilter = 'all';

    function setOfficerFilter(f) {
        _officerFilter = f;
        ['All','Registered','Unregistered'].forEach(n => {
            const el = document.getElementById('f' + n);
            if (el) el.classList.toggle('active', n.toLowerCase() === f || (f === 'all' && n === 'All'));
        });
        loadOfficerTable(1);
    }

    async function loadOfficerTable(page, filter, q) {
        _officerPage   = page  !== undefined ? page  : _officerPage;
        _officerFilter = filter !== undefined ? filter : _officerFilter;
        _officerQ      = q    !== undefined ? q    : _officerQ;
        const tbl  = document.getElementById('officerTableBody');
        const info = document.getElementById('officerTableInfo');
        if (!tbl) return;
        tbl.innerHTML = '<tr><td colspan="6" style="text-align:center;padding:20px;color:rgba(255,255,255,0.3);">Loading...</td></tr>';
        try {
            const url = `/api/officers?page=${_officerPage}&search=${encodeURIComponent(_officerQ)}&filter=${_officerFilter}`;
            const res = await fetch(url, { headers: { 'Authorization': 'Bearer ' + _adminKey } });
            const d   = await res.json();
            if (!d.officers || !d.officers.length) {
                tbl.innerHTML = '<tr><td colspan="6" style="text-align:center;padding:20px;color:rgba(255,255,255,0.3);">No officers found.</td></tr>';
                if (info) info.innerHTML = '';
                return;
            }
            tbl.innerHTML = d.officers.map(o => {
                const hasPhone = o.phone && o.phone.trim();
                const phoneDisplay = hasPhone
                    ? `<span style="color:#00cc66;font-weight:600;">${o.phone}</span>`
                    : `<span class="phone-missing">— no phone</span>`;
                const jsId = JSON.stringify(o.officer_id);
                const jsPhone = JSON.stringify(o.phone || '');
                const actions = `
                    <button class="btn-action btn-edit" onclick='editOfficerRow(${jsId},${jsPhone})'>✏️ Edit</button>
                    ${hasPhone ? `<button class="btn-action btn-del" onclick='deleteOfficer(${jsId})'>🗑 Remove</button>` : ''}`;
                return `<tr id="orow-${o.officer_id}">
                    <td style="font-weight:700;color:#ffc107;white-space:nowrap;">${o.officer_id}</td>
                    <td style="font-size:0.74rem;color:rgba(255,255,255,0.6);">${o.lga||'—'}</td>
                    <td style="font-size:0.7rem;color:rgba(255,255,255,0.5);">${o.ward||'—'}</td>
                    <td style="font-size:0.68rem;color:rgba(255,255,255,0.4);max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${o.polling_unit||''}">${o.polling_unit||'—'}</td>
                    <td id="phone-${o.officer_id}">${phoneDisplay}</td>
                    <td style="white-space:nowrap;">${actions}</td>
                </tr>`;
            }).join('');
            if (info) info.innerHTML =
                `<span>Page ${d.page} of ${d.pages} &nbsp;·&nbsp; ${d.total.toLocaleString()} officers</span>` +
                (d.pages > 1 ? `<button class="btn-action" onclick="loadOfficerTable(${_officerPage-1})" ${_officerPage<=1?'disabled':''}>◀</button>` +
                               `<button class="btn-action" onclick="loadOfficerTable(${_officerPage+1})" ${_officerPage>=d.pages?'disabled':''}>▶</button>` : '');
        } catch(e) {
            tbl.innerHTML = '<tr><td colspan="6" style="color:#ff6b6b;text-align:center;padding:16px;">Failed to load officers.</td></tr>';
        }
    }

    function editOfficerRow(officerId, currentPhone) {
        const phoneCell  = document.getElementById(`phone-${officerId}`);
        const row        = document.getElementById(`orow-${officerId}`);
        const actionCell = row.querySelector('td:last-child');
        const jsId = JSON.stringify(officerId);
        phoneCell.innerHTML = `<input type="text" id="ep-${officerId}" value="${currentPhone.replace(/"/g,'&quot;')}"
            placeholder="+2348012345678"
            style="background:rgba(255,255,255,0.08);border:1px solid rgba(0,135,81,0.5);border-radius:6px;color:#fff;padding:4px 8px;width:155px;font-size:0.78rem;"
            onkeydown="if(event.key==='Enter')saveOfficerEdit(${jsId});if(event.key==='Escape')loadOfficerTable();">`;
        actionCell.innerHTML = `
            <button class="btn-action btn-save"   onclick='saveOfficerEdit(${jsId})'>💾 Save</button>
            <button class="btn-action btn-cancel" onclick="loadOfficerTable()">✕</button>`;
        document.getElementById(`ep-${officerId}`).focus();
    }

    async function saveOfficerEdit(officerId) {
        const phone = document.getElementById(`ep-${officerId}`).value.trim();
        try {
            const res = await fetch('/api/officers/update', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + _adminKey, 'Content-Type': 'application/json' },
                body: JSON.stringify({ officer_id: officerId, phone })
            });
            const out = await res.json();
            if (!res.ok) { alert(out.detail || 'Update failed'); return; }
            loadOfficerTable();
            loadStats();
        } catch(e) { alert('Server error'); }
    }

    async function deleteOfficer(officerId) {
        if (!confirm(`Remove officer ${officerId}? They will not be able to log in via OTP.`)) return;
        try {
            const res = await fetch('/api/officers/delete', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + _adminKey, 'Content-Type': 'application/json' },
                body: JSON.stringify({ officer_id: officerId })
            });
            const out = await res.json();
            if (!res.ok) { alert(out.detail || 'Delete failed'); return; }
            loadOfficerTable();
            loadStats();
        } catch(e) { alert('Server error'); }
    }

    // loadStats also refreshes the officer table after login
    async function loadStats() {
        await _loadStatsCore();
        if (_adminKey) loadOfficerTable(1, 'all', '');
    }
</script>
</body>
</html>
"""

# ── Admin portal ─────────────────────────────────────────────────────────────

# ── Admin: list registered officers (with phone) ──────────────────────────────
@app.get("/api/admin/list-officers")
async def list_officers(request: Request, page: int = 1, q: str = "", filter: str = "all"):
    """
    filter: all | registered | unregistered
    """
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    PAGE_SIZE = 50
    offset = (page - 1) * PAGE_SIZE
    conditions = ["LOWER(state)='osun'"]
    params_q = []
    if filter == "registered":
        conditions.append("officer_phone IS NOT NULL AND officer_phone != ''")
    elif filter == "unregistered":
        conditions.append("(officer_phone IS NULL OR officer_phone = '')")
    if q:
        like = f"%{q}%"
        conditions.append("(ward_code LIKE ? OR pu_code LIKE ? OR LOWER(lg) LIKE ? OR LOWER(ward) LIKE ? OR officer_phone LIKE ?)")
        params_q = [like, like, like.lower(), like.lower(), like]
    where = "WHERE " + " AND ".join(conditions)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) as c FROM polling_units {where}", params_q)
            total = cur.fetchone()["c"]
            cur.execute(f"""SELECT ward_code, pu_code, lg, ward, location, officer_phone
                           FROM polling_units {where}
                           ORDER BY lg, ward_code, pu_code LIMIT ? OFFSET ?""",
                        params_q + [PAGE_SIZE, offset])
            rows = cur.fetchall()
    return {
        "officers": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "pages": max(1, -(-total // PAGE_SIZE))
    }


# ── Admin: update a single officer phone ─────────────────────────────────────
@app.put("/api/admin/update-officer")
async def update_officer(request: Request):
    """Body: { "ward_code": "10", "pu_code": "001", "phone": "+2348012345678" }"""
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    ward_code = str(body.get("ward_code", "")).strip()
    pu_code   = str(body.get("pu_code", "")).strip()
    phone     = _clean_phone(str(body.get("phone", "")).strip())
    if not ward_code or not pu_code:
        raise HTTPException(status_code=400, detail="ward_code and pu_code required")
    if phone and len(phone) < 10:
        raise HTTPException(status_code=400, detail="Invalid phone number")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""UPDATE polling_units SET officer_phone = ?
                           WHERE ward_code = ? AND pu_code = ? AND LOWER(state)='osun'""",
                        (phone or None, ward_code, pu_code))
            changed = conn._conn.total_changes
        conn.commit()
    if not changed:
        raise HTTPException(status_code=404, detail="Officer not found")
    return {"status": "ok", "phone": phone}


# ── Admin: delete officer phone (clear it) ───────────────────────────────────
@app.delete("/api/admin/delete-officer")
async def delete_officer(request: Request):
    """Body: { "ward_code": "10", "pu_code": "001" }"""
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    ward_code = str(body.get("ward_code", "")).strip()
    pu_code   = str(body.get("pu_code", "")).strip()
    if not ward_code or not pu_code:
        raise HTTPException(status_code=400, detail="ward_code and pu_code required")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""UPDATE polling_units SET officer_phone = NULL
                           WHERE ward_code = ? AND pu_code = ? AND LOWER(state)='osun'""",
                        (ward_code, pu_code))
            changed = conn._conn.total_changes
        conn.commit()
    if not changed:
        raise HTTPException(status_code=404, detail="Officer not found")
    return {"status": "ok"}


# ── Admin: list all submitted results (paginated) ────────────────────────────
@app.get("/api/admin/results")
async def admin_list_results(request: Request, page: int = 1, lga: str = "", status: str = ""):
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    PAGE_SIZE = 50
    offset = (page - 1) * PAGE_SIZE
    filters, params = [], []
    if lga:
        filters.append("LOWER(lg) = LOWER(%s)")
        params.append(lga)
    if status == "reviewed":
        filters.append("reviewed = 1")
    elif status == "pending":
        filters.append("(reviewed IS NULL OR reviewed = 0)")
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) as c FROM field_submissions {where}", params)
            total = cur.fetchone()["c"]
            cur.execute(f"""SELECT id, officer_id, state, lg, ward, ward_code, pu_code, location,
                                   reg_voters, total_accredited, valid_votes, rejected_votes, total_cast,
                                   votes_json, edited_votes_json, ec8e_image, timestamp,
                                   reviewed, reviewed_by, reviewed_at, edit_note
                            FROM field_submissions {where}
                            ORDER BY timestamp DESC LIMIT %s OFFSET %s""",
                        params + [PAGE_SIZE, offset])
            rows = cur.fetchall()
    results = []
    for r in rows:
        votes = json.loads(r.get("edited_votes_json") or r.get("votes_json") or "{}")
        results.append({
            "id": r["id"],
            "officer_id": r["officer_id"],
            "lg": r["lg"],
            "ward": r["ward"],
            "location": r["location"],
            "pu_code": r["pu_code"],
            "reg_voters": r["reg_voters"] or 0,
            "total_accredited": r["total_accredited"] or 0,
            "valid_votes": r["valid_votes"] or 0,
            "rejected_votes": r["rejected_votes"] or 0,
            "total_cast": r["total_cast"] or 0,
            "votes": votes,
            "ec8e_image": r["ec8e_image"],
            "timestamp": r["timestamp"],
            "reviewed": bool(r["reviewed"]),
            "reviewed_by": r["reviewed_by"],
            "reviewed_at": r["reviewed_at"],
            "edit_note": r["edit_note"],
        })
    return {"results": results, "total": total, "page": page, "pages": max(1, -(-total // PAGE_SIZE))}


# ── Admin: edit a submitted result ───────────────────────────────────────────
@app.put("/api/admin/results/{submission_id}")
async def admin_edit_result(submission_id: int, request: Request):
    """
    Editable fields: votes (dict), total_accredited, total_cast,
                     valid_votes, rejected_votes, reg_voters, edit_note
    """
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    now = datetime.now().isoformat()
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM field_submissions WHERE id=%s", (submission_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Submission not found")

            updates, vals, audit_entries = [], [], []
            old_votes = json.loads(row["edited_votes_json"] or row["votes_json"] or "{}")

            if "votes" in body:
                new_votes = {k: int(v) for k, v in body["votes"].items()}
                updates.append("edited_votes_json = %s")
                vals.append(json.dumps(new_votes))
                audit_entries.append(("edit", "votes", json.dumps(old_votes), json.dumps(new_votes)))

            for field in ["total_accredited", "total_cast", "valid_votes", "rejected_votes", "reg_voters"]:
                if field in body:
                    old_val = row[field]
                    new_val = int(body[field])
                    updates.append(f"{field} = %s")
                    vals.append(new_val)
                    audit_entries.append(("edit", field, str(old_val), str(new_val)))

            if "edit_note" in body:
                updates.append("edit_note = %s")
                vals.append(str(body["edit_note"])[:500])

            if not updates:
                raise HTTPException(status_code=400, detail="No editable fields provided")

            vals.append(submission_id)
            cur.execute(f"UPDATE field_submissions SET {', '.join(updates)} WHERE id=%s", vals)

            for action, field, old_v, new_v in audit_entries:
                cur.execute("""INSERT INTO result_audit_log
                               (submission_id, action, field, old_value, new_value, changed_by, changed_at)
                               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                            (submission_id, action, field, old_v, new_v, "admin", now))
        conn.commit()
    return {"status": "ok", "id": submission_id}


# ── Admin: approve / unapprove a result ──────────────────────────────────────
@app.post("/api/admin/results/{submission_id}/approve")
async def admin_approve_result(submission_id: int, request: Request):
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        body = {}
    approve = bool(body.get("approve", True))
    now = datetime.now().isoformat()
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT reviewed FROM field_submissions WHERE id=%s", (submission_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Submission not found")
            cur.execute("""UPDATE field_submissions
                           SET reviewed=%s, reviewed_by='admin', reviewed_at=%s
                           WHERE id=%s""",
                        (1 if approve else 0, now, submission_id))
            cur.execute("""INSERT INTO result_audit_log
                           (submission_id, action, field, old_value, new_value, changed_by, changed_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (submission_id,
                         "approve" if approve else "unapprove",
                         "reviewed",
                         str(bool(row["reviewed"])),
                         str(approve),
                         "admin", now))
        conn.commit()
    return {"status": "ok", "reviewed": approve}


# ── Admin: pending review count (for dashboard badge) ────────────────────────
@app.get("/api/admin/pending-review-count")
async def pending_review_count(request: Request):
    auth  = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    bearer_ok = auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )
    if not bearer_ok and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as c FROM field_submissions WHERE reviewed=0 OR reviewed IS NULL")
            pending = cur.fetchone()["c"]
            cur.execute("SELECT COUNT(*) as c FROM field_submissions")
            total = cur.fetchone()["c"]
    return {"pending": pending, "total": total}


# ── Admin: results portal page ───────────────────────────────────────────────
@app.get("/admin/results", response_class=HTMLResponse)
async def admin_results_page(request: Request):
    return HTMLResponse(content=RESULTS_PORTAL_HTML)


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    """Protected admin portal — requires DASHBOARD_KEY as Bearer token via
    a login form (same key as dashboard). Served as a standalone HTML page."""
    return HTMLResponse(content=ADMIN_HTML)


@app.get("/admin/template.csv")
async def download_template(request: Request):
    """Download the officer phone CSV template."""
    auth = request.headers.get("Authorization", "")
    # Also allow cookie-authenticated dashboard users
    token = request.cookies.get("ds_session")
    header_ok = auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(),
        _DASHBOARD_KEY_HASH
    )
    if not header_ok and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    csv_content = (
        "officer_id,phone\n"
        "WARD001-PU001,+2348012345678\n"
        "WARD001-PU002,+2348023456789\n"
        "WARD002-PU001,+2348034567890\n"
    )
    from fastapi.responses import Response as _R
    return _R(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=officer_phones_template.csv"}
    )
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def homepage():
    return HTMLResponse(content=HOMEPAGE_HTML)

HOMEPAGE_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta name="theme-color" content="#008751">
    <title>ACCORD — Osun 2026 Command System</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

        :root {
            --green: #008751;
            --green-light: #00b368;
            --gold: #ffc107;
            --dark: #020c06;
            --panel: rgba(255,255,255,0.04);
            --border: rgba(255,255,255,0.08);
        }

        body {
            font-family: 'Inter', sans-serif;
            background: var(--dark);
            color: #fff;
            min-height: 100vh;
            overflow-x: hidden;
        }

        /* ── Multi-layer background ── */
        body::before {
            content: '';
            position: fixed;
            inset: 0;
            z-index: 0;
            background:
                radial-gradient(ellipse 120% 80% at 50% -10%, rgba(0,135,81,0.18) 0%, transparent 60%),
                radial-gradient(ellipse 80% 60% at 80% 100%, rgba(0,80,40,0.15) 0%, transparent 55%),
                radial-gradient(ellipse 60% 50% at 10% 80%, rgba(255,193,7,0.04) 0%, transparent 50%),
                linear-gradient(160deg, #020c06 0%, #041508 40%, #020c06 100%);
            pointer-events: none;
        }

        /* ── Animated canvas background ── */
        #bg-canvas {
            position: fixed;
            inset: 0;
            z-index: 1;
            opacity: 0.65;
        }

        /* ── Precision grid overlay ── */
        .grid-overlay {
            position: fixed;
            inset: 0;
            z-index: 2;
            background-image:
                linear-gradient(rgba(0,135,81,0.06) 1px, transparent 1px),
                linear-gradient(90deg, rgba(0,135,81,0.06) 1px, transparent 1px),
                linear-gradient(rgba(0,135,81,0.02) 1px, transparent 1px),
                linear-gradient(90deg, rgba(0,135,81,0.02) 1px, transparent 1px);
            background-size: 80px 80px, 80px 80px, 20px 20px, 20px 20px;
            pointer-events: none;
        }

        /* ── Corner accent lines ── */
        .grid-overlay::before {
            content: '';
            position: absolute;
            inset: 0;
            background:
                linear-gradient(135deg, rgba(0,135,81,0.12) 0%, transparent 30%),
                linear-gradient(315deg, rgba(255,193,7,0.05) 0%, transparent 30%);
            pointer-events: none;
        }

        /* ── Radial glow ── */
        .glow-center {
            position: fixed;
            top: 50%; left: 50%;
            transform: translate(-50%, -50%);
            width: 1100px; height: 1100px;
            background: radial-gradient(circle, rgba(0,135,81,0.14) 0%, rgba(0,60,30,0.06) 40%, transparent 70%);
            z-index: 2;
            pointer-events: none;
        }

        /* ── Top scan line ── */
        .scan-line {
            position: fixed;
            top: 0; left: 0; right: 0;
            height: 2px;
            background: linear-gradient(90deg, transparent, rgba(0,135,81,0.6), rgba(255,193,7,0.4), rgba(0,135,81,0.6), transparent);
            z-index: 3;
            animation: scan 4s ease-in-out infinite;
        }

        @keyframes scan {
            0%, 100% { opacity: 0.4; transform: scaleX(0.8); }
            50% { opacity: 1; transform: scaleX(1); }
        }

        /* ── Main layout ── */
        .wrapper {
            position: relative;
            z-index: 2;
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 40px 20px;
        }

        /* ── Header ── */
        .header {
            text-align: center;
            margin-bottom: 56px;
            animation: fadeUp 0.8s ease both;
        }

        .logo-ring {
            width: 110px; height: 110px;
            border-radius: 50%;
            border: 2px solid rgba(0,135,81,0.5);
            display: flex; align-items: center; justify-content: center;
            margin: 0 auto 24px;
            background: rgba(0,135,81,0.08);
            box-shadow: 0 0 40px rgba(0,135,81,0.25), inset 0 0 20px rgba(0,135,81,0.05);
            position: relative;
            animation: pulse-ring 3s ease-in-out infinite;
        }

        .logo-ring::before {
            content: '';
            position: absolute;
            inset: -8px;
            border-radius: 50%;
            border: 1px solid rgba(0,135,81,0.2);
            animation: spin-slow 12s linear infinite;
        }

        .logo-ring::after {
            content: '';
            position: absolute;
            inset: -16px;
            border-radius: 50%;
            border: 1px dashed rgba(255,193,7,0.15);
            animation: spin-slow 20s linear infinite reverse;
        }

        .logo-ring img {
            width: 72px; height: 72px;
            object-fit: contain;
            filter: drop-shadow(0 0 12px rgba(0,135,81,0.6));
        }

        .system-tag {
            display: inline-block;
            font-size: 0.65rem;
            font-weight: 700;
            letter-spacing: 0.2em;
            color: var(--gold);
            background: rgba(255,193,7,0.1);
            border: 1px solid rgba(255,193,7,0.25);
            border-radius: 20px;
            padding: 4px 14px;
            margin-bottom: 16px;
            text-transform: uppercase;
        }

        h1 {
            font-size: clamp(1.8rem, 5vw, 3rem);
            font-weight: 900;
            line-height: 1.1;
            letter-spacing: -0.02em;
            background: linear-gradient(135deg, #fff 0%, rgba(255,255,255,0.7) 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 12px;
        }

        h1 span {
            background: linear-gradient(135deg, var(--green-light), var(--gold));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .subtitle {
            font-size: 0.95rem;
            color: rgba(255,255,255,0.45);
            font-weight: 400;
            max-width: 480px;
            margin: 0 auto;
            line-height: 1.6;
        }

        /* ── Stats bar ── */
        .stats-bar {
            display: flex;
            gap: 32px;
            justify-content: center;
            flex-wrap: wrap;
            margin-bottom: 56px;
            animation: fadeUp 0.8s 0.15s ease both;
        }

        .stat {
            text-align: center;
        }

        .stat-val {
            font-size: 1.5rem;
            font-weight: 900;
            color: var(--green-light);
            display: block;
            line-height: 1;
        }

        .stat-label {
            font-size: 0.65rem;
            color: rgba(255,255,255,0.3);
            text-transform: uppercase;
            letter-spacing: 0.1em;
            margin-top: 4px;
        }

        .stat-divider {
            width: 1px;
            background: var(--border);
            align-self: stretch;
        }

        /* ── Cards grid ── */
        .cards-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 16px;
            width: 100%;
            max-width: 780px;
            animation: fadeUp 0.8s 0.25s ease both;
        }

        .nav-card {
            position: relative;
            background: var(--panel);
            border: 1px solid var(--border);
            border-radius: 20px;
            padding: 28px 24px;
            cursor: pointer;
            text-decoration: none;
            color: #fff;
            display: flex;
            flex-direction: column;
            gap: 12px;
            overflow: hidden;
            transition: transform 0.25s ease, border-color 0.25s ease, box-shadow 0.25s ease;
            backdrop-filter: blur(12px);
        }

        .nav-card::before {
            content: '';
            position: absolute;
            inset: 0;
            border-radius: 20px;
            opacity: 0;
            transition: opacity 0.25s ease;
        }

        .nav-card:hover {
            transform: translateY(-4px);
            box-shadow: 0 20px 60px rgba(0,0,0,0.4);
        }

        .nav-card:hover::before { opacity: 1; }

        /* Card colour themes */
        .card-vote {
            border-color: rgba(0,135,81,0.3);
        }
        .card-vote:hover {
            border-color: rgba(0,135,81,0.7);
            box-shadow: 0 20px 60px rgba(0,135,81,0.15);
        }
        .card-vote::before {
            background: linear-gradient(135deg, rgba(0,135,81,0.08), transparent);
        }

        .card-report {
            border-color: rgba(220,53,69,0.3);
        }
        .card-report:hover {
            border-color: rgba(220,53,69,0.7);
            box-shadow: 0 20px 60px rgba(220,53,69,0.15);
        }
        .card-report::before {
            background: linear-gradient(135deg, rgba(220,53,69,0.08), transparent);
        }

        .card-vdash {
            border-color: rgba(255,193,7,0.3);
        }
        .card-vdash:hover {
            border-color: rgba(255,193,7,0.7);
            box-shadow: 0 20px 60px rgba(255,193,7,0.12);
        }
        .card-vdash::before {
            background: linear-gradient(135deg, rgba(255,193,7,0.06), transparent);
        }

        .card-idash {
            border-color: rgba(13,202,240,0.3);
        }
        .card-idash:hover {
            border-color: rgba(13,202,240,0.7);
            box-shadow: 0 20px 60px rgba(13,202,240,0.12);
        }
        .card-idash::before {
            background: linear-gradient(135deg, rgba(13,202,240,0.06), transparent);
        }

        .card-icon {
            width: 48px; height: 48px;
            border-radius: 14px;
            display: flex; align-items: center; justify-content: center;
            font-size: 1.4rem;
            flex-shrink: 0;
        }

        .icon-vote   { background: rgba(0,135,81,0.15);  }
        .icon-report { background: rgba(220,53,69,0.15); }
        .icon-vdash  { background: rgba(255,193,7,0.12); }
        .icon-idash  { background: rgba(13,202,240,0.12);}

        .card-body { flex: 1; }

        .card-title {
            font-size: 1rem;
            font-weight: 700;
            margin-bottom: 6px;
            display: flex; align-items: center; gap: 8px;
        }

        .card-badge {
            font-size: 0.55rem;
            font-weight: 700;
            letter-spacing: 0.08em;
            padding: 2px 8px;
            border-radius: 20px;
            text-transform: uppercase;
        }

        .badge-field    { background: rgba(0,135,81,0.2);  color: #00cc66; }
        .badge-security { background: rgba(220,53,69,0.2); color: #ff6b6b; }
        .badge-live     { background: rgba(255,193,7,0.2); color: #ffc107; }
        .badge-command  { background: rgba(13,202,240,0.2);color: #0dcaf0; }

        .card-desc {
            font-size: 0.78rem;
            color: rgba(255,255,255,0.4);
            line-height: 1.5;
        }

        .card-arrow {
            position: absolute;
            bottom: 24px; right: 24px;
            font-size: 1.1rem;
            opacity: 0.3;
            transition: opacity 0.2s, transform 0.2s;
        }

        .nav-card:hover .card-arrow {
            opacity: 0.8;
            transform: translate(3px, -3px);
        }

        /* ── Footer ── */
        .footer {
            margin-top: 48px;
            text-align: center;
            font-size: 0.7rem;
            color: rgba(255,255,255,0.2);
            animation: fadeUp 0.8s 0.4s ease both;
        }

        .footer strong { color: rgba(255,255,255,0.4); }

        /* ── Animations ── */
        @keyframes fadeUp {
            from { opacity: 0; transform: translateY(20px); }
            to   { opacity: 1; transform: translateY(0); }
        }

        @keyframes pulse-ring {
            0%, 100% { box-shadow: 0 0 40px rgba(0,135,81,0.25), inset 0 0 20px rgba(0,135,81,0.05); }
            50%       { box-shadow: 0 0 60px rgba(0,135,81,0.4),  inset 0 0 30px rgba(0,135,81,0.1); }
        }

        @keyframes spin-slow {
            from { transform: rotate(0deg); }
            to   { transform: rotate(360deg); }
        }

        /* ── Mobile ── */
        @media (max-width: 600px) {
            .cards-grid { grid-template-columns: 1fr; max-width: 400px; }
            .stats-bar { gap: 20px; }
            .stat-divider { display: none; }
            h1 { font-size: 1.7rem; }
        }
    </style>
</head>
<body>

<!-- ══════════════════════════════════════════════════════
     IYAM PREFACE SPLASH — shown once on first visit
     Press "Proceed" to dismiss and see the homepage
     ══════════════════════════════════════════════════════ -->
<div id="iyamSplash" style="
    position:fixed;inset:0;z-index:99999;
    background:linear-gradient(160deg,#0a1f0e 0%,#0d2b10 50%,#0a1f0e 100%);
    display:flex;flex-direction:column;align-items:center;justify-content:flex-start;
    overflow-y:auto;padding:40px 20px 60px;
">
  <!-- Gold top border -->
  <div style="position:fixed;top:0;left:0;right:0;height:5px;background:linear-gradient(90deg,#ffc107,#008751,#ffc107);z-index:100000;"></div>

  <!-- Logo -->
  <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAYGBgYHBgcICAcKCwoLCg8ODAwODxYQERAREBYiFRkVFRkVIh4kHhweJB42KiYmKjY+NDI0PkxERExfWl98fKcBBgYGBgcGBwgIBwoLCgsKDw4MDA4PFhAREBEQFiIVGRUVGRUiHiQeHB4kHjYqJiYqNj40MjQ+TERETF9aX3x8p//CABEIA1UDVAMBIgACEQEDEQH/xAAzAAEAAgMBAQAAAAAAAAAAAAAABAUBAgMGBwEBAAMBAQEAAAAAAAAAAAAAAAIDBAEFBv/aAAwDAQACEAMQAAAC9UAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA1hc5PUULkPVa+Ljcj7Xh5Aj6jl5xzl/pRnLtSC73oR6Hr5k76zv4x3vu+vz/ALJe6eQmdl6NUz+y7jsgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADlU8jdx/LQY1+krK1yvfR1jHksJVFlK9HIps8t29Sqn5zpfq5Ue10jKmXLim1u3VDp6FLnmePrE4+Oevj2w8wvo11dUlRr65VjRpPX2HgO0rPdPNW8rJw7MAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAr6CMPR0dOjTvokwjGzdWOS7z8+2ZL40kzXBHoA5d51RJEubiEgCHHurtEGbCWRCQADTd3lfBvmivyXP2ELVR5xZV+yiVd+YWve7+FvJW3zn0lYDoAAAAAAAAAAAAAAAAAAAAAAAAAABx89yF152vxCk2squVlhddvP0xJZi0BGQ4S53UETXR6p5y8qn3IFU4VZn0/qY/M49bTx7Mm11jg0qGz8zqp2mXEvvPKdPS+Xuh6fenuPN18KabSbs1zMp4Z6zPlPQ5rpQy3GMjn0d5U1Xq8a6fHr+o9DLpf+bXw99t4j0k77MSsAAAAAAAAAAAAAAAAAAAAAAAAHDnO9NV1kaevJKrrjWNpL87Vx7GHSNeNnPoBzuvlrzz/AKeOwsZuct/nLCzd4pbqm4h+l8v6i2A0xaN3Dt1VVNtVerh9WPJ3KC/85ro29DQX7tDy489+X1uKDv5+qDKrrnbntY0mj87XBs62R6WO86+Qusei1GLQxl3lZTes010+RW9T6WO09L4Xrd33aptp6A70AAAAAAAAAAAAAAAAAAAAAxQcjO8vyxDO2kX+ScG1y8vaFUwNfK+r8h6OTv6Pzul1fq3Dv5W6B571flPTx+u2rrHztTl18pbD1MSPaxl4/wBPVV/o5PXVdbFqnn1tZa02RvMewq+8zZ+RxdD0XnFrZCVN3gedq89f0Xrd2fz8b0flpx9BYRJnm69fKXlDuz3Vrpvg01/n/UeW9DL63pHkebrCEgEaSlzzEX2FP6WOnuafG+j3u/i/Vz0SRKYAAAAAAAAAAAAAAAAAADTXyUYd6ptDPraS7LzdOux5+o57mRzoDyXrfK+hlvKuVHnGH6Ty3S+v1tFZyfM1+Tt8wt+beBZWke7djzNiJLd5XSu6fArmDmvLuk03IkSW7ynuCfHmfTRbIVV/TWc+UHaunbsvokV5O6HT5s/Vx3OTx9wc6hb+a25/W7eX9NVPYZ7Yfn/WctdHk+vaJ62L2Fh4H1Ft1sJWgAAAAAAAAAAAAAAAAODyEa8xHaujX0PSR5OwMehjI8vNmee9fD7BW2XmbArk8v6jzW7NGs97iceFL6HGW7z3ocnRmqeHbpqjF2mZ0xi792iPLbddzXOUuYZdYxsNMdEO8dJKvsTSdjP2Cmc80o7fTJONWXiyHktPYa6qKG+2ZLwpsab0N1cTn129fC9NV2Xm69nn8dehc+mS7Wg9Di6Hj82lV7GD0114D0t1l0J3AAAAAAAAAAAAAAAOW3kOQ0iO1WfPo89vH3BluUl3SaqZk/yPopxl+a20vqj+lp41kfVuXXyNzz/oMWwrbMjIz2m4dZG3oV89843Qy58quydYmuScrWOzy7a81PdsYQ7lhxnOrrfbklzvtGXcmbQc3xnInXVHtq20R4cZrLKAlcfOs5jLMa95Aoelv6uLpQWkHvPQ0HKdztfr0s765sw8X0AhJS3S2Hjs2tT7Xn+ouPAervttBK0AAAAAAAAAAAABjPnuRiVbavNt6XSX5G0YyXKHXn6OW9742wafNwvW+Z9TFLvfP+hyX6+a9RxhKl9By6xkOsJ8u/bb1asZc9kOmkfngs7csPPmFMgAAAAAAAAAHTms5K6Qc7oTnDt6NekeYq7ASY/k28d9sU98npbTvVxcIM2lj2R6Xj0x37xKaFoq9FL8tiXPXIU3z9WKK+xOPj95cH2vP9nN8N7O6/sJWAAAAAAAAAAACPzkPynTlDM9BFu/M0B52vFD2gejkjW/aj0VewUl35e1GkoS8hbSdPRyWY8zYzvJ2w06sevTnTlwxT6czzLQr6AAAAAAAAAAAAAAzhLknrB39CEzXXp6dUXlP4+dZGxl5lnlbmdvqpecuvL3Q7ehxLqmrLNns8l6ry17vzTx5mzTzPqY+qny1hC09jB7/Pn/AEFuoOyAAAAAAAAAAx5C181ChLj+nxu2Tx95HrLYWm2+0JIM4eQsrDz3q4vW5ob7zdYVWO+3b1KsZcN0OkXV41oZZgAAAAAAAAAAAAAAAAAO/BdGdmHK9mrESdrBCba+LdE816/y/oZfT7Ulzjv2jaUF0OXpab0dteTXz9WzGedrvP8AsaD0cld7Hxsr08/tmm9mkHQAAAAAAAHDv5XkK3mmU57G0xnwvRCucLz/AK3lro85f0sPZn9erLPzNiJLR7XWJIlOvqVsIuiOeR4d4V9AAAAAAAAAAAAAAAAAAAAZwlyZvAl+xVvEmYt5Bjy9PEt8tH9hG3ZvMT7qRHum/KlzW2VFxn+jlmWvkvUZL+um7Hf5Pl6Hz3t+ff8AofA+01TlCVoAAAAAAAFf4+dBry59PV3vlaQwaoWnGl9LJ7BQ3eO/aqtke+P9BWSfSyXQ8nck6yPVqYRdcMaHhXjPGHfrrjD3kx9NbPHSyEnXl141xK2qthpHLHPQZ5AAAAAAAAAAAAAAAB1K6wJfsU7w5utnITOPDuxWyvMbc++LK40Vef8ASeUuqpw9bHz1sPXjytrzvoo99flrOtx7eD6Ar7CzUHegAAAAAKu08bGuDvpb56bXqeF6Qc6hzE4+Uz6ak9LJZZ827zb0cO3z2umk1bk4erXpyPAvZ69tcNNqmH6GSwgyY9HNO0qvzTk7R7CuyNpEzbX0T4EuS7OrjaY+gj1lz22MkR/J0hTIAAAAAAAAAAAAABnDvJm8KZ7lOkSfHp7wo7zTzpUPofKWeyjpR2Gt1cGwu+ua0csOnqOdoqv1nlfXwTfY/P8A2G7lgJ3gAAAAAV3kLSrrzberp7zydI449GlPCn+nj4ylfOPpe3kfR4tHV2Z7R1l3t0Y9/PpE218W5JzF219aNadzQ5PCryX2fTjH5Kyre9l1X7yJc1LwtszjWazYkHTEnhXLeKt5V4m0Mr0apWJkTytuBlmAAAAAAAAAAAAAA68lnJ+OXb6CiHpMh+NbD8/6xyvyd/Nhz5Mh08KyE2LMuLYd9zytqjvOF0PK2da9zzvfosqzWHegAAAI8ig5Hz+E2jNeSDwfTUl35zTT0v6e451rszW+d0t6H1sXrB5G3Mzl39qpG7Q6+pHPpW4VGtvZlZPK3cau66664k/HC1I04qZbtFPd9uSSTmKvjHg3u98KSxzjzrVfYEY1pQWnt4zvw8bcFMgAAAAAAAAAABnvMMusMhMhddkJUaTj1K4ONsfP3+cg3sr0sdPcSueS/oqqyXL2T5D0U4zxh0+dgej857Xn3npPCe52S2ErQAAAHifVeLhQv6L1nnz3HlbVPca2w8r6Kmhenj9g8yy3S+EP0t9cnOJGLX2y5e5Rw0dvDv7+es4HrYZfU8LcJNzPHGlvMZMcwAMRe1Dro9Fnzt/CW+2qmyVHxJ3wis48+eKi4jac9pGg2vo8jDxNQAABjXvNxzoAAAdM54Ww7qnnpzWnHh27zTHfYj9NtDv2rOHXpN6K99LvCPOheZoxDmUOfmsDrdehlorK6zkvjyDJeEe48p6yk256n1nk7f18vqRZpAAAA8/5+bCry2Hoa2y8Ta4d6+qfSZ5H0uiqTBnM11Jm6X1x5BntzNjyfXqQ5ESrqZG37yjtK6zz1Bj1dsdYu+AefYAAI04w5VT6DVTXuMzvJYw6W+mJclRpHLZDQYbKj0VR39zz+2O3HydwUyAYzVWxjc4WffxetzEl/PbmMay70YzWG3eOUeHty56Tunec87bUaHPfTnds65MYzjvNsYGJ8TbTGbHka74wokt4N3k/V+YuN+awHm6wAESXicfH9M8vf8z321dY26w70ABy61XI+VYk05fR9T5/1GMudoYXqfN+pi9B28tbZ7bNy65bxsSt2Po88bkfP6O9fZ0Xs4pcjGPE17dqyz1x15mWwK+gYZq7a+kLpcaascu/n6p5vufTncjPaBmRGlb64ow2a11nUb8foYs2F3QHnWDBz8zNh+3l5S7GfyXKTo83Rz216WNs6dc/FbnOzHrZFN+nPtpyZI66IxurR3pmO47447mvOVvLkLffjT2ftw7+vVE5yY3jXVXK5jdrlKiEehh+eaabS2oLuqyUMGmgrb+g9rz770XjfZbJhKwAB5z0fj4119pV3+OFkPF9CPSTo/o5a16GZKPnI3rsVz8f6LpI4378JfLunLrF9GvkPCvmeZ9P5D6Hzd4GnSvnP3Hm/TRvjDw9jCtshZKNfXL5ZtpcYRctsnz1nx20WmaTWuV6qLKqzsxmixIj9tUOevTnVJVWtbbmvI3blsnqPJvQplPphHus76DJnsa7Dj1ziXHLesty9bXGabBtGzHfMXXDrxxnDMKpAGsCyFi12jLt0i7a49JPPf0qkGfDzT5+U9ZHx887NnTbq4Uzdj0HPzNkPVMZos4eV9j5L0cuvvPA+y9OqcJ3gAY8J7Xw0KHqPMeu82zYeZsAAAAzNiy/XpxClw4dDzLZvk/Wec9/zuthl4e/MiNK1RimMNmtHY8d2azGS5kjJ5244bc9V6bxvpbYTGXm6+VVdctFXCZRXvOuvLv2dXXcoPqYZesaVOj1vPtwy+hgx5V/DON9fNs4Rnkc5hkOXTnPkfvmTKkM9mZGOGyGuDzrA50AO8h1643ZqH0FXL52WMOjMyF13QlR5HP0q4kCfx8O3ynXncezg6WfkuuW71Pkr2r528kVtli0POejpLq6n0/mLz2cnpBZpAAr/H+o8vXn7+q836TydAYdIAAAHeRx7e/RxjduPl2BknL8/e1Pv4Jzl18LalRZGvkfG+mWVfnfhrz2ZnFoYRZ8hSu3fTV5Xtceb25fYIczydyivvO6araZhmtzJjStXKfeWz951lvVac3ooUyHO9z6c8NmuzN8mWY9DnBg102xZxIiSeR2217y5iPtrV0M8gDDvM1sftso43SBCUC6rLaXMjHe21T5P1zn6PPAZx83o8l6Wjs/RyTae8YtHj3qaf0snS7qLfBpVdpChLzdnWS/d8/2ot1gAUPnbyjry2d9S3Xi7ESXAqsp/T+Z9NpprJfn7sm0V55A9Z059MeiX012+izxOe2vgXhVLryTPXz0k+nuMUW2rPpkR5PDZDnRX6hRdravvqsKyPdQly7ZZ7eVVdLIVdprWz5rrG9DdWGDT1dYu6AYLOHONdezg2j76eftxplyTJ3uTHGWANOscIdbuolW1BpoourzxftK5Rh4msw7zLAUW87bnxOYy3Z85dVGui+zy65L8sK5ZDszfl1+jzw9OvLw7qfpvG057pHkYdIR6AjyOc4+S68n0HmfQGu1usHQPL09rVV5Lq3q7Tw96ss6giei8xZ6aaf0fmPQWRl+U9DRc76vLPm65o+lzwsHzWgI9SY22iMLa0856eG4Yz4u/MmLvqjolRYdYyz9j1V7jRXB7a85xzJi8Ou8XtY9i2wx6Myca74aaGKbTernU9BHz7FHLGdPE2M4zf0OdziJAurt4tRy1UT4WuNVbDafNuvC2hLpcab+ffwHmXYpbKv25+WL/HEObjOe3FbY+Y0VXWKz0cuU/K8VyprLtTz5dIsrLdJ7cO/u0xuPfh5NsOk9OQ8dK4WfpY+1h5yPnt9g876LDpYyqn4/G2v0Pl+67RpNusO9A8nV2dZVkvrOss/D3qu0Ql5jn6trp8hr7HScfIyfRauSM4zg1zsZx9Hngj5rSHAEiJvM9jPRWMHbPVMIeLdP71MzVHfSRmUIzpzyTYygxkDfrfzhJxXa+SKvSVzRMy4+fRz58PQ+3hxE218zbjXOsJ5carTGbW8W7Nk1uqGOs6bOsZxvyWLbnZY7pPXjIz2RcZx586uRFm7c8kYdGTXqHrGlehm876fzsrRT6Jh4+/MOX5zTTNt+XWufeRHkexDhHkR/MsEPP3z3ovOa+xguqna0rlT+u59MOkM9vkufTn9B5ftpUSXdrDvQPJ1dtU15L6zqrXw94Z7QAIlfPo/Qy+ozhh0z8H0ueCPmdAc6A21S5MorOV7WWn68eOXmizi1+l2kU27l3vTSpU2Cud5Yq7h3trGrMLN+cmXXPhKQc+bvX9L71cOOLnk2MZj4rYUffT2kiNH1tdt+eex15T+/YVGZsTtfPHW4qQbTrnz78YzrXKRIjdJQ56yOFHa3eXTX03qh7Q7b11hXxlYZ59K5QPPeqhbc0qVBnZNDznWfppn4ywae8jj296jhH78PLtr/AD3qsTpo7WeqnjJntOXUA8jptr9B5ftpUeRdrDvQPMU19Q15bm4pLvxNo1z27OHOfJaBzlGX5uzqPQy+wYz5W2dnXb6TPBxvp89oCvoADOEuS4jv6lNBMuK2/N04QemG7ptv3q01uLZ22D0la1w5deMaVNjHgTNOeHYTovoxnw6vjTss+Vb0qn3j54390hb8bYayo2ZQsdq7KUuVWyqrJOusd2TmFvKEnaLtBYyqPannotufTzUiPnvqhFHnWYgWGJxpLnnVaarxWdKbJ6rknap2l6KYV3jOa7G+suzu+ddvcpjcevLwbyorrKvUPO9uLzSsR7S+u8r6zRW125YNPkzP0Hl+664zdrB0Ci856ryteax9B5r0vkakKbrlu8g9T09DL5Pp6vMZeX6ekQ7rsYtEvpw7/QUReXfh49wZpAAAAZ7cF8ZvDjvurix7XfRRRL6nsjw6Qsdtte1XiUriHCz27bTKMsYHdmOHEnlx6R7yxL7yjVYttpcqOXooUuVc7t2jKJW3Ndxw2azr3zx24kduOzt/3pbrxjfRX2RF6d9UIjfTBYxlFw1k4nGPwsEuQ5eUJYz076OY5c+d6w2w9SmJpnHzmiJDt1kKLj6NbDy3H19VoqrvTef9BRYiyoGe3zvbjN97zfZi3YABC8Z7zwcKOvrPHeu8u3cedrYefursYVZL9HL1mwYTnrc0l35uztJhzPVhyizYWWQefYAAANetnDjfybpX8dPbDhFaXTGmt/ZkCxq5yk9NekXPG2nI5216d7E1lxTPaHKc5w5dbKvr0hSnNu6NyztJrJfVjvGl8lz4SeZzh9ZPY1u87g5yziTyKxl1PnZbflHl59ON+JZM2i7X178+u/EZKcRdpGDn158ZpEXVXN30kOd9duXq0xcZi/PXd9/Hyd2f06kl57bCns+se01yc6qbaishV21Tf+3i9CLNQADw/uPJxqq/S+avMHLUeRvheeu671cV3JPM2IsrDnkfWeU9L6OWTOgyo6OsKbG0x4jxbhyk66wuG3k7lFzr7055ze0zvmXdHQc3TBrGk8+d5w9JTvKVDxyVjrFyh2xplzvG03M8tuDnTTvY1889I49r+O0ffiRHk95KvpZ8itxP4uceuecZz9eOXcWVZcZ4TNtXj9rtbWNtyd8VEvkpGeinVpvrkNTu/PfPOctumetN9YXapsjz/pfRo2jSYXLtddnjWwYltroq8/F9fi+vx/e/jWwsdzy9jzPpfI782vqvK+29WiUJ6AAFBf1/IePsa7rno9YPA9OJ531vnPQy33XyUk9HVVXKyGfWV1pRa7cc0XTue736IEer083RO4677e42zmfWWxjbAzjXmddOOvHbTTBnXGDi3idjt3iTY2XO0jPiVx893HOuta25U77WXtw5Vd1Syjy7abQnrykSOoXaTI5OHttoYzH5I9ufLdHtvpslm7pr3HzuPGmzh3m0OVm6FPtb6ac0Tvz4E/NXh205wd3OkaZIKmbMxCzrJ5dfS7pDkR/OsUkqklTzx6HpdXQ23Cqlz2CrtPL2BXOL5i7pPXwdfdeT9bv6ErQAGuxzwWlnWVZPUSai38L0WMqp1ka8aKqifI4x72509bfX6zatsst8rrEl+1VT1XqPNXWZ6a7Rs220x3nTHPU6a8sc7vrjHDGNO82wkzhHmyN76VfYYlCP367R7ttpjzLOjjpV2Tw163xj986elXjz3oKTkuDGaru3bj3d07c+EZWSs27KdE1yjyz059hjLpxve1Fz57I8mwAAOssO8znV1lhxnBwzjvbzvlz9+iNqfOaPOzIMv08dyPK2qi3rr6qn0/mPT31jXFo89B219/zPR3sSXfpDsgAAKXzPufDwzyvTeP8AVeXf2HnakWDTb81jX9puzPwtoMamz0jyV9lusJsLtOyTUW/P2qfOYmw426uea7NtWDOGhnXVKOWedtdh1g9bqu+OOp32i7cdZEPXvLXWFnvOnPGDaVX9upfLlrxvG64drG+ue/ftG6873247clvowYzrnvNdMEdu/PtOM2Zw7Tp3zrtisyy8+zAw2BHoAAACbwk+vSiyIQMeRdW0XsKn0Mulj5rndX6igjWHeSrfXbzdivsPP2QrpMa99rD6MW6wAAAHkPX08a/L3VL3zVeqYz4foxPNew01U+Xk+i5WRqa3tvro4eizX5LrhEl49E3aNJ96mJGsoGuuuh3WvJUeLvXnafNs6rte+tkOfOZjrl077uccTNSAliKlYOTt0ciJGvXFJwRdu2TjiTghQbqurnH032qt59uPLnZqEJPLHYznKXFjHsbanXXohnZnnWTncYyq7q2eXdqPOsCPQBIuj12ae/Rw5Hz+iLQ+lqNOebN8hNnH0PDPfFp49iPQj3Tyd3Reriez8v7X0IBO8AAABpuc8HzvKOrL6Od5z0fi7gzXRPPdY/s+fN61kzqJIupue2HM4cMd86ZD3t7M4ds+7RV4kcrq+WOuOuXHvy7zj17bO8nVxjO7jDbJxdHXPHUcuuN+OWOo5t8nJ0x1q3cc9O+vUKFb6O0+bLMZVWbNyVd0n9pRhy+2znN0y5rvnPDJzpnBgBgba7KJaNtfC0BRLaZrv7tCJ1jZZGvnMT0ry2NFXo6SLjRVra1OL6/Xb1dp4voCFztJGb+95vobzl1u0h2QAAAAEbxPv/LQpqPUeXsMfPQjxvQrYPoGmmNJRKpyqmu4+hmxMnWkWR5uyR3gS/Yp2rLXl6VVftjlbDhL06OZxtjncZxsZzlxgO4bYMM5NNmQyNW2DGN8GrYa6defeNeg4uupw6dHWu2dudxscY3068YZO4yAGocwHdgNcotZGsnBa1zFt5pg8C/h571GNFXj50+m9HJfZprvJoZlsl2MkJPN3PmfQyrum9t6dEgWaQAAAAAEWU5zwGLinry+nlea9L4u5jLNdRVlvK9XFW3nVh0BRaA21S5N2hy/doiQbqNsrisJxy1Gdue50HOsZwZxkM42MY2wM4zw131DIwyNdemnTONhrvqYzrkyy50HHTTd3DLhjIwyNB3mGcO5A3xK4ZcqJ6cDwbwpkA03d5SVXsI+7PBtuPbLcIEe1MNt7vm2/p48i/QHZgAAAAAAcPE+9oI1ee9B5/rlr9Y038T0Q50AAAB15LOTsxpPvUQ4lvE01wmiyOe0KYdhDrXbXrO2u3DOMmddsc7jOMmcZGGNjDOBrtg1302czjbHO4yGGQxkN9c87lgZYcZYGMbYGu2vTdKMbNISxEzp4d2zTfNIOdAAAA18vY0/q4l9Ue29CvcT0AAAAAAAANdh4mL7Hx9eWyvvHei8zRPHnawABjvM6U9Xtz+n7eQ2th6/tRW+XROzElevXHrbvlqhQWMaVKG4rk121G+mxnOHO7A1zgbDjXIZAxkaZw63xnHGNd8jBxnGcGc4zzoAADXbBjpv2kxly53MXHPxLs+d157Mkr0PDvg0hRaAAAjd/M6aY+Fr7WG2tSzUHegAAAAAAAAPOej15HwPSTCqy+r7eZ9L4voZGe0DHn59B6WRP63so10a6Y9Hj7hUejk9fvx38nfN2gyfYq6R5LZXAS48WmM45zXbA3HO5zjPGMZwZzjI121Ns65dDjXG2vebDjLXV3cGcBs5dOMgMDLHbrn32zaNY1PekdjxbuXnHD0sXf0ekjFoU2lbfVf2FfYY9AU2ACunGJVM+553X2kOyvuCVgAAAAAAAAAAEbxnvK2NXkbOtxRV7FU23iegFc/OxJXP2/O9LseL6Ac7B85aVns+f6WXy6+TuOGvU+RA21xnY5dvVq4x52O8r8S48DbTeHWcZ4a7YMZ1y7tg4xnGTLGRptgywMsZDGeGM4OfQN2OnWm/ba7muzWbblx5+bZnHGjyLHSjttdEzzvsK+mdPIr8ehl29Fid52trjz1M5dv5T0t9cg44tGnmenH2fPX0P1uvuRO8AAAAAAAAAAAACj839A85CmjvqDOeHsFfYeL6FVSev876GW4l+Qmx76KqrYsuLHh6Scd6i38llu5p1ztz0N9TRIy9dvG7eZsldYG/oQmOPX0K9Och1DxM1q7FdtK3NnWPdmM8DBkDGcGM4yMZGW2/XLEne3kfp2WcxlpNvrw44Z9uJ5lmKyJX7c2dLWr2Z/QWHnvQ+VtDNdWUPsar0MsC/wDJ7aKpPHb0fGJDTytuPN5i+riS+fsdsOnQs0A6AAAAAAAAAAAAAxkeXp/f+XhRU+g89nPH2GK2z8bfU13p19fmJ1y612U9Flx5e+1vr8/6fykjZn9D5X1dRRbItYM3LdH7eZ300+mzhh09u0NrjPQeuyElx6ao7a7LOc9eyLhiQijpDnY+e7vOG3V1ptlIY046I/PNKVzjMc99DFNFp4G7L7FSXWPRG2kOdUt1r3nkPQ0uvrYfWOfTxfQDnaeu9Tpsz69XLLdnzusb1cTq9ftrzLJ6Q70AAAAAAAAAAAAAABjI8zS+/wDPwooLinUQ9hnz195G/cUWPIev8v6GWVfeY9LCUCg9h5u2Ha+8l62uSHMoKbK28pfV7s3QeTuhR4MX1MXq9ta3BqtHn7yXOu2iE+u3BbGRmMsSkV1JRnEjXig6a6qukGBZC95+a4aKfQxKmRdC+lQZ3m6/NRpb1sMG060x7DNBfeVtyKbIfm/YUXoZcX3jvSE0efqFfOPfzumnr4Xff12qOksnoDvQAAAAAAAAAAAAAAAAAKfzPvoMavGyMcKqfUSfIXXma7WJKzjv8j19HT+pjxX9pFsMei5dfL2aeTvaDdnsfQQpuLQ49qaPai1qfUelkk+e9D5XLdwu6SRvzepY5+J6Pmbajv8A1cVkPI3CHOMep4WnrYazpeVpMtvHenyXwInpEJUVX7DyOyi6taC/xaAy3ef4SoHtef6yk0rqLNbCHz2Uewz530Hj79tN1M6a32WRY4UFkJlSetiWMz0V0+fUncDoAAAAAAAAAAAAAAAAAAAHLy/rcch4B6bzleeTe+W2y2evU1t5e3cVTAo6q+ofZ8/1+1Na+Zs38ra0m/NL9NW2WS/XyPpfM6qZsL0lDbXf61srLfSeo8x66yGw8zYrbLSyHkb6o4exg9fr5PvkvvZNRcZrthnteb9JTa6K71XjvWXV9R52ukqbmm9rz5NtY5wacVNups8nfTc3ViFntmVNdG9HJtqst+eF6eb3neErQAAAAAAAAAAAAAAAAAAAAAAEeQ5zyVZ9Aq40+T69I0Kr2z8f2w6PVquy8/VmmunXkNfWwvQy+fl2s+Pc5PN11lDbV3sYfUVVzx83X5abBx7PnyPU+f8AQeZsDFoEaXO8CJbaKqqJ6ZKPjrSNr6Gb1A8P0UGdpZHyPofP2vq4bseN6FTSenehl7dDBpEXvJUangbs86Az6GXHW29BZOsuCdwdkAAAAAAAAAAAAAAAAAAAAAAAAABrTXbkfC8fe08afNdOsaFdvZ+Vzku9g83Y4tNm59M1oc7Q8vRtlAY7/OwPTVHr4ZVxCm+fqCixw7pc8fKtab2MFzwp+ke87LN1XPceXsAppU9oqGKLModffXdQKPnszzIbbZRrm2vrJUHoJSdwdmAAAAAAAAAAAAAAAAAAAAAAAAAAAAABistHOeRrvfx41eHeiquVw5cNXy5m+ZZ7fYZ8hJz2+mUUmiy0Qu1U+7XaEg50AO8GDLnxlyUro9sLl56LdX6WHQNFVlC5NNR1s7OU/f09jKygue6VodkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABHrbpyPlIHusch4B7aHyHlV9F5GrS4/I6b6OO20dHsnMVxJ14OuvPCXDPbrgnynaZ6SZ2XkJvr9+z87ZWCU9djswAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOfGU5yFpYHK1ZCu2niJ07nddjoHQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP/8QAAv/aAAwDAQACAAMAAAAhAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAODftu9MtuMqCPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPSdZ28xDQzQgQ04wfoFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOecvBiCCGJCCHCCCCDg22shJAAAAAAAAAAAAAAAAAAAAAAAAAAAAJ20zjCEGACAsJcxAC5msCKTw4vrAAAAAAAAAAAAAAAAAAAAAAAAAHXYKAFJCdgACjFE7BTIxRuUBuCBU2LKAAAAAAAAAAAAAAAAAAAAAANEjRCCpXAZJRCBdxFs4v1NRkg+pCCS3kmAAAAAAAAAAAAAAAAAAAAW4zhDCCAhsblXiThABgDDARg3HcBCZkCC0sAAAAAAAAAAAAAAAAAAbYRCSKICVT3QFFSiADyxiQHDKwlyCTWKILUkjAAAAAAAAAAAAAAAAAYSCQuXJLCgF7VFaiDyxgjRzKbTLCF2/ZoBCCsMAAAAAAAAAAAAAAQ4wFKB7XhIC4FYzCCCCCCCCCCCAUAAIS6OrdWJh/qAAAAAAAAAAAAHkhC7/KCyCWnTgCCCCCCCCCCCCCCCRSZBLaMYTDCU3BAAAAAAAAAAA0hBG5BIhACmDCCCCCCCCCCCCCCCCCCCBuXLgvQHBKXpJAAAAAAAAABYLCyzBTghTiCCCCCCCCCCCCCCCCCCCCCRIQTQqLSgD8EAAAAAAAAA97CgGzKCCJiFKmkRsKCCCCCCCCCCCCCCCCQqCDhLHhBE8KAAAAAAATvjCBOBaZAiSqJ3T/uhGPCCCCCCCCCCCCCCCDiFDC9jCCzrAAAAAAA2iAaduBEDhANlwDxR8Q2ZDCCCCCCCCCCCCCCCAQDj1fzhEhDAAAAAQ6gCqRsDYCCqBmfjjxFwiCtJCCCCCCCCCCCCLPJZQR/HCNC0tAAAAAH8CQKB1Un0CjWKiCC9CTmDVECCCCDCCCCENfnPr1LKBiygC04AAAAAcDTtRAgIaQyHZCCCCufCHDCZrCCCAKMIHUkxLVBhFCBGCCCV1qAAAHJiSjHAHDo5GJqCCx/u2CChAwoCGM+qhLjB0hKYIjpDJNB9MC2qAAAW4Cc2yxWCA5ryCa4Sfe28KBADjCGdmbpziA1JCCLASHGk5QOB2MAAA/uCCCCCTpAzgDGuNhNOx1iSOZeEEsV9BQDDDCCBLFDHTXqKQKGsAAA+rCCCCCKrC9JAc95Yo7/1nRQijTyoBHfoUgCCKKgkCRaLrT9jT+AAAdLApfRKrrCPrTpia7RToBHfCLoKVHPyb7gCNKTGf3PCqh1CCCFsJAAdgAr90j8DCCWJ1mTmzCjOAgDBFBwe3EfO6Cs2pujvWABpwjfKRcqAAfLRw03WWACCZHCeNGRhGsPD0h10Kt9RAT2V8/GVWNXjqLFRDrBfKAAdBCCCiXcLCCBveaBdKKB8HzrWlfV18YClA4sC+d3IaTjJ7jhLB3qAA1jFMOHLjiCCCRioP7AxECFq1bsp0vrpCxhvLmaKEDJpiCEOPYU+pAA/hR+oBBLICCCCCRAopC78b1F6FJYJLc/qjC7KRnBSypuCAAApQ/AAAcrCauzqCKCCCCCBU5rhE3gkpj7e0WGxpHeVR1+NLxlJATsCJgTsAAAA7CFiXREWCCHs8oEJY5I07YWxELh6l+Y3g0kOZyeBq5UBPx2jF8AAAVsCmhzAFduy0EEMWww/C5B1JyzYC/vjQhCDCfnMFn4DjQ5WJCfYAAABXpBBTboq6iHjOOvhstMgPP8AiiMEWEjF+wgggEYcww5Y1gkywdYAAAAGIAnNTPhBW61xTQ64qhZMiHpSs6kSO0j3gQQggggowRlLmyEbwAAAAB7CZoGq0YeBs3Z5T4mcxyZffI934R55EwCFbiwgk5YOxQAAlqwAAAAEfowuZoDwSScW+n76+UZ+tvLgG34iQEHTNAnlAgMCRy2x4xfiwAAAAAM6Qsw1PwhQSJ2qp3gmupEtMlRX9cDi8JKK5iFSWQ1Z0EwV8wAAAAAAA/gUpIwggb0wghvAuiUf+8w5XBZkvwI08w/sAY4Qgg1AhvqQAAAAAAAALwQggggkoI+nvmNxepqWIFLa/SWbDTz0lIIcQQggggvOQAAAAAAAAAOSQgggWTSb8Zfvp7oSM8/cTNYS6+gwwfa5axQwgggj+AgAAAAAAAAAFP6ggvKwkwhA+rd/Ny4VxDgezvGDzTSeDx6HY6QghmEgAAAAAAAAAAAMfgwloQsQgDWd/YKUXHmuGhOOt7h0YGqNEpan6xl+AAAAAAAAAAAAAAMnApRXqU/1gYTQmrxGhMLeTEhExVc6N4wsjtw77gQAAAAAAAAAAAAAENKQ0kAiPOBNkYCiAEcQoYwgTYYhRgFqSwlw5+YwAAAAAAAAAAAAAAAEH7rAlbJiRpQhyAEI8AIk4wGdAGrYwgjghtqgwAAAAAAAAAAAAAAAAAAELoW7Zy1IhzRaSLAhXSbs1KAkA/BAYjfIIQAAAAAAAAAAAAAAAAAAAAFd+Rggv5zajM2Lgk/UwggrwtsUARV3QQAAAAAAAAAAAAAAAAAAAAAAAA1/bx8FKQcFhYghCkaAkMQvgh2vfowAAAAAAAAAAAAAAAAAAAAAAAAAAI+dewgwsAkgAgOAdAgkRSf+qwQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAh+DrRSxywggRwjQfbNuAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIBOMNbPr3XcP8AeMGAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABIPJNKEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD/xAAC/9oADAMBAAIAAwAAABDzzzzzzzzzzzzzzzzzzzzzjDDDCDDDDDTzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzCAM17x1424x0gQ0DDTzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzziAxPzrP/JGLCGDLDvR/sRDDzzzzzzzzzzzzzzzzzzzzzzzzzzzzziAY1riIAIIMYIIA0IIIBB/D28nDzzzzzzzzzzzzzzzzzzzzzzzzzyANTZsMIUI8GJ2FfGEK3e0IpBba8xDTzzzzzzzzzzzzzzzzzzzzzzgdR4IEMoKqjMcAcfEEEp4aFaMEJGbe8DTzzzzzzzzzzzzzzzzzzzzgq3YAIK8UmSsdhgaieLe2AZJKP6EIIBu4pDzzzzzzzzzzzzzzzzzyA9jCE4IIPYlbO0NPAAPAJOEM2cVAEJg8JOcFDzzzzzzzzzzzzzzzyB/rsJKwAJLqaEcjMGBLMILAc0/KfMJTkUwiTaJDzzzzzzzzzzzzzzg15AIeT3YgUAwvQF3BLELEKBEkKEsIUPkPsEICURDzzzzzzzzzzzzgNwgFxFE7OoXGBSEIIIIIIIIIILFYzEVBzBh+cjaQ1TzzzzzzzzzzwYSMLG4BCUI4EcEIIIIIIIIIIIIIIIFBkQEESAGwKX+rTzzzzzzzzyB+cE+oLuIQmBkIIIIIIIIIIIIIIIIIILGBAo6gowULI8nzzzzzzzzgBkkIdUMORp2IIIIIIIIIIIIIIIIIIIIIIEoZhqEoeMD6hTzzzzzzyDcsKcOLsJp+EcLvRswMIIIIIIIIIIIIIIIIKVMuoQKkEfwPzzzzzzhOiAIDoGssYJUnyVS2BLkIIIIIIIIIIIIIIIIElEQCJsMLW1TzzzzyDxMEZadIQdfDb5qD8bZpiG4IIIIIIIIIIIIIIINixNBSKMCsrzzzzhJsIcBEoPFZ5iHqrCECfAoLOMIIIIIIIIIIIIMEMdFlHAP0J4JTzzyhWcIeILBUeMWNfKIIKI0CwEX8IIII0IIIIYI4u1cSOVYqAIJOpDzzwDSJK9IOJstjkXAIIIHpQIaEK8kIIc+BQohFrJw58PZdqAIIIW4vzzxWsNK6UkaSoQcuEIIez/cIJAgaoIY5wRN0o4Elj3/ACRhhKK/1CcoU84ArCKz0nEEAbcuCf1y6Bb4KDiB0CKhim5Je90bCCLb0vUhSCGJ0/U8o11CCCCCTwAzipGxNgDly35Q1cuApgyubl/zjCCBTSD1BFwzNIDqQ8oWICCCCCAICMCQFI6Zce0koPCwSV8Y6F2qprCCFeseCGRAAwsiQpA8oXjXLyKqjLCNrCHAPuzRtSG6CoAHfFs7icDCNN2Pj1OCIoKACCStJ8A1DTloqw9oCVrBStDFDMXuLjU1XT7D+F4QvCvtop4M+PprByTLDfL8AWrTSVjBTCCCkNH0ENRnKtIlQBKu3RUtsxaQV+D7VMViiiCDnIBVo8A1JCCCRGdDCCD25U07JCwnlXZQPFsZbEkGObyNtUGGSArA0yBCB2D8IWqFAHIJCBCCCR61oRxD9EZeituUhAsVzRevBzuKF9ixJhIDBYS6h8o8gC/YhDoCCCCCCgShGmo5J0laXXxraMSvc/fJRGSy3INDhCSoX9A8oWKCbwtggjCCCCGLXFvCKvo0P+VWVaNejqquRnRPOqXaCHjFChCpE8oVrCUAXFzECGb712yASNlEvBXy831zzWCIVLSr8A7cZRCDtAqgYpU88T7CVsnjBOZSCYuZRB6Z/jzGmyVH/wA0r6EAkbpcXh1hoY19nQwtUFPPEdagcYWCqhZ5XZKgckvFxwlbifEIq9Z8jAggkA4IkURouQhiA3qvPPKHRAlqk+8FTdnGYXXKA5xe3M7SpEAfptS/hSgggggm1Rl96ax6gPPPLAqyZcjMywLbmfntIhJlOzlgnF8Dws+snZYugAAgkgqAJ6AQrsVPPPPAdyAp/AzpivYDOPl1BODbzmUxyX67Tv6NreFISBMWDRf1CgmKPPPPPKIqwkBiEwjNpMtWFVJrsHsK1PSx/JpF8ZfEWsjEjY4e1YoXNxPPPPPPAlC0owQgg75Dr/i1RPEa+Z/8l1OOgxcUtvZPuZ4QggEogf63PPPPPPKFeQggggggKYTHD5ZGDVbH/eTwL+mZjDnG3MYAQwggggh8BPPPPPPPPEN7wgggUriKtLXw1rxBzggywQan1b8sMgkASBawgggkuoHPPPPPPPPLE5aggp4cdgqE/A7Nya8zSOW5yPTSAzTqCUO+Y6ggguoFPPPPPPPPPPKM9AgkQwu4zSkx2aU5zovjyX+f+nwGQyzSsz6ilh19hPPPPPPPPPPPPCECyfYRs1veG5D4asmrS3YPJYvc5Go+UgAkenZ78YfPPPPPPPPPPPPPGP7EQUJEK2JsQYiSgUE8AUMTSeoXRggvSAtQ7fwnPPPPPPPPPPPPPPPCHP5gkIxAUYQkqyo04QIAk6Wbw17bApLghTOYnPPPPPPPPPPPPPPPPPCJfii54I7wgDA6D+ghT/7wPYQt2Pik4B/coXPPPPPPPPPPPPPPPPPPPCBP2xggnppADMHYwkbzMQhSAlok+RLeexPPPPPPPPPPPPPPPPPPPPPPLEkZzSsvboVMPAgglQ9AhJwtQwWXswTPPPPPPPPPPPPPPPPPPPPPPPPPCAPZ+QCAkglIggVdMAggQQNW5EBHPPPPPPPPPPPPPPPPPPPPPPPPPPPLCIEtP/CCSQAggRBxB/TcZ0hHPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPDAE5eeMv3bfjPe/6oIDHPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPDDCAE0IYAoQDDHPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPPP/EAE0RAAIBAwIDBQUEBwUECAYDAAECAwAEBQYREiExEyJBUWEQFDJCcVJygZEHFSAjYqHBMDNAgrEWJDSSNVBTVGNzstElNkNgotJEo8L/2gAIAQIBAT8A/wDsosB1NGeMeNG5HgtG5bwArt5K7aT7VdrJ9qu2k+1Qnk86Fw/kKFz5rQnjPpQdT0I/6nZ1Xqaa4HyimmkPjUk0cYLSSKo82O1XGpMNb/Fdqx8k71Ta5x67iK3lf67LUuu7g79nZoPqxNPrXLnoIV/y02rs4ek6j6IK/wBrM7/3r/8ABaGrc4P/AOSD/kFJrTMr1MTfVKi11ej+8tYm+hIqHXdqf760kX7pBq31XhZ9h7xwHycEVDd2043hnRx/CwNLI46NS3B+YUsqN0P/AFE8qL6mnndunKp7q3t0LzTIg82O1X2tMdButujTN59Fq81hlrjcRusK+SDn+Zqa6uZ24ppnc+bEn2xQyytwxxs58lG9S2N5CvFLbSoPNlI9sOBy8yB0spCp8SNqurC8tDtcW7x/UfsRyyRsGR2UjxB2qz1RmLXYC4Mijwk71WWubd9lu7coftJzFWeSsbxd7e4R/QHn+VLK6+NJOp68j/jnkVOtPM7egq+yljYJx3E6r5DqT9BWS1tPJullF2a/bbm1XN3c3Tl55ndvNjv7IYZZ5VjiQu7HYKKttEZGRA0sscXpzY1ltO3+MAeQB4/trVhaPe3cFunWRgKvbzHaZtIoYbcPKw/E+rGsbq60vXeG9hSIFTsSd1Pod6zQx4v5TYvxQnn02APiBWjcPDLx306hgjbRg+Y6msnrSdLl47OOPgQ7cTc96w2Zt89DNaXcCcYXcjwYeYrM484/IT2/VQd0PmprS+JsJcRJPdwI4Z2O7Doq0dO6avv+EuwreSuD/I1mtM3WMTtg4lh324hyI+vsIIqOSSJgyOysOhB2NY7WGRttkn2nT15NWN1BjcgAI5eGT/s35GkkZehpJlb0P+KJAqSfwX86ubqC2jaWeVUUdSxrLa0duKKwXhH/AGrdfwFTTzTyGSWRnY9STuaVWdgqqST0AqSN43KOpVh1BGx9mi7KKO0uL513fcqp8gtX+pMpcXLuly8aBu6inYAVeaqyN1Yi1cJzXaR9ubVpMqM5bb+TbflWuUYZG3bwMOw/A0iM7KqqSSdgBU1vPAQJYnQ/xAitN/8Ay13Pi4ZfzpuprRSMcuWA5LC29azIbM7DqIkBq3x8o08lnGQjtb7bnzapNI5yJxwIrc/iV6zbi0000Vy4aQxKn1atP2Hv2Ut4iN0B43+i1lLfTk9wLa77JJyoIPwn86zGkZbSJri1k7WIDcj5gPYCVIIJBrE6uvbThjuN54vX4hWPytlkI+O3lB816MPqKjmK8jzFKwYbg/4d3VBuaklZz6VmtS2eNBjUiWfwQeH1rI5S9yMpkuJSfJflH0Ht0IkTG9LIpdeDYkcxWawlpmYWntWUToSNx4kfK1TwSwSvFKhV1OxBrRM6S464tyeaOTt6MKyVo9nfXEDAjhc7eo8K09g0y0s6vIyKig7geJq4X9TZoiGQv2Eg5nlvV9aWWpcfFJBMA681PkT1VqxGj5LS7S4u54yIzxKq1rDKw3l1HBCwZId92HQsa0nnobMvaXLbRO26MegNX2jrS7mM9rdBFc7lduIfhVtb4rTVnI7zBpGHM/MxHgBUckmXzkbOOc043Hktazvpba3tIoZWRy5bdTsdlrTeTmyeLkR5iJ4+6X8fQ1mZ8g95LFeTM7xMV59K0XZCCynvZBtx8gf4VrLXpvshcXG/Jn7v0HIVoyaabGTpKSyK+y7+RFXqLHd3CL8KysB9Afbb3M9tKssMjI46EGsJq+Kfhgvtkk6CT5TSORsymo5Q/wBf8LJIEHrUknVnbYDqT4Vn9XHv21g3o03/AOtMzMxZiSSdyTRilEayFGCE7BtuRI9ug2/f3q/wLWOts1Bnrx4Ij7s07cfHyUgnqKz2AgykRZdkuFHdbz9DVheXmCyRLxkFe7Ih8RU11pbMoktxIiuB8x4GFTZ7B4i1eHHKrufs9N/MmpppJpZJZDuzsST6mre7ubZuKCZ4z5qdqny2TuFKy3crL5cVcLHwNcD/AGTUdxdwjaOWVB6EipJJJGLSMzHzJ3q1up7SdJ4X4XXoayWVu8lJG9wwJReEbDatNZaPG3xaYkROhV61EbLIZK3ms5lftwFbwIbpzrIWdzFgTaWKcTiMJ125eNfqvIduITayhydtuE0ix6f0+QxHaBD+MjUzFmLHqTufZprTi5BZJ7kMIdiE8CTWdwUuJlXdw8Tk8DePswWqLiwKwzkyW/8ANPpVrdQXUKTQSBkboRUU3Fybr/g5ZQg9annjiR5ZXCqo3LGtQamlv2aC3JS3H5v7AdiKx7Y7PYX3fs0RkUAoPkbwIrIWE9hdSW8y7FTyPgR5j2aGbbI3C+cP+hrUmpb+2vJrO3Cxhdu/1Y7isZn76xuzMZGlD/3isfirMZmbKzK8kUaBfhCjn+JqG3nncJFE7sfBQTVlovN3OxaJYV85DVr+jyAbG5vWb0QbVb6NwMPW2Mh83YmosJiIvgsIB/kFLa2ifDBEPooFcEP2UpoLZusUZ+oFSYnGS/HZQH/IKn0lgJutkqnzQlauv0fY99zBcyxn12YVeaDy8O5haOYeh2P86usfe2bbXFvJGfUEVYagyljsI7gsg+R+8Kj123B37EF/MNWXzV5lJQ0xARfhQdB7MLipcnepENwg5yN5Co5YoriDH2wAEacT7fKvgPqa1Zete5Ywx7sIu4oHi3jQ0dkzY9vy7Xr2PjtUkbxuyOpVgdiDyIrD5q6xcwaM8UZPfjPQ1jsjbZC3WaB9x4jxU+RqGbfut/gZJAg9annjijeWVwqqN2JrUOoZclKYoiVt1PIfa9T7NKjGtkQt4u7H+63+Hi9a1LpyS2uRNaRFopW+FR8LGsFh1wsT399P2bcHwb8gPXzNZSxs9Q41Z7ZgZFBMbf8A+TUsUkUjxyKVZTsQawuVOLunnEfGTGVA3261kb+bIXT3EoUM23IdOVYzAZPJMPd4DweMjclFY3QNlDs97KZm+wOS1b2dlZR8MMMcSjyAFXup8LZ7iS7VmHyp3jV3+kOEbi1smb1c7Vca6zcu/AYovur/AO9S6kzkvxX834Hb/SmymRb4ryc/V2r368/7zL/zmlyV+vw3cw/ztUWfzMXw5Cf8WJqDWuei6zpJ99RVr+kOTkLmyB9UarLWODutgZzEx8JBtX+63cX/ANOWNvowNZLROJu92hBt5PNfh/KsppLLY/ifs+2iHzx8/wAxRBFKvEwG45mojZ6cwgkDK8sg33Hzuf6CsDFcx4u4vWUvc3JLj/RRVljLLCp73eMJbuRuQ6nibwUVe5IY/Htc3OwfbuoPM9Fq7upbu5lnlO7u259mKytzjLkSwty+dPBhWNyVvkbZJ4W+8vip8jUMu/dPX+3dwi7mpJB3ndtgOZNam1A1/KbeBiLdD/zmgCSABuTWGwVpjrN7/JhdynJG+UH+pq5khN1I9uhjTjJQb8wK0xnhkIBbzsO3jH/OK1gckL/a4cmA84duS1p7OPi7nZyTA/xr5eorPZKHI37TxQhF2238W9TWPxl7kZhFbQs58T4D6msNoezteGW9Inl+z8gqe5srCDilkjhjUct9gKymv4U4ksIeM/8AaPyFX+dyt+T2905U/IDsv5D+0tMjfWThre4kjPoaxmv7hOFL+ESD7acmrH5jHZKPitp1fzXow+orMaTxmSDOE7Gb7af1FZjTuRxTnto+KLwlXmtCdm7JJXdo0PJd+gPXaptVYq2xsT25424eFIuhG3nWCt57p3zGSfwJiU9EXzrPZWbL34SIExqeGJB4+tY3RIaISX0xUn5F8PqaudF4+WJjZ3JDjzIYVeWc9ncPBOhV1NYfL3GLuhKhJQ8nTwYVZXkF5bxzwPurCopOMbHr/akgAmpJC7elatz/ABFrC2fkP75h/wCn2acw1ta2wyt8y7BeKMeAHnQnxepbGWIMQQeh5Mp8DWVxN1jLgxTLy+R/BhVtcTW08c0TFXQ7g1fZvEZLBFrrlL0CD4g48R7NP6Qucjwz3O8Vv/8Ak/0q0srLHW3ZwxpFGo5n+pNZvXNvbl4bACWTxkPwCr7I3l/KZLmdnb1PIfQf4CGeaCQSRSMjjoynY1hddyx8EORHGvTtR1H1qKayyFtxI0c0Tj6g1qDRGwe5xo9Wh/8A1oh4pNnTZlbmrDyrJ6omvcdFapEIvCUr0IHQCtE46J+2vZACUPAnp5mtR6iuLy4kggcpAhI5fPWPyNzY3KTQyEEEbjfkR5GtZ20c2Ptb0Ls4Kg/Rhv7NOZx8ZchJCTbyHZx5eoqKQMFkRgQQCCPEGkcOu4/tJ5PlH41qbNDHWnZxt+/lBC+g86ZizEkkkncmrPBZG8s5LqGLdEO23i30p7m7WD3VpHEatv2Z6A1ZXtxZXCTwOVZf5+hq0vMbqSwaGZQJQO8vip81rMYifF3Rik5qeaN5ilUsQACSTyArTOjQgS7yKbt1SE+Hq1ZHJ2WLtjLcOFUfCo6n0ArO6ovsq7ICYrfwjB6/e/wmKzN9i5hJbyED5kPwtWC1JZZeMBTwTgd6I/0rUWlbbKI00IEdyBybwf0NXVpcWk7wTxlHU7EGtEXsZhubJjsxPGvqCNjWWxN3YXUiPGxXiPA+3IitP6auLydJriMpApB5jYvWtclEVhsIiCVIZ/TyFKrOwVQST0AoggkGtH5zpj53/wDJJ/8ATUUnA3of7OR+BSau7qK1t5Z5W2VFJNZO/lyF5LcSH4jyHkPAezBaomxwSCYcdv8AzWr7D4rPW/vNrIolI5Ov+jCshjrrHzmG4jKnwPgR5irW6ntJ0mhcq6nkRV5eXWQujLMxd2OwA/0ArSulFtFS8vEBnI3RD8lZ3P2uIt+JyGlYdyPxNZLKXeSuWnuJCx8B4KPIf4aCeWCVJYnKOp3DDkRWmdWR5ELbXRCXIHI+D1qHT1tl7c8gs6juSf0NSR3uLvircUU0TVZa4i7NVvLZi4+ZKyOtmeMx2UJQn536j6Csdhcll5i4B4Sd3lfpVnjMLgY0eaRO1Ow7R+pPoK1diBbXIvIR+5mPPbwao5HjdXQkMpBBHgRWAyy5OwSQkdqndkHrUD7jhPh/ZTPxP6CtZ5ftJVsIm7qc5PVvL2YjBYbI4dUWUGfqzjkyt5bVlcJe4yTaVN4ye7IOhrHZO7x0wlgkI818G+tZ8w3+mxdugVuBHX0J9mjtMiNUyF2nePOFD4fxGs/nrfEWpc7NM3KOOr29ub65ee4kLOxqC3nuJBHDE7segUbmrHQ2YuAGm4IF/iO5/IVB+j6wQD3i9kY+gC0NE6eA5tKf89PobBOO5NMv+cVc/o85E2t8D6Ov9RWQ0xmLDdpLYsg+dO8P8CjsjBlYgg7gitJ6oF+i2l2wFwo7rfbFam07FlrYugAuYx3G8/Q1LFJDI8cilXUkMD4EVpTCY+/DzzvxtG39z/U1kNVWOOnW1gh4uBgr7DhCj0rV2ON1DDk7Yl0KDiA58vA1hVkyunJILpDyBRHPoORojYkVp3KnG5BGJ/dP3ZBUb/C6nl1FKQQCP7CZ+FPU1l79MfYTXDdQNlHmx6VLK8sjyOd2ZiSfU+y0vLmzmWWCRkceIrF6lsMnF7rfoiOw2PF8DUdGYv3gTCVxFvv2fh+daszUUwSwtSOyj+MjoSPAVo/T36wuPep0/wB3iPIfbaspkrbF2T3EpACjZV8WPgBWSyNzkrx7iZiWY8h4AeAFYLRVxeBZ70mGHqF+dqDYXB2+0axRKOrHqf6mn1Vc3hdcdatKF5F2IRKTM567u54ZJlthCrNKQu/CBV3k70tB7pl5ZjIduEjgYGlkyaSvbRZsvdopJiKbqSBuQCaxWe1BLAZhDFKgfg68DE+lWOrrOWQwXIMMoOxSXlWT01iMsheMCGYjcOnj9RWXwd/iZeCdN0J7sg+Fv8BFLJFIskbFWU7gjqCK0vqBMta8EhAuYx3x5j7QrWunhPEchbJ+8QfvVHzDzrC5STG30cwJ4DykXzWtTYRr4RZCwTtC6jjC+PkaxAu8Xp53u05xhmCE/L5VkdaGa1aG1tjEWXYsT0+m1JDLKGKRs3CN22G+w9mkMp73Ydg7byQcvqvhVu/Vf7CZuJz6VrXJdtdpZoe7Fzf7xqCGSeWOKNd3dgAPU1bYbB4a2SS/aN5T1L8/wUUkWlcqOzjWAt6Dgas/pmTGjt4WLwb/AIrXvt52fZ+8S8G23DxHasVjpslfQ20Y5seZ8h4mrW2tsdZJEmyRRJ1+nUmtS5uTL35Ee/YRnhiXz9a0xpSK0jS+v0Bl6pGeiVqHVsdpxQW+zy/yX61BZjNWXbz3MnvTzlEJ5pyG4Gw6UtlL+rpMbIywXMU/aAO3CJFI25GrSzkt7wpBBNdxy25juWQbjibrwn0p9O30S23uOMl4o5e0MkpUMfJdgalxlxa3M2RTHXZuHViI9lZVZhsTuDVq9p/uGMlt9/3bSEtxIRMee3hXFFIb3LZCAEFuyih6cTdP5Cv1tcYSa1aFne2niEghc7sgPgDVnf43O2JRwsiMNmU9Qa1JpybET8abvbOe4/l6H/AY7IT4+8iuYW2ZD+Y8Qax1/b5OxjuI9irrzHkfEGtWYM4y/Lxr+4mJKeh8VrDaou8bF2JQSxeAJ2Iq/wA7lc6Pdba2KoeqrzJ+prGaJ6SX8vr2a/1NXmZwOJge2gjRzsQY4/6mnKs7FV2BJIHlWAyJx+ShlJ7hPC/3TSP8LCgdwD+1I3ChNXtylrazzv0RC1XE73E8szndnYsfxrRsKSZgM3yRswrV00z5mZX34UChB6bUrMjBlJBB5EVhrt8pp2cXPeKq6Fj47DcGj1NaJwws7H3uVf3s43Hola5zfYwjHwt35BvKR4L5VorT6v8A/ErpO4p/cqfEj5q1ZqU24NrbN+9I5n7ApmLEkkknqa07HkruyurO1DpxyKxn6KgHXc1Pe4HEhEYtkrqMbBnO8aVdaxzM3KKRIE8FjUCnzWWc7tf3B/zmodQ5qE7pfzfi3F/rUGsXmAjydlDcp57bMKS2sb+z3xhjuETdvdJ+TIf4TWUuLue8ka6QpIO7wbbcIHgBWNyVzjrlZoW+8vgwqyubHPYwo4DJIuxB6g1msVNir6S3k5jqjfaX+1EMp6Ia92m+xRt5h8hrRuZfH3wtZiRDOdufyvWcxceUx01uQOLbeM+TCpYngmeORdmRiGB8xUWpsLYY+H3aAdoyAmJBtsfU1f5/LZR+zUsqHpFHVho/KXWzTAQIftc2/Ks/pu0xuPjlimLSB9n4iOYPkPZpe/8AfcVFxHd4u434Vbtuu3l+1cN0Wtb33ZWUVqp5ytu33V9mIyLY6/huANwDsw81NX2MxeobdJ4pQH25SL1+jCv9hsj2u3bw8G/xc6y01rgsJ7hC+8sikevPqxrTuLOTykEJHcB45Puir26hx9lLO+wSJN9vp0FWsVxns4A5JM0nE58lrN38GHxfDGAAiBY19fCpppJ5XlkYlmJJNYHCyZW6IJ4IIxxTSeAFZrPxiL9XYsdlaJyLDkZP2MLiLO4tZri7cqgICmsxhTZcMsTiSFujCra5ntZkmgkZHU7gigbXVVm24WLJxJ9BKBUsbxSPG6lWUkEHwIrTGXbH3yqzbQykBvQ+BrVmMXJ4k3Ea7zQAsPUeI/ZEUpQuEYqOrbcv2URnOyjerXFyzMFVGdvJRUOnDEge5nhgX1O5rs9OQ8mnmmP8Io3enV5e5TH6mu104/WG4j9aGLxdxsbXIqG8Fk5Vi3uGs4ln2MiDhJB3B28a13iewvEvY17k3J/vitO2NlfZFYLosAVJUA7bkeFXN9gcAOyWELJw7hVXdj9Sav8AWl/Putsiwr59Wqe4nuHLzSs7HxY7+zRd92ORa3Y92ZeX3lqFtnHr+1K3E7GtV3nvOYmAPdi2QfhWJgjuMlaRSFQjSrxb1qXTzWEhuLdd7dj/AMhq1vru0fjt53jPoaOrs2U4e3X73AN6nuJriRpJpGdz1JO9aExnu+Oe7de/OeX3BWv8psIMejde/J/QVoCwCQ3V+469xPoOZrWmRNxfi3B7sQ3P3jUcbySIiDdmIAHqazky4XFQYi3O0sih7px6+H7FnZz3cqxxISSazDR2WOgsEcF995NquOKPTMazHvM26A+y0uprS5inhYq6MCDWpbeG+srXNWygCUBJ1Hg49mksh79jEVzuyjgf8Kz1j7hlbu3A2UPuv3TzHtw+Lnyl9FbRDqe8fIUmnrCHFNYLEvAU2J8SfOr+1a0vJ4G6o5FWlqHieRhy6CnXhdh61BA0p9KsMJFHALi8fsYfAfM1XGbMcZjx8Ahi6ce3eNPLI47SRjIxPMsd6jB3fYbEjlXe3QufGu/xPxfDsaVAFXkPM1pjKTW+RWF5mMD93YnkCelZ7HDI4u5g23bh4k+8OYq1nks7uKUcnikB/KtX2yXePtMlFzAADH+Fv2LO4a2uoJl6o4aoZVkjjkU8mUMD9aU7qD+w52UmrmZYLeaVuiIzH8BU0jSyySMebMSfxoEggitNZaTIW7WN7Czjh2EhUkMPJqzGlL22ndrSJpYDzG3UelSwzQsVkjZD5MNqs7Z7q6ggT4pHCj8TUEUVnaRxjkkUYH4KKzF81/krm4J5O54fujkKwkIs9OWg6bx8Z/zd6ruV7u9nlAJLyE/ma0rhLr9ZR3VzbukMKmTdxtzFZS9e9yFzcMfjkJHoPD243GzX86og5fM3gKu7+zxEJtbLYykbPJWDs2v71pJtyid5iaz2RN1cmNOUSclHt0o4vLLKYp+YliLx+jCmUqxB6g7VoS7KXdxCTyIDj8K/SFbBL61nA/vIip+q+wAkgCtI4uLEY43dyQsso35+C1ldW9Y7bmftVeQteXJld++x5nzpgkUKwAdKkty9048AedWNlbY+1W9vF/8AJi86v8hcX0peVuXyqOgFRN3OhJDdB471aaeu5I+3uGS1gPMmWpJtIWbd+ae7cfY5LR1DpsclwpI82ahmNJz92XHzw+qtvSYjF327YzJqz7f3UvI1dWF3YuiTIYyDvueh9QaxN4t7j7eYHfddm+o5GtW2HuWauABskv7xf81WeYv5sScZDZmbqOLYnYGrTR+Xn2LokK/xnn+Qq20RYxANdXLP5gdwVqW2wMFlHHZNEJkk5hTxEj2aWuvecNb7nnHuh/CoDun0/YnOyfU1qu47DC3Gx5ybIPx9mCiwNnjY7i+MJmclgG7xA8OVXOtrKEcFnal/InuCoNdXgc9taxsu/wApINZ/UlnksekUMTLIZAW4gOQFaGsveMyJSO7Ahb8TyFatvvdMJckHZpNo1/zUvNh9aMK/q62gPwmIKfpwgVBYYzHx7xxRRADmx23/ADNXuTtJcPlWtplcxwkEr4Fh7Mfjbi/lMcIG4G5JpNK3vGON4wvid6vsha423NpZEFyNpJBVrbTXt0sa7lnNY02mLju7S6l4JGO2+3hT4GwuNzb5FCx8DV5g7+13Jj4k+0vMezSEpjz9n5MWU/iKzMYiyt8g6Cd/9a0i5XMxeqMK/SCgNjYSeUhH5j2aasoZrz3i4/uIO83qfAVlsvNezEBiIh8K0TSMwcHboaaQyz78IG58Kx1jE9xLcz8oIe83qfKsnkJL65aQ8kHJF8hUEEtxKkUSFnY7ACpXxumohxqtxkCNwvyx1kctf5GUvczs3kvRR9B7bTF3t3G7wxFlXrTpNby7EMjqfoaxeqiUFplU94tzy4jzdKwFolrDKLeYS2shEkLeW/UGv0hWXHa2t2BzRyjfRq03mGxs0wEDzGVQFRevEKyepdQxoD7n7sjdCVJNXOSv7okz3Mj+hPKsNgJcrHO6TonZ9QQSaZSrEHwNaEuO7eW58CrirY82H7FyfhFa7n2t7OH7TlvyH7X6PrXgsLm4I5yS8I+i1+kO771lag+BkP8AoKX4l+tX00iYhZ4viW34l+vDvV5kb28ctcTu/oTyrTv7zAagiHXsg3sxn+54G5ul+N+6DRvLk7/vn/OiSTWm1gsjFcz9ZX4UrWWNEkaXcQ+9SuyncEisRlb0XEUJYyIxAKnnWo7SK1yDCMbBgDt5VpNC+fsQPBifyFTYDEPdTTvaq7u5LFiTUVtZW+ywwxIfJQAa/SC22OsV85f9FoAkgCoVMFokI+repo770wG9cflVsBsWNXt8wsUtV5AuWc+fsjePT2JF26g3tyNoVPyL51NNJNK8sjlnY7knqT7QCSAKNxLiMNb8HKWVuI1mxFeYq3v+ACUnZvZovNT2t/HZO28Ex2APytWpbX3rCXqbcxGXH1XnWLn7DI2ku+3DKu/03rI5DH25hhvNgkwIBYbr9DWS0da3KmbHShSeYTfdD9DWkre7sMjdWtzEyF49xv0PCazMHYZS8j8pW2+h51o2fs8wqeEkbL/WoDtIP2Lg9/8ACtcy8WQt4/sQ/wCp9hBHUfsaUg7DA2Q25shc/wCY71rS47bPXA8I1VB7Ma63mnbQ9d4Ap/AcJq4iaGeWNuqOQfwNaLnT9YT2kh7lzAyVd27211NA42aNyp/A1Epl0sypzKvzoqw8DVhaPdXMcSjqavchjklSFrdyIe6pDbdKx1zb5fFPGFPQjY1kbKSzu5IXHQnatJRRNPMzxkuq7ofAVkriW4vJnkbc8RrRsYhkv8i/wW0DbfeNT5zKzsxa9l2J6Btq0VHLcZGaZ3ZuFNtyd+bGv0h3ANxY24PwozH8ax8XHOGI5Lzp23ot5e2LlGtXsbdkkm3InasBYi9ycMbDuL33+i1qTJHIZWeQH92h4Ix/CvtSNpGCqCSassXbY2EXd/tx7bpHUk11nb9EA2UcgB0UVqC6iihhsIT3Yh3vU+yynNvd28w+SVW/I0yrNbkHmrp/Iip4zBcyoeqSEfka1MPe9O2N0PDgJ/zDasbm8hjmHYykp4xtzU1i9WY684RPtBL/ABfD+BrVUlrLlnlt5VdXRSSp35jlWAl7LMWLf+KB+fKkOzr9f2Jj+8Nawk483MPsogrDR9plbJfOZK1vwC/tkVQNot+Q8zWbxFhaYKzuI4eGZ+zDNufFdzWj8daXgvTcQLIFC7b1cBPeZQg2XtDsPTesdH2NhaR/ZhQfkKzcply1+/nO/wDI+zQd6J8dcWbHvRNuPo1atsTa5V3A7s3eH18atLmS1uYZ4zs0bhh+FaptI7qK2zNsN4p1AlA+VxWMzU1gjxhFdG6q1W+Vx+R4re5t44uId11G2xqSK/wju0IVkbo+29Xd09zKZHUBvHYbVZZ6+soRHAVUfSiuVzZWTsw3CebAbVlLz9VWcFtAEEjJtIR1oCSaQAAszNsAPEmsyVw2Ct8UpHbz7SXH/t7NE482+PErjZpTxn6dBWqL733NXTqd1VuBfotWkYihHmeZp2J9gFRwySMAiEk+VYzTksvC0/dX7NZPAWlzaQxc0VG37tS2Fnh8Rkbi3Qh+wK8RO55+wAmkid2CqpJNWVpb4e296ugDMR+7Sr6+ub+dnck+QrTls0WOu7lUJkO6rV3b3vaO8sLgk9SDRBFCsRL22LsZPtQJ/pWoYuyzWQX/AMZj+fOoUzN7pmKCCCKSIgqeff7rVPbXFu5SaJ0byYbe2yfgu7Z/syof50p+E/sS/wB41anO+bvfvAfyrTKcebsh5MT+QrVGIyd5lBLBbM8YjUbitW2tzJjbGKGF3KsOIKCdtl2rSNvNa4/IPLEyNv0YbdFqJeO4QfakH8zQ5RfRauX47iZvN2P5n2aayn6tysMrHaNu5J9DWq8SMhYF4gC69+M0QQSCNiK05mYbftbC9HFZ3HJt/kJ8azuBnxk3Ev7y2fnFKOYIoHasfnZrdeymUSwnqrUf9m5Txkypv8te7aalACTyRkeJq4ydlYY9rawlZnY83p5JZn3ZmZifrWKx8GCtf1rkV/fEf7tAeu/mavr2e+upbiZt3dtzWExr5G/jiAPADvIfJRWbvo8NhHK7LIy8EQ9SP6Vaxmacb+e5phQQtyAJNWeAv7kgmPgXzarPStqhBmcuf5VDYWlsAIo1H0FJWanmt8VPPFw8SjlvUt/dXuk8pJM/EwlQfQbj2YO2t7awuL64jVwOSA0NSWynijx8St4Gr2/uLyUySvv6VpPGLc3XayDur038TWde+xF4ywSMsTniAHSodUXo5TIki+RFIcNliUEXYzt0I6b1eWklpcPC45qa0s/HgbA/+Ht+RrWCcOoLz14T+a1iNTX2OjSBVR4gfhPrWSusWkEBv1Tgl5DiXcb7b1PpTD36GSxuAhP2TxrWXwt1ipEWYowffhZfHakOzqfUVEeKOM+aih0Htk/vG+taj/6bvv8AzKx1/Nj7pLmJVLrvsG6c6j11ej47WI/QkUmvPt2P5PVxrWzmtZ4xbSqzxsoPI8yKsed5bf8Amp/qKk5Qt9w0/wAbfU+3RucS9tf1fcP+9jHcJ+Za1bp14ZXvLdO6ecijw9as7Nrp2UOFCrxMx8BVllrrEhrDIwdvZt8p8B5rU+mLPIIbjC3aSA8zA52dau8XkLNiLi1kT1K8q2NKjsdgpJqw0xmL0grbGNPF5O6Ks7PEYiQLDtf3/wD/AFxmr+4hSGW4yDiaSTkVI6fwoPAVDBJczrFDGSztsqitPYaDE2ReQji24pHNanzZy1+ShPYR7rGP61YQlY+Lbm1Y/C3F2wLdyP7RqxxFjaAcEYLeLGlUeAoA1tyoyKnxHatQZqOeJ7WIgj5iKwy+8YXN2g+Ls+0UfT2XRK6WtwPFufsijaSRUUbkkCo4bqzezS3KcMfN+8BuTWfsRkcV2q7F1G9EEEitNY8zXHvLHaOE7ms5dpd5CWRBy32FaQ/+X7L6N/6jWtv+n7j7if6VaWl1dShIIXc7+ArJ4Vcpa2cUshjEexbYc/h2oPp3T6nhcdrt4Hjc1qHOrlpIuCHgSPfbc7k70vxD61b/APDw/cX/AEpfhH09sn9431rUo2zd79/+n7GAsrO409fO1vG0q9oAxG5Hd3q1bguYD5SL/rR5xH1WpV4ZZB5MR7be4mtpo5oXKuh3UisFnbTOWvZS8K3Krs6efqKzWmLqyna7xu48WjH9Kt81b3MZt76NVk6EuOVPg0Ui4sLwoWO6AHz9RRzOqcdGTOUniUgbuA1LqC6kbZ9PW7v9yos7mGkeK1xNrA4AJ3XwNXSZe629/wAg5QkApF3VA8SenKpszYWNv2Vmqs5JO4BA/Go4slmbvuqXY/gqisBpu2xkXayEGTbvyGtWapF3xWNk20A5O4+erKznu54440J3YVaYywtYY0ECswA3JG9KpAGy7UNx1pXFB1rI5R7QdyB3PoOVZDMX905Dkov2RypN6wF6LPLRlz+7k3jf6NWfxzY7KXEO3cLcUfqrVjp7K8w/udxOImVuRNT6bXsHktbpZyvUCsKLG3maW6kKuu4C7VfTf7y5SYsCd96wefsbWwWK5nLE+FZcWUl4XtX3RzvttttVpFb4SykeWfd54u6lO3E7HzNaVjaPA2KsCDwE/ma1m2+oLr0CD+VYPUdni8UYzGzzGRiAOX5mshqnKXu6iTsYz8qcv50SWO5JJNSQyxHaSNl+oIpfiX61ANoYh5Iv+lL8I9s3941asThzlz6hT/KlUsQACSfAUuNyD/DaTH/IajwGYfpYy/iNq0xjb2zx17DcwlC5JXcjxXavhl+jVbP2trC/2o1P5ispH2eRvE+zO4/n+xBPNbypLFIUdTuGFYPWttcqtvktkk6CX5TWT0zjcmnaKBxEcpE61PpnOWDE2c5dR8u+x/I1c3ebRRHdWjFVcMd0I32+lDU14Nt7ZOL5jzG9Pk76aR+yjK8cfBsgJ5b71HjM/fbDspuH+LuisZoVmKvdy7/wJ/71Z46zsIgkcaqB4AVlrK8yQ7J7kwQeKJ8TfU1b6QxEPNkeQ/xGobK1gUJDCqD0FQwRx7MTvRHEd6deXSnQg0pau8w2KA1Lj7aYEPCp/CrrTcexaA8Pp4VOCk8g8QxqSJdR4hUBHv8Aar3fORKZXjdkYFWB2INWGSubKYSRt9RTLjcyhZCsNz5eBp9PZNXIEBNS4HKR7b27Hfy51a2lnjMaLi9tgZS3dU1lso+RmVygVVGyrWm8IchcGabu2kPelc9Dt4VbPG8ETRjZCgKj0rVEgkzt+fKTb8htVhpK6vrGK6juIxx791gal0bmU+FYn+jVDpvLpcwiWzcL2i7kbEbb1rp9msYh4Kxq3QvPCnm6j8zSdFoe2cfvK1tHw5ZG+3CtWNwLa8t5yCRHIrEfQ0+urQfBZyH6kCn15J8liv4vUmuMi24W3hH5mncs7MepJNaen7fC2D/+Co/LlWq4Owz16NvicOP8w/axufymNI93uDwfYbmtWX6QIHAW9syPNk5j8jUOpNN3I/4pUPkwK1F+p7rcxSxSfTY12VrHttEdvTakeDbux/nTXBPIcqLEnnTUa4iDQkLdajNcSjbc1Lws1IBvQCbUdhXLgNZy0NrkZ12PCx4l+hq0up7SdJ4XKup5Vc2WP1LGZrcpBkAO/GeSyVeWF3YzGK5haNh5ilZlO4JFLlb9VCi5k2+tR5zJR77XD1c313dkdtIz1iNL3Fyoub1vdrReZZuRYelX9+k0UePx6dlaqwAHjIfM1Egit40HREA/IVkZhPf3Uv25nP5mrXP5a0jSOG6YIvRdgRUWtMwm3F2T/Vaw+rrm+vYLV7VBxnmwJrW0vFlUT7EK1h4zLlLJPOZaQd5R6/sXI5qa15DzspvvKfZhsNcZW4Mad1F5u58KTAacxyA3JQnzlajiNMZNWS37EOB1ibYisti5sZeNBJzHVG8xWhLntcMYt+cUrD8Dzr9INrwZC2uAOUkWx+q/sxwyyHZI2Y+g3q101mLnYrbFQfFuVWmhJTs1zcgeiVa6YxFtsew7Q+b86hiih2CRIo9BtUwAYMOm1FgSdq3IrionehUccbdRUqQp4UGXwFFeLbauEDrWw35GuIiuKrmZYbWWRjyRCaS8t81G1vckLLuTDJ/Sp7WaxnaKdCD4GuDZhJE5Ug7giodRO8Qgydol1F5sO8KbG6RvOcV3NasflYbiv9kse3NM7bbetDTWBh53GcVvSMVFd6ex3/AY9p5R0llq+yV3esHupt1B7qDkorCobrL2cQ+ESBj9F51mLgWuLvZt9uGFtvrtsKtYDc3UMI6ySBfzNXOhZxube7VvRxtVzpnM2++9qXHmnerHXU+Jv0ne2JZAe6269ayuRbI3slyU4OIDu777bCtJw9rm7fyQM35CoRvIv7FwO4D5GtZ2/a4guBzikVvz5ezTbC001cXMa7v+8Y/VauLme5laWaRndjuSatLiW3uYpYmIdWBG1a3jV7KxnI2fi2/Mb1+j684L25tieUkYYfVa11ZdviBMBzgkDfgeR9ltaXNy4SGJnY+AFWGh7mRQ93KIx9kczVrpTCwAcUJkbzc1Ba2luNoreNPoPaQKY0XA2JJNFTxHhobmuCiVHLiG/lQBIq6yNxbXtnFHEWEjbMfIVJ3mrh2pOVOAdtxvXID4RTClFatmeDBz8PzkLVvctER5VaZe2u4Ft79eNPllHxLVzhLmJe1s5BPCfs9RRmIPDIhBHWn7FttthzoJB9qt4lLcgelNOPlWraxvr1gIomb16AVpnCQWk0kzTCSZV4Tt0Xetd3nY4gQg85pAPwXnSOyMGViGB5EVaZrUMCBo5ZnQfaUuKttc3SbC5tUb1UlTQ1ZgrqJhPEQdjydAwqRuJ3bYDck7CtCW+9zdzkfCgUfjVuO+T6fsSDdGFZS296x11D4tE231ojYkVo7IwGOfHTkbSblN/HcbEVd6Huu2Y208ZjJ5B9wRWK0dDaSpPdzCQodwo5LWr8xFeTx28DBo4d92HQtWEvfccpaXG+wWQBvoeRq7t47yzmgb4ZYyPzFWmhUjk3u7jiAPwpVnY2dlGEghVQPECid64a2okUTXOlRmPSmg5beI57VeApbTNxcO0RPF5ECnzWUDMBeSdfOmzGUYbG8l2+9Wl7ieXNQB5XbcHqSatLVmI4hyq5tYV2IQb+dcFdBQANEVwk0U5URWuZAmHRPtyj2RyPGd1NWWYnt2BSRkP8jS521ulAvLOOX+NeRowacm5rNNCfIjcV+qcOemVH4iv1dgY+b5Fm9FFC6wFtzhtHmYeMh5VeajuHXgRliT7EdaXtTBiIWYbPL+8b8a13f9vlVt1O6wJt/matL4S1Ns2SvQCg3KBumw6sabWeHSTs1ikKDlxBRtT22A1BAxi4O02+JRwuv1rJWE2PvJbeXqp5HzHgfZoy17HE9oRzlkLfgOVW47pPmf2XXZmFZ2090yt3Fty4yy/RudKzKQQSCOhFW+qc1AgQXPEB9oAmpstmsm3ZGaWTf5EG38hWO0XdzbPduIU+yObVkbT3K+nt+LcI+wPmK0nkRfYaAk7vEOzf8ACrhNm4vOt6G1F+VFq50FqGzlkI2FQWaxKN13NNBGeqVNaWs0TRSRAqw2Iq//AEbRzTySWt12aMdwhHSn/Rjf/JeRmsL+j17C6juZ7vdkO4CikQKABV0h7PfaiK25UNwa4qFEcqWMk1+kSG57Gy4YyYgSSQPH9gMR0JFC4mHz173N6UbqY+NNLI3VjWDxzZHKW1uB3S27/dHM1czxWVnLK2wSKMn8hV3cvdXM07nvSOWP41kY3Oj4RBuQIYy23l4+zTcs8eZtOyJ7z7MPNTWuez/WFvt8XY9786RGd1VRuSQBVhbC1sreAfJGoqNeFFH7NwuzA+da5stntrtR1HA3swekoryCK6uLjeNhuET+ppRa41DFYY55H/gXb82NXVpqvJcmeK1iPyK3P8SKzGlp8dZ+8mcS7MA4A22BrRGVFpkvd3baO45fRx0rhViA3SriweMbrzFEEdaO1bGooGcgAVBZAEFhUYUDYLTk0WNDmOdBBXSiNxQXapQHQipYypNCip3rY0o9kCd4DapoIpEKSIrKfAjesloXB3e7JGYXPilZXQGTteJ7YidPyapreeByksbIw6gjb9rQeJ7C0e+kXvzck+4K19leytorBG70vek+6K2NaZ1HBDALG9OydEc9Nj4GrrSeHvT2tvIY+Ln3CCtWmJxGn1e6lm4nAIDNtv8ARRWWyD5G+muG5Bjso8lFaXsve8vBuN1i77fhUa8TqP2pl3Q+lZ2x9+xlxCBu3DxJ95aIIJBrBanbF2skDQmUFt057bUdZ5W5kWO1tYwzHYDYsatZLuytWu8tegHb4AAFX05dTV9d5nUTmK0gdLXfqeQP1NTwz2N28b92SJ+o8x4itPZVcpjYZt/3gHDIPJhVtJxLwnwqWxhlO+21fqqI/OaXFxA9Sajto4l5LXBzoBVG9FlNBVrhFbgUCDXEKJFbrVzBxjdaaJgelHlXEPKgSaSJj0FQxhBuetOedE+zI4bHZGMpc26t67c6zP6PbmHiksH7RfsN1q4tbi2lMc8TI46gj24LFSZTIw26g8G+8h8lFO9vYWZY7JFDH+QUVlshJkb+e5f527o8l8BWmL7CJbyWl5GA8rc2cbqayujQQZsc4IPPsyf9DRe/sZGjLzQsDzG5WpZ5pjvLK7nzYk+zRVh2NjJdMO9M3L7q1br1b9uReFiK1RjjZZSQqNo5e+n49R7NPYays7OOeHaWWRAe0Pr4Cry0xyP71lLhX4fhRjsi/RfGr/WkUamLHwDlyDsNh+Aq4t8rfxz5GWNmUbcT7bVpPNnGZALI37ibZX9D4Go5NirqaSQMAaDDzpCNuKmkrjNcVb1xVx0zUHrjririoPRVH6imtI/OvdBSwRrW4HQVxUWrf9jNafsMtCUmjAf5ZB1FZ3T17h5ysqloye5IOhpVZ2CqCSTsAK0rgxi7AGQfv5QGk9PJa11nNyMbA/rMR/Jaw+Bu8qZTEyoqdWal0JeeN3EPwNYnCZfHEBcijxeMbKSKyGKschHwXEQJ8GHJh9DWVtYLS+nghlMiI23FVlavd3UMCDvO4FW1ulvBDDGNlRQo/CkXhUD9u4TkGrVWM9+xrOi7yw99fp4j2Wep8hZ48WkPCNidnPMgGgMhlLkD95PK341idI21qouMi6sw58HyL9azWrLVIntLGNHBBUsR3APQezRWoBcwCwuH/exj92T8y1bzdm2x+E0BxfSnYbADoKBo0T7N63o1vW/Kg1b0DXFtXHvRc7AUWrenP7d5ZW15A8M8QdGHMGsfou0sMs9yX4415xIfA1qLNx4mxZ9wZn5RL6+dTSyTSvLIxZ3YlifEmsLqK6xf7sKHhJ3KdD+Bq3ydhmrcpBdPFJ5A8LisvZamsCzrezyw/bVjuPqKOWyh63s//OaJJJJO5NaJxnOS/kX+CP8AqagTdt/L+wIBBFOmxKmtS4o4/IPwj91L3k/qKG243q2zWAxeMhe3Qcbpv2Y5vv8AxGsrn7/JMRI/BF4Rr09tvPLbzRzROVdGBUitO52HL2YbcCdBtKn9RVvckL2bHl4GvAVtTcl/YJrf2A+zegaNA0em+9bn2N19u9bit63FSyhRsOtZC/t7C1kuJ32RR+JPkKzOWuMrevcSnl0RfBV9scjxuHRirA8iDsaxOspogIr9e0T7Y+IVmbu3u8hNLbxKkZOygDbf1qxs5b27ht4x3nbb6CrO1jtLaG3jGyooAqNOBAP7GdNxxDqKz2KXJ2Dxcu0XvRn1qSN43ZHBDKSCD4EftY3I3ONu0uIG2Zeo8GHkaw2ZtctarNEdmHJ08VNW9ztsjn6Gu7Uuwj/GuMUHFFxRkFCQUZBQl50ZaEtCWjLQmozV2tdrzoy867Wu1rta7WhLRkNXd3BaQSTzyBUUbkmtQZ65zV2FQMIVO0UY8fU06PG7I6lWB2IPIj9rR2H93tzeyr+8lHc9FqBNzxHoP7OWPgb0Naxwmx/WECcjymA/9X7CqWIABJJ5CsZou6uEWS6k7FT0Xbdqn0JBwHsbtw/8QBFL+ttN5BX5qfzR1rB560y9uGjIWUDvxnqKinZOR6VPOCi7Gu1oS00tdrXa12tdrXa8q7Wu2oy8q7XrXbHbau1rtaMvOu1rta7WhLUak8zV9f2thbvPcSBEX8z6Cs5nr3OXSxRqwi4to4h4+prB4C2xMBvb0r2oXfc9IxWoMjb5C/aWCEIoG3F4v6n9nTWGbJXgZx+4iILnz9KjT4UUbDoKVQqgD+zdA6kVNEro8cigqwIIPiDWocK+MuzwgmBzvG39Pbo7CIU/WE67knaIH/WtRaqNrI1rZEGQcnk6hfQVBqbMwzCT3pn581bmDRFrqPC8XCAxB280cVBcXePu+OKRo5Y2I3HpWn9X2uQCQXJEVx+SvRG4pw6/ShLRloy0ZaEtGWhLRlrtaE3Ku1rtDtvXa12tGQ12tdrUSySnZR+NRQKnM8zWa1DY4mImRuKUjuxDqay2Zvcvc8c78t+4g5Ktad0/Bj4VuH2knZd+Icwo8lrUWbu7+5a34HjiRthH4k+tad0wkKLe36jfbdI26L6tWb/V/wCsZvcd+y3/AA38dvT22FjPfXUdvCu7MfyHmaxuPhx9pHbxDkBzPiT51DHwjc9T/azR8Q3HWsjj4L+1kt5l5Ecj4g+YrJ424x108Ew+63gw8x7LNzFpyJ4eq2e6/ULTszMzMdyTuT7NKWclph1Mo2MjGTY+ANZGRZb+6dOjTOR+JpcfflBItrNw9QwQ1g9aXdjwwXoaaEct/nWrDJWWQhEttMrr4jxH1FSWytzXkaljlj6ry8xRlrtaEtGWu1oy0ZaEtdrXa+tGWu1ozb1FHNMdkUmoMcBsZW3PkKllt7aIvI6Roo5knYVnNdAB4MaPQzH+grH4bKZqcysW4Se/M9ZDCYLG4mWOVwJWHdkPNyw8hWmNSG1ZbO6feEnZHPyU+Kx8t4l60KmULyPgfWtTajlunks7fdIVOznoWIrCYO4yk/IFYVPfk/oK1VjMRZQ26wMEnA24Bz4h5moYZJpUjjUs7HYAVp7Bx4u23YAzuO+39BUEW/eP9vNF8w/GsxiLfKWpjfk45o/iDV9Y3FjcPBOmzL+RHmK0hmongGPuGAYb9mT4g+FZTRUzTtJZSJwMd+BuW1YfRqwSrNeurlTuI16fjWqc/HaW72du4MzjZtvkWtHY2G8vpZZlDLCAQp8WNZPU1ljbsW0kLnkCSoGw3q/xGLztp7zalRIR3XXlufJhUc19jbtuzkeKWNiDsfEVidfEcMeQi3/8VP6irLJ2F+nHbXCSDyB5j6ipbOCXqux8xUuKkHOOQH0NSWt3H1ib8OdF2G4IIoy12tdrQloy12vKo4LqT4IWP4VDiblucjBf5mocZbR8yC59amntrWIvLIkaDxJAFZXXllBxJZIZn+2eSVkszkclJxXM5YeCDko+grA6VtTDHeXkiupUMqA938TVjf4+5MsNpIh7LYELyA+laos7q2ychmkd1fvRsfLy9ml9SGMpZXb9zpG58PQ1l9L2uRuYp1bszv8AvNvmFZbMWWCtFtbRF7Xh2VB8vq1SSXN5cl3LSSyN9SSa03p1cfGLicA3DD/kFRRcZ3PT/AzQ7d5azGGtspblJBtIPgk8QavrC7xtyYplKsDurDofUGrLWGUtkCPwTKPt9avNZZS4QpGEhB8V61YY29yk0wh77qpdix61gLyXC5QxXcbRrIAr7+Hka1HgkykAng27dF7p8HHlWkclaWPvFpdSGORpeQbkKz8sc+ZvGiO4Mmw28SOVTaPgTFrcNcNHKsPHJvzWoppoHDxSMjDoVOxqw1vmLXZZSs6D7fX8xVlr3FTbC4jkhP8AzCrbM4u6A7G8hYnw4gD+RopE45qrU9hZv1gWjibE/wD0yPxNfqax+y350MPYj5D+ZpcbZL0gX8edLDBH8MaL+FXGQsbYbzXMUf1YCrzW+Et9xG7zN5IP6mr/AF9kJt1tYUhHme81Xd/eXj8dxcPIf4jWL0XHNbJNc3B76AqqeG4rK4m5xlwYpV3U/A/gwr9ZXvuYtO3YQgk8FYfJyY2+jnX4d9nXzU1lrCDNYsGMgsV44WqSN43ZHUhlJBB8CPZZ6uvLbHPbleOQco5D4Cv95vLj5pJZG+pJNad03Hj1FxcANcEfglRxFz6UAANh/gpYfmX8qyWLtcjAYp0+63iprMYO7xcu0g4oye5IOh9mhJY1ubuMnvsilfwrWltcLkxOykxuihW+laMy9xKzWMu7KqcSN5AeFa2s4ob6GZBsZUJYeorTVh77loFI3RDxv9FrWl/2Fglsp70x5/dX2WGmsacJE15HtIUMjODsRvzqyx7ZC/8Adrc7cRbhLeQ89qymCyGLVHnC8LHYMp3qDJ5C327K7mT6Oaj1XqCLba9c/eANJrnOr1kib6pX+3mb8oP+SjrrOt0MQ+iVLq7UT9boqPRAKmzWWn37S+nI++atNOZi+VZFi2RxuHdqttCSHY3F2B6IKTTWnbEBrhg23jK+1ZdLVMjci2ZTFx7oV6bGra6eDT0NwgBMdqrbfdFRvjdRY0ggEHqPmRqzGHucXcFJBuh+B/Aj2aNzPA/uEzd1ucRPgfKtZ4cIwv4hyY7Sj18D7LKxub6dYYIyzH8h6msHp62xcYc7POR3n8vQVFEX5npQAA2H+Elh4ua9auLaKeN4powynkVIrOaRmtuKeyBki6lPmWra5ns7hJomKyIasNTYrJQCC+VEcjmH5oajkwGLjeWJ7eMEcypBJrP5Y5S+MqgiNRwxj0rRNkIrGW6I5ytsPurWpr/3zLTEHuR9xfwrB2Jvsnbw7brxcT/dWtXXwtMV2KHZpjwD7o61oeDjyM8vgkX82NZ2wF9i7iIDdgvEn3lqGJpJ44x1Zwv5mtUx28GDb90nH3EU7Df2Wls91cwwJ8UjhRS2eDwFojzKm/TjYcTMas9QYTIzC2C825KHQbGtW4KCz4Lu2UKjts6DoDVpq2+tLCG1iij7g2423NaUzN1kRdrcuGdCCOW3I1qyBoczPzOzgOPZjUNxpmJFG5a1Kj67EVp3TmTsp1uZZxEPGMcyw9avLayyMMttLwuPEA81NZrCXGLn4WBaJj3JKjd43V0JDKQQfUVkMvf5FgbiUkDoo5LWG05eZJg5Bjg8XPj9Kx2MtMdCIrePbzbxY1FBvzb/AA7xq4p42TqKzGmLLIcUiART/aA5H6islh7/ABzlZ4jw+Djmp9ulZ4psJAiEboCjDyNZTEX1lcyLJE5UseFwCQ1aNxMltDLdzIVeTkgI5ha1ZkPe8o6K26QjgH18a0LBw2l3N9qQL+QrDZYXWRylqx+CUlPu9DUuJ7HV0MYXuPL2q/TrWu59ra0h+05b8h7MXdizyFtcEckkBP0q/wAfYZyzi3k3X4kdDUmh7iJxJa3w4lO67jYg/UVmbXOWpC30kroTyYsWU+zR1z2OXVCeUqFf61ru252dwB5ofZpd98DaHyDfyNZTVeTumeNG7BNyNk6/nWOyt3YXQuInJPzgnkw9azOqbnJRGAQpHEf8xq0s7m7lEcETOx8qw+joYOGa+IkfwjHwio4xsERQAOgFRwheZ5n/ABJANPB4r+VTQxyI0csYZTyKsNxWU0XbzcUlk/ZN9g81q+xd/YOVuIGXybqp/GsZlrzGTdpA/I/Ep6NVprawkUC5heNvQcQrK60t+weOxVi7DbjYbBaJLEkncmtMxi20/G7DbdXkNYnJNa5iO5Y8mkPH9GPOntIZbq3uvmjVgp9GrW8/HkoovCOIfm3st7ea5mSGJeJ3OyipLTP4OBJu0aJGbbZW35+oq21nloiva9nKPHcbH+VX4iyOCld02D2/aAH5TtuPZjbj3a/tZt/glUmtWW/vGElcDcxlXHsxGpbLG4WKFgzzAv3BUziSV3224mJ2+tW9rcXMgSCJnY+CjesXomR+GS+k4R/2a9fxNWdja2cYjt4VQenU0kDHm3IUqhRsB/jGRWHMU9uR8POpYUdSkkYYHqGG9ZDR2Nud2g3gf05rV9pLLWu5SMTJ5p1/KpIpImKyIykeBG3sTVky4s2Puy7dj2YcH2af1BYyY6CO4uUSZBwEMdt9q1Fcrc5i7kVgy8WwI8gPZbzvbzRzIdmRgw+oq0ymJzdl2UxTdh34nOxB9KTS+n7dxK/MA7gO/KtR6ks1tHs7Nw7uvCzL8Kr7bjUeVntxbtOBHwBCABzFAEnYAmrLTmWvCClsUX7T90VYaItY9mu5jIfsLyWraztbVAkEKIvkopIHbryFJEqf9QFVbqKa3HymmidfCriytLleGeCNx/EoNXWjcRNuYw8R/hO4/nVxoW5XcwXaN6MCKn0pm4j/AMNxjzQg1Ji8jEe/ZzD/ACGmjkU7FGH1FbHyrY0N6/en7RqOwvpduC1lb6Iah01mpttrNh6tstW+h8g/99PFGPTdjVrojGx7GaWSU/8AKKtcVjrTbsLWNT57bmlRm6Cltz8xpY0XoP8AqVkRuoFG3U9CRRt38CDRikHy0VPitNFG3xIp+oprKzb4raI/5BX6tx3/AHOH/kFLYWK9LWEf5BSwQr8MSD6KBQXyWhHIflNC3c+QoWw8WpYkHy/9W7A+ArgT7IrgT7IrgT7IoKo8B/8AaH//xABCEQACAQMCAwYCBggFAwUBAAABAgMABBEFIRIxUQYQE0FhcSIyFCBCUoGRIzAzQGJyobEVQ1OSwTRQghYlNWDRY//aAAgBAwEBPwD/AOlPLGnzOB7mn1C2X7ZPsKbVU+zGT7mm1WY8kUUdRuj9oD8KN/dH/Mr6bdf6pr6ddf6lDUbofbB/Cl1Scc1U0urfei/I0mpWzcyV9xSTwv8ALIp/7PLcQxfO4FS6qo2jTPqakvbl+bkDoNqCySHADMfzqLSNQl5QEDq21R9nLpvnlRf60nZqL7dwx9gBS9nrAczIfxoaFpw/yj/uNf4Jpv8Aof1NHQ9NP+UfzNN2fsDyDj8afs3bn5ZnHvg1J2amHyTqfcYqXRNQj/yuIfwnNSQzRHDxsp9QRUd1cR/LIai1VhtIgPqKivLeXk+D0O3/AGKe+gi2zxN0FTahPJsp4R6VHFNM2ERnY9N6tuz11JgysIx+ZqDQbGLBZTIfWo4YohhI1Ueg73dEGXYKOpOKS5t5DhJkY9AR3vqdjG3C1wmahubecZilVvY/UZEcYZQR61Po9hNzi4T1Xarjs5Ku8Eob0bY1PaXNucSxMvr5VDeTxcmyOhqDUonwH+A/0oEEZB/fZ7qKAfEd+gqe/mlyAeFegq2s7m5bEUZbqfKrTs7EuGuH4j90bCooIYV4Y41Ueg7pJEjQu7BVHMmpe0VojYRHf15VZara3h4UJV/umrmdbeCSVuSrmreC71ed3klKop/L0Aq70Ke3VZLeRnII2Awa083X0VBcrhxWv38icNtGxBYZYirPs/E0KvcO3EwzgbYq/wBPl0x0nglbhzz8wasLoXVrHL5kYPuK1m+uUvkjgkZSFAwPMmhqur237eDI6lcf1Fafq8F43BwlJOh72RXBDKCOhq70G0myYv0benKrvS7u1yWTK/eXcVDdTQn4W26Gre/ilwG+Fv3pmVQSTgCrnUicrD/uqOKaeThRWdjVj2fVcPdHJ+4KSOONQqKFA8hRIUZJAFKysoZSCDyI7u0Nw7TxWynbAJ9SattIsooVVoldsbsag0W1huTMvFscqudhWuA/4dNjqK7NsDayjzElMQoJJwBzNJLHIMo6t7HNat/8x8XLKflQ5Cu0LKLEA8zIMV2fBGn5Pm7VLdIdVadwWUS5wOgpNd051PExHoVrTh4+rh4lwgct7CtUufo1nK4PxEcK+5qyl1aOLxYONowcEcxVhriTuIpk4HOwPkT3EAjBq+0K2ny0X6N/6GrqyuLR+GVCOh8jVtqEkWFf4l/qKiljlXiRsj93nnjhXic+wq5u5JzucL5CtP0ie7IZspH97r7Va2dvapwxIB1Pme/tIzgW4DEKc5FafqM+nyCOYHwmwcdM+YqORJUV0YFSNiK7RRMl1FKOTL/UVaTrPbRSqeajNarqJsUjKoGLHkaiP0/T8yKF8VDtVtPcaRdOkkZKnmOvqKvteWeBooI2BcYJNaDZSQQPJIMNJjA6Ctb02S44Z4Rl1GCOoq316eCMRTQFiu2eRqaW91e4RVjwo5DyHqaZUsdOYKf2cZ/Ouz9sk0s7yIGAXGCOtavZx2d4jKgMT78P9xVhHarbo8EYVXUGu0FwZLiO3T7PP3NWNuLe1ii6Lv712gjjju4mQAMy5OKgYtBEzcygJ/LvliimQpIgZT5GtR0J4syW2WXzTzFRSywvlSQfMVa3kc4xyfp+63V2kC9WPIVJJJNJlskmtM0L5ZboeoT/APaAAAAGBQdCxQMOIcx39pR+itz/ABGrubTpdMt1kf8ASiMcPDuR71pmpy2TgNloid16eoq5gt9StBwsCDurdDUcOtaezJErFT0HEKj0zUb+ZZLslV9eeOgFRosaKijCqMCpYYZlxJGrD1FR2NnE3EkCA9cd2RTRQv8ANGh9wDSqijCqAPSpoY5o2jkXKnmKtLKC0VliBAY5NavYteWwCAcanK1pIuLW0lSeMr4eSPUVa3EL6mJ7lsLxlq+mWvhmTxk4cZzmmL6pqgwDwlvyUUAAAO7V9WNqVihIMnNvQVpupJfIcKVdfmHdqWjRXQLx4SX+jVJFNbylXUq6mrO/EmEkOG8j1/c7u7WBcDdzyFASzygDLOxrS9HS2AllAaX+i990LvTNQ8XiZgxyGP2h0NWtzFcwrKh2P9D3doxm0iPSStI0i1lgjuJcuT9nyq80y2uYBHwhCvykDlVhYR2UZVXZs880WVRkkCnvIV5En2pr9vsoKa7nP2sUZpTzdvzribqaya4m6mhLIOTt+dLdTj7dLfyD5lBpL6I8wRSyI4+FgaudLsrnJeIBvvLsabs0OL4bkhfUVY6fBZIQgyx5seZ7tQvUs7dnO7HZR1NOjvFJdTEku2F9TWiW4t7ESNsX+Ik9KOvWYufC34OXieWaVldQykEHkRV/p8N7Hhhhx8rVdWs1pKY5BgjketWN9xYjkO/kf3G7ulgTqx5CgJZ5cAFnY1pWlpaIHcAykbnp6Du1k3gtCYDgD58c8VpGrJNCY53AdBzPmK1K/OoOttbR8Q4vmxuasrmfS7sxTAhD8w/5FI6uqspyCMg1qFkL2FYi/CAwOatbaO1gWJM4HWpJ44/mbfpUl852QYpndzuSaS2mfklJYH7T0tlCOeTQt4ByjFCKMfYX8q4E+6K8OP7i/lRghPONaazgPkRTaePsv+dPaTL9nPtXxIfMGo7yVOfxCoruKTbOD0PcTgE04n1bUeEgqin/AGrWpvC17FbKQsMOF/8A2ri8uNQbwIBwQINzy2Hmat7T6VdCGHJXO7enWoIUghSJBsox3XtlDeQlHG/2W8wau7Sa0mMcg9j1FWF7xgRyH4vI9f19xOsEZZvwFO8k8mTuxOwrSNLFqglkGZWH+0USACTV/qU93cLbWZPzbsPMioVcQosrBm4cMetazphtZDLGP0TH8jWg/QzbZiXEn28861XTVvIsrgSr8prTbSS0tljeQsefoPQVJIkYyxxU167bJsKVXdtgSaisWO8hx6Ckgij+VR+seNHGGUGpLBTujY9DUkUkZwy1FdSR7ZyOhqK4jlGx36UY1HGyKodhzpNFvZrt1lGBnLP1z0rUpYoVWwtB5/GRzY9K0yySxtuJ8ByMu1XfaEhylsgP8RqHtBdI4+kQgqegwagniniWSNsqav7GK8hKNsw+VuhqeCW2maNxhlNWV2Jk4W+cfrWYKpYnAFXVw08hP2RyFaHpmALqVd/sA/37tV1Caeb6FbA5Jwx60Y7zSLlHIG45+RHSrK9hvIg6Hf7S+YNSxJNG0brlWGCKttOvrTUsQ/J5seRXunu1jyq7tTO8jZYkmobJmwX2HSkjRBhVx+4FQRgjIqaxB3j29KIdG3yCKgveSyfnWzLseYqz0aO3u3nZ+P7meYrtFduvh26nAYcTVpOlRW8SyyKGlYZ3+zV1aw3MLRyKCCNj0rQJnjuprcnK7n8R3atpy3kJZRiVR8J6+lK0kEuRkMpq3nWaMOPxH6zUrrJ8FTsPmrR9PN3PxOP0abn1PSgAAAKn1K1gnSGR8Mw/Ae9LDAZPGVFLEY4h0q4t4riJo5FypqeC70i5EkZJQnY+RHQ1YX0d5DxrsRsw6HuubvOUjPuajjeRsKKgtUi35t1/dJYUlGGH41PbvEeo61b3LRHB3WlZXUMpyK7RW78cNwBkAcJqyvoLmBGVxnHxL5g1qmrxQRtHEwaUjG3Ja7PWjgyXLjmMLRIAJJwO7XtN53Ua/wA4/wCasrkwSjPynY0CCAf1V3OIIS3mdlqGKS4mVF3Z2qztUtYEiTyG56nu1LRo7rMkZ4Zf6Gra/vdMl8GZSU+6f+KtbuG6jDxNkeY8xU0Mc8bRyKCpq3t4baIRxrhRV1dF8oh+HzPWoIGlbbl5mo4kjXCj92YBgQRkVc2pj+Jd1/tUE7RN1U8xTLFcREMAyMNxVx2cfjJgmGOjVadnVVw1xIG/hFXeoWligUkZA2RanvNQ1NmWNW4Bvwry/GtDv/GhMEh/SR8vUUyqylWGQRgitTsjZ3LL9ht0Nabc8aeEx3Xl7fqr6fxpjg/Cuwrs/Y8EZuXG7bJ7d1/qWoWl8WKYi5BTyIqy1G3vEyjYbzU8xV3ZwXUZSVc9D5itL8S21YwK2V4mU+oHdd3OcxodvM1BA0rY8hzNIiooVRgUzKoySBT3sK8smpdYij+Z4092o9orUH/qoqTX7Zjtcwn8cVFqaOM4BHVTmo7mF+TYPQ/uJGaurbwzxqPh/tVtcGJsH5TzoEEAjka1vUbq24Y414Qw+erXRbm6jM8smOJcr5k1oV2IXks5gFYMce/StQKWWrJLA3PDMB68xQ3ArVbIXdqwA+Nd1qN2hlDDmpqN1kRXHIj9Rfz+FAQPmbYVY2rXVzHEOROWPpSIqIqKMADA7p4Ip4ykiBlNXmkXNm/jWrMyjp8y1/6gvPC8PgXj5cVaJp7xlrmYHjf5Qau7jw14VPxGoo2kcKKjjWNAq1cXyRA8JG3MnkKn1uW4mMVnE9xJ1Hyin0zUH4TqOoeCG5RRDJo6DosEEcgje4aQgJxOdyat9I04LJ4+lxR8Azn5gRTWGlOiyvpSrAxwHDYO/nirvQdLgmCw3U8DleLIJZR700Wu2Sh8LeQfeT5sVp2uxT/CrniHON9iKinSUZU79P3AgEEGrmAxNt8p5VZ3HCfDY7HlWoWa3ls0Z+bmp6GtH1EWxe1uW4eEnhJ8vSr4wXuqqIG2cgFh161a9nxHOsk0wcKcgCmdEwGYDJwMnu12y8C58VR8Em/41pc/zRE+q/qL+bxZyAdl2FdnrTggadhu+w9hUkiRozucKoyTUuoalqErJahlQeS7fmaZ9asjxOZMevxCtM1dLw+HIAsn9DX0eDj4/CTi64GalkWNCxpmaRyTuSatoREm/wAx51qWpxQROxfhQcz19BVrY32tsJJi0NnnZRsz1xppc3gwW6CBYwzY2Y74zRuE+lpdqpkheLgJUZKkHpT8dzAGbhhZJeKIHoOopJHk8X6ROuHThCJkgetBJpIY7V5IhEpGX3yVFXEVyhubtXK/GFAGGzGK4XQW9laykEDjd+gqbSINTWZjiK5icqJo9gxFW19eWFytrfjhf7Evk1W1wsq9G/cJI1kQqakRo3KnmKtZvETB+YVf6NBeP4gYo/mRuDVtptlpp8aWYFhyLbYq87Q80tU/8zVvp+p30qzSuyjOQzf8ClBCgE5ONzWp2ourSRMfEBlfcVG7RSq3mppGDqrDkRn611L4UDt542qCJp544xzZgKijWKNI15KABWvyMlgQPtOAa0OONdPjK82JLGiAwIIBBq/gWz1WIw7AlWA6ZNCrybjfhHJasocnxDyHKtRvEhjcFsADLmtNsZNZufpdypFrG36KP7x6mgoUAAAAchV5BCbqGaTBCqRw9TUcMj8RVRErHJwME0tnCOYLH1oQQ/6a0beE84xTWYG8blTUkDLKXZ3idhgyJyI9RVrBFBCqR7jnnqTWpabb6hbtDKv8reamtPubmyuzYXR/SJ+yfydahlEqBh+P63IriFcQq8hDpxDmtQymKQN+dAhlyDzFPo+oXV1J4sh4A2zseY9BVtpdjZLxkAsObvV1rtnDkJmRvTlWl6tPeXTo8YCFcrgd2s2v0e9fAwr/ABCtMm4oSh5qf6fW1WX5Ix7muzttx3DzEbIMD3PdfWouraSI8zyPQira8vNKlaJ0yud1P9xX/qO14M+FJxdKsUm1LUfpMi4RCD+XIVcSeHET58hSKXcKOZNMVgh28htV94upajFp6MeEnjnYdKhhjhiSKNQFUAAVPMIl6seQqG3JPiS7uf6fUv76eGZY4VDE7kVY6itxlHHDIOYplV1wRkV8dq/WMn8qBBAIrtDphu7XxohieD4kI5nHlWh6iJ4Y36/C46MPqllBxkZ+rmri8gt0Lyyqi9WNS9p45HKWVrNcN1AwteP2pn3S2ggH8RyaFp2pbf6fAPZa8PtVHyntpfSv8Z1i1/6zS2K+bRnNW19bXgZ4CcA7qRgirGXiQoeY5VqtzcW1qZIQMg758hUNtqWqHjMhKZ5sdqtez9rHgzMZD05Co4o4lCoiqOgHd2gtvEtFlA3jP9DVhL4dwvRtj9a7k8S4kPrgfhWiQeDYRkjd/iNXsjRWk7qCWCHGK0jVRcqIpTiUD/dU1vBOvDLGrD1oaHp3Fnwj7ZOKjijiQJGgVR5Cr6TikC+S1YRbtIfYVqtwsUZJOyKWNdlbZjBNfSD9JcOSP5RRIAJNQAzStK3IbKPqXFxHAhZ2AqxzcXUlyw+HkuaiCvqrNGNgMN3OqspUjY1bMUdoW8t17hH/AIfrtxbjaOceJH71A/HEjenfczrBCzt5Davp1w10JWY5DbVFIJI0ceYrGe5mAq/11zObTT4/Hn5E/ZWrfs/40gm1Odp5efBn4BUcMUX6OJFRQNgoApiMLvkZ3rbDcPSvhwuOdM2WO56CpraEAyLEqyHmQMZqCTw5Vap4lmhkjPJlIrQpWgup7R+px7j6k8QmhkjPJlIpg0chB2Kt/aonDxo3UA/UnfgiduimoUMsyIObMB+dRoERUHIACjWr2KWsoubeQLvkqDuD1FWGt28sSiZwkg2OeRpJEcZRww6g5p2CqWPkKYl2J8yahTgjVegrtNMfo8qjnJIqCrOFLWzgiyAEjUVPeQOrRRyqznmAc1EgSNV6DvvLpLaMux38hUNvPeuJZ8hOYWr+YWttwoMFthWnWvhQqT8zbse+6+B4pR5HB7u1SeE+nXg5xzcLexrT2yjDoe/UpXu5hFFkqpqz0fdWkoRrGoCDalGBTHFajf3Oo3R0+wbAH7eboOgrTtMttPhEcS7/AGnPNjTrl/cU0yg4GXb0pVumGyqooW9z/rV4N0u4dTRkkQYkiIGeYpXViWHxVIhRyDVq/HCvUbVPYWqXwvHn8PkcbDJFT69YRbKxc+gqbtFcucQwhffc1pE2pyXDNcBzGV2JGAD3a1D4WoS9Gww/GtNfit8eanH1NSfhtiPvECtEi8TUIui5bu1J9Tnu3ithIEUAZGwz71D2euJDxTzAe25qTs3AVHhzOD671pekz2d0zu4K8O2DV6/DDjqatU45l9N+7tFIwezx/r5/KpLm8umwzu56CtIs5oZQZYyvERjPddXUdsvE52ptah4TwK5PtVtby3Uv0i4G32UqaZLaIs3ICrrxbt4ZoF4lHlQ1O4hwJLVgPSoNRgmYANg9D3XYzA9QnMSH+EV2sQNo8p+66GtHfijU9Y1PdeSOI+BPmbarOzjhQEj4jz7sbVnGK7Q6hJCqWltvcTnhX0HWtK02LT7VY13c7yN1amYKCSdq/S3J2+FP71HEkYwq98t5BCwEjhTSuki5Ugg1La78cR4Wqdy5HEMMNjVg+GZeozWrWC3ccZMqpwHJY9Ks9I0p2I8fxmHMA4qG0toR+jhRfYVf6mlk0atGzcdA5ANdpYt4JfdTWlPh5E6jP1NWb9kvua7NR5lnk6KB9a/bMir0FWC/O34d2txq13pwf5TcYb8agtbeBQIolWrjaeA+vddn6RqUMJ+Vd6WCPlwCgoWtXElwrRR8lGTXZ+84Wa2aiikcqvrKAxvIBwuBkEVpUzy2ylzkirs4gep9Vvg7ok5CqSABWpTXElsxkkdhkcya0ZSsSDpEvdHFxOXb8K4NqFYplFHTkOrPevuRGFQdO45uJeAfIvOgAAAB3kgDJowpf38vHuiDAqwMkF5Lb8ZKAZWhV5CrIXHMVbtwzIfWryPxbWdOqGrS1upQ8lvnijxkA4NWmvTRHw7pCcbcXJhWuSwXNpDNE4YK+D1GasJPFsrd+qCtfj47At91wa09uG6T1yPqam2bgDoors4mLWVur/2H1rpuKd6s1xAvr3dpY2WHxBzimV6t5VlgikXk6Aj8RV4p8NXHNWzQYOgI8xTfBrAY8ilAg4INXEyxRMxNQ21yVLCT596vIZLC+VgfPNWtws8KOp5itbdxGio2xIBq0iSG3RV6Crw5CRjmzUmn2Sbi3TPXGa7V8HgWVqgAMs42HQVpyYRj7CiM7Uo4dqye8nJri/SMKnfgiJ8+Qq3j8OIDzO572dVBJOAKubua6cw2/LzakSOwtWYnJ/qTWmwuxa5k+Zzt7dzrxIw6ihkGgQyg9RWj/oNVuYevEPyNXenWt2P0ifF5MNjV5ol3b5MeZE9OdaKkyWKpKhUqxwD0rU047C4H8BP5VbtwzxH+IfUvzm6krQlxp0fqzGr9uGyuD/8AzNdneI2srEk5etOvrmfUriJnzGvFgY6Gteu54Po4ikKkk5xUXF4UfEcnhGac5dj1JqEYijH8I7tZtRNFIh5SIR+Ndl7szad4D/tLdijCmUMpU8iKtWKFoW5ryq7sEuWDElSPMVJZXFriWKV2xzU0stvfqqyZVhzHKooljUBScCrnTbe5fikBoPaacnBxYBqzgN7cSTSFuANlAaGFX0FQ5mnaU/Kuy913J9O7QnG8dqmP/I1apwQqPM70nMn6jSIgyzCrrUkGQlNqU9vxSABi229add3N7L+lYEBhgAY7iwppEUZY4FTzSahKYYiRGPmare2igjAUYrVJVe6hhJwvMmoZYOFVV12GOdBgT3SjEjj+I1bnMMZ9Kkawt9XeSSR0fY8vh3FRyxSrxRurD0Oe+4XiglXqjD+lA4cehoch33hzcy/zVo4xp1v7VrDcOnXHsBWjX9nb2ZSSUK5YnFaHPCl3cySSKuRtk45mtdljmurVUcMPQ55muSewrzpRhR7d1zF4kRHmNxQlOk6wlxyguPhl9G60CCAQcg1cQlsOmzrUE4lGOTDmKIyKudOSQh0PC48xQGqRjh+BvWjLqg/y1NR2dzcXCy3KDC8lpEVBhQAKlkadvCj+X7TUiKihRyFazqK6fYyS/bb4Yx1Y1oNg6RAybySHjkPcvKiwHPYVPqdvFkZyan1idj8Gwp7meVss5oGoYI7iRI5M8JPlUFrBbXESRLgEd2oTSSXMdvE5UndiK/wqRhhrlyKtrZIIwiritbu2hh4UPxGtNW3vrdWlAZxsak0aAkFGZT6Gm+nWRLluOMfnVvOs8SuvIirkYnk96tDmBKv9ItrtmkJZXxzFWkF60kn0YtxJzwcVHrd/bMEuYuL3HCasdQhvUYxhgV5g03yn2phh2HRqjOUT2Hfdf9RL/Ma0n/462/kq6tkuoWickKelN2btj8szim7M/dufzWouz08c0b+MhCsCaf5G9jQ50OXfq2nxzxOrD4X/AKHrWhanJbyf4ZeNh12hc8mWoYjKSMgADJJq7sHSTPyv5HyNLcvGeGZSPUUksb/KwNbGthT3MKfayegqOK7vWCIpVDUVtaWVsQQGJ2PUmri4ht4pJZWCooySaRptZvxdyqRbxnECdfWraHwk3+Y86XkTVxcmNdlJNXV1dOeZA6UxJO9E0tBC2MCrCxdSJGqb4ZoX9cUx2qAA6xLnyWgMGpHCIWPlUjRTrM0itlsgbVpdw1nelG5McUGBxWr3Qii8PGS+wrT4WitkVueKu/271Z/sF9zU88MKFpJFUepqz1E2c1w6IGL5A/OiuraoRlTwe3CtaVppsUfMnEz4z0GKPI1L+2k/mNQ/so/5R33f/Uy/zGtIOdOt/wCX6mp3E8WqWyrKwQ8GVB25026n2rzoch3soZSCMg1q+kJOm+QQcxyDmprSu0NxZSra6iSjDZJ/Jh61DqEFxEBJwnP2uYIqbTYnGY3GG5A1PpBQFimADzBr/DJ84DS0mkSu5Vi5I6nFRaNFFgv18t8eu9PdwQR8MIBOfKtT1WC2QzXUoHRfM+gFSPea5MrzK0Vopykfm/vVlZrCikqBgYVenckZx8TURBuowetNbQv9kU+kRNuKm0WQbpvUOkuSOI1BYwQgHhyabHCauE44jjmNxUDiSJT+Bq6ingvfHijL5G9RaqwkVJoTHnkTV8Z5EVYQCDuTUMeEUFMHFappdxNdGSJKsTOtuBMMMBUzvqFyqpH8Mb7tSr8IBq6OZ396sx+gWtR0me9vQ4cLHwgZq00ayt8Epxt1atgKV0f5WB9jmjyNSHMjn+I1FtGn8o774YupfetEbOnQ+mRRIAyTTXdqvOeMf+QptTsF53KVrF3bz3VvJC/Fwjf86G6j2phhiPWojmND/CPqMoYEEZFX+lRTxspQOnTzHtUcOr6WxNnL4sXnC9WfbGFCEuUltn8wQStRa9bXKgLcwyDIPMUurP0T1qTUMcRLogK45+uau+0enQg+Leq38Kni/tU3aO+u8pp1mwB/zZNhVroks0wnvJGnl9flFW9mIwCQM0qHzNAKgJqeaSRSqUjeHlSDmo7hetRMGpsLRfei9cZwQaHKgfo8ufsN/StiM1c2sU8ZVhQN1YsMgvF18xS6nalcmQChqVof8wVNPPeXIjt5cLj4jVjZraxlc5JOSauJvDXA3c8qbIYg86thiCP2q61yC2uXheJjw+YpNf09uZdfdafVrBoZClwueE4Fdm12uX9QKlbhic9FNc2/GlGFA9O/Ulxck9QK7OvmxZekhq5iM0EsYO7KRS9m5j81wv5Gl7NJ9q5P4LSdnLQc5ZDSgKoHQVOvDM49atWzAn1pIIpPmXfrU+lRygghXHRhU3Zm1Y/9OV9Uak7MIWwstyv/AJVB2It5ck3EjMByZqg7L2MDfskyPM7/AN6jsIY/LNKMbACsVyo8LDFeCgzinio2zEfLVvEY1xVwGI2o+P6UpbO4pTlsVFNHLxhGBKsVb0IpkV1KkUryWxw2Wj8j0pJEcZUg0VBGCKaztyc+EtHT7U84xUVrBD8iAVLcqp4UHE9Q2rYaWQ/Hiickmoxwog6AVNpllO7PJCCx5mn7P2DcuNfY1f6FDbW8kyzMeEciK7OpixZvvSGr5+CzuG6RtUQ4pYx1YfU1VcSRt1GK7NSbXEfse6/v4rKLibdj8q9abU9Wu2PghgOiChfaxaMGl4+Ho42qxvI7yASLt5MOhq+XE2eoqwbMbL0P1s0Aa8MYzmljWrdDuM79aZHHzCuHNcOO7NTySJxEGoJbmT7QxQ8Qc3B/CkuCpbxGA6UH4+W4okcjRRWNGIDkKK8L1fwXmlXr6hbAvC5/Txf8itO1C1vIRLC+QefVfQ08KOMrjentCrZQlDQku05qGr6XJ5wNX0mc/LD+dCG6m+ZsDoKhtoYRmrub9C4B2xUS8UiDqamkEUMkh+ypNRdpIztLAR6qc1DrGny8pgp6NtV1DHfWzRCUYbG64NWVqLS3SENnGd61uTg06X+LC1ZLxXMfvn6mqJmFW6NXZ+Xgv+HydCO7V1M+rwwsfh+AfnUUUcKKkahVAqeJJYnRwCpBzXZ1ytxcxg5XGfyNX6ZRW6GrJ+GbHUd+/kKC55muAVgVjuBwRSuWBCgDqakEskQZ6LAV4gouDRO9SxK8ZBqAcCVx/Fzp0V+YzSoyfIxFFJC2TK1K3rRcYokNIKZAwNXei3NrO13pb+G/NoT8rVY9pYi/g3iG2mHMN8ppLlJFByCOo3FN4TYxgb1wQ9azGpOwNGYeQq+1WztF4ridQfJeZP4VBqlxfu7CAx24+Qtzc1ZJmbPQUwDAggEGptO0uQ4ZI1b0PCal7NwNvFMy++9f4JqULgxyAjPNWxSjCqOgrtJLiGCPqxP5VpaZnZui/Uu047eQemfyqyl8G7gk6OM1zrX7WQPFdxA5XZseWORqDtHD4YE0bBx05Gr3X3nRooIyobYk860Kwe3ieWQYd+Q6Cpk44nX0pGKOD0NIxdQwGAaVTnPdnuHdkUzADnUN0gmC+XImo5+KKRSccJOO7Ao7VeXAjXY71bXMjkgmuLNcAY5JINRRbbvXgvxEAVIXTmjflQkJPykUGyKi3k7iAavNNtLxOGeFXHU8xTdnrm2JbT9Qki/gbda8btVBs0EE46g4Nf4xro56Mfwav8S7SSbR6WierNX0LtJd7T3qQL0jG9WfZuwgbxJA08n35DmrggyEDkNhVinDEW+8a1nUZvFW0tyeI/MRz38qGgX7LxF0DHyJpZtT0uRQ/Fw9Dupq0uUuoElTkeY6Hu7QTcd9weSIBWlpiJ36n+31CMgipkMcrr0atNn8eygfz4cH3FEAggjIqXRtPkYsYcH0OKjstPsxxiNFx9pqu+0MEeVgXxG68hVpP9It4pcY4l3FXUfBM3Q7irCXij4DzXuOfqFgKnvIkB3qe+kk5HFRzyxsTxV9Kuc/tDUWqEKA6ZPWhqsfmhqXVgwIRKeV5DljVo/xkUOVIaDUtzwjFTXDuazvRcCraRWd99/q8IrhFcIrAqeTw4mb8qVS7geZNIoVQo8hVo6jXpDL99wM9e7Vkjawn4/Jcj3rs3xfRZc8vE2pmCqWPIDNXMpmuJZPvOTVrH4cEa+n1dTj4Zg/kwrs3cZWWAnkeId2pa49vK8MUXxLzZqJnu247m7VF/iP9gKgn0S03CvM/UirDWo7ufwRFwbZXer2Lij4hzWo5WjbiWrbUY5dm2NZBGxoUWFSTqoO9S3OV2NOMkk0irTKKZRjuxvTKM5oVF8DA1EQwFZxSOKLCnYHulfCk5rjbjyDUWo3CczxD1qHUYn2b4TSsrDIII+tfS8ThByHOrGLLFzyHLu1jSZZJDc24y32lHP3FQ65f244JUD4+8CDU99faoywpHhSdwP+TVjara20cQ8uZ6mtZuPAsZMHd/hH41ax+LOi+u/1tQi8S3Yjmu9abc/RryKTO2cN7GgcitS0db2ZJBIEOMNtnNDQLKJS80zkDmdgKmSC4mEFjb7Z+Y7k1bQWGlKHnkVpse5HtUMsc8KyJurCp4jFIV8vKplKtxDbNQ6jNGMc6/xZ/NabUnI2prhnOS1eL600udqDHrQlxRmB8qJzRauIVxVxNnnVrclTg0rhhzoZxsa+Lud1HnV1NxnAocqHdDPLGcqxqHUkbaQYPWlZWGVOR3zyiKMt+VAM79STUUYjjVRWsW+otKk8DEhBsq8xVlr+CI7tcHlxj/kUFtrhQ/CjqeRwDSRxxjCIqj0GO7tDc+JcJCDtGN/c1pUXzyH2H1iARg1cRGKZ06HatGuvpFmmT8SfC3dqt/cTzvE+URGxwiree6ZfBsoiuebDdj7mrXs87njupf8AxFRS2Vs0dqjqD5LmrqHxI8j5hyp1DAg0y8JIrAo9KXux34rAogVgVisdyTMvnSXhr6dUl655UZGPnWM0BWO7r3QXMkJ2O3Sre5jmXY79O66n8V9vlHKrKH/MP4Vf6lBZcHGCxbkBR7S2/lA/5ir3UbC7BJtWV/JwRVre3Nq+YnI6jyNWU0k9tHJInCzDOKuJlghklbkqk1LI88zOd2ds/nVvEIoUToN/r6nBlFlA5bGtFvPo92FY/BJse6fR7We6M8mdwMqNgTRNrZxZ+CNBV7rs0xMVqpAO3F9o1p+iTM6z3LlTnIXO/deQcLcajY86mj4xkc6JxS9TRrNCsCsVt3AVjeivdtWKxWBWKxQ7jkUO7z7kd0YFTg0168sIXGCeZq3hMrgeQ50AAAByFahpUN78RJWQDAapbS50+XikhV06kZU1Y3GkXOFNvGj/AHSKFlZjlbx/7RQAAAFdorzZLZT/ABPWnQeJNxkbJ/f9Q6B1ZTyIqaNoZWQ+RrSL36Vark/GmzUam0/U728dZWPCrfMdlx6CrLTLa0UcK5fzc8+9lDAgjY1cQGJ/Q8jU0OTxCsYPcOdY27wKwO4is9xFDFNQ50AKNAUAO4D6kaZOTyqNGdgqioYliQKO9lV1KsAQeYNX2gRvl7Y8Dfd8qsIJYLWNJXLPjfJzVzOlvA8rclFTyvcTvI27O1WkHgwqvnzP6nUbfjTxFG68/atMvTZ3Kv8AYOzj0pWV1VlOQRkH60kayKVYVNC0TYP4GpYftLRzQPxd4FAVisURWK4axQG9cNcNYrFAVisVisCsUqZpELEKoqGFIEJYjONzSsrKGUgg8j9bXr/xZfo6H4UPxerVptvxv4jDZeXv+qIBFXtuYJdh8LbitA1HI+iyH+Q/8fUJABJq87QQxMUhTxCPPkKj7SycQ8SBcehqGe1v4co2eo8wamgeJsHl5GpIg2SOdKh4jkVw1igKC0VrFcNYrFcNAYNY3rh3rhrhoCsVisViuEUiM7BVFRRR28ZZiOW7GtS1OW+kFvbg8BONubVpdpLa2oSSQseePJfq6vqAtICFP6R9l9PWo0eaUKNyxqKNYo1ReQH6u4gWeIofwNESQy9GU1peoLeQDO0i7MO/XtRYN9FjOPvn/itK0UTKJrjPAflXrUukWEicPghehGxoGbSdQxnIB/3Ka+CWMeasM1PaNHll3WioNFa4aAoCsVisVigKxRXesVisVigKxWO+GB5TtsOtRRJCuw9zWq6pLdSGJcpGDjHIn3rStOgtoRLkM7DJfyA9K1XWGdjb2pPRmHn6CtO+lfRI/pPz/wBcevfc3EdtC8shwBV3dSXU7SvzPIdBVha+CnGw+Nv6D9bf2nirxqPjH9RVrcy2s6yIdwdx1qzu4ruFZEPuOh7p149WdZORnwfbNKAAAOQHdrc6T354NwoC5q1UpbQq3MIoP5Ubq2DcBmTPTiFTWaP8SEA/0NPG8ZwykVigtcNFaArFBa4aArhrFYoCsUFo4FFulAFjgDJqCy5NJ+VXmpWtkmCQW8kFWuo6ld3yOikoDuo2UCtY0kTAzwr+kHzL1pb26SBrcSEITuK0jSUhVZ5cM5GV8wK1HUorKPfdz8q1ot5fXLymUZjJzxdD0FO6RozuQFAyTWqai15NgbRr8orT7TiIlcbD5R+vv7PnLGP5hVhfS2cwdd1PzL1FW1zFcxLJG2Qa13T3En0qIEg/PjyPWrPtDGIwtwrcQHzDfNX+vmVDHbqVB5uedaNpjzyrPKv6NTkZ+0a167kgtkSM4MhIJHQVZ6PcXcBmV1G5ABq2vr3TZ/BnBKean+4oGKeJW2ZWGRUtj5xn8DTxuhwykUGIoOKBWsVisVisVisb0So86LrRcmgrMcAEmorF23c4FRwxxj4VrUtamEj29uhUg4LEb/hVza3UISSdGHHuCa0aeCazURoqldmA692s6RxhriBfi5uo8/WrHWZ7SF4iOMY+DPkasbC41KczzseDO7dfQUiRW8QCgIiitX1U3TmKI4iB/wB1WVoZm4m2Qf1oAAAAYH7hfWPDmSMbeYqw1CWyl4l3U/MvWra5gu4Q8ZBB5irjQbKZiy8UZP3eVQaBZRMGbikI68qubu3so0Mmyk8IwK1S3TULIPAwYocrjz6itJ1JrOUxyZ8Njv8AwmtdtJ7nwp4UDoE3xzrTEeOwgD8wtR69I16YhEHQvwrjY0QGGCKeyibllaexlHykNTQyr8yEVvXG3WvEfrXiPXiPXG3WsmljduSk0lnM3MY96SxjHzEmlREGFUCrztA0crRwxD4WwS3pVlfQ3kQdDv8AaXzBr6Jb+OZ/DHiEfNV/Zpd27xnnzU9DVjcy6fe4cEDPC4pWV1DKcgjIPdcaFby3Syg8KHd0Hma/RQReSIo9gBWq6u10TFESIh+bVaWjTtk7IOZpEVFCqMAfuV7YYzJEPdatLye0lDxt7jyNWGowXiZU4cc0Pd2lRzDA45BiDXZ+aI2fhgjjVjkV2gsYkAuUwCWww612dneS2kjY5CNt7GtXufo9lIQcM3wr+NdnrbxLlpmG0Y29z3XOr3Y1F1gfK8QULzBq5ultbbxZRnGMgdTVnqVreFljJyBkgijHG3NAaNrbn7Ao2UHQ19Bh6tX0KAdaFrbfdB/GlhiXki1cavYW7FGkyw2KgVN2lTlFAT6sabV9VuTiIEfyLViZmtITMCH4fizUsKyarJExwGmI/M0y3elXex9j5MKsL+G8i4kOGHzL07tfsOJfpUY3Gz+3Wuz9+WBtXO43Q91xcw20ZklcACtS1WW8YqMrEDsvX3q0smnPE2yf3pEVFCqMAful3YCTLx7N5jrSPNBKCpKOprTtcjm4Y7ghX8m8jU0MU8TRuAVYVc6Pe2kniWxZl8ivzCnTVLx1R1lcjqMAVpliLO2CHdycsa7RXBe5SEHZBk+5rR7b6PYxgj4m+I/jWo3P0azlkzvjC+5rQrbx73xGGRH8X412jk4bSNPvP/atMuTbXkT52Jw3sakcLGz+QUmtFeaXUR8bY+JiM7d08qwwySNyVSaM+panOyxs2Pug4AFT6XqNohmJ2HMq24rQ9Sln4oJjllGVbqKn0O2nuZJndviOeEVrenw2ngGFcKQQfetEkEmnx9VJU914wi1h2OwEwNarq1ncRGJIy58n5YqCa4tJEmTK9OhrT9QivYsjZx8y0yqylWGQRgirWxtrUHwkAJ5nzq/1a3tAVzxyeSirq8nu5OOVs9B5CrTTy2HlGB5LQAAAAwP3a4tI5xuMN5NU9tLA3xDbyNWGs3FrhW+OPoeY9qtL+2u1zG+/mp5jv1qN49RkZhs2GFWd9bXEKFHUHAypOCK1++SWRII2BVd2I61olr4FkrEfFJ8RrtJJmaCPopP51f2Pg2tnMB8yAN786S+8TQpHJ+JU4DXZqPMs8nRQO68gM9rNEObKQKtrq502d8Lg8mVqXtHE6lZrY4IwcHNafPp02TbKisBuMYPdr8PiWBbzRga7NTbTxezDu1lf/cph1Iqz0WzhCuw8RsZy3KruyguoTE6jHkR5VYaNDaOJDIzP+QqaeGBC8rhR61f6+8mUtgVX7/maRJZnwAWY1a2CRYZ/ib+g/eWVWBDAEVc6ZzaH/aaBlhkyCyMPwNWXaGVMLcLxj7w51bXttcrmKQH086vLGC7j4ZV9iOYqfs7dIT4MisPXY1Zdn5fEV7llCg54Qc5oAAACtXYzaoyjyKqKvrQTWDQgbhBw+60s8iQyw+TkE+612dj4bN3+8/8AbullSKNpHOFUZJpJ9L1KRo+AOwGckYqbs/YuDwcSH0OatS9pqSKrZKy8JI8xnHddxeLbTR/eQitDl8LUUU8mBXuvtIuLvUHkBCphfiNIvCirnkAKlmihUtI4UetXnaJFytsuT981PcT3D8UjljVvp0kmGk+Ff61FDHEvCi4/fJYIphh1BqfTJF3jPEOnnX6WJ/tIw/A1a6/dxYEmJF9edW2uWM2AzGNujUro4yrAjqD3NokZvBc+Mc+JxFSO7VNLuFupXihZkY5BA61pUJhsIFIwcZI9+6WNZY3RhswINT2V9p1xxxhsA/C4ptZ1SVeAfmq71pOkzmdbidSoU5APMnvi0myil8UR5fizkmiQOZq41axg+aUMei71ddopnyIIwg6nc1LPPO/FI7Oag0+aTBYcK+tQWcMPJcnqf+wSRRyDDoDUulqd43x6GpLO4j5oSOo3qKeeE5jkZT6Godfvo8Byrj1FRdpIT+0hYexzUetac/8AncPuCKS8tH+WeM/+QoMp5MKyO7avgHSmubdPmmQe7CpNX09Oc6n23qXtHar+zjdv6VN2iu3yI0RP6mpr27n/AGkzt6Z2qO3ml+VCai0tjvI+PQVDbQRfKgz1/wCyyW8EnzRg0+lwn5WZafS5x8rK1NZXK84j+G9GOReasPwoM68mYULm4HKZ/wDca+mXX+vJ/uNG6uTzmk/3GjJI3N2P4mgrtyBNLa3DcompNMuG58K0mlIPnkJ9qjs7aPlGPc70AB/2wxxnmg/Kvo8B5xJ+VfRoP9Jfyr6PB/pJ+VCKIckX8qwP/p//xABHEAABAgMFBQQGCgECBAcBAQABAgMABBEFEBIhMRMgMkFRImFxgTBCUpGhsRQVIzNAQ1BTYnI0gpIkYMHRY3OAouHw8SVU/9oACAEBAAE/Av8A0BkgamFTkqnV9HvhVqyQ/Mr5QbalfZWYNtt8mT74+vD+x/7o+u3P2U++Prt79pMfXb/7SI+u3f2kx9eK/YHvgW4ObH/ugW2xzbXCbXkzzUPKE2hJq/OHyhLzS+FxJ8D/AMrlSUipNIctKTR+bXwzhdto9RonxyhdsTR0wp8oXOza9XlfKCSdTX8Ah95HC6oecItSdT+ZXxEIttfrtA+GUN2xKK4sSfKG5hhzgcSfP/k92ZYZ43AIdtpofdoKvHKHbVm16KCfCFrWs1Uonx3ghStEkwJSZP5SoFnzXsfGBZkx1RH1U5+4mPqo/u/CPqr/AMb/ANsfVX/jf+2Pqr/xvhH1Uv8AcEfVb3tog2bM93vgyU0Py4LDydW1e7fbnZprhdV84atpwfeNg+GUNWnKOevhP8soBBFQf+SHHWmhVawIetlpOTSSrv0EPWjNu+vhHROW6hp1fCgmEWdMq1omE2Un1nPdCbPlR6tfEwlhlOjafd6cpSrVIMKk5ZX5Q8soVZjB0UoQuy3PVWD8IXJzKNWz5Z7rb7zR7CyIZtp0feoCu8ZQxaEq9ouh6HL/AJEfnpZjiXn0GsP2w+vJsYB8YUtSzVSiT37jcjMr9SnjDdlp9ddfCESkujRsfP8AEqbQviSDC7Ol1aVTDlmPDgIV8IW043xoI3GJ6ZY4V5dDDFsMrydGA9eUJWlYqlQI7v1+ZtGXYyriV0ETFpzL2VcCeg3GpR93hRl1MNWWgfeKr3CG2Wm+BAHolvNI4lgR9Nlf3RCFoWKpUDuuT0u3lir4R9aM+wqEWhLK5lPjAIIqPRkA6w7IS6+WE90O2a8ng7UKSpJooUN7T7rJq2siJa2Qcn00/kIQ4hxOJCgR+tzM8xL8Rqr2REzacw/kDgT0F4BJoIZs15ea+yPjDUmw1omp6n0k3PqJKGjQdYzUepgtOpFS2oeUIWpBqk0MSc1t0Z8Q1vnZwrJbQezz74SlSzRIqYFnTXsj3w6w61xopEvNOMKy4eYhC0uICk6G6Yc2bLiughFpTA1oqEWo2eJBHxhE3Lr0cHnl6BbaHBRSQYesxJzaVTuMOsOtcab2nnWVYm1kGJa2Eq7L4p/IaQlSVCoNR+rvPtMpxOKoImrWcc7LXYT153gFRoBUwzZq1ZuHCOnOGmGmh2E77z7bKarMO2k8rg7Ij6TMfur98NT8wjU4h3ww+h9GJPmLrQd2bFBqrKEJK1JSNSYYl0MpokeJumbPUXAWgKHXuiVkdirGV53Truzl1nmchdKS4ZaHtHW5SUrSUqFQYmGtk6pEWW5xt+YutNdGQnqYQgrWlI5mkLs6YTpRULacRxIIhp91o9hXlErMB9uvMaj0BAORh6zml5o7J+EPSzzPEnLrfLzb8uaoV5colLTZf7Kuwvp+qzdrNt1S12ldeUOuuOqxLVU3sWc4vNfZHxhphpkdhPn6BSglJUdBD7ynnCoxKyG1GNZonlH1dK04T74mJJxtYCaqB0iSk3GTjUr/AE3Wqe00O4xIf5Tfn8t61T2Gx3wwnE82P5DctP8AyB/SLN/yR4G61F1eSnoPnFnoxTI/iK3axONpbmFpTpFln7ZY/jdNO7JhaufKGph5rhWfCGrUH5iPMQ2806OwoHdoDrD9mtrzb7J6codZcaNFppfKWo6zRK+2j4iGX2n04m1V/UXn2mEYnFUictJ2Y7Keyjp1vYlHXtBQdYYk2mdBVXXdKgkVJoIQ4hwVQqo3LRVSWPeQISKqA74AoKDdtUfdHxiUVhmWj33rWlCSpWghMzLq0dTdao7DZ74YVhebP8huWgvFMq7sosxNXyeibppeOYcPfCHFtmqFEQi0pga0VH1rl91n4wtalrKlamLMZKUqcPraXWo7VaW+mZiXZ2zoRDtnPo4e0O6AVJPMERJTxWdm5ryO8pCViihURMWaR2mv9sEEGhFzTrjS8SFUMSdqNvUQ52V/A/p87aLcv2R2l9OnjDzzjy8Tiqm5KVKNAKmJazgO07/tjTdOhhx1x01WomGJhbC6p8xDLyHkYkm+0RWWPcRANDWEnEkEcxeXmgaFxNfGErSrhUDE+1tJc01TndLPpebB587rSfAb2Q1OsISVKCRzMJGFIHQRNs7VhSeeouk51C0hCzRQ+N0zONspNDVfSCSSSYs5nAziOqvlDq8Da1dBdLSjKpZGNAqc4nZTYEFPCYZQhawlS8NecN2ayk1USq5RCUknQQ4suOKWeZizGqNqc66XT0ulbSl07SRAJBBENqxtoV1Fd9+VafHaGfWJiUdY1zT1vkrVW1RD3aR15iELStIUk1B/TJ61dW2D4r/7Xy8q4+ctOsMSzbA7Iz6+gMOS6ktIc1Soe6GX1srxJhh9DyMSfMXPI2jS0dRGkWc9jZwc0/K99eN5xXVUWWn7NxXU3TkkUErbHZ+UJWpBqkkGPp00RTaxmT1MSEmUfaODPkL5uQ2hxt8XMdYWhaDRSSIxK6mACYlbPUSFujL2brRXhlqe0aQBUgdYSMKQOgiZa2rK08+V0i9tWBXUZG60ncLODmqEpKlBI1JhtAQhKRyF0yrDLun+N0qKS7X9R6AgEUMTNnesz/tjS6VnHZZXZzTzTEtMtTCMSD4jp+krWlCSpRoBE9aSn6oRk3875Wzyqi3ch0gAJFAKD0LmTi/7GJKipNAPfE5Jlk4k8Hyhp1bS8STEvMIfRUa8xdaEsUq2qRkdYadW0sKSc4TaqKdps17omLQW4kpSMIhCFOKCUjMww0Gmko6XuScu5mUZ90fVst/L3w3LstcCBukA6isfRpf9pHuhLbaOFAHhfNSv0gJ7dKQ1Zzjb6FFQKQb51rZzCuhzEWY5heKPaHyunndpMK6JyizWsTpX7N9qO9lLfXMw2grWlI5mAKADdenWmnQg+fdCVBQqDUbkzJtv56K6w60tpWFYuaecZWFoNDElPtzIpovmP0d15tlBWs0ETk85Mq6I5JuAKiABnEpIBui3M1fLcccQ2nEs0EA1FRvTH37v9zEnNMsy3aOdTlD9oOOApSKJubcW0sKSc4lplD6f5cxGsPWYlWbZp3R9XTPQe+EWW4eNYEMy7TI7I8/wkxKofpUkER9XOtrStCwaHwh9ZQytQGdLpSd2AwlFRXzhucl3PXp45Q7NstJriBPQQ64p1ZWrUxZjGZdP+ndmpgMN158oJKiSdTEvMuMHs6dISagGlNx1lDqcKxEzKrYPVPI3JUpKgpJoREhaIf7DmTnz/RX322GytZibm3JldVach0ubbW4oJSKmJWUQwOquZ3ZpbynTtdRyiz5vD9ks5ervTf8Aku+MIQtZolJJhmzFHN1VO4QiXZQnClAickyycSeD5RKys0VBaez3nfCFHlGxVGx742SY2aOkYU9IoNyg6RgT0jZo6RsUxse+NkuKEbrsqw7xIz6w7Zix92qvcYW043xIIulpBxw1c7KfjAASAAMtxSkoSVE5CJh9T7hUfIRgo3jPM5RJtbWYSOQzO8pIUCCKiJuSLXaRmj5XaRZ1pbWjTp7fI9f0N99thsrWYmppyZcxK05Dpcyyt5eFIiXl0MIoNeZ3p2V2yKp4xdITe0GBZ7Q+O7P/AOU55fKLKP3w8N8JJ5QGTzgNJigH4IoSeUFnoYLahu0A5b1oTWNWzTwjXvMS0uX3MPLmYnFhT2FPCjsiLMao2V+1ClJSCSaCJueU72UZI+cSs8trsqzT8obcQ4nEk1G5OSOCrjY7PMdL7OtHaUadPa5Hr+gvPIZbK1nIRNzS5lzEdOQ6XMMLeXhT5mGGEMowp8zuTdoE9hrT2okprbJwq4x8biQBUxOONOPEtjxPWEqKSCDmIlZkPt19Yajcn0KM2QkVJAiQlnWcSl5V5boaPOA2kbtQOcbVEbbujbGNquMausYldTGJXUxiPWMSupjGrrG0X1jaqjbd0bZMY0nnulIPKCz0goUN+fmtkjAniPwhKSpQSNTGFMnKK9r/AKwlJWoAakwpbUsyKnIDKJmaW+c8k8hGAhGI89O+6Tl9g3nxHXdnZGlXGxlzF9m2htRsnD2+R6/j1rS2kqUaAROziplz+A4Rcyyt5eFMMMoZRhTu2hKYftUDL1oQtTagpJzENzjSmdoTTqImpxb+WiOkS0q4+eieZick9hQp4IZeUy4FJhp1LqAtN9BWt4BOkJZ6wABuFxIgvHkIK1Hn+BqRAdVAeHOAQee4W0mFNqG4a0NNYe2u1VtOKLPlcCdqrU6RajvaS30zMS7iWl4yKkDId8OurdViUYYlAlG2fySPVh50urKj5DpFnSv5yv8ATvz0lhq42MuYuBINRFnzwmE4VfeD4/jdItGe+kLwI+7HxubbU6sJSM4l5dDCMI8zeSAKmJifcU6NmaJT8Ylnw+3ipTrBFRSJyV2C6jgOl0nI7Xtr4PnCUhIAAoIWlK0lKhkYmZdTDlOXIxZy3Q7hSKpOu4BWEtdYpS8rSIL3SCSdT+IDihAdSdxTYMKQU3uMtOUxprSCQkEnQQ6suOKWeZgAk0AziUkA3RbmaunSJ6a2q8KeEfGJSWL7n8RrAAAoLlzssjVfuzgWjLdT7oQ4hwVQoHcnpPZ/aIHZ591yFqbWFJNCIkptMy1X1hxD8Zas9qw2f7n/AKXJSVEADMxKSwYR/I6m/SJ2c2pwI4PnEpKqfV/EamEpShISkUAucbS4gpVoYfZUy4UnyiSmtirCrgPwvfZQ8jCqGmkNJwoF6WidYCQNL1OJEKdUfxoURpCXusVvU0DpBBGtxAIoYmLNNas+6JWUQwK6q6xaE1gTsk6nWEIU4sJTqYYZSy2EDzha0tpKlGgiZnXHshkjpds3KVwK90IcW2rEk0MSk2H055LGovIBFDE5K7FVRwG6XfXLuhaf/wBhh5D7YWjQ/irSnfo7eBPGr4XyMpsk41Dtn4bk7ObTsI4fndITgSA0vLodyZl0vt058jCkqQopUMxFnzf5Kz/XcAJ0hLYF6lhMKcJ/QQSNIS6Od5AMLbI0vNaGmsPodS4dpqYkJXZIxq4j8Lp2Z2zlBwDSGWVvLwphiVaZGQqet07JJWkrQKKHxhtxTawtOohtYWhKhzF60JWkpUMjEwwphzCdORus+cMs7nwK1/7wCCKj8RMvol2lOK//AGHXVuuKWo5m6z5XEdqsZeruWjMLB2QFBzPWJaXW+ug05mJiz0FsbLJQ+MEEGh1iSnqUbdOXI7k9KbVONPGPjEpIBui3M1dOl6GyqAAL1O9P0VKymErCr1tV0vUhCqYkg3Ta8Eu4e753SbAZZHU5nceADzgHtGLONZYdxO5MMJfbKT5GFoUhRSoZi6yZ38hZ/p/2/EWjOfSHcuBOn/e6Uly+5/EawAAKDcfYQ8jCryMNNIaQEpum5MPDEnJfzhSSkkEUMSU9go25w8j03kNczepYTClFX6Qh3repAVCkkXzqcUq5cCCARzvmXww0Tz5XSSMEsjvz3Z+V2qMaeIfG4Eg1ESE2Jln+Y4vw1rTmBOwSc1cXhchKlqCU6mJdlLLYSPO+ZfDDZVz5CJe0jo9/ugEKFQajcmpRL4rovrC0KQopUKERJTikENqzSdO7cQ3TM3rd5D9LQspgEG4gEQpBTcRUUiYYLLpTy5RJToSNm5pyMBSSKgiH51loa4ldBDzy3l4lRJyxecz4BrepSUipNBuWhLbNe0Twq+d0pMql3gsacx3QhSVpCknI/hJh9LDKnFcoccU4tS1ak3WdLYU7VWp08NyblNuBnRQ0hxpbasKxQxLzTjBy06QxMtvjsnPpuTUqh9PRXIxJyWx7S+P5Xobp43rcrkP00KIhKwq4isLRhumJdD6MJ8jD0u6ye0PO+WkXHc1dlMIQltISkUF0xOtM5cSukPTDrxqs+UWdMYk7JWo08L3EJcQUq0MPNKacKDdY83+Qo/0/CWrNbZ7AnhR87pOX2zufCNd51lt1OFYiZknGcxmjrCVKSag0MSk+F0Q5krr13m0U8btIWvF4fp4NIQvF43EVhaMNxAOsGSlT+UIRLMI4Wxc4620mq1UiYtBbmSOyn43N2e8tBUcsshCFKbWFDUGGXUutpWOd8/L7VvEOJNyVFKgoHMRKTAmGEr58/H8FaM19HYNOJWSbgKmkSrGxaCefO+cmdgjLiOkS0+h3sr7KvnuTVng1U1r7MaRZ75dbKVap3G0UzN1aQteLeoY2a+kbJUbJXdGyVGzX0ih6foyF4vG4iohScJ3Ji0UI7LfaPXlDji3FYlqqYwKwlVDTrEi4ht8YgM+fS60ZfCrajQ6+MWc/gcwHRXz3J+X2TtRwqusua2L+E8K8vwU/M/SJgkcIyTdZrGJe1OidPHcnpNxai4g17rpa0FI7LmaevMQhaVpxJNRfaGH6SqnQViyh2nT3C9pHM3rXi8NwJJ5QGepgISOUKWhOqgIVOyqfzBBtKW/l7o+s2PZXAtOX/lAn5U/mQl1pfCtJgpSeUFkcoKFD9EQvFcpOIQRQ3TTRdZUkHO6TkkOJDizUdILaC2UU7NNIeaLTikGJGaDiMCj2x8Yn1oTLqSdToIFaim5MMh5pSfdBBBIN1nTO3lxXiTkfwFqTOxl8I4l5XISVqCRqYabDTaUDluzMm2/norrDzLjKqLEMzDjKqpPlDE8y6MzhPQw/OstJyUFHoIUpS1knUmJNjYsgHU5m5tFTe4uuQvCSdIS0BrC3G2xVSgIdtRI+7TXvMGam3tCf9MCUfVr8YEj1XCZRg+uT5x9CZ74+hM9VQZEclwZJ0aEGA5OM81fMQ1ah/MR5iGphl3hV5QptJhSCn9CBoYSrELnEYhfaMvhVtU6HXxiQmNk5hPCq602gWg5zTGcEkxIyqlrC1Dsj47tpMYVh0aK18brOmdhMCvCrI/gJ+Y28yo+qMk3WYzmXT4DfW2hxOFQqImLPWjNvtJ+N8jJFP2jgz5C5IqaQBQXOr5C9DVdYJQhNSaCH7T5Nf7oSy++cSveYTKstiqs/GDaI/LZKkjnpDD6X0YhEwVzE1sMVEjWFWdhKSysg98LXs2yo8hDLCpyrrqzSuQg45F5HbJbVyiYd2TK1+6EWgsJBdZND6wgIlphOJPwhco4jNBr84YtB1vJztD4w0828mqDWFtdP0JKsJgGtzqOdy0JWgpVoYeaUy4UGJa0UhAS7XLnE5O7YYEDsxZw/4keBjYs1rsk+6/bNbTZ4xi6XvNh1tSDzhSSlRB1F1mzG2lhXiTkfTWlMbGWVTiVkLkpKlBI1MNNhttKByF8xNNsDPXpC7SmFcNEwJ6aH5kNWor8xHmIaebdFUKrcuXZc4mwYQwy3woAvbThFy1YRehumZiYmm2BnmeQhS35tf/2gjZMSyMbmcMvtvJqgxaFfoqqd0SeH6M1TpG0ZSsNVAPSHTsJ8OHhVExOttt1QpKlcocDjsmajtFOkWe+3scBUARE4sTD7TSM6RaSirZMp1JhM4419iuXzA0EWcytCVqUKYuUPzykP4EJxU4oSqWm05a/GFNvSysST5xKzyXeyvJXzhaAqCCD+gtrob1pwm6blQ+n+Q0hxlxs0UgiG2XXD2UExJyuwTnxHW56dZa51PQQ9PPO5VwjoIQsoWlQ1BhCsSUq6i+02aLDg9bXxusuY2UyAdF5emtV/azOHkjK6zGarLh9XS994MtKWYWtTiipRqTDVmurFVHDCrKV6rgPlDrDrXGmELW2oKSaGJWZD6K+sNRuNJqa3E0hSsRubRTxibnAyMKc1/KGmVvqxqOXWEoSgUSIn01lV91IKdi00+05QkZiGZlqaQW1ZKI0/7QgTsrVCUY08olpZ9T+3d15CFS20TRaQRCLMYQahA+cbHvh2zGHDUjPuhqRQzwJ84VJq+lB5R8otFJSpp5PKJiYDbGPrwxZ7GFBdVxL+ULwptJGz/wBVz8rTtI90Sc9o26fBUKSFCCKH9BaXXK5acQ3Xp1lrnU9BD0887lXCOguZkHnMz2R3wzKMs6Cp6ncmGtqypHuvk39vLoXz5+PpJl7YsOOdBBJJqbpZrZMoT777Uc7aG+grFmshbpWfV+d5AUKEVETsrsVYk8B+ESLmCYR35XgVNIAoLnVcrmkczE5NBhNBxmGGC8rGvT5wBS59OJlwfxMWfLtOJK1ZkHSPoLS3A5gzgNpGsbRA0gvHpG0X1jGvrG0X1jbKgOp5xgQqJyz9skZ6aRtZ9oYC1i6KiTlVoJdd4zfMy3ro8xEhOVo04f6mFoxD9BBoYSaitzqed8xOPrJTwjoLmJB5zM9kd8MyjLOgqepvdnZdv1qnoISoKSFDQi+0GsD9eSs7rFfotbJ55j0ltPfdsj+xukWtpMJ6Jz3LQ/yleAiylCjqfPctCn0VflDP3zf9he0nnctWEXNpqYfeSy2VHyhCVzLpUrzgAAUG4zLobTSgHdCnQNIKidfQVpCXesKQlWYggjXcmmMJxp0iRmtsiiuIQ6jn+LpFIpFIpc0qhpcRUQRQ0umJDav4gaA6wzKMs6DPqbluNtiq1AQ7aiB92mveYdmn3eJeXS6z9psKLSRTS+faxy5PNOdzLpadQseqYSoKSCND6Oce20y4vlXLwus1rCyV+1uWozwujwMNOqaWFphibaeGRorpfaUwFUaSdNYs5rG/i5JzuSKml7iqm5IwiJp4zD9E6aCGmw2gJF+sJSECphbhV4eg+my+MpxwCCKg3pUUx2XBCklJvIBFDBxyr4I8vCELS42FDQwpOE/hqRS5UwynVYgzzP8AIx9YJ9gx9Yf+H8Y+sE+wYE8z3iEzDKtFi5BqLnU873p1hqorU9BDtpPL4eyIKio1JrDbTjhohJMNWWr8xVO4Q1LMtcKM+u5rDzezdWjobrJe2kqE80Zein3dlKuK50oPO5KSpQSOZhCQhKUjkNxSQpJBGRiaklsmozRcidmUaOe/OFzsysUK/dCEKWoJSKkxLMBhvDz53NJyrc4qibmU51i0H9m1hGqokmstofLcQnCKmFrxegeWG21K6CJWTMwFqxUhKn5J2h06dYQsLSFDQ3gkGMnEwRQ3zLW0b7xpFmP0UWjz0h1NRX0JNIStKtD6Klzky03qc+ghc84eAUjZTLutfOEyB9ZfugSTI1rAlpb2fjH0Zj2BBlpb2fjBkmToTCpA+qv3xsppnSvlErPL2gS5oedxgihutNrC6F+1DDJecwAgQ1ZzCOLtGAABQCm/ajfaQ51yN1ju4Jko9sfL0Vtu/dNf6jdZzeKYr7IrfMObNla+giUmg+n+Q1F7sgw5nTCe6FWUvk4ITZSvWcHlDMu0yOwPO4CpgXOKqq5IomJhZmJrLrQQAEgAXtJ5w6uuXobUdo2lHUxJt7OXQPOLSQCxi5gxZp/4f/VuIVhMOJqK7j4LL+JPiIQsONpUOYhQoSN8kAVibmys4U6RKPltzugZ+gpDryGh2j5Q5NPOnCnLuENSKjms07oS000MkxiX0gGojiVGAQo0EJSKRwqhZoIxKGohTTTozTDfALnhzunWtpLq6jMQy4W3EL6GAQQCPQTre0l192fuuaWW3ELHqmsJIUARz9DaDu0m3T0NPddZjdGSr2j8r7SCyx2RlXOG3FNrCknMRLTCX0VGvMb7I53LNE3NiqhE25s2Fn3RIo7Sl9LwKmFHAj0MxMoYTnryENtPTjuNWl1pu9lLQ56xJtbJhIOpzO6yrlC00VfOIxNV6RZjlWlI9k/OHhmDvz8xQYBfKOY2hcTQQndmJsI7KMzDUu4+cSjl1hLbbKeyICqw5pA0EYs6Rwqgr6QoVTCVimccSoVmqkVVpCRQQys4+65QqKXzTWyfWnlyizncbGHmnL0LyNm6tHQ3WY7tJNH8ez6B5zZtOL6JJvZRgaQnoNydktn9o2OzzHSGXlsrCkwy8h5AUneQKJFzpzpcyNTFqrybR5xKpwsp7872RnWHTVXoJqdSz2U5r+US8o4+rau6fOAAkUAyiYeDLZV7ok2lTDxdXoN5JoYdFU1vUMSSOsSCsE1h65Q6OzvOuBtBMOLK1E3Myi3O4QwyloUFyzCMoF8zN+o2fExLynrOe65fDBplTWAquRjCoaQhpUbHrAZTGyHWNimNl0jZlJrCuIGFnlCRSEmoBudFFXWm1VCXByyMWc0+hzFgoki9S0p4lAQu0JZOhKvCF2o4eBAHxhhzaMoX1F9poo8Fe0PldYjnadb8/QWu5hlCPaIF0ojHMNjv+W6opA7RFIm0MJc+yWCOnSJeYWwvEnzEItNk8QKfjCHmnOFYN6BVQuMHO5rgi0DimiOlBAFABBIGZNIXPyyPXr4RLrCmErpSufoJ2dwdhs9r5RKSNftHfdfPOl6YDY0Bp5wy0Gmwkb6O03uOfZzdf5AwcxvTsxjVhGkNsuOnsiGJFCM1ZmKXHSBncm6bmfy0eZiVlqdtfkLzoYbGUBquZjsIgvdBBcXGJXUxU9YxK6xtVwHuojsLhbEUXpDGQpc6Ozeucl0auDyzhdqj1G/fC5+ZX69PCCSTUmsAE6CESMyv1KeMSrKmWsClVzvtNFWAr2TdZzmCca78vf6C219tpHQVustFXVq6D53zL+wax0rC7QmVc8PhClKUaqUTCUKUaJSTH0CapXBC21o4kkXWfj+j1USanK5kam509m9HCIfcAmlrPJfyh20nlcAwwpa1mqlE3YdmwlPRIG/OTuz7COL5RJyX5juvIXrVhQo9BEgUfSMS1QCDod9nnC+I3zw+1HhDRxNIPVIg6ncnHtm3TmYl5YvKqdIQhKBQDcVpDQVTtQYbiaf2aaDiMSjGI7RWnLcAJgJSgQp08vRJd6wUpVpAqlVxFRe8jA6tPQwiWfc4WzCLLcPGsD4wizpdOoKvGEoQjhSBuzCMbDif43JUUqBHIwlQUkKHMb9prxzrndldZiaMFXVV9oJddLaEIJ5w3Zbp41AfGG7Pl0ajF4wEpSKAUuIB1hclLL9SnhlCEBCEpGgFzYokXOntXjSJnVzxuQ245wpJhmzXipOMhOcPcO5NzxZXgQAesfWj37YhVoTDgwpRQnpEnI/mO69L5t/Ysk8zpEq79JlylRz0MfVTf7hhVlujgcEUtFjqfjCLTpk63Twhp5t3gVXcZ1MO8Zvn9URK/wCO1/UQviN5NBBQqZer6ohCAkUG+qAoJBJ5QkKmX8//AKIAAAAvSmpglKBClFW+SACTCLSYJzqISpKhVJrelRTFUrF7goo3OfR2ztFhIJ5x9Yy1aVPjSEqSsVSai9zFs14daGkfSXyoEuKPnAzF7qcDq09FXWcvHJs+FPdvurxuuK6qJulE4ZZod3z9GMzeTUm8aQpouuFA1JhmzmUcXaMAACgEDUQ9wi9aglJUeQiSRt5kuK5ZxQdIoOm5PvbR2g0TlEk9snh0OR3FstOcSAYfklMfasqOUSczt0Z8Q1vZ4onp/YvlGzr5wq1HuSUiFTsyr8w+WUNqUrESaxKf4zX9YXxG96p7A5whISKD0HOJn7uJVrZt95vArGSEwSSfQTzmCXV35RLSCXWMSiQTpCFOyT1Dp84SoKSCOd4JBhKqi54aG60U1lj3EG5t5xo1QqkMWkhWTnZPXlAIIqLnU4HFp6ExKqxS7R/j8r7QThmVd9DdYq6sLT0XvTKsEu6rog3AVNIAoAPRtDt3LNEncTwiB2Z0/wBzuOcF9oKpLK74stP2Sj1VuzTuBrLiVkIdlB9DwDiGfndJP7VkdU5HcdUEtLJ6RZddqv8Arezzi0lYp17xvaBAzES4ow1/UQriN41r6A6XYQrXcbFBWFqxH0NpOY3UNDl/1htGBCU9BFqAYGzzrEl/jN7jJzucHZufTjZcT/EwNYes0HtMnyMLQtBopNDDMy6yeyryhi0GnMldk/CJ9NJlffnFmKqwR0Vfaqe22rupdYivtXU9U1929aasMk730F0sMUw0P5D0jPO57h3G+ARNdicUe8HcHab8r7S/xv8AVFm/4w8TurmUGbqquFGnjH09j+XuiYwbVRQOyYkX9k8K6KyO5OTZdVskaV98ScvsG6HiOt7XDCrPYU4paiokmsJkpUflwlttPCgDyia7UxTwEaC9foVbqRUw6rKnoZueDfYbzV8ok5Ralh5zyunnNs+lpGdPnDSA22lHQbiOIXHMXupwOLT0JiVVil2j/H5Q4024KLTWH7NUM2jXuggg0IoYqTFlK7Tie6t9qJ+xQeirrJVScT3gjetlVJZI6rus8Vmkd1fSNcNz2o3GdDFqIo4hXUQyrE0g917J1ELFFG6eTill92cWWv7FSeh3Jx/ZNHqchEu0G2kj33TrO0YPUZi6Sf2rIrxDI3zwCJuqe4wNBerst7kuNrOJ/tWF5JPpFXLUEisIzSLmhlWFGqj6CansX2bPviUkMFFucXS6fmdi3RPEqLOl/wA5Xl6BWpun04ZpffnFmKrL06KvdYaeHbT5w9ZrqT9n2hElJvtOhaqAUvtAVlV91LpFWGbY/t8963D2WB43WWPt1H+N86aSrkSql7doYjxC5+0Nk8pGzrTviXe27ePDS42k+FKyTrDSittCjzFbm+AXO8e40e1Fot4pevsmsSK+wpPS9BoqHU5VuUMSSOsCTnWiS384/wD6g9qJSeJVge163PSDjqsSnvhDSHUgBS6+VziXFCiVAR9VH934QxIuMrql74XT00824EI6RLyC8e0ePle0mprDpzpfMLwNK90WW3mtflDxyAuPojDr6GxmYXMqWrPSEzTKGk43BC7VaHAgn4RmGRXWnoJqaW+vZNafOJSTSyKnNd3KM5uc7q/CAKCg3kcIuc4zdaTK1rbKUk5UyizmXmseNNAfQTQrLu/1uZOF1s9FDets/aMj+N1lDN0+F9on/hj4iJEVmm7ps1mXfGJAUlUed7Yo2gfxFydBcviO4MjBAWkjkRCKsTND1odxtVU0hScJ3JuSDvaRkr5xLzq2Ts3gYSpKhVJqN0mgJg2o1yQqG9rNzQWRlfrGSEbk65VQR0iUa2TCBz1MOGqrj6FTiUjMw9O6hPvgqUdTWBTlC9YYTjfaT1WIe4d+am1PK2LWnziVlAwM+LrfNqwyzh7os51ltS8aqHlAfZOjqffvNcFz3FdaDrjSWyhVM4atQ/mI8xDU0w7wrz6b7gq2sfxN4NQDu20f+KR/5d1lcLviL7UP2KB/KLNH/EjwNzxq84f5GJQUlmv6w6aNOH+JgZkbp1O60rKkWmzmHR4GJZ3G33i8GhrBAWmNNx6XbeFFDzgtzUmqqO0mGp9hY7RwmPpct+6mPrCVrxfCA80oVCxE1NtJbUAqqiOUWfKoUgrWmueUBIToKXtopmYWrEb1rCElR5RKNl+YxHQZmFmibjvqWhOphc62OHOHJt5XLKFKxakxnyMVHO5Mq68ex74lLN2bqFqXUg6CHuW4/Ovl4oZOQj6TaPsH/bCjPzAwFJp7olJNDArqrrfaL+BAQnUwwpM1LUX4GDZst3wqyh6rnvj6LPM8CvcYTaDzZo83/wBIZmGnuFXlezw3PcrpyWL6AAqlDDknMN6oy6i5qcmG9F5dDDVqIP3iad4ht1tzgWDvsGrLR/gN22P8v/QLrL+6X/a+1Tk0PGJWYDDhVhrlSBajXNCoJqSYZnZUNoTtNEjlExNMFhzC4K4YZFXmx/IXDUegSaGHEJcbKToRCSqWfIV4HcQvCYUkLFRBFN1clLrNSj3R9XS3Q++Pq6V9k++PqpH7hhNmNDVRMJSEgADK9tvmYccrkNycdxKwDlEmxsWh7RzMOqqaXHdcnGkd8OTq18JpBUs65wadDGXJUHF4x2eYpGuhrHwiz1doiEcQh7lfNObNhaostvjc8t0kAEmJla3HCsjXSLPe2b1OStxSEKFFCsTUpsftWTSkSkxt2q8+dzOhue5bi+NXjH1bjaQpC9Ug0MOSj7eqPMQCRpDM/MJIBOId+4riV43Sn+Kx/wCWndtf/MP9RdZf3C/73z8u88UYE1pBkpoflGDLvj8pfugpUNUm+UFZlr+1ydR6FpXKJ+W2iMaeJPyiUf8Ay1eW4hZTHZcEKQU+iCSdISgJzMTE0lIpWEvOpOLXrWBmLpl/ZpoOIxZ8tiVtVaDSFqwjdJAh2eQnIZw5Mqc9YxX+UeQMUHeI7XJUZ80xlWPjGXhGdOsSKMPagaiHtBfah+yQP5RIJpKo7892bJUUMD1tfCJ2XCpbsjg0ulHtsyk89DuTqwmWX35RZQOBw99zPO57luOfeL/sYadbblmitQHYEO2okfdpr3mHXVOqxKpCEqUoBIJO4594v+xukv8AEY/oN21v81XgLrL+4V/ffKEnVIgy7B/KR7oTLMJUFJbANyeIePokKxCJ+UwHao0590S0xtBhVxfO5ZWt0oC8ITDMxWiVHOsA00hLo5wW0nSC0obwbUeUBoc4LiU6RMTfIGp+UGqu1UmGGBTEseFzzyWk1PkIYZXNO1OnMwAlCaDICFKxG43OPJbHf0iYedWc8SY8zFf5CM/ZEdnvEDF6pgnqmKjkqkZ9xig8ISlfLOJaU9ZY8oApcrNu+1eBrxiT/wAZrw3CQASeUMvubVbuxUqukfS1/wD+ZcOpUlZqgproDFnvbN7CdFXnIGHHXJt8I0zyENNBpsIHK5nnc9oL3Z1hv1qnoIWcS1Hqbmpd53gR5wzZiRm4qvcIQ2hsUSkDcd+8X/Y3SX+Ix/Qbtr/5h/qLrL+5X/f0E46tpnEjWsSs4+5MISpWR7rhqPRJVhMZLTE3KKZONHD8ol5kL7KuL5w4wheZ1hbamynrDTi0aHyMNvoWmukBRGhgPHnG1RH2XdFGu6Psu6NogQXughcwM6r8ocfcUnkAYQhSswkHlSGWMNCrM/K559LQ7+kNNOzbvzPSG20NICU6Q4vF4Xvz1FYWxUxtJtfrUhpKUdpWaupibWlzCIRLp9ox9F/lBl3k8q+EEka1HjGXT3RXv98FPdARU5VhmTdPFpDbDbeg3Gz2aQRQkXWmisvXoYs5eKXA9nLcnXh2GQeI5wkBKQBoLrRaxsYuabpR/bMg8xkbzT6f9n7d7PO57ldaX+N/quZkH3OWEd8M2ew3r2j377n3i/E3Sf8AiMf0G7bP+X/oF1lfduf29BPisq55RKGky1/a4a+jSopMZKETcgU9trTpDM3Tsue+OytPUQuXUB2Dl0MLQUFQ5CElaBUV/wChhM0qmYrDTocBypSNu3RRrwwZhsUz5Qh9CzQQqaVn2I2zp9eOxswnD26x9FOQK8oShKdBc9NhOSMz1iXlHZhWJXD1hCENIokUELcr4XvOAJIrnCUpTCnAmHJnpCDzOsbSA4YQ4YqDrC5dk6CkKZeGhrDYQVfaJI8IaVLJAwmAoHnuDSEGhh5PO5xG0bUnqI+rX08Lggyc8nPH/wC6JOdNdm8c+Rg5jWFWa2TUrXWG0FHrk+NzrO1BG0IHSPqtn2lQzJJZVVK1XTS31zRaCzSukSsilk4iaqvZGVz3K59kPIwE0zhqWZa4U59d0PNHRxPv3Fam6V/xmP8Ay0/Ldtof8Qg/wusrR7y9BNCss7/WGTR1s/yG6riPoUqKYSsKiZkW3cx2VQRMSqs//gw3ONq4uyYISociIVLN8soMqoaGvUQEONsqFMzBadRUUrUco2LmXZPDDaXCtslNMIhLH2q1ECnKNgjt1zxGAkJ0Fzky0jnU9BCnX3zhSPIRLWcB2nc/4wSlAhx9J9aC+31gzXQQp5ZgnnBdpDjpVAhPeYTsz6wgtjlANDAMVjFSNrTlG2bPKMXsmkCZeT3widQrXKAoHSE6XJOJNIUmh3JuSD3aTkr5wxOOsHZvA0hC0rFUmo3XXNm2pfSFWo4eFvOJJh1b22XelOI3vajfcOFtZ6A3JFABccgb2hRpsfxG7bg7bB7jdZR7bg7ryQMyaQZmXH5qPfBnZUfmiDaMr1Pug2ox7K4ctNKkKTszmKawNb08Iuc4z6NLvWCErFDQiHrMQc2zTuhTU1L8iB8IRPK9ZMJm2TzpAcbOixuFSRqoQqaZHrQqe9lHvis1MZCp8NIZsw6uq8hCQwwmgoIdnwOCFvLXqbqGt7juHIZmCjmYw1g5QCekbVXQQHSIK6w2qCaQpyBUxgMaRiuSVozSYaNUJNwNDCgFp3XmG3k0UIMvMyqsTRqmG7TaI7YIMG0ZbqfdH1ozXgVAnJb90ROzqFoLbedecSLAQwCR2jneBWAAhMJNRc7xXWm4pJaCVEawJuZH5qoFozI5g+UC1XObaYFqo5tmHrQZWytIrUiGBiebH8he8aMuH+JuAqQN62x9myf5G6zD/wAQe9N8+P8AhV+V9CYDLx0bV7oEpMn8pUCQmvY+MJrhTXpc3wC53i9ICRpAe6wFAw5KS7mrYhVlt+q4R8YVZb3JSTH0CbGifjH0Sd9k++Poc3zHvMJs1880wmy0jjd90BmRa6HxzgzrSckiHJ1xXdBUo6ncEVuUkQppRj6OvpBZX0gBSeUHCfUhaQqEMmsIbhxMLyMbcjSNufajadbhAiXWFNil6FYYUkLFRBFN1Uuwo1LaY+iS37SYMpLEU2Qj6tlv5e+ESMsg1we+8CsJSECFKxQnQXOcZumJRt8gqJ05QbKHJ34QbLe5LTBs+aHqV84MtMD8pXugpUNQYkRWabvnDSWd8LpcYn2R/Mb1sJrKeCxdImk03epKVpKVCogS0uPyk+6A22NEJ92+zobnuXpwtQ5wHj0jbJjao6xPzrmMIZ98TTz+FFSYROOj1jG3cWK4jdTcrBhOcP4hmIS7GKEERjEYkwEoMENU0EYUE5RghAhwCFt1MYGxxCNm2dDGzI0MI6UgCFmiYsx3NSTz03EqKY7KxCmyPRJbJjsoEOOVhNagd95137VP2bY74swVmCeib7RNJY95F1nJxTrPjvWgnFJveFfdcyrC82eihvO2gwjIdo90KtR31UJEC05jomGbSbXksYflezqbnR2fTlSRzhUwgQZlUF1fWGAMaqw81jaIiWYSSuphCaZRSDfSsLqICoYPF4wqFtNhGIGh6QtwgQmYpyj6UI+kZ6QmaTC5snSGArDFDFDCkxSMAUIVLCuUbA9YCcMCFCogyqmm0rH/AORLTO0GFXF87ta1MJVAMB3rBQhUFoxhV03MKjygNHnAQhMKd6QpdecUhpOYuVwm9K0q4VA7tppcUpuiCQBFloILpI6X2qrsNp77rHTWbr0Qd51ONtaeqSL21YkIV1G5OThdJSk9j5wATpCJGZV6lPGDZsz/ABPnC0LQaKFDFnTRrslHL1bkGihcoVB9GVJHOC+gQZroIU8s84JO4DRYgOQSWZkxXO5QvELEFHSCdmqsYgqFw7rAENoEZDlCsNdIpzhE2EQicCoDyaaQXEwcEF3A5AUlUFIhSTHOGU4lpigpSJiWLRxo0+US80F9lfF84IEEUgEiMQgGA4qNqOkY0RibjGiNqOkFwwV98GsAEwBSGudzp7N04vBLOHup74BI0hE5Mo/MPnnCLUX67YPhCLSl1a1TCHml8Kwd21FfbIHRN1ho++X4DfmkYJh5P8zdIqxSrfdlfPuYJZX8soQgrUEjUxLyzbCchnzN80wHmiOfKASkgjUQhWJCVdRW4aXLFFHeKgOcF9AgzJ5CC6sxmYpFIpFIwxhjDDyecB7s1h9wOLqIQpZGQhDsY4xRijakCMVRDajVULoRSGl0EKWVKoImJZaO1yhMJgkmEpEFAgy3SCwtOYgTKhqmFPuGA871jNWcNrjHBMKiTT612sTEn6zfuhicKey5p1gKSoVBqIwQRSPCMRgHKMQpBXAVGI3a0FIwXKUlIqo0hghTYUOdzxzFykpWKKAIhdnyyvVI8IXZR9Rz3wuRmU+pXwhSVJ1BFyJl9HC4YRXCnFrTO+dVimXPddY6KSlfaUd+10YZwn2kg3WWvsOJ6Gt9qfco/tFn0+lI891ea1eMSn+M1/W5o9m54aHcU6lMKmFHSKk+jxQsiFooqA2gpKucNet2qCMXagExncdISV6UimGCqOZpEun7QGJ1Sfo9IF1aQlysJMApgoBGRgsd0FoQWxcaVgGKxWJP7vcelUO56K6wUvyyv/tIanknjFIBChkaxhEYOkYcqRhMYDnABrUwBSKC4qSkVJpDs8kcAr3wlD8yr/ryhKQlISOQuWaqN6loTxKAgPNHRxPvuIB1ELkpZf5dPDKPqxsLSQs0robyaCsKViUT1N0mjBKsp/jv223kyvyus1dJintC+db2kuvuzhtZbWlY5GGnUOoCkm+dnU4S22ak6mEgqIA1MNpwISnoLmTnS5wVTcuZbTlrCphSt+l9YxxjjFFYJh2G1ZUjRREMNhbiQYEo1H0ZrpH0ZrpH0dvpAbQPVieSBSkFNYaaKlAQ+hLKE4BD7i16wIELEYFjMQHFcxCSpWkfaI4oLp6xtO+MQgqgqrCDlcIlPu9ysaw5JNq4eyYLEy0aiviITOup4gDCZ5o6giBMsH1xG0b9tPvjEn2hG0bHrp98GZYHriFTzY0BMKnXVcIpAl5h01VXxMNyTaeLtGGxnco0Sb5q0DUpaOXtQSSak3NzDzfCsxL2ilfZc7J68t2cXglnPCnvuZRtHW0dVAegtNvHJud2fuuZXgdQrodydlSyvEB2D8IQ442aoURH0+a9v4CFvvOcSyY1iRky39o5xch0vBoQb5szAmHGyrsg5eEIFN0X1isFUYordWK3bFbuSYT9mshUV7VYZOJxIG9PDhilTEqzh7UTgqye6DnFLgKwBlCCOYhLLa84mJd4nirCdrQjZecFD/sQoPdIo5zhKIpcIlxRsb1blNoXxJBhUiydKiDZ/Rz4R9Ad9pMfQHuqY+gO+0mBZ/Vz4QmSZGtTCW0I4UgXVhodm546C60X8DWAaq+UNNLdWEJhuzmEjtdowqRlT6lIesxQzaVXuMEEGhizpkrTs1ajTw3LUX2G0dTW6yW8c4D7IJ9ApIUkpPMQtJQtSTyNLpVeOXbPd8ryARQiHbMaVmg4Y+qnf3Ewiyh6znuhqWZa4U59bluIbFVqAhJCgCNDc0aputNrgd8jAgXC+sYoxRXeSlStBDcr7UBIGkOyONwqhuVbR3whpCDUJgHdeb2ghDATc4KoIjQkXovbcpG2jaCFOJhRxQoQkZXCEJhAokD0NYrFYrFYrFbwKml6zVRutFVZkjoAIstAwLX303LTZGEOjXQxJqwzLXjT37loLxTJ/jldYjf2brnU093obVawTajyUK3WW52Vo89919prjVD1prOTQp3mFLUs1USTFmrUWcJBy0NzRoq51sONqQeYjCQSDuVjFBVFdysVgVMIlyeKEYEDIRjEYhGNMDCecYbqwVRtI2sVgxWMUTCcLp770xWCIKVDSNo6IDx5pjFdSDckQ2KrSPwbI53OGib58Uml99Ist0dpo+I3LSUBL06mJJOKZb8a+69SglJUeQhSipRJ5m6Sa2Uq0nuqfP0NtNVZQ57J+d0k5s5hHfl7916bYa1Vn0EPWi6vJHZHxgkk5wzIvu8sI6mGZBhvUYj37iTUVunmDjxpGt9Yru1ityaEwjCmMcbQRirAXFYxEQ2/1jGIKoWuKwlcJVGKCqKxMpxCt6IBuoIwiCiKXGKwmNBEsigxXV/ApFBS501NL7RlytAcTqnXwgEg1GsMWnydHmITNS6tHU/KHJuXQM3B5ZxMzCn115chFmM5KdPgL7Rcwy9PaNLpRrbTDSOpz9FMNbVlxHURpcw5tWUL6i+0nH0KACqIIuQ2g5rdCR7zCJiSZ4G1KPUwq1HPVbA+MKnppXr08IU44riWTFnvY2cJ1Tc0rlc4jGkiFs11hbakRWKxWKxWCb6RSixBOcAxzhWRisVvSowHM4WY5XJVG0it7icKr0qxXYoxRWK3YidIAhIhAxrAhO7T0rSc63KNBXdmbOCqqayPswtpxs0Wki+WkVukFeSPnAASABoL7ScxP4fZF1is1W470yHo7TZ2U2vortXWW7ktvzF84ztWFDmMxcmXfVo2r3Qmzpk8gPEx9WYRiceAEObPF2K074ZYceVRI84l5ZDCcteZvSaitzyM6wRD8vzRBbc9mKL9mO30iiukBCzyjY4RUxSKQU9tMKTCUxghaIwxhikUhIgpgiBBEUikUikUh5FUwIpCqjSEv8lCAUnQxSKQpaU84BU5rpAGUUjuhhvCmBrv09EBWAKClzqs6XTb+xaJ58oZmnmTkryMMTzT2XCrpcQDrBlZY/lJhEuwjhbTuLWEIUo8hClFSio6k3SDOxlWxzOZ8/R2wzjYDnNB+Bul3dk8hfv3AhCdEgXOvIaTiWYmZpb56J5CJWTW/nojrH2Es3SoSIetPk0nzMSb+2ZqeIZG5pVDS4iopChQ0vpFBFLlqxmAmAiCjNMFEJRGGCmMEYYwxhgCMEFEYYw5RhgogJjDGGCmHpY1xI1jFTJQobiiNnGzV1jZnrCW4F8u0VGpTlFISPQ09A0nnctWEXrQhYooVETFmkZtf7YIKTQihiXn3WsldpMMzDTw7CvLetN2jYb9q6SZ20y2nlWp8B6RaAtCknQikOtlpxaDqk0ukHtowBzTluTM0hgdVchDry3V4lGGmGkduYNOiOZh20lcLScIhSlKNVGphmUfd0Tl1MSsoJevaqTehWIXPIqK7y60hKYwRhjDnFIAuIikUikUikUikYYAikERSKRSKQUw4whWohUnThURGzdB1hacMUjDGCEtEwiWHOEtITyvA9HTdQnEb3FVO69LtPDtp84mJB1rNPaTAJSag0MS9pEZO/7oQtKxVJqNyad2r6lctBdYrFG1vH1sh6W2WKLS8PWyN0g9s3wOSstycbUiYXXnnAUUmojMwzIPucsI74ZkWG+WI99zz7TIqs+USs4HyoUoeXhchWE3uooe6MMYDGCFYUiphvt9oxhikUinoKRSBFIpdSKblIpBTBTBbjZAwWRGwjYiEoAikUikUinohcdxCcIudVQUuUoJSVHQQ5OzClkhZA6R9Kmf3VR9ImP3l++Nu9+6v3xtHPbV777NZKUFw+tpfPPbJg9VZC5tCnFpQNSaQ02Gm0IGgHpZpgPsLb6jLxgggkG6Ue2zKTz0N9osY2wtIzTDNnOrzX2R8YZlGGeFOfU3EgCpiYtIDss/7oUpSjVRqYk2H1OJWjIDne0vlcpIUKQrGhVIqqKw59oqkAACm4PSn0NIIikONlXOAmnPep+AaRzuJoKwTU1umGds2UYqQ/Kus8Qy63S/0NWTlUnrXKBZ8r0Pvj6vlfY+MfQJT9v4mPoUr+1AFBQXzz21fNNE5C6xpfE4p4+rkPH09ry+zf2g0X87rPf2b2E6K3picaZ71dIfmnXj2jl0hCFrVhSKmJezkp7TuZ6ct1CsQucbxjvg5QTiyEITQ7o9KfQ0ikaCNbtdwfgEIrne4vEdzWH7NQrNrsnpyhxpxtVFppDE26zocukMTrT2Wium7OP7Fk9TkLgCogDUxKsBhhDfTXx9POS/0hhSOfLxggg0N0o/tmQeYyO5M2itVUt9kded0vIOO5q7KfjDTLbSaITvJVhMA1F00ypaKo4ol2yBUxoN1Pphvi+l1Nwwn06E1N7q+Q31tocTRSaiJizVDNrPuggg0ORiznn11CjVI57k6/tXsuFOQuseWxul06I08fwNry2zd2o4V6+N0k/sXs+E5Hcm5Fanqtji17ol5FtrM9pXoULwm95quYhRzG63p+JXWmUJrTPcMJ9Gb0pxQBS5xdPH0T0s08O0M+sMMpZbCBfPv7JrCOJVyEqWoJTqTEswGGUNjl+BmGEvsqbPOFoUhakq1Bus+Y2jeA8Sfl6ZtdMje+z6yYreVQ1w+mPo8oy3DA9GbkpxQABcteH0yiEgk6CJh4vOlfuuseVzL6vBP4O2JX89Pgq5l1TTiVjlDa0uICk6H0zblMje8xXtJ1urCzDPAPTH8AYHpEJKoApcteGCawpSUJKlGgEIWlxIUk1B9JaUx+Sn/VdLsKfeS2nnDaEtoShOgH4NSQpJSRkYnJYy7ykcvV8LrPmcC9mrRWnj6NS0J4lAeMfSpf8AdTCVJVwqBubcpkb3mAvMawoFOREHMw3wj8SYQFUz3D6RKK3rXhgmsLWltJUo0ETU0p9XRPIRZ7jodwpFQdR6OafDDRPPlBJJJN1lymxaxqHbX8B+Fn5T6Szlxp4Y0ukZnaowq40+gJAFTEzaJPZayHtQVFRqTW5K1INUkgxK2jXsu/7rkOU8L3GkuDP3w4wttWenWEaD8cfRpb63rcw+Maw66hpBUo5RMzK31Z6chDDC3l4U+ZhhhDKMKfM+iJCQSdBE0+X3a8uV1mSe2dxqHYR8T+HtaTodugf3/wC9zbim1hadRDLqXmwtO/Oze1OBPAPjdLyLj3aPZTAsyX/lD1mECrSq9xg5RZ81X7JZ/rclZTCVA3EAihhTVOH9HCSYSkC9bvIXPPIZRiUYmJhb66nTkIl5db66DTmYZZQyjCkXTk/XsNHxVEnOB4YVcY+PoLQmsR2STkNbmGVvOpbTqYYZQy0ltOg/DqSFAgjIxPShlnqeqeE3SkyWHP4nWAQQCNN20X8DWAaq+V0hJ4/tFjs8h13LTYGTo8DCSUkEaiG140JV1FwJEIcBvUgKhSCn8GdwwlVd0uUNM99LfW8kCFuE3GJlx1bp2mo5dLpXY7FOy0unJ7HVDfDzPWACogAZxJyYZGJXH8t+emtknAnjPwvs6T+jtVVxq17u78TNS6JhooV5HpDrS2nFIWMxdITeA7NZ7J07t20VVmSOgENI2jiEdTCQEgAaDctD/FX5XSP+K1uJdprAIN6mhyhSSNbx+AHoPWV4QjTcDZ5wABepwCConW6Yn228kdpXwiRnC6Shw56iJyUDyajjHxggg0OsS0yphdRpzETc6XeyjJHzhKVLUEpFSYlJRLAqc13KUEgknKDaf22Q+zhKkqSFA1BvmZhLCK8+QhSlLUVKOZusqRrR9Y/oP+v4u0ZL6QjEn7xOnf3QQQaG6QnK0aWc/VO5aApNL76RJGk0147tpTGNQbGidfG5hGzZQnoLlzDDfE4ITNyytHR8rgojSEug67imQdIUhSeX4kDO8NnnASBeVAQpwm515tpNVmJmecdyT2U3SkgvElxZw00HO6dk9qMaeP5xpdJMNNthSTiJ9a5a0oSVKNAIm5tT5oMkdIbbW6sJSM4lmAw3hrW555DKCpUPPLeWVKus6SMw5iV92nXv7oAAFB+MtSQx1ebHa9Ydb5Kc2owL4/nfabOJAcHq6wDTOJWaS+j+XMXzk8E1Q0c+Z6XWfL7RzGeFPzun5spOyQfE3szTzPCcukS802+MsjzFyVqTCXEncLaTygtHlBBH4AmAa7qUKgNDnFALyQNYU70vmbRSnstZnryha1rViUam6RZYDaVpzJ57k7J7T7RHFzHW6Umywqh4DrCnm0t7Qq7PWJqaU+ronkIZZW8vCkRLy6GEUGvM3OOJbQVKOUTMwp9dTpyF0pKrmXcI09Y9IabQ0hKECgH4607Pw1eaGXrDpcCQaiJOcDwwq4/ncQCKGJyULKsSeD5QFFJqDQwm0ZkcwfKHZyYcyK8ugulpVb6v48zCEJbSEpGQucUVOLPUxIModeOPQCtIXJyyh92B4ZRNSamMxmiELUhQUk5iGHQ80lYvStQgOg7pQmNl3xgVBHpsKukBowGhAAHLcK0iC8eW5aDr+0LZyTy74AKjQCphFmLLZKjRVMhBBBoYkJnZOYTwq3Z6SxVcbGfMXYlYQmuXSGGFvLwp8zDDCGUYU+ZuWtKElSjQCJqaU+r+I0F0uw4+4EI//IlpdEu2EJ8z1/QLRs/ZEutjscx0uBINREnOh0YV8fzuIBFDD9mA5tGncYVJzKfyz5QJOZP5RhmzObqvIQlKUgBIoL5potvrHfUQy6ppwLENOodQFJiYpsHa+ybrKP2bg77nZplpYSs6wlaViqSDcCRpAe6wFpPPewp6Rs0xsh1jZd8bI9Y2RjZGNkesbLvjY98bJMYE9IpukgQXhyguKN5nJcLCMed81LB9I5Ec4YlmmR2Rn1utKX/OT/qukJjat4DxJ+W7PSWrrY/sIl5db66DTmYaaQ0jCkXOOIbSVKNBE1NKfV0TyFzLLjzgQgZxKSqJZvCNeZ6/oJFcjFo2cWauNjsfK7SJOeC6Ic4uR6771oLbmFAAFIyhibZe0OfQxOS23RlxDSCCDQxLTKmF15cxH2b7XVKomLOWnNvtDpziz2i2xnqo1umXdq8tXuiSStUwgJJHXdCiOcB4wHUwFJPP09YLiesF7oILijuzrzpeWgq7IOl0lPYaNuHLkd1SQpJB0MTDJZdKfdDTqmnErHKG1pcQFJ0O6lCUiiRS515DSMSjExMrfVU6chcyy484EIFTEnJolm6DNR1P6HrFoWbs6utDscx0vlJ/RDp8FbyzVSj3xZaftlHom6dk9qMaOP53Wc/gc2Z0V87553Zy6upyF1ltUQpzrkL52ZLCE4eImGrTQfvE07xCVBQBGh3KnrG0X1jbKjbd0bYdI2yY2qY2qY2qY2w6Rtu6NsekbVUY1ddxTjaOJYELtGWTzJ8IXap9Rv3xKPF5nEda53T3+U55fKEIUuuEVyrdJTuCjbnDyPTdnJfbNZcQ0us6YwL2atFaeO9MTLbCc9eQh55by8Sjcww4+4EIH/xEpKNyyKJ15nr+jWhZerrA8Uf9r5WeU12VZo+UIWlaQpJqNyba2b6x31ESj+xeCjpoYBCgCDUXWgEiZVTzhr71FPaF9pO4ngj2YSCogDnDaA22lA5C+fd2kweicoaQXHEoHMwAEgAcotCYcaU2EKpCbSmBrhMNOBxtKxzHp5ieQyrDhJMLtR08KQPjC5qYXq4fleiUmV6NnzyiRl3GEqC6Z3Wj/lK8BFnn/ik+cTslq42PEXSc7s+wvg+Ua7loy+BW1TodfG6Sf2zWfEMjuTU8lnspzX8oWtS1FSjU3Sso7MronTmekS8s1LowoHiev6RP2YHquNZL6dYUlSVFKhQi5iYcYVVOnMQxMtvjs68xfNSwfR/IaGFtrbUUqFDDMy8zwK8oNpvkZJSIJKiSTnEgztHgrknO5awhClHkIUoqUVHUmLNaxvYuSb33dkytfddZbVVqc6ZC60lVmSOgF1lu9lTfTMXrdXtFqSoiqjFmuOubTGskCm7M2gEEpbzPWFzL6+Jw3IedRwrIiUtDGQh3XkbrSZUpxtSUk1FMoRITKvVp4wqzVIaWorzA0ustQ2axzB3LT/yP9MSRpNNeN1oSzaftEkDqm6TnS12F8HygEEVF60JWgpVoYasztHaHLoOcJSlIokUFxIAqYmrQr2Wv918lZ7kyanJvr/2hpptpAQhNB+lTki1Mp6L5Kh+Xdl14Vj/5uSpSSCDQxLWiFdl3I+1e6w28mixDtmOp+7OIfGPosx+0r3QzZzyz2uyIaaQ0gJSLrTcwshPtH5XSDWCXT1VnfajvA35m6Va2TCE8+d0yrE+6f5XSzuyeQr33PKwsuK6JN1mJowT1VuTrpal1EanIXS1nY0hbiqA8o+rZbv8AfE3J7DtJNU3Sju1YQTrobyKikLThUpPQxZi6PlPUblqffI/pDK8DqFdFQ9aTqskDCPjBJUak1MIacXXCgml0pOFk4VZo+UJUFAEHLdfmG2R2jn0iYm3HznknpfI2UVUW/pyT/wB4AAFAP0x5lt5GFaaiJyzXJeqk9pHyvl5x1nLVPSGZhp4dk+W/av3jf9bhSgppcTQVh9zaurX1MSTW0mE9BmblqwpUroLm5fHJur5hXyukXdpLp6pyi0FUlVd9BdJpwyzQ7q+/ctMVlweirpWeaLaUrOEgUhKkq4VAw42lxBSrQwJGVH5cIQhAolNBuT6MMyvvziWXgfbV37lq8bXgbmZR93hTl1MM2a0jNfaPwgAAUAidkq1cbGfMQ2y46aISTEnLuspOJf8ApvJAFTExaQHZZ/3QpSlGpNTc0y48vChNTElZrbFFL7S/l+oTlkpXVbPZV7PKFtrbUUrTQ3BRSag0MS9pcnv90JWlYqk1G7aTJW2Fj1fldKT6UpCHeWhgONqFQsGJ+bRgLaFVJ1us1nA1jOqvldOqwyznhS6RRSVR31h9vZOrR0MWa7hewclRaquw2nvrAFTCRQAdNxxAcQpJ0MPy62V0V5G9E1MI0dVDNpq0dT5iEqStIUk1B3LVR92vyuZXjaQrqL7V/J87hmBuUAvfnWmcuJXSH5p17iOXS+Us12Yoo9lHWGJdphGFtNP1KYlmZhNHE+fOJuzHmKqT20db2nnGjVCqQxaSFZOdk9eUAgioO5M2calTP+2CkpNCKG+VllPr/iNTAFBQXWoqjKR1Vc0nC0hPRIi1GuBzyMJJSoEaiJ94OrbI9iJVOKYaH8t5SUqFFCohdmsK4aphVlL9VwHxhySmUZ4Kjuus14hzZ8lfPctBGKWV3Z3WavFL09k32rwN+NzX3Tf9RuvTTLPErPpD8+65knspvaZdeVhbTUxKWS21RTvbV8P1aastl6qk9hcTEo/Lntp8+V7T7rR7CvKGbTQrJwYT15QlSVCqTUXrabcFFpBhVmy50xCE2bLjXEYSlKRRIoL7VV22090MpxOtp6qFz7e1aWjqL7NTWYr0TuzMxsEBWGucKtRz1WwIYd2rSF9db50JEy5hiU/yWv7bjicaFJ6i6yl9txHUVvmpfbpSnFShhmUYa0Tn1O49NsNaqz6CHrReXknsj43gEmgFTErY61dp7sj2ecNNNtJwoSAP1hSQoUIqImbHQrNk4T05Q9LvMqo4gi9t1xs1QoiGrUOjqfMQ1MMu8Cx4b8+0+p9Sg2oiJFB+lJqNKm+0GsD5PJWd1lJ+9V4DdmWtqypHuggg0MSs2tg9UnlAtOX/AJQ7agp9mnzMEkkkxZrOJzack/PdXZri3VnEAkmJeRQyvHjJO6SAKk0h20WUcPaMPTz7vOg6DclrJfdzX2E/GJeUYlx2E59ef64pCViikgiJixm1ZsqwnodIflX2PvEEd/Lcanphv1sQ74btNo8YKfjCHW3OBYPoJuW26ABqDCrOmRyB8DEg0ptiihQ13puRD3bRkr5wuXfb4mzcll1XChR8oZs1xWbnZHTnCEJQkJSKAehcnpdHr18IctRw8CaQt1xw1Wom8AqNAKmJeyH3M3OwPjEvJS8vwIz9o6/8gEA6w/ZMs5mnsHu0h+y5prROMd0aXgkQ3PTKPXr4wi1R67fuhE7LL/Mp45QCDofwy32UcTgELtNgcIKoXabx4QE/GFuuOcaydxqWfe+7bJhixeby/JMMy7LIo2gD/kV6WYe+8bBh6xU/lOU7jDshNtatkjqM90KUnQkQmdmU/me+E2o76yEmE2o16yFCBPyp9enlAmZc6Op98BSToR6IutDVxI84M5Kj80Qq05caYjCrV9lr3mFWjMnSg8oW+8vicUd1tl100Qgq8IZsaYV94Qj4mGbLlGvVxn+UAAaf8lOy0u7xtpMO2KyeBZT8YcsibTw0V4Q4w83xtqHl6ALWNFGPpD4/NX74+lTP7qo+mTP7pj6ZM/umPpcz+6qPpMx+6v3wXnT+Yr3xUnfAJNAIbs6cc0aI8coasRX5jvuhqzJNv1MX9s4AAFAP+UXJOVc4mUwux5U8OJMLsRXqPDzEKsicHqpV4GFSU2nVhfurCkLTxJI9HrCZaYVoyv3QmzJ1X5XvMIsWYPEtA+MIsRr13VHwyhFmSaPy6+OcJbQjhSB4f8tlppWqEnygycqfyEe6DZ0kfyRH1XI/tfEx9VSX7fxMfVUl+2feY+qpH9r4mPq2S/Z+JgSMmPyEwJaXGjKB5QABoP8A0D//xAAsEAEAAgAFAwMEAgMBAQAAAAABABEQITFBUSBhcYGRobHB0fAw4UBQ8WCA/9oACAEBAAE/If8A4DHsA7z7XD9JufwcJp6Y/M3T8gjsH9eI7fuJ+qz9Bh/fsN2Hcvy+02v8UzVPL/GaN69/VPirH/y9QY5WpqYfH4of579lzT/is/M+id+yP2i5c/8AA+HtE4l7TD/PfuufUDY+J7Kpm9v/AB/t33n7S/HuciZOD2/mW5/Kvq+KguaT6xX1ith5EVqXq/iG78kN8z+//U/f/qX4fvzHY9uO38kDoPxGuv0Rmkfyuvbtw5PmZJ33OmX+F+hAAEdE/wDEVnfdl8l5QtRkCrr0fC8TRHmfxN5vYV9ZrL8z7TQd6P5/noLmpD+m0/OSQfz8agjn8ERGkp6Ll49y9pVgPkSiD8+/8Jel3nxcC5NYauW6txBWgtnEvOWDzfsyTefy5vn/ACSq8+XNNp2cvmZqN7o1XkTorh3zyVifUQVctFWf7+293/d2lwP2Ob0USh4BMwV+AgNeA19/4vYeXOZ1fAynN7N9LyreM8zvxx2nxP4gJBHRP4wEAnEt07z8ZaqD7MVJ4ErG/wAe2j5JTfEPqQolbjf+7PT1Qf6lx8x/LiCBV0CUCHvRXPqi/wAlLB1Or4gZQU9WVhHKiB3huQEKmm57410jkj9ZQy1NiILo7Ii+Z3be8IWteGONYWYIK55PlyJovzFPxPhvjf3x+SGfXVG9yW1j5D3i+aHO3vj4wFv5ldZ/oeIBOmiNn+3oZ+t4l9b7j+Iqqra4HFJoGcruyGcU4Tvu+vX6ThuxNAe8xRv5sVPDflL4fUHB7LN9m80a0HrA+c9VmsJm0V0KU6WtUaZ4IQ/cMMvMu39sDAAzIu3Dk9mPnv7Z4ft2E1gi900Z4n8z5OyW5Q9ntMnayv4IBAR1GXyzIe0ZmOgXu816Sm8g0fD/ALXKX7b+Yve3P2wM5V+P3SsAc7n1/gbXIV9J6E5wcRxU2g1YpSh5tcHvVYZ+syC1Kp98G7RIBv8A1bq8vb7EBfRJ9+gh3R9WJO9hfoexcdk+zBAIljtMqHI1xZG4Df2cMwNFeTFMoeojMqv10lw+qe3SgQCOzLh/1PSWecHZxzM9z4GVmjfk8/7FwQbcviWVnG18sW/MvSVb3L9uloMN3KV2XVZ0Ic/YZ3+BAAKAoOnP8aejH3yxQigzZ99lfWaxe3Z7wH9Bn36ArbSO8b84eS4eDKXwPafVcV9Jf/mx1rS2CXWjxMKTcj6jHuKKVeAly0957Sxit4SAl+v89TUqbMtFs5a+kRIE1HAc4uPvMpH7lf68dp9lBcz4DxgDcmgSlzX2esAABQaHS0Y1pqfbhCXSyfAkGecbj3xvXK+0QBs3B0MCeuKQMaiLlnkHDcSi2PpBRshtcpR4cLMafgE1XAD1gh6AHpKK/IIiNME3khdMBIGwNvMRC1bWIwZt+jSd79wuPBm3z7yxS8+zxKKTRS8+8LITZyIAFEUbIV9JriMy+DN14GB5pcvkNmMhSNjO2P7uvSBtOsQycB0xzAbH7LIHMWSf6tQLcIgqqra4ZF0dXpKL3D1f4BSneDvcHl3hZfJskvxn6g4d0MPMRSJSawLDt/ZiQmiV4jHAT2wv6PND9ZTvXgjmh7BFvNo+VieU9t3cbisXg/tLUnuQIoq4uIACrtKgQ03PnC44P3RzdUB6wQ9AD0mUui/IwS2w6nDm/BCUyAes0GiMCQ5nvlgg3+CQAI6jNf1vwiKRKTUwv6s1tH+56rRr5f6kP4rVjdv5vPtjVh2d7Dpg0D+E0OPqQmYlBHzMzi3/AIYerO5snDNOY8pg5+k7/coDD5mkT1CXp3VvNl9RMoCtmby4rUTlkgTvjPhPOr89I9AOHOKP2k+IsGIAVn1lesuEQ75aY0UPu0tBy+bBlx2PpL8Mjl5caRufsDSajxEAjQKOm1d5kBzJonQQ+Hz8xfRfh8YMTD57M2Gz+0/09RN+exOJz/c98HrKcgmRjZNv7dFOMABLEseo0f63KTlqjrDQtk7rhWGHzL2ZD9fEQCJY6kafL9Im68NZF2zZWs91Wr/ia/gpO8Kk6aYfEjIGeeFmVbKQdkVxnRotoa1mqYRUPIy+56Uu/ke8eu0tZbVb1ejEWsQU46G9g2dzxM/5n6uCIEWJtDyr4P7/AOlpzBobrwS31Hx/2wtbEsOqf8ehumoF6iuB4lgzjVw8dRqfsvIlJQet7xkEJTvfmODVQBuXtfG8LovXfq0Zw3EIHdw77ANs7WdiUYUcS/ZOzjI8jF7QhtcdYPTaUdjJl4mfAx6vNkBWgggnhdUrrAoOisUFrMoo04iLHgjzWrLiPoR1B2TUY+3fJgFQjSaMDIX6/n/R1pg0N14Ioeh4xgM852DvKZ2vK9Q1WTl3OIiNOSTY5ZPDpFfu0TK8n69ehqI1VA7XDQB/hasIENuvx0mmB1Zl5ub9NIVGRz4iPMgKfEu8zVHggNg1WWl+47/0lLb8/hAA0xQREyjfoi07YaT17/6+/wDoaNnydiL8k+mwDB9AIGH6i9Fiqbtz4h3mTn2c4IEANVn0oguZR2Sxm1mR0TIchBcYIUnd0AuhH8Jt1+el1giG9xOzjshHknfz/tT/AKU7r3n/AEp3eEO3B7yJrZDRDp1KC3VNZOvMDL9kLG0oJdDZm8uZ1wg8sBaqQ3amzx+rEzKIAVALWeuv8enzWHt3MBRslNkp4P5/zzvitWX7M/6L3wMl5dg5gsvLuvTYc40cPMrOJlKL65bw8RSna5eZkvI/UjK0rLPUe82otTk4iWZPx2xzFC3fFWhOT6QTI6N2vxgDUl/gmkUgdc4nRU0YdHbGdxOhhhdMjllylG53MocvLwf3LRv3DpHvvikX2nbg8Rdkizc+ZkUbbANiWIPh9+v0CPbuYAkRGxJe8Bz7eT/NUCrQasewzffzhfGXx3gHV9xxdIAFqxQZGX5QSWrod+0BEWJSRqHlcdsEI1bDeArBoEtEBSRg83nzEtZ0Hjv0IqCbvtgBQY6yxtnv/ksb9fmaxlBHHsbzO7HOOW2yyOVQW+Caql0NIVoEqibOz+0zqz/dzKi8z+0ABQFBgzQrx+CLcz8xXC9uiw5ro5fjB56rGbO5XC/j/Mt0TYwWdkoJy7/lGKgVaDViMlBjNeX/ACCGABkYHtYT183JHtM3Pu5giWYK/Sdxg2ifXzjnGQhNDHcc+JoGR/m62gORrvAJY45xkY7QwRAIlIziZc26eJug6/jM4cvPwf3DGtKJ67Ll5ldYjKj8zzgnYRzaEFhxKNT1B3MUAWJSR6/Ny7PGDz6anDiLd9h4f8rNHKy7OZmuG0YyOGOQS+vl1f3phTGP3vocPI58TKxCUk0PI/bodoTPXNx1J9Jo+R/oX80dllxFpJmuYx27kyvS5X9uu+fEyAy/ZgiJm13d4ZPy7BD3cXrhqD0h+s4zmZNDonG94GZHGYs+YwpD4vEAksSx/wAjZI0OXEuTP+mFZlGjl56L5HL+w7TREPSlOKP7e8dAgaRnqQe3Z6Micj2TIjs7P7Y5hoQ2gwUC1m374t/6TQtOJ57jEM2RiI0mGkgbL5wQfXKfDABrLt6NAQI9443ZffoywHXiZU8SnDS85/jlAV0Ilb2u/BSXkf2gIKAoOhMP1Bhp0Hy8uGVgDX7onYHMYlxf6HiCII5dP4nEjvxFs/b/AFGx74N4Ed+YjTiwTgfZvDTaFnrjfHPkeWKqrrHK65nr05B5fswBJSNjAC8I+/r/AI20qzccPXATLSgmq9quXFadnMwrrDs28w6ZNE6K6yDLl2YmDUEufRW5X4x1nfGNHLzFXX/Vd0OIbY4VpjHbnAERYlMbfc+SH3oen2ZYiORibL+lcWPnsbB2hpGbm57YozBqsEQRycckMzPtgzntHNFpwrHs/wCJtaMjl4ic3ecMiMv0/wBugGDD4Run6niZl3vejL7QNXqdHCn9WFonZgBWiFncChH4H+tZshHfjABSRF2w0Ap6DK3l7HRxXCeV1fErrG2F4fC28zSZ2GhM/Mn1/wBMRGsKZtu6PJzhQuuz+4/xM0s6vO9wqgbv4zTpzGT5PEvftTzAbg0SZKdjZ/bp1lO3AoLY+Q/15FYwxgAIxF22wAQCOzGbfQU+kfsbzq/OFYg+viX1/Nums78lav4mTxNsIszh4xyX3/JxhRqQR4SHnsPA/wALMT9i8EAFq0EDfub74n6PfeVFHw+HRSut2x8RFIlJqRaLrL5HopbmCBbFXbqE0Fgk+PDeOJQnqv8ASijZDFOAmGPQ+mOkuKfZPzFiO5BS1WmmUyR2QW7nCrHNlm9x9uhkXunZ4wo33nZ2f8K430439cLQdjolSOs/HiIiiUkq7tqAMNNzEjmeoicSjH7NgtFx1R0NPj+owYDZHdqao3i36QGio/5hE7D0nEvIk+IBuaoJvFTXT/RijZCHfAKGOw4PyNR3rZiIomZM/iZHtzK8xUpNvlyeTmE8g+yPKZJ6kYbUJXQ7ddVwkBCkaTDPj9+/8C4XZeN3AbbSiacI9+/SI+j/AHmSLw7PiaWW60Zkx5z6MQcBS/eZyX10/JKcMydDG/sYt5JyxlbT3lkL6QjSU+BUVup5Wwfxk0TwBP0seI9SbtPJc/GQzZc7wHlZ+ujBctfQzsLPDc/6JASHYYaBqY5eZvp/tM9/tPOG4or0Ybi8s4haq95T8e893TUHYwLOvve/+BXb9GMLxtP7XrYDSX93BswBWiIHfPbu4KQhEGFHzYpmynEdCPeNmfq+xLaOf6ka6YtdELe/Gh9JXdM6R2YBKNp3yuVvtzVNR+e96m3XAhJE2qcgB7nSbMxaD7zK75yI949pPGQilSHslEo3NzzAc/bERp/0LWEACYbH1wJawpm2bo8nMzWmg52d4RILWrqxmjQunOPNcRqEz5O2OzYyeGBPSUnjDNr7Po/zZ2fdtXA9bQD1mhpVjmTb0OsYzD2LfmLa/UICgE8LKeORuYNW1zWcct7ms8dV1dcPOOk1msPkTfJ0ZnXMNoFhWN637EzDOTc8x+at4uaNy1+d5VWwkm1v19Kl4Yqg371MhW/T3qFLnclq7iW52aaZ/iPbhV8EALVU3CvWJruKXaEqY5P3iWLITPYQq2tj95XVbfECd+ZWn/Q5U6OGs8btgTqjWfZib6Igo9Iy94LqvXcdjC7H99nL4fpM2b2YJkYlDTqY0+yyeDC7W/8AO3814Xb+d8KPZZPJx2oNDlllBQqcuyWwpmnen5j9KcOz6xwYbza7I+90Zp0MACsRVwp2xl4T9XE6rOa1fEI1AnrQ+ZSYZa6pk1NF0Hl3h2521EQ1tdtoXmohW4NLv6pf/mK6i1ckvUbdbW/MzFQMq6ZQpM3S/JKd1JTuxOY2f61gBFKZDvr8RBESyPc8t+HiX+oH3n0wxmH/AENC2phYntEproux/fZy+H6TNwrx7+r2lc+uz0DvSe40iIomZBRsnYdXi1/kJjtedogS1bXDn2r8nFUDlY8sAWyteWL0yajDLVcu7iLnsnf1xYhCIML2u2uFPxTXKGRx3iRHPm7qAACg0MO4/wBKHXlItDeFKaeh5mb5o5XN4mxBHGWAagxWiojyr0mQGJ0ae01w9AL+kzjPjvG+y/TUm6Dc+kLuTT/QICQSGGfX1wdI/wBqUfvOFOPf1e0rn12cbkfXM0lAnrjfE+63wsxkfXP5MtH65YZwNz6dF/ZPaco5dAK7Zp7wqJr97GovvphnO+2GbOhPTacsc3K7f2ISCg0MavKoMBDZMgzxvN/AJWNTZ98F3OY7Q6Mq9WdbMzGy8+5zNr6/5VMtg8sFpTM1aOAIMZltg9XJ7+0rm/znCsF7syh34CWFnYyMFTcGvcca6Nn43w1nhf6iD2Aj2f48ys1PBkYXQZv4OhGpp/XG8zPk4hLuD1xzYhW+/EQaZNvO2FPAUVh4M0gKgQRIWzy08956mDy4gqibsOZkBl/AzTRGrdICAjuYqZe0/TtKA4ogsTOcC2+7hM5cZcnt/jWlJkHE+jLP6Q2h6Edj3pX9vxDc96O1PSn1ZZfWZOc8vhlB9cXVRwugnazY0Undbno4sZT74feaNdzN6EAiZMbxh42wsRzPbt/FTzm9Rkw1nAD1miQQenRZmCkiQ3m3POAFIn6ay5EOMn0iIEZEM6lm++FS54eacLFxmc236byhBrl0B3yIu38Cn3JYxVyy1Zrn37B2jf5FmN6J+3aIg42FG9+EsHlm8+J3EfwgFWiH2T/EQFTKrfIZlIn3YZ815dTZPomrn5Zsi+qfvM3RPVNHPhubNfCo5gE5dkFSKVsTACIxkHbClTI5+SHllXnKxv3sj2h4QNjLrpIaPUMLQcqfXMfxaI/pRhXPT3mhjR2uR5ciZny/7Ri2juPxlnPncSXMsdl/WV7NdVq4MBzAABhZcGFcRgLL2ZvNMoKMb2+2ksV0Nf4QN1ufBLk1S3rLOMlT5jNeF0LbtvCr6nQLys5aUV38BiBORH60fmALc2sQBN8b6SLJmbbmUJQ+RlBScNZcCO+8szMsaxKlWjoRsEylyxVnO5VNaM0TXBF8J76MdhaoVbho+/Cmh9mgM5vTeIhYln8FK3FcGqQPsilWAT1/hzMyo/DCz9fhxalAXxJRGho2HpPXmuC+52w+RSn3Oq8mWo2UeuKAG8Cg8H8OfNs7rLD5XN28ECioYhmrhvywOmwXtpLo22x75d+jL5c8jx/ABn3nvgNJNZ5mMK8+gKlpRuuxLLBavV8T5g7sKfqx34pTvRzV0ZpDziUbwAMlQyXYljG0vupc1OXHLC3xVAar8GZsc32bfwJZTO6GemGcXOy9NP4Cd/ERVVdXDs2j5x1lx/RHtELzNTk4jleTh46qZhf2sPpEyTyr6E7qzPXG1cJknH8Byp+jOXdbt56/0hwwNiP/ADdnLM1lLe71U7KjhiKegSOS2rlj7dTH4jF3ZUztO5Pc7hdlMisQVLr/AEO0vovH5zSHNMhKA5sMkspqKzZZ1EFO6i0uaiXlHOBAHmbHVlEqCZvha98K6Zv0GJEqRvLxiHYPdqa6nZ+ZlndXNOaszzvj+vrkwz3uB+j/AAcu/fMOD878ulAg3XpLAw1H3TUAPgZr/eye3cOeNawVCxWXnAUO8U3YH1gAaBUPpDlanNPbcvaByPEW1ev3MMZc1XMf3mRlWGx2hypsTmfd69A8V0LxfvYLTk6VqZyZJ8oZQQABQQwKqVUVwKlVp+2kyxZ+x3xN+KBmrOXtCeKBBW9RT80773lW+Ab3AhWlU3iO+mZvphfbjHWw8fgmlZ7uvpOIeMsvyXLnEaReCcY85ZXhmOW14/pSOF82d/4F/PX65YeAfFcZiwDuzTSOz8y3A5W5ZgcBcE+PZcTrz5WBouhmboMsPoGFVOccnjgKqO5d0tAD7svCe7eBR/xOXWW+erDCFmzT6uII6IxFQZKW7s0YfXreRgrz40LzPdMe5Mvm6MsQzh1feUlB0C14mZ6l0mhm+Z28PY5mXORycvQ7RLA+8dllIt/wCmkcyz94Da9YR3hcHJEpms7sp6TWCc6HuzOvB5prYd34hlG9iunvEq8mGoMCek0lAT16+2an0MP3KDGusWq04M5mIvBmmbKu6VBjgKwAoCcTXSuZbWaOgV1ODHR8RWvL+uDleNIuBUy1Zoeeg6kDVDd+eIfLBYog3rF95WFQO1AaHXwu8f6mZmr4mS0vbGV5b/0YPY9m/R8LoAz+zFc75eIIu0VrWgIqCV0Xjf7IWxI6HN7QaFAUGKVE/R5xC3rZDILZSj39SCyh3MVLIfp5Y+ZZ4WSA1mcTNs9CHSpuY6wrPO9RoCEa0RAJomPfMMLVx9t1/wDaYcPVj9X8YoOWGFy746HiNOCde2cpn12ntBwANjKfMnyMdBxH2jDbsry6TtvaUaDoZ+CeZm56MM94M5lSak3IGSrIH3x1PEEu0DPJrE/fma2X6bS4NZZrc+Cny8WO5q8Q2WR1VNIqza4c5ytU338YogJ+7eXp/gPm5XrLtCe2ol802HJEGsFmN6IFxh9Iwt/7lYXWL4ZVn2H6qAEEdEw7LL2Z48H0Y2vjgeZ70Tq7pE9sEAarUIzQK9v47B2wsXbozeCZR2/noN38Yp34JV257dLMaz12IoZOfndM5nZ/Yvo0ghs8H+WJzcsHAPYrFfYXzKBx9OfJcFouBU9WHRXTNA0X0Zpma7bfwpnXd5gj9gmWuzca+N+vRQzkwufbDutREAUvPSE7kv8AExW7gZsS3WYyu9X1esuxoQe07q3znj5E/Z/3Cvhnuf31fuBbh+oVfyHN4LIcvQrgf3Lr1l9J1CzcCSpflhVjgSyypQl0Q0Lae6eVD8cTV+WMoms8s2aKXyrPhSBLVexADNgjm4LQgMDCuuakNMK8lQO/8Nknd2/tLQdbDqvLg2tCvVBFwdCrz4Gw7Y9ll7M8eD6JQ4Jf1PPWPlBqMaFq0UeJ5ePZjd+lTCv/APOv7dXaZ+Bw8IL4/kOrlweR0LK4Z4nfadyc3pjkQpmH78ygO5t9+gK0wchdWuVlHEOqfbMM+P1LxtyXmeYrTtgC5TSdq6LNtf25x+ixc3CuvVgjTFb7YUKL51tBbHbZrSd/EJZvZwwzC2Dscysozd/16hsMBRd8O2AfCW3O+c8awC7bifgySBzcCXnnj5QXzhcu0+jq8pX7VhZwn9TG+9g92LynnF4PlbTPJqXK3cpV3pgbDgqyfzBOpNHkwFYKtdjoqpySyGoRZyVnhxpmXA7YCjoKikV41Rcz7WWyrZo9GbZT0uTZCRkK2ODS0FNlztIpm7m7B6wEumeccbG+5fOPaRL+1jc90p6y5+xX6s8gwWWAfwLObmOJxJcpVga01fYnzB5IKAqmzu9agK6Ezkt6y/WUGV0a8eMFBLoS89f0jALIArqV+PA1gLXVql6TMrRrPj+Ct977Z4d1V7PV4on3cPFgY1FyU9Rr7GFz8fbKeZ2+YtC8EW1Z2qD4wFF2wd+foVDxEpnQfWXW2vFz0Fd2jKe3Rp9u8RkeDK3UgsiaJ0khsXDXYfQgyAkvgDbEFAT9XMW28ai7nyywTR6jLXtlhmgddxoiTSaQstGZhmOzh9sb7sWQ79SgW6TWqmlP1lLrRev4Y+nU9coukgCZr5gz6VkwHxwWYr3B0XfrpNAe5k9feYPjHvID05BwPq4HNxPJPtZY3DYd6vryg9j75ztI3xBQ5cTD5z03dqVia/0sq7dlxchLmekRVPRrdbbiWq5ePUnnLjEC/kiVLveLoZ5hfmAM2sJ/ahe0Koh2yxo708YaY6Rgmv1Pxktn2wVHWLZk1IOWmjsmyb3gVyHtKHIj2g33mUhW6aIObZUZZT7/AEZ5goAu6lUFYDdlSWFm68PErB01Z74mvp+U3nAPrN/+iX9b9N5nid9UNdY1WuOp5w14Q3KzPeXNr5BhQjPkEyBX5CD34HXoSxI4d5F+Ole3AOb+8sfWz+kZObvVNm/FM7wNxHUQObYgOtIC8853qP5w+Ri69Ne4K62qn0OYIgjk4t2Jvg5iKk6cvx7qw8oVR3htyP0j1/HSjsCgxo+nO0Oim01n3ZTk/RqZC0MNXQoFrMiu3aXwZi11lPUvEHcnmU9kvcdxKpk+esAE1XeJf4wX38eWqo8s1U7Pv0uxQFsYsHxkp27b56KERwzXwLM4gA6rlg+Z1fyEMhqxoZnJLu9PIRAUibkFDZ5e/QaLuwVr+ldbfkPoYudQN5hr5moehTNS97NGHkxqfl7Z4fLMHTry7+ky6yM+8aSef26EO3ESTx3P8TeSD2Z8xBkNZHMWpeBfSKh5MMzODt3j1eZ63meddMHEC1l+KolmZxM+h9SI7+E4FoyEjxfMvmPEz5Ip6l4+JHorz0jrzQZ3fGvl/RPWj5dOqq/SJUswX4Qu7lo6fc6L11FPWcJIr06PT59H7zmCJdzxLo9sHtEuY7AS/CNAvo/Yc4fp+Ou/y30OvUh5JqXtYfQ6Jh8Fg6PWKIkGzfeI453IboFhkwU6Jaw8xNA8x1aissszfJ4m1XETU6NZ+SQ2u4VWbxKMkntBzClDTWuIixu8jxhuAfUj9XPf2iAwAUEazB4Z8z2DWUIBxpLrZ8lztTySkZp4iboA7szJG6drG+p3Sd31TQxSyhAIow13a8fmOlXaoFsRt9ZsEd2dZglBtM8tv1xVjgmhpk4JpSj3cPt4fMxvz3TMkasa8xV1YjmJ6D3lJaeA95VU9uj9Zz/D5v7nY/gbsUGveX07rKG2HzjB/gSwg6QRMyZ78/q5S5ME6HuIm6NaraLVl+gywXXWfMziANFwTXKWsAyTsvgn9hMiDIcjxPEnpzgqEMwc/MzTp0NsGZc1pCRXL0B2hsUP22WqNGNj3NNyT2l/m3cjPC7ShkPrEJpXuTNSruitegS7XR8oU49hGXQ+GLRQ8SrvXdBd7nDchpLlwnmzCi477wd3T0GOaGdpDGoADAW7n03go2MD83OKrVlX+8dOHThSKnZeFGve/jKBPxPaABQUdX6LnD9vx0jJ3+9gsrs+n8HgVvmWHsPfDQ84ur/BljTiDzx1Ioiu7eeI9cw9/rO2SKC/2XiZ1fqS7go5uBlx6TL3jpmLMg+iaoerN2KJn+Ha94LICmtc7lwzjdVVEzXH14bYdqzmh5hfV+wJYtE5vV8Q4QZpTEMMyHOWbEJY0pRzpY0wHIlPIPmGrY5JoGeGe0OQ6C++sCsD0aEreGaB9cCT3iOe98kConwpmkzUv0YLwpe5LPSdbglODOirqKn65PPfMqcDXEA0Gc+x9GKDeXD7+F7Ao2dpp/3M3p1s+B0K37uAoukebfe4PI7/AH/wVHvfadhG+cTQw+S/woWQ7vxLXzuz5lNCHvDKfsJmI9xFDnd2PEKsWtpkMKlmyDa4bsXuiSLBsZaWaStjrt3gjMm/KHELK4hQUVhekBa5n+zKelez1maaCKZgOIDS8KG+xs9TDsrnxPFy9yuyQlDXqLrS0xXmQ1AjzPCa/lmVZ/eB2hmjgVj1iOdLayHTd1IPPyHTdxdbqCIRsbuVhQu891xYSAABg/Y6+6K9iGbOxwGDtcGPagPjp8pH2rDyYvtifSHLlNQ9pPsgLBaP4UBo3oQtZ3m5RUHhhmYO/DgK/iRSbfvicFbOcubflmRO+6GcAyN7mU1d+RNZPr0aOPLOT+M4f3oGUlxkhWg/bWUZOWQM+Y9axz3jvIXgG5I1q2yzlskub3vNr44mOfiWRJZKk24NOc0zgtGmLia4icoYPcS1Gu0RGno9AZ3I25l/ZO5RlXBmS+IKogcxZcO5waIKvVYoqJnr1iCu+D+OC8DM01NP9bP6zXvK/CbyeFJtf4b/ABKPbQE5ndP62Pap/jBAN2oFAdPjoe5hSXKffG7t2+cTQCzUX4U0D1iovj5EILUC8HeAfj/IranF9poaMsM5yZPxHfNEB+/Efswz/n/zHizszMUzzczD2VfWa0L3Qd6GWQNJro4EcFRpljL11nNJlZlAZ1K83WfxKShKmuYbBNzDTAg+32jen3EyqxjuaM5Z7QrFV2m+DmIqTpdLPNSj8EvQfBU7ECQsnK8XVBL467sddoKLtgrwR60oRv8APNobW+5NseAmv+nb6TSV5KnhqvsY27x98sO+p/PVnPM+2HnqnuY3gDUmjfOmjLwJp1LK74HX/P8Alk3UHA4YSXkTRz5gAniuF4zo27y0qVEmqLKIGUyLeZlRvqzX3CPcJQNE8XSypvAkGMJtkLtp3jruNr2o+SVRsmEL6ugp24iTd7Of4u1mGVoSrmArFWnv1+Vt9id0H8Y1rhffCg+fsX1V/j5l4d4V89KgWx1t9p7xegd7YBzX0jRe86wIlmCyu2Fni/z6oJomcdoVFbpYdTpCrDKbLJVQS8WE5JcJlYS8zlyGCZiZS2qKlRRTGRoy+a1QZqkIq8WVdYSAjOcRcVSicYjy4ow5TNCxze6FcZcDmylaEtodaiGYza98zE9yD0biWq6DSKL1VMxfdgB9UdfJmZ0zdrlfNg6TtgoCuhA7A7N9NrUMwszjGCgZ4+TN9v8AuF9yn26v+rwY9jF9zoq7GEACrsTN8ruqALDsEK38DAHPNXDx0SoXb+PUBO/RocGidWZ4Vga5raZg5LKt+YEzpWGmAjAG4QBowJc0xTZN3KLRGlULEpK7WaEEVmh+Wk1aYtBpAdM46GmkUrvLMrKqqZ3M/rFagbHaFbZxfZmzMNPpBgvK4hmMFrnDcUu1+k8XtKv+R2FF9oLbmY1VsEvLPWHBzWFVeXDmJo+EQtI9pojH6awfySp9Sx+J7Hw59Nn6RcMzxX16+xlHi8sO+Lexxao5o9+sD/Moggu9a404ylvvMshLPJDK2/dDJisOFC79WpCaZnhTdous4DqIQA6kpGSoA1KloSjVmRSxCRMYzbaayI47gUpYDqQFvyqRmgziqLUQ+0WR1tYGE01xnLKFjfaxzJhbfD1EtWBqZDtFLPLf8ZR2PuEJiSI2cuIlr5mZsvtAzbXWW2cprK3gG0ssqmX8Zy13c+ZTwusAOq4KyA3YZjWYvDIODDt2pLmjp7vzNg+B+Jw/3XFKc4SsNCfa7Pmasur1Y3/h+GWHmq9suv8AW4y+2F3FfdiW795M1Lint0oQaKqXz8Fh7YZnQb1OIRRm4VAlSpUqVgpExBUqqcrmREpoQaHmkc6tMNd5mwZ5my02gtRnAFuCLRZ3EKt0OcDKGZc0BOac2ZGIDrBpwIASqMsjDH1cLqDcuvRPvLIWHJmpRl3k0l8Ids4rtXiPJFVL5idhmR0HpNANJnvMCKrCoIcssi5y0g9bS9WiaTgB6YWDvj8iDUTpfgYD0CcOc1YL9NJlDgULvEEWgWxl9UffDLjavlzevxVX65mHmqe2eKCNcj0muZdLAw/HZw0g9NgdA7QSrSg8wR9s9sKHywte2cUNWOUbdppTUVZUCBKlYKlxrGGG0cJmnjMOyCxYA0hhADOjCMMaxIAu7FyhvOIFocOUhA8CZJEqzys7ISlkpfeLiVQhhVwH59BCApzJfv2k99JmXfAs/KdNJ9bL6wbRPRP+5HWB6JrvpZ/SfhumXB45sysO7M79RpDrlkYXTF9GGu58RUhXdw0nHGp7RkEXTc6bju0fDD/qkMCiuvR+dT9WHa9Xx0LM0ZP0S8F7TLk0juLy9oCgC12igWKdpGGCTj1JaQYIQJUEJWBw6pml4iy18ktsLJlncjWnWGh0s3g6CUVastHK5lEKQYqgkAaJigCnkymXGOGWNN1loh3OA6sjRqjpdwpCHP8AgEsnzyE+hN/M/uodr5p/038Q3vmn9VH1cv4nxSEswU25wycBiWImc34OYWq8zkexBfmFIa/I3vEQIjmMQNY5uehXyBemHAX23+DSUQfDNcZ16Yc95XzkxVAR1GIKXxqT9Bm53sK+s0/7mbhXY94tdhY+cKXtlhph+2WMGK4ywwysZcuLEKug6+2A0KJUiUw1sEyqIZZ0gJuQu0zlTu0ROxMYYNoxNMtvLIDSXsOrNgJfXdni7+C2WwUlJTBaXhUcoFAYWzBf3ouG377Rn0V0yPlIijf7ToruwMFKDTP8KrNEfo4XL2a+vWPYnbd9Ja0fkY15oWJaZ52iOGSc4f3OEoxSNMJcuOIM3LixhgJkXK5VEFad0SwtlFIlRmmLhmzhbA1iZ7/S4M1yksI1am4LiYZ9oDEasSuFrGE2Yaf4Wa+3C25xfzQ9oK5mv5ug1vOvTOKPj7jHScRfSagyL64ZT5+8zfw02Z1vjBaryV+lYl3nMtvQs4QKVd2UD+02lK9z0e0yDAylPgFpmvzHLWXiLixhi4KFxGUZzPNiV0SnWId5s2LpFmjhIZRdTNzg1rHFac6kDBSiGApAw1FeUc1QZwhnbWDD/ABbUp8HjWJIWEDSIGxJSH77chd+ofuiJT2VviXuUMuIjgGv9uN8XP2mrhkrkPoM3+IHd8PO0RSJSazSeuR53xWXsWWfnD9xS2JnfbfHfJL+E+2QJ8iaszu2/TbCpeemAey+YdzIznk5xqSkpgLAvAeSjEZUrVLkkzwUW4LKURoy5JrDdykznZFsuOZUVOJUzGyeuJdQhwRy8rjdNPMpmZH2SGsiEISv5l3Zwv4W3ocp7mz0lwL3xOJ72vhDroKDHJrlV6uuFxGj1nX+PKBsvXXDMLp/djRB9ZMNZ3wmv/vNpmQLWj8ytF533T1nth5lezv1cBREgngqpNdYMqKvJA748qVyy3dNGcVorxDCT1I93MqEWGG8vL4ObZEgzxeqEuHeoZkY4BnKVfKw/wDOeRBuXEQ2cIJSapq0mcOrhECGG2Kv4UQEMhhdRtgB+vLzlpuFtzBlUnuN/DgBQE4Y7b6RX0iFieaz6NJJmaygX1wvUyPX/jVsbjA4oH2OsMy8fiIKwb0DY3fEzfkfqxS3e5eIVKH7v5mr+i4jEuZ/bhnjfA2W8R1tGJE8EeKU4JkRtHI0icYHu0uhVK9O3D0wMrM4ynjMzSdjEyIN+kmfYhkw3aWucCncQRtgrBlsBsuFcAIGL0Uide/9MLF9prgrOmzLR7OWvox4oNRlJ61qeGWDM3Wp1U85q3wYX6fyCFnal6zUrOCxHffbozfmfqxtadjY8Q0rtfqBElI9Hf8AqK2putyhavGIQ7hz4xzHffDLup9IxwYsuyRYCBlKZRvKZXUeZ5mjnjGxhtgZWJxoNAzVdglUNyZLK4ds8ILibJNUIRkZUrACBKwei5rHpZDtvAorDR9DpoGZsNSWnpWp5IcUGiSkOzhr6kDnTc6K6fsTCmGa9E/lr5l7s0wsh33nboBGtqLuMuBTs8RUtVWUK93+EpV7n8MMknjczJ/sxzg8bviN+6Z9Iy8oVJAh0HSBMYzSpUqVKlSsQRzYAphDFSpUegAFiiBLgYBNulYGUhJFdDiRxDEFaINO++GSNXBtaC1ii05DYi355d+TFdfcxfX3mN7HB2MaiP7DgJuSHrNOtB/Kztcx2aQEKRpMLVcvumKXQ2huMpk+aMws8pwRIAarLjMeWnpHbU1WZ8U16eO+O59MEK3hY2DbzIZxJHSGBtKx1SpUqETCpUqGFQQiZxJUqJCJGMmGiAZMlQzMqBKw3h1HEjCMuJhRn9MDZR2WGZi+594nq7TpgKz9aPxFgiEd17p+n8k7T3YABQFBjWF+wuFAcvcv56Q+fx1wzA2/XbqBRe1+80kbTpBS02JWUe1fmAAAUYijZB7m+BdtpM1HUw0KPEOmGDCONYEqVDlCJKlYVKzgXgckysWoULI0ZsK0SoESZmV1OJHBwJbtpjkBodCARLHaXLfIfiWKL90ietz6Sqt3v26bmP7t4GXaAHKwEajNytf5y8ufZEAUjSYXE/WcXIlAX3D8RVbdZXX/ADJRocu756kBIRJgw0xpzEevEA74DjvxcWEqV0JliqVKiRMVZSmiSjjBRgTRNX8VRl4p2oAFGH3brYG7ktXrz19ImBDUcmN1j1a3x0X1v0LwoHteX/BpT+g/vhSW/VvoI5f0lzK/zroeP4fEMGy8GHfzJlGFy4sV43AiQ6DF16mHRS6S4UlYE0fwt+oyr3gCjCnR/E9LY6kzUd15cc9Nk7G7gJ1iA7s2KGby7v8Ag7QrJ4dmElThPGGc+T7/AMzszG3Kz3MFy5UTTxYsIwwqPSMKxTAlTKCo0l8YH8S9AiKj3htGADvFVt/lWugtY+tNBwYXBZH9r/h2gen9LhqIrTk4ijWFn83YHOOfHuOY2YFznw8WMHEwcDDRCEcXAwDXDJlGB/EMMGOduYAowAd+IiVhgAZrNusH+TdO/wBhhvaM3g3YJFVj/Ds1AicjE53LnBnVnej+38YNg91TPr54PYHI3h2BziLofXFGkkcDR9uhhDAwcDoOhjDDeXrKExCaofwOFx87kQAKDAB34iK2V1jecOX+xg93prv/AB7oeR7x2LVtcNrhf2P8UlBuL7esRSJSamGfmRn3Of4ESABmsWfqh8RopO63gV5oGpZKx22euDZWBEscK/q24SlHM5DRg6JwMDFhiw6HAhgIxQgjp0D+Gs5Vn7cSyEKq1lAYfPYlryH0f7gYfoBKcZ+ov8T10FrFm3ke2G9Y/o/x9ujo74H8zJvyanDx1q65nuwOPIOr4IEzW83HDPde8CkSk1I9PMfbBjtxAbHBUFjPCcdDgPS4GDgdDCE36z+DSZ5XnH8zgky2xuvaaMx4SaMx4SBMtu7r3wRNcNDfxLj0H8DetZ+Xj0wBHM14OYCOR7vP+OF5Ck7MTZ3P7HHphnDPS/eOhaLHpYn5cAVyx6IIHnf9sUqksfEA/bffB2xmQuTj255mpGXODgQ6HAxYdAwJvFW0QadIZReII59NLpN32wAwFtZkJkYWpotrSZuhKkVec/QW+/eKAq0Ea6r9jxHrKaAgytX6rrz1yfZgCtEOiy/Q/wAl5XuZzKkj54U2cM3LpXjoe1xxNogl0FB0ILt2nvglu4fr0NkzkAscVZ5GJ5IwxGL0PSx/gMM4ll6GcJppjlpmxm1he1fAnNYv2mSsDJ+hjoEDSM1MP1prtucw4BGQTVMGbx2MArALVmX7J5e8NECxxvbN+sy8QlrhcyY1P8sNZkehCAERpHCj9SHbodHBHtUBj2e5XSCvMhrOUxvzg5RXjV+I7XqfymsZtTKMjig5M1JU1bJzDpcDrYwjAreJcqug88MtIC6EfqqaWYn5sy8yMMvXg3fEt/Emr5gKgFsGissPqwBhoEIpEpNTASg8/sMAAAzZmkBy5d2Xhl8eYBLWq9+2CFZGhy8RitdDg4wpRmepAAKDIP8AMpZO09nPnAazIQNkMn964qGzyeDEQGkbIQzAn9xiJaaR+jvgonm++DLydr6GLJm8+k2ndXDdMuJ2V6OPeIHVc1ArpOg6HECUt+lSZV5m6XDQGIuaK5e+KrazILlnR7d+YqWm7gNdLNajx0Vkcv6O+GdBTJx3JRBkvunDn/qwJnt3YO81ZryuFGYTRsPSMDWWc+FKSDyP867mvM/kYAEpGxhB6BgQBYlJFwK3J+hhVwaJlB6fL/CAtLwMMp5D9fMruBlgydVYeC/LJQnvhb4mfVzrueZX8TKb4BmcOO65TXsoI9C9q8R4wntENTA63oCDaKblCbm2aMDo1ljNFRVbXG1vCNneHFJoE+YePvEAUjSTMPm7PPT6FPv3MFkOrZsLlOPoBAw/UXDf7BnB/wBJwFrN1dhywa/qLl/z0Epj5xf09sASIjYkBMAwogEdRiT8Ve8Wzb3fSN0D5y+szz9VzDahoGNT8n0DNtXU5OIzeT8PDC9v9rBXYBfcwRaIu6ylUJyN4auqNs6+LamPHUedP3qdpO+Tvk7SXmnOBcwOAGgdOqIQui5v1eMauVNWaHlxHrtZfVPVQerhS0ZOX5YZ45Xvy6feAfqTTiPCQ9WN3de+FBYnBz9XC5cvg5YPzX51/oQCFjqRBZep+9MBUI0mjLFV6T+3XonDdzXOZDV66D6r/aOgRGklWZv1p42lrf7J+YJTWQdtsOBbrwJmAS6NZHToSgdQYrWyaIP50GrBx/aTdq8S76H26AZGH4Jfs9IZ2FJH3pquSawK9ziKNkX01Yi7o74NqxsbviaXT6OFi8+Dln1Nj/r/AEaARLHUj56/X8Y1/rHPUimqmAr7Hq4E4ZD2xEaY9txdsWRH7thYZm/QMcqmivg1leK/IRO7Cx6AtFAoOAhPup2Geeeeeed1Ey7Qiu5Fd/R85bU0BvZ+ZsU7q/pD1ecUwNef6EXFgVOxgtx+h7QRBHovqb/4zSZtZ3o/t1XB2/Kx54RseML8Tu7Dlmr5+df6ajyJYqu/5vCBhTROhK3J9AzehPRh1k0TDl+D5MSLqyPfGuHI/LBOzQHrNMQjGojsfvNUSqGVQKPSXS9KzT3lPxNBy7+dO62+DOfkvT4bn4RbgK0Fs/UZ6xLMwQNsBXd+hKu6D4l+ran9TB2Pex/ekEAjY6PRkpn+n+2D/lp36Lqn4vKJjTVcKD0NfSCf1gu/+oogbuz+8VCikdTC+GbV0ZfVR1dTGvaf8JjEw2iGxusyVlLmmPWU2rHUN952w0kmYymYX1lqzJv1dMS3AyeXSKqrLTMj6jh2Y33wtW/YOuOarDJrVmTBKX36TbB1ehE/ZjR8RV1YhfhmZChZcvnBoSipek4f7qhhVbQ4wRVNz2egUe5lg8Pcwzki6mvcwREVSCQRLExNawpmfZdy+ohsoaBgiQA1WLd6N9z4xIXE584FEO3+qtTlGR9+0yx+HbwwRsGiSqr2Nj5g3mYZ0vDuRhQ4HKEWvnMCJ8+vtKUwfPdwoXX482FkTc+umOiv7ZYZAaL8nDy6r0w4NuvBw78D4w773x0OBWZecEeAWDWpk6ecPRCtZ6jgyr7gxBFolMRzVB9J+1adBzv1mwH7olriWvdWsPFJqucHRQzQwz8J+qgVlFidNv7Q6szo4D98aRJ61gEAAoD/AFiwyfHiZ0+ffyxrfV/tL3n7vU601bW+uHigrxgCJoC2Kp2PG0pqfZsO4g+0W25UJkk8DP64Zsdz6Sm8/kw/XLV0KBsr9MKZqZtGoPYHZuFnepOeeVZS9xHR2NUes7HG/Dl0DMwqFo8YlEvxQCABoE89h79yfahCGho7NB84oEANVlxmPLT0iNyargUc37nMlXPt4f7DKzubn4jVD1HAq4NEilF6PuQqVNzpA3O9+WF5wZOz3lYV2ZStqJoHGFiGbl4YWvn5MsCEm8+rFU7HjaW6cn5J5UvZ/wBiAGq1BP0Ae3QMeRTGxZeg4CjY1PgItnzBIBPU9oYDQJ0ZnmX2w7Fq4nV+9MHa5OgBQBjfW7H7xPR2nTHOjztXxKNBu7vn/ZZpnA0eGehgankxs8+hlGfYfqoCAjomOs1xG/4RooNnLEiUjf1QACgKDD9XKw7FD2Jop+2UUbME9JoCVfVnjw+2fUrKmznL5Tdmz5hfhlfmCqob5sHbcizw6LtuzBQ8k++Iz+/6YNWdftdJHsObL/wRq+uJlv0vMypcO38/7b18ho+SUHL2Oa9cb8pzsfSVXZjOAQU3G8agXuRe14X8x+/BXL4gsoaBjZxm+/8AydnB84Cp2PO0RFEzMPPp98ukGXemWUf8kr+IW+D3b4m2iz3TOXyeHQK++e8RFHaUcAHpjy8q6ylbvOei1LvOZc+j/dFVtwPKTQM1lGvZavxB/br/AHCU6aiWMv8Avz+qlij8PhxsRe0JVn9dIBmH0Pt10sCAhe0deFczxjTpsvO+GT4Tp5aT3EQBSNJERW4/GJMwvFS4Hv2fSIhatrHVnB5dJHWU3c4aoonBn0vhA3cpfF+1ke8sj27iCoBbKD17X6T1Znmv941ctRLJcesr+kWoT1L1xFGyUx4GeZe5zohV+Bf4LsBcFml/tN5mIrU+OpJQ3+IcyPerPclPE+LRRwXsM42OAP4FAtaJzbxnmUmeXNlwL3xOOTQC2Vj9vWQPqT/wABAI6jLkX9zKWT3HO+0RSJSbYoWKPM4F4zzSN5f5mnBcfkgtgnJ/jew1ec1kvFHzMrO93zHvc3oZrue3vND9dyzy4u76/wDhR/U9/cirW9Se5L18YXx0sW5yNTT0+6G+EsivhKZvryU+0sTUV4f4VDWfP0JrvoW/SaE8B+Z+Qhrrx/lPZAvLpsDuy5Qo3EKxe8zfEAAAGx/4rVd5rP3mZv8ADkmmj7qfmfIjk/g0Z+GGi97Avz4qv/ai/wCZNSX1R1CvXagvBnNITn80TmZ2N/WZrnc/oQ8ADYy/8jqFeQp9ya+8TZ8weq9D6XNL/Q71Pp4+xPlmE/jBVBc1G+HN4ngJ9wVQXwg/Kc197BVH8Cv/ADfzmBmre3J/YZFP3/Wd3hZdx+/eB7P07zRPVLnxNGDZB4/+B//EACsQAQACAAUDAwQCAwEAAAAAAAEAERAhMUFRIGFxgZGhMLHB8NHxQFDhYP/aAAgBAQABPxD/AF9y5bLZbLcLZbLZcv8A81ct/wAa/wDyNvX3/wDICWlFNhvtGgJ4l+Bo/vBpSbbef4ztt5SO178m57k7n0Jb2+GTbHwPuMR+q/JPk8v5zT/4/Ykrs0/VH6F/+Kt6XArqIHqy9OGr88kuzv4/CFmH7Le42Wfo6/tWI3zVFfP+BRURsZKESW0CrO7L8IUJa/YbyrvnYPq1dV/+EvpCVrnb0i2cAOYnZpe357M7pzb+epqvPf2TQL+vZNb/AFvRZ8oP4o2t8CmxPhP5g94XjC7fqS3a8shdZ5T+J+y55qas/wBvJmdA8iIiKIj1VAI6P8d5Xmt2YVaabf3ZweBLRYnRf/gF6OENpF+DecYbnXBr2vvrEFSq2r0Vnm1T3lImed9rw1PPD+Vpqpc/YUJVXfJZ7wAKD6xlFcF90vfMW/DL1a8krNC4KvzLl7EB98fKDUSk6Lx+7RhedDORewbUR/4BV6K7TlK7X59bpTE9XE0pGgC1lKiff+3WUK7c6+7KtYNv5LwAAP8AIr5O3F8+b3si74KjaRbLA+Hop5UvtXlc9L4SJ6n+7vot+pvFH2ozfnfpTob0BLot/I/lpxm8rN+ll7/BPZB4DygRKnmulqJ6h9+k7RkMkeye9iHDa0CJ2T6aFQUosZYm25R+E7UvPtsKptXL5x5q8L3DJmZb0qXyjp6uMP8Ad8ENr9XaMyh3MxjvDagWr2CdgBv2JXAf06UC2PjRXzlsSmpGgKM7SXZ7pNG7iVDSoqNBscXfpdSZGjew7YDXnFAgWtGteBZGRWZrm7nDM6ex+HuYVNGX4/MMrQrn5KKoa5YPxKIIu4+2IAiI6J18NxLq8O0tR7n6MV2/Ua34GI/cK2XYHJjRG2Srw1gH22chyJjf+xXFLtZeq4Gqyn+42sumRararg52KEp8BPTPJPwS4PqnX5l15Mt5Hm/Ym2AlB77L8u7BO7W/jgCoTJME9pel64Ei/ISqF4alhnxAIgiUkNsysIlUaaOq7G9cKCwDd8Nwmju7ftMK7JEiPVtNvnEXmQB9jBS7mfs3Gcoo8Kq5fJnjfasfrvzh6MOWdmm35UrAM7B5OziIgj0gxagWJ3GeNSM16bR3eWozn1xuaG379xZ0XIXP/wBpKAqziXnWsb7rlocDQMAoAVWgJe15RleNpdapSZ+b6Clwt9hcVtza29sTO+xovnsSq+IyK5NzDwh7etZo+eCNl7okFuT9k6kNpOgUvwo9BhNVvtQQNBOFOOV72SX+Wj+z74HWBSixILBkp0zSQZ9avI4V7RfHyBudLN9GUzl5j1UoHdwaHlZnSRUqQsSUEesf5i0s8ngcXlh2WVA7sG0XgbOI/wCrXHTwsarwN2Od8b3uJNzGsn6OWF8oa/o26dbtDgPVmc5HcHQrlWnv/CA20e8NQtAwNgyDpXxX+GAm0Vbj5i3Klo9JQ2V2S/akEAiIljGbL2rGnpU7HQIVZ6+FsfZveIMKTbFj0UftJSrU7yrCXne9Y2gJtW32S5a6RvKB804f8MgcPzYGXQlUh75PXIfXgFQQnl+x493VrP8AAsl1Wsrk848pqBSPccN9fWhOBuTOrMjZ/wDWFx4B2X1If6tNg4Gxg5+aFazJvrhcnnvDrgoCgDY6dK03lUtKdhfYDIhC5UPFc4NH9Ficrfff+U1lCHkj/X4yFmK/MoWHuLAkPqTpfiZrjA8mqAERGxhua27/AIXCvtdE31wN7Mu6qfDGEKiE7IR0AIjSOzCLoJ0J+ZrGCYR125lyiyNVW1lfc3jAFpo/kMoqqrH1XrTV2ZILdyG7mRnzKpAWhmKJptgo9kgAAAoDQJTmS+AWz4oCjoYxys2SROuqgpQYGomYwSOG8L66LSFZc/yRi3jQ/lw4sO0mocL5uH2P+pXBAgAWrKfHl9pfshVW1Xdw10ajZOxyzOvRvOuQRGI9siWXiIeFI7RV/RMB0hQvonAXtewNjHvIQOokP/dpasFAVYm9svm2Uc2gvpwaS51ZweCtFRlwU9ne4RGmcltDBOUKfUuuJvcJck5OIWlOzEotcCCI3CgFqy9hrb+mQAACV25+hGf7Qvb8lKp8MYQqGdZfPzBQtWzdrRwyqqb77LbQl3VTR3j3o1w0/KXk0w1BLPcv6AqWoFiOyMqCd1/OPWQiFImGeObV/XA7OlGWbgf6m+BpqAjQ77bv8MW6k9Pz8EMQNCoD6PZceyjKWOWJeODyU6rYVunJPKCClyQzm/kwd0tZPf8AEfBmg3E3Iw88uJ71KU9prFGN1gPuxc79Sc1xdqWrr+JeF7DDxyHv3bPS7fNgHzL83din6hLxi3zSoAsXht0ssLyOOX4noTw8T3jDUJ/Mn3mYtjWOex8YJptF7W6w66EHAFHSWNheaz6WQbY2qx6BLQmQeyXibbqHKw9re43Dch11fcuKP+kJ6H1WwbrKs3bfo3grAKFasohqtX+foPwpq3d4AgaRANiOidXZ772CNPRtkJ/rZRnB4ftRwm5MkJe8dzmDrgRCxHZhl262vpdSVUByGGHjd5P8tVnv8SPZCK+mxuNjRRU06bxDn6lrZDlxFVV1lqAmqswR+3jOV2dlKIll2uwbB2JSVB7y5dJlqb+Xy8EcCwrVWZvM/cZMTkmOakuugOutoLypbbvoT44OBqeLUoizHMnTx8Yh/wBFpGVqbIN2Ib1Q3I/OBufsQ5XYg7L/AMw4nRsC6yvS4qyOA051Gul3vtequd77lwAgbNyobzPqhfFrSxdkxYzeVuz2iLirMonbUGEmxoWChex0grQKz7kMoz+ShdZ4KgdTyM0D1M4Dp7MC0M7BKOCL6iOsT0i2vtxfDwsVocLpvkqaQfBmlbydN2Iv781nDooaXuhAfDowAhXQIZ2NuREwBQtAOgyzytgl5AwdsoJ+5g0maY/Up1zoViQtkue/l7YPWQINIm5ADNBNO1xj/oNCyg8IN1mdwAbgLmTq/qoBZILGb1VCH8zYiBRSOSJKPea1P8nT5cv3h6wxCCUlnRr0mfD2Gc19PuwyjHY/waGfIIyj80OznODuYRFEpxQSksjlvcgHVd7HGgfiLtKH6nNgCgoNA14P3cy52qgIw/RaH8cfdu5N3+hs8Js4ggUUiWIy8J1Fr3YAqEURsSEKglfGf6FKGb1WwbrHjbobl+Rd3DuFk9UzuFU9U4rRbFNNWl2JoIMJIKSVFAG7GrAlaXBEe8YLUSBMlDhPJ2eheGIpOVQ4UYvYp36H6ReCV6g+7ORuc0ANDo0qeWaQ12IXWeWo/R/MRoTwRfWCv8if2Cf26f3CAae7Af8Akgtb+k/iDNKQ/wCYoI9BNExtr9jmTWeuTM69I+uNeXyxdZot1gjR+Fp6EAp7gkqDLEJqjYmZVqxcjvyYAt5C601Tscx0iAAzVZ429jjpdxb3bggERGxNRhF5Sba/za3glxejBG+dU/SzYZmpzfylO7Kf1T06jXY70tSvF+GU+GcN8Eb3tEt4Pd7ptBVUpl45MQuttm4cNtnk295S7Gem63XcxDARA0zQxuhWBk+glGMMVAtZzVxmjsgO7Psiyir/AIHxWM10HuSoF+RDbZ0X6HeJaNd46Clk1mjYFjsLEj7duJQbmGr7+Y0Nj8ZDZGOJyLewR+75DQOBDvbTZcMH0DIHYoaUYt3z1/K7+yGB33IUiZiMM9sWf5ZLgdYCo0ARI0y/3eF1w+g3XYhA26d/GT+5CgCWjWn7zCQG00py3IGpRBYjsy2hRebzwzJRkuj9iBjGhUBBaOIlrdn9RmT0qvuPRCKWHkvohQQY6T3xvH2V94dtH/HFERRlKNHEVhd99IAIiY37XYS+U7DFFJmrqVt4dyXWBTgLZrmGnBsehE/KALVdibig1v541nupNN2AQpT9kO7AUjAKANAwal+ovyiv9xT8XObwquvJtigiJNfxyd78sAmDHskPCqvyg7v8y7l5rdduDXjBaqymtFH9DIxOsBUaAN2Lx8Ms3e0sFnfubygSQtAwyG4PI7J3IJdmuztpLH2QQCCJYmiYDC62ei3IDw9Xdcrdxq7PmZRgY2I24JYi7RrFX/MUtTttKLwDSFhI7mNtTxbMsoHB77gLEdmZsDpGz7p2g9UbxdpsNjDV9vMZU9j8vYmZJWe6+sH5Hav2JsHkXPvkFQCd9YwJve0+x5I8IC9gYgKVEFiOzL+nOfm4Zp7JtrqpadDM33u4YD/kFQ/X9+VQqqtq4AG6Rq/5cVQqAFqx10bL3GDMktaRbt0amLP6jJjlHFbMqe1L0c3hlX5HYxC3thrL4XaNf9DaFJQhbnaCJY2YMRJLS75THuAfEXW0Bb9b0vuppH0zrw+XDP7odnvNml1f1TK26cna8cGA8gI9BIokWzhNx7MVfKPa9sQ5VkL/ACxsMO9RYObsIYk4DYjmJgP+It4KDYa310Ez3wvAbDsYahHm7H2nRRVFG12gDtpuyf5ZWrVg6CCf1AUibMXPun9VIIgjjkT2Xx+SVi1Gu/5xIC8vPiDaxg+ABuxm8khEqqu7/pHs17lAMmtyxuq+HZj5Amo4Cgk610NzBf6zLhR+eFJlTdW7enR8WPo5oUA+L6FFY9xNaDwYe6q/P+PVAAVVoAiNdqe3PAaJK+z2HdhnzgZAGgdHdKh6om5VTvuYD8MjbAbRXXMakYgnovr2sAgUWI2I9NlDlt/OABQYZub2CXh5bDQ/04oiKJNH0/5wAI2OGfitkUd8OziF+Z6UlBpGMpZU5BZi4wi7v+EiIKm1d1h10SXnZ06wc8Go7eTBzpwGkTMSO2FPf7Ht/jNJV0tfzsCpT4KS0eX6uNFaoEnnbIM/AQxI2qxOjKb2LhU+VJDIO0ih98QKALWFUb2nGJXQd2wiCpV1X/VLh5YP3jB0dkqbmtMAalgdxySECtu5Zp+RtiBzhsJ7kETYFH3bS5noD0UMHoO0tjAAAKDARcWqggIELE0Rx0J5SaN+HCxT1dQVBZ9HAP8AAcNMCjffQS4bT93Di+i7m/noP3AWVd7JASjnQcrclmvm/wBwMyqltP5jozDQf+XZAql2bg4AAtdCEh2vjABVoJYqm93f9aXfybMyvkNcD+8M5+9HBOOaT1kxZdrOY+oaaVH2w5I6B93C1oG/ku+Wgoby3jJea3J9jzisuuH5O5NRV9htBho7WZ9/8R12u7NLAVwyn+3qgAAAAoDpYB2nROVEL/EM/HFPjaqSJoTlpt+OkCgC1gkdr4wcIARZsBtz/rwrhJRnIamAd2M5u9WANWpCxjsV7T7Mg8h0RV4bYMNo71XA3nB8Oz/EBVFqsEsdkefCit0XyajNMvuAarHWmCOd/A0I99UWMolEV/OMD6y4Auu/mOfTgmoQGaroSvI+TP8ABiLEOU2xu5xeJtjEERBEpIcPracL1lIhSJsxyKrGuiXoriz6HGDFKCVcyOh1fE4XNAXqhBdj5Z3vcx/vzkPhJpo9P9KaRE0ZlABqc4DHYxPobucVAqgBasrboGYiDb7OwbESCNZUXQuPTMus20GB0PWBty9Y2r+vg+/Ra1WePcwZQbnD/g8oCrHtm3u6nqYa9N5OhS1QXUIogRSJSJLm6PUD8kGQ+SY5bbR5SP2P1iuOgXj+cARNBNIg079CWc99pu+lGjny5zudYh8xBPCfsmfGjD+5/eah5c1QB8Tst7aampjvFWr2OZM+tOTP/RmkpNGUtyOpg19J4YEeZhpnjjQ/gMRAIiOokMe6F77HFC4VAA7TO6btp9IMfki9T0Y5Ee5KBuXvXXkOXRUEV9ko5VZGomSYWe+rla09ZgfUXDKVu5zDU/jLuzYud8t/U9JRCvINex3i1nDzPlTKHXO85ApvUQepGtaZZm7jQi7rC93YRRmf4Ft6YZcc3dhkUYWDZdXlxrVnLsSofwpRQdLVfg3jS5tCOu5QHqZy6bur9kuGL8p+VYsiZqfgmD1P2Xkj/imM78OvyTZ72/MslAcb5DDvLyyfRlq13CXyl8H+ifOkhfW3OHALBxREUSkwrNLge54jRmoPG3gOABW8zKUbGybHMQcsK2K9MLstA6fs4iMMpunhDR6H/Azh0/J+rho93871m1/Z+47Mf9+v88RFEpIABVaAlb056zBbvGrwQY8gwpLZurGgu2NzDaTmmgjWnNKYZI/UzadkVLPZOQ3qJnSlyDwKZUfuaFr+25qIGsuwy5KIS0T1lWG3mV8Gvwt2ASqWp2QQfUlXVlw75UFzlkU9CHxaqfFgl6AVkAe9qVyl6cIOwJaVj7GIgRNR/wBCN0tzkjdWOG1+P5wvWuE1xXe0+ghY16JkIGjzkKkXhafilQKj1ZmmAVAFZbyGt8dNqvgMxjJEeMqnDW/VzOAPpOFSp8mYCp4b7qpoEi3Lu+ribWJevd3gjfgoXrM3ATsv8Rfdt9gZV+HY8g4Pd4XReUnane3ucc0G9gPL4hFUq2usBQBayhEdpsTO2las3u8EzYUZHI5mIoBWzbSDVRr5PwIDPUlIMhCOTfT8oMiTPDLXYqVdTMed8urvInuwCyv/ADaItCHioV2XC3hTP2ZfsJmGHca/AnLC4VYmX6PqQvnULK6kK8tJpDDaR+tIIAH+GJwtOzxTgR0hAdJ/ocw872cEAiWMRA154La2t+jBKWOqZvDoyyPblB5WRAlVKNAaYGVp91Hu0JnZe7SkGpoXkIIhakqwunH7E1h+TBKtq7O/9Z7bN+S8NLix7MSRXO6Ec5ZL+CGiyxHrkWvHG+1ilvWtXxCKWPIfZ5JlkVfsh2ejLDg7uCFUBPSCOCAqAWwSO18Rw4Wmwd4zZca54gqG2Puy77sfQEEUSzTOA3LMyF2N7OAwKIIZ+lM/mIZNO33lZjL+itKjwKdrBRVWpvA6JUiU8w9hTNrFCmJ4y47MF7j3oGyBAuQdQm/eDPEbWwsDzIUiWJL0k5nr3Q67H9pNSdYDHSf6HM/Iy7mA7hqu8REUjSdGXp91Hu0JnYe7SkM1meB7WfxGXK/8DY6G0GbXGdDthER2SAERGxNRhtIvxsnAfQcKeLVLu8h7xlCyNVc1gKgFrAoaPm5z0TjaeQaLvoYnJGhWJL1FQHWLh5nk0/OOqAs0Qgwzcy9xwoBZuiWjPBXnGuU1mIKwKAoDDKm0h5tUXGnpcBapU83RVTRBqkMtsc5EIoybDKK+Yzi9K+AnewH/AJJpAe0y1fkR+628BZlDFnszZmItQ8sI9maFtOZXu4kM71PeDdi913y5Bl0YikSk/wBA2VIzcYMKgWTlgsKgWsiLKbZTlkkAVAmaF7GfxBlK/wDB4xoSHuYttqXYXjl7HX44f+rOP6leuH8GG+XfxI9+gm/RXwrM6KuXc06HHWvzqTVzGezG9DPJ4YCu5lCqqzLzm7szp0y3X0I/VX4GA3xgGwYjdLXlU0FZBR5eWZbC52JYFfoXRLtHKGzhLGgtBoxfUToXKysfdEyK1j407H/lRtsK/OFfnCS2mZH4XAkskqapiwOHs3La7iEUd/weMONiZC+DVljtBDr11/XkBWiBsviNarHOkb0GDajDyGvqIWQNtwsfp1jZOGmQlfPbOgK0B+RSmlrjomqh4NWbr0cmKolTtKaSrs2TuyGDlbufiEQKAowt6diDQtWicbS1l+Sey5TskZ8xq4nStdCM0Mmf4EYbODd8w6ytb9BGnDDhNYljjf3lvsYPe/Kjj0nnExCwHcYUlS4dG1UfTNT8jG2XXw/xhIBFcqA1XIl+LeH/ADSzPGAPlh9d5JO5k/TeCyjPPBPhZVlvh3+yCAETZJSLoycLwMzLwxUmdJ2j3Z+g2ebNavTk+rLAbemR5dCUbnf5yCxGP7t6DrIIjuMdd053LNe2GbU9fPCdThk0YPwy2Zd1UCv2iFdBznFaIzVXEOfZkURGk0YXLtifKHL3qB84ovkCH9Mld3+DC4DPIeMLwHZMLAMsh5i5bv8AHeh5kW8e7iFxtWIteCcfOh9DRNV5diIYlTbGu3DoquovMiwwPEJPT94+R8qHnSOJ5+t/N6outB2xrGUj+H6IdgLVjYacfSRgRc06Qus2jNcYWlAeBj6SdrPlqBX41D4CaZzuv5n91F2kOw/mCX4lHyQ1vak+RcUTTX3AEq+k62HRcBYySmarywyKqsboA7cTit5kK37RNBDQAPAden8/vJhlrOH0ozueDDolbPUxHaU39xuyu0P+ltglx4L9gfMBX7acrjS7wHXvSOsjVnsIH6qaLAUYZC4CawOLM/Mt/VdrWgxKMHYxtxy+SZi5nqfooN7aiupN55+Khp0ktbQp0EtzIQVmhfk6MmwnjG8yZwBQeTSKhs9ZogLVicVQOurQilWCzEu9ALANZk3oyM/ETSK4tfMkF9Tn1seLg736so1oAJUzdc3KWpaxANdCHTtmmTtycomepOUojecksFHbk+pHTKlqNbuGaXjBnSF6k6rYo50D1IZQZDcSx+hU5frnBo7W72upbaVcgsfo1Xb0vhVSzVe0YixWGbSTY6x/D2YGopG7/E9dr9sjC9jmywUzYzekydH8ZIBjL1PFqBKoS1a+iWYLwGDNBnEoExAFAUEMAkLsaQykBB2eg6S3WebwlUas8QUT+MGb9QmrBuU9armmuWrcdQw71hpwJFZbpBwBWiBD79Dq/wArEiqvMxVClcnX5GFZZO5AviLhIOEtVX7Iq088JLJPEWaqD6xKlCok0ZlQMkWM1Kj0LPVij7fhNMCW3IiKOo4ZMi+fkobVp55/QAgWJSRXtg77HCq2P+m4wOraYg5RYR2lRVd1gKgFrArc/TjPFAIgiUkRIU2XzIqC5CabynabPrvLqylzq3y4V4cj8uHwQlLmvocNG5L9h8Y2PoKPLL0ORr16nDiuLbujZDh+sg/CUBQQewCuV0iCa3abJL6UD2c/EEXX7HHVhJ4Spka+pTP8Sz7oTqYIsVRzbUgnQlOzdRDNi1mcL5tgiVjAVqAIyyzQt+0Vsv1f7yAAAAKCWdhGU12hYQxKll5iaNUtrTqLkQHVDirmjtPYCfukfu7eUGJ6ocVUuZvdH6Vlkj0d0lHWZnfeP3YMLAGWfBrE+wuOWuvroHHvfwh8y9OBfmpLgVysL8RH6eXyMa8MjvCUsaf8YR0OFUtejTPhaRdHizemnwqTCne47su20XbQkD5Ap2hlSc6/hhBc8D4a48EXb6YC5oFsVXVLhed0s/Kh7/qaTCHgi8N1ED1ZaFi2T56QkuZ1OERDVVetwnWZfwO8Rc0GPeQAAAKA0JtMxMG2RSwtss9VO4sOljA5jiIo41JyPeEUPkg6QCrQRdRKQYMm60g3j2xC4gNAgjRZy2HUYAMqMqRtW6F9sE5/zf5O+NQ7ovkquMAH3SgtQ+7DZO92aKfAmqQf2aGg92ayfIjsmO5KCpPJkxG895NYgUa0WUzVaz1wqiZv4cGqzqiXYa76C3wb7RF2eEffrFSbqin1YdctAVZVo13Pt1htGAQ0NmNYmY37OF3tN71aOpwpt0J50+2FzGQjysSplt2hhcjzb81md8DpfMLcppfE+Ss5tpO/Ao2Rjuda2IXhr/hhk7V1iKHbFnEDUlE761n3sa2A/wCEMOleB6oWgmxr60QKwbqkj4LjVWPkDICx+wcTot4KPvKZ3Pvnj2R/cWOw/wAoQUeGdBsuyTKmsLAwSNoaGFRKZaWDvCoKZ4Eyy0Rs8q+dFPYzhtbvY6KW3zLtlm78Sx8huxEqqu79BAUibkogp7pW5boIRmry8jgC+iERB1GmIBEsSkiN7d3DkyqXGjPtBK8TwWFc82oPak7FkP7OkKy1vgWYM9XnJWRDL8ZCzryA3gSFozb2axM5hYyLqKVG60GqB97j2KISF9BD2MELRqixl8qtz7DKXtULtWt3Clbpb64V235nE0PZF3NfvhtPt7kPLDKFG/24qPk4sDtVZWl2hNR7IcbIWA+Ja2HVqeZY2hV97+7x3NWjVCiNgXkzKh7CrglHHLJVVXJflDIui3QPI9DouZFdwMfJjly7HsQV5OLBUC2WKFJYFMGXQuMFMQAShc2F7AluxbEDuMDYMdjG7xAZvsbqX18GxA6qSjk4CJDl0Mhp4KSzG6uW5szvI91C6Lw7KyMGZsGZaEKDDSHSaJfHsxztQYNjkmYNsLQ3pFGsCPZx44L8DlhZbmy9b1Vomct/PTh36L8/9fp9swgACM70rHNINyMdJZUBt+n9MaQLgAeAmTxZ8D9sVWpT4Fyg7uW71F+vtoNyHgrDPC8jP4Xcx31U+pozUwZVu5/cS4Fazk/ki1Cn9kI1gI7jT10ZLlmegpLYB4B8oYSjmS+YK6DgytariB9HyQwMFg71i0jhXKJ6Eh9EKhqe2bYgKzWL3fyoie1+hosnV8pZbYjMDLMi42vs2FWZU7OJN6T5hf1jhw0LusOafmP54czeDPyDkykvyy/MCha0WJ2TDIHTfFEz/ts+f+ONSFAntThZzm3gLA6FLafMB1gSlgHlmmMx4FfT8FL1RHZ5MOyV7xGI0jLobKwWUu0+/s5X9gYzXGzREJrFZdKNwZxq4YQbvWa06Dyg+vRyhShRz+XF6ACCPp7H4nGzCUSzkiv6iYdphGy2J/AAGElwz6DpQIsa6FOa6BfJU9iMmxkPpRQtCn20jUiiPYkVnYvZdCbP8hh3vRwygtN5VZCUACrcNpk3FBoG22oVLsxe/hKZD35exA/fA0XY5m9AYq/++t4P1vwwmDh3GA9HDLKxR8K36lvYgYU8A6PESvaceBvUOhkQzWnqYOC7bCMaKng/fh4Yqm0VlnGpn/lMtl4NyBe3tlCAvcfGK2e4wF4+ktdsTVW5+8sq/wB+NI1656pz/MMmQw8BEo8uCUHOeCDCkAjjsIEywVhw4k1fEshnq8fRps0k5/yQsRo4bh4qFlzGGiyfk7vRfvD3ywHlkY9pd8UTN+2z5zftGe196nh1Ja14U+HeK9ihInkYYTM4brVRKuNx51+ceS6PRsP6NzoThzL7VWF32+Xn1K3ycLeEL79F/wDRGLsLvy7/ADDtrRPLI43t2bI2yXZ4Yk76B70uV9mM0xQ0ab51Y3V3jmmZmiU50tRnyIKIkA1on4YKgwTYKRtKIaovBABmtERVc6DFQI2qrwC4lr8PfHPsqoJ8QK6lZg5VBRE74hwtbdy8EbYbo8HUxKWgLVaoi1u92dkgma+qd7y4KFQ1N96EVWeD5dIURNm4SDcw7XPC4BQr1oywnM/oDFXaCgy8TLwhXSzzrmoAD7B2xs+/wcw7i9HXXCv/AInYD/6GcRhmn4yQXg8w0g2mFCPJ/I4MEFBuMAbXu1lD2gqQsaDdhQO1++FQ4Do7O+Yjnr9M6Mshzxx470fDKoZ6vGAyilHZiKRpZiIDZqcElZ7acM0G1WLrJdIiEugVB7wuugliuW8C/wC6Wi+M8C1uByfKe+cLouWe1BZqg/1ZnXOVjmA/llcXLX5caVa9Q5Y47IF7sK+Qt9MKoqWOcDEcWWJlgLhuhZCXAR5o/X90Zei/LIH7eq2LPrdgAVXtDDszNTgNkZtnbFx8gBVdghQXL0QnACANgh03Dtw8hpwI/CLS9l15lBNncXf0HY72XB/SxrDF1wtwrFn/AGsuP79zcrPHulw7YODvl+84b+iL6RENVtn9KQDDsMcLF4e3Q4eqGGXZByCNeoccrT8uhFmoryRw6arkxuGo+nQcLLQv7yH5C0WdOiwy8BcPG0rolSpt7/YMTKtWieqD3URE2rbjcvIzMnRfJ5lXTI9MFZwDKHQ4CRRywijJHojaOcaC5vpQ1oqAkWI+BGUcrp3GSAC1dAgm0bWlwRVYfxdovLDnpfmkBRQGq3leXG1IgWNkYY3DwuFAeTgydBuErcZQKd8h90pi6/r3r/uiEYCiJqQjtPfDpu2PycOnmj7DjXxF9pfrGZYd9t7ud/ID+qqbHF1A94AAGhAtCCgjHa89NdNz0+GLk9Tzwqt1X43HW4IGTPVcMelSamNsch4slrNVGbPgl/AP6bFi9KDHjiqGFzRDGayUUpVsyd2DYEdqWkHxiTjSN9iPluMTrUeXYlZcztru4DvdV5YgQiQwWOPJ2Xh5hUWItqM6981uFsL3Q8xqCY35bwiUxoRKTomjbuY9OKsyKqaE29TDUr5gImRgNJ+x2Qw6RUuYTUMLglQjS5hCxDwp1xu35SZi5/eyb70ZIGCUSV5I4ux4wjO5HBHHjRSxUvhb9YwoRU6j38h02W4snk1OgnNESBFHUcP7eg9N/icKud9gxo5XtCGdh6GRaNx/7g8kRTVPcwDpydUDiF81NFSp/e2HA2XZDSOjFaeXpfb9HxC0uh/CQ1mqnl6CGgQETcceVPUgdDJly7MSuE6EuNLPqpJbb7iOwrYKJVR4MPjzbIQqxQNAjAbmTDPbhMw/e89CK1032fSFS1PldPRM+P5cHdOIY18HdirfwCfchsZD3NY1ckFXtQSoqLpPWD/oHabQOMkHl8PMQ28Sq8s1vFXrX29l6Muv8qZR1meBJFkdAImf3MUPFEoejtbGMrBkn6jZFBYlfn3kMELWc4LJwHJ+eg0PH34dTLG7EtQvhX9IiNWsSkjPmCjfsdHaQ/Zw72fb9Kvt9JKKgP3RRrL8vsM02d4fO2GPZgYD9VzhqddtvMh8z5QayP3W6OSs3rFtHPnciSpfB0101Kzl2JVgjVaEPZRu8trrQhKqjJjN2TLNKGnUvAtBadqCaXLH91QkrijWKhcKuOzAbsV5ZkuO3qZQYznofeWrM8DKlU+Ugnw0v4Y65oeH8RySP7CXeqe1kWZnujNBREcIrQZceyXcDAsfbGX0S472vWGGLYufRFQ2Vs4lJGCGmxGbSk8LougUj5ZCkQvyDeGn04fKdBoOPuJlWZ2zaGhLB3t8ZGbG/ukoJoil0Gj/AEtgr6afouPo3vg/Wa7O9DLUAtgqysP0HOHwOt0KRsgfYyERiXt57nZhkgfsN/ODmhQ6syPAcyBosCVXkhINudoTeJ30TQAe0fpDyV0AqgVmloOckr3wtCULEbaCZBFUU+9Eb41D3nxqCFs9KoNLMGOZlbild4H2cgsHBoBEL9DgwEylhqwxYfVWoO2ZNQJt9+A+IJfRf5IpInvlIo3chTSeL/DLxVe6FM3XeQm2zkpnergxWQB2M5m2dSCegHIhtDcNaYHAetwmBoFEWqCOodTctsRIix1KjhTlBEx8l0eO1mcGZarY0atS4QttpzWbWyXfcXrxX7X/ALLoQy6D+FrlBY0AXwQwbmWT5Utj736sCOOTV+ej9/yw/Q8dJruYSy+F8/QhOA1FlQC9RHRNMMv72cJofoHfWOSJBThCxHUSK7+eJAYw0x0INkA1KvzKelHZgGRnFIG+XqKO0rIoWMrwZRqPGkyoe4ygtF8ia/50T97P2UxNWftB9K+8JW6xBeS0qERAFuXBrPswEaZIC90zfgBmwMLll2Xq93gigApeeBARz5u67pE8Id8FAVlhgKOwZYz1tFYnWNk8TWupnCyxTNJlad/VsNC25l+8Uoub1RRr8h8xitreRh+1M9nuthivLnviTSCseIDH9GOhtCTIzP0jJLbZk+50EfvXnYLfgQAhrBGbf4QAQJuR6Oh4G/rhlvpWcDc2onLC4DM7mCyu7hbGWNHUgKgFrMya7T6RbOe/25BAAoAoOp2nOENfp306X5PyMLOEnv8AQd4/AM7ifaMFS8CGhHRhouF+hesy1h/kGRC/Ik1drPPuznwDI1CEJp+SNiPCS+4aUNndMV+YEXISUTV3AE7TKtyL5ZhltuyYBFSM38wkOwHXLVkbQNaMppQtg16JkgFd3HhG7qtoGaShNelhnDwbWpkXqDAEuaujWsC0aN1fzMzlee0ZEeun5WM3LHziPxCqIihUtZdFAj5kvSpnEv4IAqLQKzTN4j0HAGGh1xalqgdm2GA03SoUQDUwSEdmKMWKMd+OZm4GWAu2DTj8LcUSXtC1oZdXIaoCBVSWMqrPFxHdIsLNCvoyfgMC9XWRM/uoolcrgB3m5yBvUrz20K8EcEIcGD04LuykLYHGk/vXp9n1UESzHuY3u4dui9h018ZPbCv4qe59B2Y9kuf19Qxdp2wNB9Gsr5NmUIa3KZX+GezlyKYqjzjm/WDR1mXQZopYCtiPaaV+LH8P6usOhi4hBSOcu6qgs1luG4u+8pnQmdUZNos3KFam0vRdod4Cjdry4WJX7nuw3T6N87oq7VgzP5QooZQfgi4L6WguR4MzAvlg6ZTtM2XRAwwQg21aGljFesMEZz5lBD2lTitbnlM14iVoDLkt9yGZobKQjJYDnWVJsw0Y7mkd4DrNQqGT6PJ0XtJj4XeClHo0P8kXpNE6SoVqlXFCbUUvTxFNOOAOm7wQ0KAowsPhdf8AeKjAgGq0QRdDPQrDvmvsY/1rwxdcKP8Ao/gr/wCsmKsN1UHuz7TmmrJ+lkT9lV3qfbpfzKK6yAqtRS9QT0iATRLMLlycKB3v6SAiibxCvZQ6Z5gBlv8A3+KdAa5/KTultKmx4b7lk+FkvR8dYSzsvgsEsZ4T8EUrzSFHllMM/raiTdOtXy6sCtOSO1GNvag7ClwXHWnIlSeiCZzfmhoRRQogVBB2O6Fkoz+xh8Ry9PSWu1iAFmYMLO0buDJYtaYYgXMGU25z2pHYZi4EPWOSGF5KPwpNToTW+HJuzG8a4C7Jh69oUxaq8JGb0LqCyPmxi0JRloAbol3gWJmDoSqwAnaxqrPVcs1grOFleCYVS851sE12ef8ANPj/AP8AjC/uzzcd8b/fH8Vu0yzhQr4Rj/SWPDWYIesITQAMXXC/Bzf1SyYu9V7RxcpzsXPnDc0Y/o1munujNqi/IM8PELMKrcn6l8Iza9WCvTJZLvX79F69hIfiItB3Vm4PifdJsxzT1qn94eJ3f9hM2Vbg/JMz0Xv9ppCaGigFBKEuxHb84xJSZWIgvgd9AOUUJtKJXraJV41RjpbmGmkKMo3My2nblPraJdEACsldYMvUiOW1BAvlyhcIFWUpialjyl6kGIZMvwYvcJ0JEgHXIWCsvWhiCd7kXdDsQ4Z1ikGRRGB3KxG5k/4E4iNCdhhhYO9e2Gaa9AKu9xjfaj+xJ8wsNL/V9UnzsQe8I0zwn3Stbe77j3m+aYP7Uo9DhRfs3DanT3NMTADpdGffWPunwoiACgA6r+Bh+VPrCjZNOacZoDTfDURqc5KeRjAHeQwCsLaAx0aMkNIg5mqbMe9CkpKiFIGGLBDLjRHqpnKdpRCtrSttHfgiiZxWirhGouSiAoQaxzyTWrhtKFfaRbPhMcQpCBcG4gCGYpKtj2Xjoyxnujuf5JZtdt01006nfYGT1+WJk3sGhEErOLW57wjozvcnrp/o7fmcaN9VMX7/AO/xw7Ar3/V7SvssPBaXg6SBAAtWLiLw96Nvq4/iX4/hZ9mKTXIvAAgiWJhXybe2Fy8D9c+zS1LKXoQmcsr/AFiNo1k1Q0BWEh7inajtUowEdDAbRBqUd1o68RZdLaUVxptYzSdWG/VkGpMo2Zqq7hJeNLGTXmZcLtgGkFYXZJaClkpfZg3W9YAKt5ZqpjAS1l3SBTrsMKgB9hud8GKW2gZV0SyxErOyQmg9EEyBYadH2ZpL9Oj8Q6mfAfdgWQpBwZTlMhKnyxoU2xX8mFbBFq4dmlg6QAqugE70+MfHSYtmvBUSojgrVXGj/i/WDgX3tOqLOnvEREUTMgpP6E8cVAVQAtYtRtZas3YjUKAtYGBtuXx1l/B5HyEJrtQqI8bbsDc9lr3w7go+nokm+vhNJPcw7Ru0ZW+IyyR4Vvmdph6I+KwANZYESLwu46h8SWysdSan3nPGUqvtFayyEaCBbmTFHiVFrFjlHMgQHwqLjCCQuPigoUSL2jMYNaEA2ZMQS7lJXAeLKLM2BdiLE4lk67CasO0xaH8DLVRkyUjpm0yZUUa0lojGhW+JQLO4zZAitJ8yt+WANt8QLWecoc0hh1wjeGA26SjIuOrIyBKquIPlmuq6st4pRh4gGFKNeroAyjRVMrji1YV53FH83KwSec97yst3Yvs6a5cr/XCscaHylga9OTtHrS2C2LYJ6wY5MRB7ZvgS60Mu7AqKZpmfgxfZvbs4eGJUgpwlk0GgvAuJBNRuCXog4esT36tPkcqyjMgJf5zxELRi8E7Y9spxDxi6gY5tviNYDuZnmgXLDuXsEFxBwcGoWQuAHfJJajsE9JCDqdqblghtoWuGU0Q22S5wvcsp9RlQJm22cQawI2SmK3m8yyfFi4tJKhArBnQUDQoiKUwQUUjGWHVunecmayNXm5mlwhMyaBhIuya4A0IU2gud2eIJ1WgX5mU3dBndsqJtJyH3lmyJdMy4ZrFGpKRYgRVkso2B1NqhRQGguRhrYSSCkW6BS6HC7vT3wULGph8y4eZP9rELazin8xfuQ3L46zufpL5wqQI0XwrEBPUNVXXPHgQx6RYKzMzfFMJ08AexmFdrmB4NfjFGyD8y/ZZv0KAq0E1578FjSaqYVzdphWTcp6Nds8EvSgj5ZZTCVhJBJIYgQzObHfIiKRzqdJW7KDTJcoo2ejCuyrSPNZVpYOpLQYTgkDSH0os24kRpPIXDQMt0aCKqo8VSTcWoMheJC6wJYeI9fqS3Eg4uwIQumcLKcOyoSzgErIMBAdg18JfJT+YwJTyz/JBjbugQ0NBsckprJypsu4XeGaaoCm8jSBQLKzX3l/UcgEFuxu0YDJd04Oh3cAgvZsQ/ljVIpo+GGJXjIVh6PD0xEEd3P7oUU8LfhwYv2oA+ZcKTdvwyTI2vCgbSysW9pSdiam6fKvCzqIZwPTOFxTV/hgW7cvVsUz7I8+dmnzCc1qQ2Aszdb9zBQKoBqwwgly90Ldi/gTbqojuWWfNKw7JLPJh3rkekAUA5Z8maJYFHtFbWClJBhAbhFCMDulW8POPuwL4ouac2UFb5Y7mqMHA4LdgdqPmP2V8QHgxv0I+9Z7S5YUbuyxtDaZsoCMBJzYBS72JQ+ZmsRpUsgsHEiX3JdCFwmu0b0lkpFy0js9Yx0XLEoCFIljLhPEZ+2KjQbp9wjgYNbPtysEfAZTUvVL4vEm/8OfPYE1JeuV8O+hjWctKffhwgd0exrK53tK9sywAcg0MOcQy84KAqyztKGFoTWoq+VwXKtvflTLugOp+OnNWvV2AgP4wBgACgKDpcMiWPsrX8MOG/XjOCJZihpvFonfOdwqq/JvNNXznfaGJz7T2Mo9ZFALVlYFGbsd3vi8gQgmCU1ZO4igC4YcDRFu2CNG7iQ1nejG8ZecHDCAYzlXBwKs1mRFZ2doirnzXD2zDQ0hQnYOlbYLgQtWAF2jxF4ZgjJcwBUoARcU2alBP0W0H1/SBac07COPhbgNS8wQHmXJBECZpeZ53n0ikHuQbeFV3oL95cvuE+6bvo/kGE1vlH4cEjNL4V+Cbvo/mWVyPSz7INXegH1YlvF7Rrurv0wqDu24UebxTYay0TrroGqnPyVfQgqKrZZKBg4D0w99wCkTZjL7dteD06Ktc/CjR98LzLQfb9/oBbZnyFM+XvZ04X62B74xKc1AsTuMfnOM2zKnxA0Ttx/K0zk277y4NOSVr4N4L0E24LMKA652F/6TmoNILqUEBBCJJSZRY8RiM0MylRwgtimUe2KLOwYGzrLKlSNucIZlnRvxMpXGyAKDKG2XbL81UwGEKmlJlzhVO5txuQmWgWOllR0JZApm7hIhwBOyx+gBowkds7TPNO0zxjJTq4KdqoZhQFYcSXR4MHU5Gep/OVCyx7CdBB0D3DRY+WivHQU22F+78uGrJ4CLcDrptWCxebM/Dkeu4vqzX4hKge3eiaEaGtVLC1MqE2IezhnlyNeuGmAwPGofRmfM4cI0ypA5QJvBiuZqZxlg+YwM1IsqzLgjG6VbygG92HZ4BagrAJIPeBolGVShdwekZAIRGI5gamPsIy9lSZoQxZGC1UOF3dkD5cNVADZsWwFp4miIFISkhlWWMYj/CtUaZYLWObIxenQeEklPB5blQ6HC5eO2cwFG+8DE4leMhcZ63LurcL4pX7nwOlwS8kYfc9PTPSrTKDXk58x6xErWpasvmx2faMqY28niewAIIgjYxKEaRshHbmeAK4eU2MC6FOBhhGNspioleDZgNguVZg3iMzvC5EWt6TUlxHD1IWW5KQ9ymIRXSZcg23EoOVscbO9E3rEEXDUiQUC84iLJV3xIp2VCrVQpD0IRQ5xNTdsOcasFNfkRiqYkv66AC1aIAWxhShy+9xYKsA1dfxA/qQpE3ICypMNeGNe1YvV2MnpGQd/wBBqzSXfTG1jkSE9DATNDX6zQ+lW9nh21r0Y95CB1EgqEaRsYG4D6eXyx7qxUjUcFdwOCni2vWVYDtfC9Jb/tRpFqXHER2/2szldYU83guRyzeWCU1S1wI0ClS9xjWZw0pAzI1jyR5p3ZYxlDJSRnXqlcuOnfCO+oak50UQjZB4AGQILRAJOdtg9UVz0SBVNnJhCASk0jByrIZXvgy6CLzlhjVLTeMCHcpXztlYdiCrRAZpEMMGBVRE+pfIyyHnAXtjLzERW1bcBHRwRhGaZN79p48YyfDo461XKK7QhgSBaAY09g3y8Gksfvv01qqr8v8ANhfn5ei5Ys/b+GPUgK0Csp653se5lD6SfzgDu7dp7ogS8gK77wAogi9zZHyo1yqXa/wMHYzGyALcwMMpl8okSrNxg2y2IbcTZI+jOaX/ACUUCM7mmkEbTwjECc9Eyes7UqstJ4wCdqHBLTMldSI1I9LI92EEkuIqMmkWtIK9JnGnZADncmZjmWQhMyPCZ2Hhith7kefvRtZeCa5joICBlC8aELIe/CogjQYaRlEOkUfRArNaJtGGFO2WvzgrRp/fv6Ropy2UzOsd/JYSNq1Cxii1hCiPStHq9DW519CLbal3V4ZRnzTwOtwzdvs04PaaXuyoQAREscWFeda/swProai8CZlqslyO/JgW8Hu9kpVypzfup+7viBOqti3bBYNlo84aXogL5qZoDqYcFoZ/US7VoIlzWVgxHUlqG01NIQEChsYrKzsTNqo1ZwLlrKjkkEkJNvFE0iTaGqiZSG7awHMAMU8kbhKw6FQiDoPMHc/cmqmHkCoDFWlvjOJznVkVlAJQ3jAQDoCXFsyKddQlm5YA3XQd4qlW1bcNQrgXLb9ZXJB7s0JEZZVHZ5UK/EF5LzH0K7Pm2H5A/UqvOewqa8IO9OvhwzqvYzV0V1gsl+eBEzvkGjwILms1feDMJWyFKDXsaRrYcco6m0Gsnhty8GAoiOZDHYyGGQ+Xn3i1xRcpRgkaDIu0yVmELtDyqDshQGWmkzOHxieJWWoUaYa0mXaUzmUeiyTDZU2Q7DIoEumrNJtEQby6w+4dh1KSHSQJoB07CmAywdIIzIQ1XKtOlQ2M1AAFAYZUts6a/EFZLxMoqjucqD3YtSIzhvgckNGriX0Z9hcD/wCkcvA+np/x8uwy0qO2tXRnHda2MpSfJ3HDHKitVtWZ25t0viMve74eMFbqTKz8ZG9Y2W3AQHdlAiCaOFOZLl2eIjMrnIizWNCAQaMzLNAhTbBkYySQ2IyYauIZplYTPLGXrKbIVYKM5EbYItws8MtibR8gl26I5nUpIIg2jIHaCbQkwqDBzMCogycB1MQAWrQQluZrDNz8ZgIFN2wRVlXSQ1D0mR1kfUB5jqg8wVVVtcHnDl+exyUrfW1YWAjnuqnuRtVu936bhQBmG2zFHKrK1EaRgo2QLEHEn41iFsdeXG5npKgJ4CctrRQHdZWQ0ksnjE+farWE5m7rjcYri/Mw6T49u803Tk8nJNalSKl9JemGrkKMKESGC0MLQmAFy+HXNMQgLlasFUykStgJSZLBng2zNZRayTWtpkRyNA3jnlMFAohn0qsTrHSaptIKIqwWYUg83RwYaVBEgzcBXvB8Gw3I/cwrNvDb/n/ChQOLG8C1fzML+cwenqr+YGgYGgGQGJ2JfxKYGn7veP8AB9fJxuezAVVet9t7q4M6+fq2ibJG8uf5Zpq+CX91Yfkg0AFAFAYgEpGyG3ZkMNuGa/xBQ6SkY6s11nDxUy54OsdIbZRgKwi4UQLlJTgzEKSyBYEsCNjBREUxpiS9RyiBux3CyUdiayDCVEETJKlSug6sdJqmohNuAGFS5X3cFomcnF3eeg6wKUWJKrdVf9kMtr70e60YWOfz/TxCwv7uvlv0hQ3vTr6MHuktqiglFrZDuFgfTcKbGW9tpxFCyCkTJGCiI0w729tb+uKEW8jYuK6MV5eKESKm1c1Z65yZPYlaz5E5XV61RyRY7HDLkzHD/MR1XzALdXMqJRFizPi0wym0FMzBNEIYaYWxkUgZJLS8vEMbUreCFBOGVUACU3MoFpKyxAuKrorBjqx0hGqELeHCjqwAFBhqN5/jrSNth8nDLHd3T+W8diNAQZQ93dK0D0a37OF5wZv/AKiH4P8ABek18AwW2mvgOPR0DuWXbuV2Zl3Laz+z6L1PePzAANjhQnMsW/cmUOIuWCsrFlw94TfFmgYgjg0TAUmHQhBmw0wzjowjesYsqGsFQQ5oY9SRKWBDVQwprINXEPnQYFf5j2Iqtr9GgHQrIyaloqylt3G0pk82CKpMbdKJVm+v+eAfVcC+A7+ecFz8Qzg1ZgDnZfrXhM+jxBHBG2hme/cw1hJLEX8546JQxwZdAM4a4VBnNcTOKOiNZpjQghbmh5xDbpKuejAR51Dl1XGHOBKqZorJp0DXhAJ0Yc/ehHaWur9UXrKtglhons3Qw3e3z6f4fru0E9sFC9xsTVSi6f5B7n1so/bwx0zatvuO8TIiI0jFEpHt0FomqNcGyFXNWBhKYAwM2c0Y6JozTNSBdpWmUAtkiBkGPVMgddwbxZsZlijYucHzoMLPqtIfu1l8kmgIdkVh9n6l9cNT3wdQDtZ6kp10fYwPruBwn30BSSztfm2GnPOHRvx9N3k4I+Zyl4zt3ZD4wWr9vCCII4NKz9vOPY1wxk6BDphdMdjGrmrE5MeDmRyY+gzIxwgQL7g1A1i1Y2mGQwNP0NE3louX/Kw8ADQw5e9Ifvaw5I7V9iVwv9bN4+E72j9Osabef/EiZ3laq5rhoqAh12f8UZrM5zv4w9ZCIUibOAZ9Actv6F57lFAG7F/2FHsNprV6cn1cABrRShP05aawomf9kPARwYHQMjWMyiDFHixdHE6jpgMBHDA5xdGiJnNkWU3iuTdQBVmrI6wmrA9boSyALWHQs9sQTn/ZHqldWBD94tg3ZnyflOXc8x3CqeqYTtKlfVP0jesK0AijszXt/Nw1sFUdNn/H3efwHbZgqVWOybj2YoGQrcLVdSgKtBHpm02t3tgtN+hXEqe+kSVOnit4hayEQpE2Y3aIZ/Brhy9mtYH+UbmBQ9UJZHN2T6yaWMywMcImkGbQ4LLDRN4sXAs5pwUGhEHUgGxEwZp+gnQy3dpm4XuWIFrbvCqqtrKTQyD0QjJcwE5MBDkFjJlJoZp6qFAVY1hVOp7JJA8IE3PoWtrIWh/ZgtVUnZvLsSlVSvfcXdwP8NwTs1+Y5CMGRZO33LvgHN6J9sdyB3KBmI9NQm8RsNcE7bMOiN3sQAAMQuiPd4in8PGVZNK+j4pdYAazKXwOz4xzBOyjWpsGnQcWWBDTAmEwEWcHHIYGc0mBRqZcVa4ODrCvdEWHmAAbHpEAFZoJ6IAAAGDEQS1t+VwReoVWq3i47lYNg4JkM1XnUbYhv+eRpAKq0ARlfQ3e0gwCgLVYKA3kLt1hbmg39/LgABVaA1WVr1Hnsf45wyaV7gdISBUDZ4Ts4c2U9G/D0s5y9dfyxWKtXgdWD9BFsBR0HOo/lVwQP6b6K67k3IKEjggiJZLte1tK8pw7OIugwQhiLuEJc0zVFlLzwKDBxhFzF3aly5VuRK9yfmDVVy743fgdiZ3H2MMt793uwJq1oBZvNW+zIQT+oCkTZhW6ubI/kl/7ceXtK7JIDlPZvCpiUlQBEM1MKCogtExe1Wz5x7Eco4rdw1l69w/5ejsRf3FEUmApEyRw0rhS7OfQrDL29jRlX6mHTZTcs34+kBQBatBET0b0LcH5XrejyWYFvXa/4YIBER0ZkYfZlP8AhmWOACAjqMzy7xqRRW7GZHDC+sSEOkgbzY+4wBTAaMFi5Si5EBl5ROkXtM0qcGbAN7nfG5GcG7LT8u4IhnyPwJmmbey+6PiRoAtVlSKNaJ8MB5eGCbPePWQiFImzgamaH6AYWfyJF/mNQvD9oOVsRTgbzlbhsYVFsgNdoTMLZD02hhamEv7AgKRACgDQMA/yszYb2Bt7MEQiI2JDBma9his3Q5LQCGoks4GP1MnE+r3oXZyhVVWZSGJxtGC19Y64vSSqqtuF5O5nuMVU7VmdzkwrAtySgF7T0NLmO+SaKnjRilP5QwYmcWLiIzVBlwXMYYNDFgXlUSMrnJAop2GRCKIO2N8El8FHKOkK7sVCQAtWUs0Tq8Eaynkw5n2fdA26DrsrDaQiKIiaku73YqWgTwGy2gTcqf2N5YaWaeqgkOQW1cDmHzd14O7E1ts2f5nC+CrsP5WD1oPyPdwC/wDJTHdutyb9rB1IwGkTchlvDBNzvgChRBYjqM5YChTBtcWpR4SE+9M77IrZdQq+awCBX9o7HMGSrB+Xu4Mfab1YABgbFYRHxSkQ7foDeu0oyvEQbqyuJyTGlC3BlUZ/fSACImKCUlzlbnJG/Ek534mpp5MFLjKcCMGBEuUUBP8AkqJ/kGD1/AQWuiNJ741ZcFTlzY+QvLjzx2TF3hzg0K1ewQ/b2Dbux1CyNRNRmltQ4bfSqZOz/ZHBMFSls1IQlkzNPVM7rVPVOAORWksLbn+XuwtV5j+TDI9GaeT/AJomCABEpGX41s97hO+5CkTchZG8ANzvgclkBYjskF23mfRDo7tQwS7uIfsihSv3uA6HQqDFkUZ3Xsmp+62G1UtGnmb7ncITAb/qWMEvozzgVnm7EF0XUCveyGDlrA8r+5KWgPDl0oM/HmL0fgs2Ax4wdrGdzdhzD2hv+gTUl+WaMfXOZGLwdId+cZYCfEliZXGBQFWKK8mY2I/zd13Q6JV72VncKnzYCbO2HVWI77XSs7/8hIUWSWMn8srdGaeqWAMR5r9jllsbX/27sFAv9NxNghxlVcr8ANj/AD0wCyJELEdmK2a9df5wvWQINIm5DR9LR/h6zR2E8l1BFY+vPK9PMv8AUJeTlxP7kFImzAt9c2RydyDqTY1LPwkt/unImFRrU0CKArHPbfxYhMIttu0a6dPk/ismjIagvWWfWLsCcme2cDRfMZVlcZIpWqvQZnS0R3rXBs3aF/YQRBGx6CGsq3GWpvrboxm9wbbil/wY7cj3OlEvpCBeZcKlZlqLwJeS61OQ/Lg4D2e4mwQNQAXM3+B/oUwOsBELEdmJp6g69xzj2xaL9pEQRESxOnUdN5Wa3nvsAIpng2NvMQAiNIwZm/G2uj+cvt0uSNEvogWakcoyhnfIFuPRpY8M5j5CA1WLv84b0yL+wnd9keL2x247BerE6Ly3NGDwTWfRaiq2t4l34Di7ONP96RlnBL+FJlmlcosbPhwp/Ke6Rzy4NXUrBVN0vr2sAARLEzEegqFmtzz64ikREaSaKc4dP4+pkvTKf6CW5tj0eBhelc28xIBLKrn/AAH+jTBBESxid1fcMbPFG8wgnvoCao7vvZBuC6jZbw8A2qxMBCpXjYppaX8gY5g2f99ltYr5VRPmgKmrjmmH6pGi4bcG76EqNKfAKIqZa8GywNZSHpa/ImlIKcO56P1z/YZKp7pdg/LcLkpuq09qxEqquqw8pNALWUiYd6D7Ys2O1baN4Wrk/AlX/vrO/wDbtzhtrRrXf5QdYQg2I7nRW6FA9zxAoiMFtrRfh0CncDWZyh3gCV9FLKnfLP6xf6VMd2raf8cso4PSYZoaQeUyL/VJidCX37nYZrJjX3OSMrRt5p9I8YzSieBY4BSlqszvGJztYa19e9EttKfdXMxb7N4mUlrzyoQBVbV3Wf8ASIXhwJ+79cNT5fGYluElfcbRdjIXedulw+JnG4OWMFI/8FQi1ovKwN2WtepEsBgchcDANMQiz2aeZUKU3P4FsKPhfT5OGV6a7qei18v5Sdwn7gwUc94At28G4N5U7naHfchYjuY3LXCMB15rAbraB8+hUGBz3tFAcrDntPThVVVtcPLWi7YJwuQ3eV3X/TpiiEojfHOHq3t5lysHDjaqSW80eh+JAAgiWJhVHXRy8bNoKmJXNf2J7kDHvOd+BNasl3W6YKuqg+8wytF9D+DHO/WawX7T7vuGYlifgqMHRaPxZgiCNkzqpE82rCyDP2QDoQEs4DhFB7V63VmhmH5656E4XVCKc5N4l7ag7M+WlpVLRcn92+inlg9pHuibUhzqVBepI9WLWp8rGyIWEKwKunmbt3gMCkrEenP9RvJpiqbdl55OAKgFrNz9aPdg7zgqANAD/Wcj9tVytmPdhZ7fF23mXTz2mntLynXL2dDy4CTrTK6ZMsA1II7BLYqZDschMyRvRwg9o/4FxETatrMwfvR+IwaxP4k+05qE9bYZBan3P59GnJ4ojgGFbGjFCM7sGMfEV7SoNOTc0Urmdn5S6aL6MvlB6JnM0KBL1Hw9FfND2cL0Wh3QcbkQ4TUAA8BO+9GFlbuhXkXImfmcC+ULEULWigO6wohppZPGOLm1WuGwPo0DlbE5dExcL/1iCIlkpzrNDx8ppo5dODa5tVJM800DlpF8ezpvKOiYDJAGrDYSW5VFILJzrvd25cFpy12aGGd2YD1mAa6EbuMS2KmFy81MqKme+Sj/AI+1ADWAeWaDBfArosrLg2Va1mTgOQjRGmJGQNPgVpumWdDzByyvRPRl81HfOALa3zxM8cvunAIegvudFQZa0FFuJRVd3R79ozyhvJnF4KtrL7YY8kT8rEP9YmNMteN3YVAR2cvGtLNzXyDKy+/L/MHCa0CJ2TFAIgiUjAgC5q0k63eDUejjT4B/G7sDQMDQDIMOQTfQuGXNXPkCfszupTmS+FWRO9q4VUmQdlzxn9WtgwwRtBbfARZeIX7YZnqHPiZ4NvWDwF9FJFj/AEafhwtVz9Df+8beOfvgIREqOoo6V1TtzWDuu3crvidhNa0HK0CWR1hvYAP9qDV1kc+Nyw2vavFSk3aZ+ZSyDvCfkmiUwwfUx4sZUp4do47LvPiwmvhfwkIRlDoMf7V3UcWKeHNhr9Jfg5qP2ERHZMOw/W6dMqUVUUUWWv7U6QFEo0NjkMSuCwhoAYR3nQaBnsaoCFKRO5LWflDr84gxQNS1UlEqTLoMkMv+jxCfV215hAiq2rq4I9mhqOAJc/rC35toFE9hq8rquIf7KsQ1fQwOEZRZ39f5ji3Sl+AyceWwWyfJozKh+J9VAT6KPq6xQl3mjshzeABNKGOVkf2mDOxv8hekDqzj4ziOpWRkiajBdfWzWfK2YBd9IsQfsoq9gjl3kNquqxLVaPkOnXKxdCvSCrSgHOlqoaQB5WUa/XbTUHcnudXF0CNAFqsau9vfgmtCIr56/wB4Xw6GT0ZZ13B/zCOwaD7AYgERGxJohP3usqdzhGgy3CU8n0G2981OpLfz4fxlvyeamnUrInI/PwxwFG37kJmVa+IqHdEEG3uoJ+CGPFQfQVAAtVoJbiD/AHukuttp/FjLWHg0MXtzSyPYJQTysyHEndmfRX+8Bq1AsSJpPc36pqcL/ua49ZFKKRxAnGgaSUgMdr79Y2gubnxFAJNl+WSFXLRBP8a/sjcX2GcszayfKi/5Cr5YtnZEQ8HRVQXSSj5VEf3X80tPqpBfmWb0V/4GkKbVo+KGX6m0wrDE4iNPR3/0l8SuB+CfNlKedr+ZV+pR/E15nBZ8CVnsz4zR+iBaA5Z8Kimunx+yZY+hofImueP8AJf+735yzDHcz2HTzQS0Hl2iG6uKwYd6n4wGoUBQHRX/AIGuk1vG+ntUy32UpnYv6dCsSTv7HofofKik0y9o6J6zYBh+rr6Eah6RT5DGjlucrfWUU9AU+hKB7GfjNJyN3yioyfe+zKNFHwAPAdNf+LQREsZaqjUvxrLVX4H4rL57G/zUe9wUvlI35Za/r5r9M4pOwWyi86slGgeYdOkdmCpe1/fZr0PJ8ItTtOsn46q/8ZX0KJcRs+1mbUPTfsM/hpQzcmQOrefxRp/rp941V/q1nwbOUofgB9Gv/J1K/wAKmV/5+pUqVK/2n//Z"
       alt="IYAM Logo"
       style="width:200px;height:200px;object-fit:contain;border-radius:50%;
              box-shadow:0 0 60px rgba(255,193,7,0.35);margin-bottom:28px;margin-top:16px;">

  <!-- Title block -->
  <div style="text-align:center;margin-bottom:28px;">
    <div style="font-size:0.6rem;font-weight:800;letter-spacing:0.25em;color:#ffc107;
                text-transform:uppercase;background:rgba(255,193,7,0.1);
                border:1px solid rgba(255,193,7,0.25);border-radius:20px;
                padding:4px 16px;display:inline-block;margin-bottom:14px;">
      Imole Youth Accord Mobilization — Osun State
    </div>
    <h1 style="font-family:'Inter',sans-serif;font-size:clamp(1.3rem,4vw,2rem);
               font-weight:900;color:#fff;line-height:1.25;margin-bottom:6px;">
      A Message from <span style="color:#ffc107;">IYAM</span>
    </h1>
    <p style="color:rgba(255,255,255,0.4);font-size:0.82rem;letter-spacing:0.04em;">
      Official Field Collation System — 2026 Osun Governorship Election
    </p>
  </div>

  <!-- 4 Paragraphs -->
  <div style="max-width:680px;width:100%;">

    <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(0,135,81,0.25);
                border-left:4px solid #008751;border-radius:12px;padding:22px 24px;
                margin-bottom:16px;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0;">
        The <strong style="color:#ffc107;">Imole Youth Accord Mobilization (IYAM)</strong> stands firmly and proudly behind
        the continued leadership of <strong style="color:#00cc66;">His Excellency, Governor Ademola Adeleke Nurudeen Jackson</strong>
        — the Dancing Senator, the People's Governor, the voice of the voiceless in Osun State.
        Since his election in 2022, Governor Adeleke has demonstrated that governance, when rooted in
        compassion and courage, can truly transform lives. IYAM was formed to ensure that this
        transformation does not stop — and that the people of Osun State speak with one voice in 2026.
      </p>
    </div>

    <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(0,135,81,0.25);
                border-left:4px solid #ffc107;border-radius:12px;padding:22px 24px;
                margin-bottom:16px;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0;">
        Under Governor Adeleke's administration, Osun State has witnessed unprecedented investment in
        education, youth empowerment, infrastructure revival, and grassroots welfare. Salaries are being
        paid. Schools are being rehabilitated. The dignity of workers and pensioners has been restored.
        These are not mere promises — they are <strong style="color:#00cc66;">deliverables</strong>, evidenced by communities across
        all 30 LGAs of Osun. IYAM recognises that continuity of good governance is not a political
        luxury — it is a <strong style="color:#ffc107;">moral obligation</strong> to the future of our children.
      </p>
    </div>

    <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(0,135,81,0.25);
                border-left:4px solid #00cc66;border-radius:12px;padding:22px 24px;
                margin-bottom:16px;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0;">
        As the mobilization arm of Accord in Osun State, IYAM has deployed thousands of youth volunteers,
        collation officers, and field agents across every ward and polling unit to ensure a free, fair,
        and credible election. This system — built with precision and transparency — is our commitment to
        defending every vote cast for <strong style="color:#ffc107;">His Excellency Governor Adeleke</strong>. We will count every
        ballot. We will report every result. We will protect the mandate of the people of Osun State
        with the same energy and dedication with which the Governor has served them.
      </p>
    </div>

    <div style="background:linear-gradient(135deg,rgba(0,135,81,0.15),rgba(255,193,7,0.08));
                border:1px solid rgba(255,193,7,0.3);border-radius:12px;padding:22px 24px;
                margin-bottom:32px;text-align:center;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0 0 12px 0;">
        IYAM calls on every youth, every mother, every elder, and every patriot across Osun State:
        let your vote speak on election day. Vote <strong style="color:#008751;">ACCORD</strong>.
        Vote continuity. Vote prosperity. Vote <strong style="color:#ffc107;">Adeleke</strong>.
        Together, we shall secure four more years of purposeful, inclusive, and people-centred governance
        for the <em>State of the Living Spring</em>. <strong style="color:#00cc66;">Imole ti de! The light has come!</strong>
      </p>
      <div style="display:inline-block;background:rgba(255,193,7,0.15);border:1px solid rgba(255,193,7,0.4);
                  border-radius:20px;padding:6px 20px;font-size:0.72rem;font-weight:800;
                  letter-spacing:0.12em;color:#ffc107;text-transform:uppercase;">
        — Imole Youth Accord Mobilization, Osun State
      </div>
    </div>

    <!-- Proceed button -->
    <div style="text-align:center;">
      <button onclick="document.getElementById('iyamSplash').style.display='none'"
              style="background:linear-gradient(135deg,#008751,#00b368);
                     border:none;border-radius:14px;color:#fff;font-family:'Inter',sans-serif;
                     font-size:1rem;font-weight:800;padding:16px 52px;cursor:pointer;
                     box-shadow:0 0 40px rgba(0,135,81,0.4);letter-spacing:0.04em;
                     transition:opacity 0.2s;"
              onmouseover="this.style.opacity='0.88'"
              onmouseout="this.style.opacity='1'">
        ✅ &nbsp; Proceed to Command System &nbsp; →
      </button>
      <p style="color:rgba(255,255,255,0.25);font-size:0.7rem;margin-top:12px;">
        Oneness and Progress — ACCORD 2026
      </p>
    </div>
  </div>

  <!-- Gold bottom border -->
  <div style="position:fixed;bottom:0;left:0;right:0;height:4px;background:linear-gradient(90deg,#ffc107,#008751,#ffc107);z-index:100000;"></div>
</div>

<script>
  // Splash now shows on every page load/refresh (no longer suppressed by sessionStorage)
</script>


<div class="scan-line"></div>
<canvas id="bg-canvas"></canvas>
<div class="grid-overlay"></div>
<div class="glow-center"></div>

<div class="wrapper">

    <!-- Header -->
    <div class="header">
        <div class="logo-ring">
            <img src="/logos/ACCORD.png" alt="ACCORD" onerror="this.src='https://via.placeholder.com/72x72/008751/ffffff?text=A'">
        </div>
        <div class="system-tag">Osun 2026 Governorship · Command System</div>
        <h1>ACCORD <span>Intelligence</span><br>Operations Centre</h1>
        <p class="subtitle">Real-time election collation, incident reporting and live analytics for the 2026 Osun Governorship Election.</p>
    </div>

    <!-- Stats bar -->
    <div class="stats-bar">
        <div class="stat">
            <span class="stat-val">3,763</span>
            <span class="stat-label">Polling Units</span>
        </div>
        <div class="stat-divider"></div>
        <div class="stat">
            <span class="stat-val">30</span>
            <span class="stat-label">LGAs</span>
        </div>
        <div class="stat-divider"></div>
        <div class="stat">
            <span class="stat-val">332</span>
            <span class="stat-label">Wards</span>
        </div>
        <div class="stat-divider"></div>
        <div class="stat">
            <span class="stat-val">14</span>
            <span class="stat-label">Parties</span>
        </div>
    </div>

    <!-- Navigation cards -->
    <div class="cards-grid">

        <a href="/vote" class="nav-card card-vote">
            <div class="card-icon icon-vote">🗳️</div>
            <div class="card-body">
                <div class="card-title">
                    Vote Submission
                    <span class="card-badge badge-field">Field</span>
                </div>
                <p class="card-desc">Field officers submit polling unit results in real-time. Secure officer ID validation with auto-filled PU data.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

        <a href="/report" class="nav-card card-report">
            <div class="card-icon icon-report">🚨</div>
            <div class="card-body">
                <div class="card-title">
                    Incident Report
                    <span class="card-badge badge-security">Security</span>
                </div>
                <p class="card-desc">Report election irregularities, security threats, or misconduct. Attach photo evidence and GPS location.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

        <a href="/dashboard" class="nav-card card-vdash">
            <div class="card-icon icon-vdash">📊</div>
            <div class="card-body">
                <div class="card-title">
                    Vote Dashboard
                    <span class="card-badge badge-live">Live</span>
                </div>
                <p class="card-desc">Live vote tallies, party breakdowns, LGA completion tracker, swing PU analysis and AI analytics log.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

        <a href="/incident-dashboard" class="nav-card card-idash">
            <div class="card-icon icon-idash">🛡️</div>
            <div class="card-body">
                <div class="card-title">
                    Incident Dashboard
                    <span class="card-badge badge-command">Command</span>
                </div>
                <p class="card-desc">Security command centre — live incident feed, severity heatmap, evidence viewer and real-time map.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

    </div>

    <!-- Footer -->
    <div class="footer">
        <img src="/static/logos/popson-logo.png" style="height:18px;vertical-align:middle;margin-right:6px;opacity:0.4;" onerror="this.style.display='none'">
        Powered by <strong>Popson Geospatial Services</strong> &nbsp;·&nbsp; ACCORD Party Osun 2026
    </div>

</div>

<script>
// ── Enhanced particle + ring canvas animation ────────────────────────────────
(function() {
    const canvas = document.getElementById('bg-canvas');
    const ctx = canvas.getContext('2d');
    let W, H, particles = [];
    let rings = [
        { r: 0, maxR: 0, alpha: 0, speed: 0.6 },
        { r: 0, maxR: 0, alpha: 0, speed: 0.4 },
        { r: 0, maxR: 0, alpha: 0, speed: 0.5 }
    ];
    let ringTimer = 0;
    const COUNT = 90;
    const GREEN = '0,135,81';
    const GOLD  = '255,193,7';
    const TEAL  = '0,179,104';

    function resize() {
        W = canvas.width  = window.innerWidth;
        H = canvas.height = window.innerHeight;
        rings.forEach((r, i) => { r.maxR = Math.max(W, H) * 0.7; });
    }

    function Particle() {
        this.reset = function() {
            this.x  = Math.random() * W;
            this.y  = Math.random() * H;
            this.vx = (Math.random() - 0.5) * 0.35;
            this.vy = (Math.random() - 0.5) * 0.35;
            this.r  = Math.random() * 2 + 0.3;
            const rnd = Math.random();
            this.color = rnd > 0.92 ? GOLD : rnd > 0.75 ? TEAL : GREEN;
            this.alpha = Math.random() * 0.55 + 0.15;
            this.pulse = Math.random() * Math.PI * 2;
        };
        this.reset();
    }

    function spawnRing() {
        const inactive = rings.find(r => r.alpha <= 0);
        if (!inactive) return;
        inactive.r = 0;
        inactive.alpha = 0.35;
        inactive.cx = W * 0.5 + (Math.random() - 0.5) * W * 0.3;
        inactive.cy = H * 0.5 + (Math.random() - 0.5) * H * 0.3;
    }

    function init() {
        particles = [];
        for (let i = 0; i < COUNT; i++) particles.push(new Particle());
    }

    function draw() {
        ctx.clearRect(0, 0, W, H);
        ringTimer++;
        if (ringTimer % 180 === 0) spawnRing();

        // Draw expanding rings
        rings.forEach(ring => {
            if (ring.alpha <= 0) return;
            ring.r += ring.speed;
            ring.alpha -= 0.001;
            if (ring.r > ring.maxR) { ring.alpha = 0; return; }
            ctx.beginPath();
            ctx.arc(ring.cx || W/2, ring.cy || H/2, ring.r, 0, Math.PI * 2);
            ctx.strokeStyle = `rgba(${GREEN},${ring.alpha * 0.5})`;
            ctx.lineWidth = 1;
            ctx.stroke();
        });

        // Connecting lines
        for (let i = 0; i < particles.length; i++) {
            for (let j = i + 1; j < particles.length; j++) {
                const dx = particles[i].x - particles[j].x;
                const dy = particles[i].y - particles[j].y;
                const dist = Math.sqrt(dx*dx + dy*dy);
                if (dist < 120) {
                    const alpha = (1 - dist / 120) * 0.15;
                    ctx.beginPath();
                    ctx.strokeStyle = `rgba(${GREEN},${alpha})`;
                    ctx.lineWidth = 0.5;
                    ctx.moveTo(particles[i].x, particles[i].y);
                    ctx.lineTo(particles[j].x, particles[j].y);
                    ctx.stroke();
                }
            }
        }

        // Particles with subtle pulse
        particles.forEach(p => {
            p.pulse += 0.02;
            const pr = p.r + Math.sin(p.pulse) * 0.4;
            ctx.beginPath();
            ctx.arc(p.x, p.y, pr, 0, Math.PI * 2);
            ctx.fillStyle = `rgba(${p.color},${p.alpha})`;
            ctx.fill();
            p.x += p.vx; p.y += p.vy;
            if (p.x < -10) p.x = W + 10;
            if (p.x > W + 10) p.x = -10;
            if (p.y < -10) p.y = H + 10;
            if (p.y > H + 10) p.y = -10;
        });

        requestAnimationFrame(draw);
    }

    window.addEventListener('resize', () => { resize(); init(); });
    resize();
    init();
    spawnRing();
    draw();
})();
</script>

</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/vote", response_class=HTMLResponse)
async def vote_form():
    parties = ["ACCORD", "AA", "AAC", "ADC", "ADP", "APGA", "APC", "APM", "APP", "BP", "NNPP", "PRP", "YPP", "ZLP"]
    party_cards = "".join([f'''
        <div class="col-4 col-md-2 mb-2">
            <div class="p-2 border rounded text-center bg-white shadow-sm">
                <img src="/logos/{p}.png" onerror="this.src='https://via.placeholder.com/30?text={p}'" style="height:30px">
                <small class="d-block fw-bold">{p}</small>
                <input type="number" class="form-control form-control-sm party-v text-center" data-p="{p}" value="0" inputmode="numeric" oninput="calculateTotals()">
            </div>
        </div>''' for p in parties])

    return f"""
<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <meta name="mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="theme-color" content="#008751">
    <title>IMOLE YOUTH ACCORD MOBILIZATION</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body {{ background: linear-gradient(rgba(0,0,0,0.6), rgba(0,0,0,0.6)), url('/static/bg.png'); background-size: cover; background-position: center; background-attachment: fixed; min-height: 100vh; margin: 0; }}
        .navbar {{ background: rgba(0, 135, 81, 0.9) !important; backdrop-filter: blur(10px); color: white; border-bottom: 4px solid #ffc107; }}
        .card {{ background: rgba(255, 255, 255, 0.95) !important; border-radius: 12px; border: none; box-shadow: 0 10px 30px rgba(0,0,0,0.3) !important; margin-bottom: 20px; color: #222; }}
        .section-label {{ font-size: 0.75rem; font-weight: bold; color: #008751; text-transform: uppercase; border-left: 3px solid #ffc107; padding-left: 10px; margin-bottom: 15px; display: block; }}
        input[readonly] {{ background-color: #e9ecef !important; font-weight: bold; }}
        #loginArea {{ margin-top: 100px; }}

        /* ── MOBILE OPTIMISATIONS ── */
        @media (max-width: 768px) {{
            body {{ background-attachment: scroll; }}
            .navbar h5 {{ font-size: 0.75rem; padding: 0 8px; }}
            .container {{ padding-left: 12px !important; padding-right: 12px !important; }}
            .card {{ border-radius: 14px; margin-bottom: 14px; }}
            .card.p-4 {{ padding: 14px !important; }}
            .card.p-5 {{ padding: 22px 16px !important; }}

            /* Stack the 3-col selects vertically */
            .row .col-4 {{ width: 100% !important; flex: 0 0 100%; max-width: 100%; }}
            /* But keep party vote grid at 3-col */
            .row .col-4.col-md-2 {{ width: 33.33% !important; flex: 0 0 33.33%; max-width: 33.33%; }}

            /* Big touch targets */
            .form-control, .form-select {{
                min-height: 48px !important;
                font-size: 1rem !important;
                border-radius: 10px !important;
            }}
            /* Number pad for vote inputs */
            .party-v {{ inputmode: numeric; min-height: 44px !important; font-size: 1rem !important; font-weight: 700; }}

            /* Big submit button */
            .btn.btn-success.btn-lg {{
                min-height: 56px !important;
                font-size: 1.05rem !important;
                border-radius: 14px !important;
            }}
            .btn.btn-outline-dark {{
                min-height: 50px !important;
                font-size: 0.95rem !important;
                border-radius: 12px !important;
            }}
            /* Modal full-width on mobile */
            .modal-dialog {{ margin: 8px; }}
            .modal-dialog-centered {{ align-items: flex-end; }}
            .modal-content {{ border-radius: 16px 16px 0 0 !important; }}
            /* Login card */
            #loginArea {{ margin-top: 50px; }}
            /* Section labels */
            .section-label {{ font-size: 0.72rem; }}
            /* Party card images */
            .col-4.col-md-2 img {{ height: 26px !important; }}
        }}
    </style>
</head>
<body>
    <nav class="navbar py-2 mb-4 text-center"><h5>IMOLE YOUTH ACCORD MOBILIZATION OFFICIAL FIELD COLLATION</h5></nav>
    <div class="container pb-5" style="max-width: 850px;">
        <!-- ── Step 1: LGA + Officer ID ── -->
        <div id="step1Area" class="card p-5 text-center mx-auto" style="max-width: 420px;">
            <div style="width:56px;height:56px;border-radius:50%;background:rgba(0,135,81,0.12);border:1px solid rgba(0,135,81,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">🗳️</div>
            <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ffc107;background:rgba(255,193,7,0.1);border:1px solid rgba(255,193,7,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 1 of 2 — Officer Verification</div>
            <h6 class="fw-bold mb-1">Select Your LGA &amp; Enter Officer ID</h6>
            <p class="small text-muted mb-3">Both fields must match your official assignment</p>
            <select id="lgaSelect" class="form-select mb-2 text-center fw-bold" style="background:#1a1a1a;color:#fff;border:1px solid rgba(0,135,81,0.4);border-radius:8px;">
                <option value="">— Select Your LGA —</option>
                <option value="Atakumosa East">Atakumosa East</option>
                <option value="Atakumosa West">Atakumosa West</option>
                <option value="Ayedaade">Ayedaade</option>
                <option value="Ayedire">Ayedire</option>
                <option value="Boluwaduro">Boluwaduro</option>
                <option value="Boripe">Boripe</option>
                <option value="Ede North">Ede North</option>
                <option value="Ede South">Ede South</option>
                <option value="Egbedore">Egbedore</option>
                <option value="Ejigbo">Ejigbo</option>
                <option value="Ife Central">Ife Central</option>
                <option value="Ife East">Ife East</option>
                <option value="Ife North">Ife North</option>
                <option value="Ife South">Ife South</option>
                <option value="Ifedayo">Ifedayo</option>
                <option value="Ifelodun">Ifelodun</option>
                <option value="Ila">Ila</option>
                <option value="Ilesa East">Ilesa East</option>
                <option value="Ilesa West">Ilesa West</option>
                <option value="Irepodun">Irepodun</option>
                <option value="Irewole">Irewole</option>
                <option value="Isokan">Isokan</option>
                <option value="Iwo">Iwo</option>
                <option value="Obokun">Obokun</option>
                <option value="Odo-Otin">Odo-Otin</option>
                <option value="Ola-Oluwa">Ola-Oluwa</option>
                <option value="Olorunda">Olorunda</option>
                <option value="Oriade">Oriade</option>
                <option value="Orolu">Orolu</option>
                <option value="Osogbo">Osogbo</option>
            </select>
            <input type="text" id="oid" class="form-control mb-2 text-center fw-bold" placeholder="e.g. 10-001" autocomplete="off" style="letter-spacing:0.08em;" onkeydown="if(event.key==='Enter')requestOTP()">
            <div id="step1Error" class="alert alert-danger d-none small py-2 mb-2"></div>
            <button class="btn btn-success w-100 fw-bold" id="step1Btn" onclick="requestOTP()">Send OTP to My WhatsApp →</button>
        </div>

        <!-- ── Step 2: OTP verification ── -->
        <div id="step2Area" class="card p-5 text-center mx-auto d-none" style="max-width: 420px;">
            <div style="width:56px;height:56px;border-radius:50%;background:rgba(0,135,81,0.12);border:1px solid rgba(0,135,81,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">📱</div>
            <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ffc107;background:rgba(255,193,7,0.1);border:1px solid rgba(255,193,7,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 2 of 2 — OTP Verification</div>
            <h6 class="fw-bold mb-1">Check Your WhatsApp</h6>
            <p class="small text-muted mb-1">A 6-digit code was sent to</p>
            <p class="fw-bold mb-3" id="phoneHintDisplay" style="color:#008751;font-size:1rem;letter-spacing:0.1em;">+234***0000</p>
            <input type="text" id="otpInput" class="form-control mb-2 text-center fw-bold" placeholder="000000" maxlength="6" inputmode="numeric" autocomplete="one-time-code" style="font-size:1.4rem;letter-spacing:0.3em;" onkeydown="if(event.key==='Enter')verifyOTP()">
            <div id="step2Error" class="alert alert-danger d-none small py-2 mb-2"></div>
            <button class="btn btn-success w-100 fw-bold mb-2" id="step2Btn" onclick="verifyOTP()">Verify OTP &amp; Unlock Form →</button>
            <div class="d-flex justify-content-between align-items-center">
                <button class="btn btn-link btn-sm text-muted p-0" onclick="backToStep1()">← Change ID</button>
                <button class="btn btn-link btn-sm p-0" id="resendBtn" onclick="resendOTP()" style="color:#008751;">Resend OTP</button>
            </div>
            <div id="resendCountdown" class="small text-muted mt-1 d-none"></div>
        </div>

        <div id="formArea" class="d-none">
            <div class="card p-4">
                <span class="section-label">1. Polling Unit — Auto-filled from Officer ID</span>
                <div class="row g-2">
                    <div class="col-6"><small class="text-muted">State</small><input type="text" id="s" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">LGA</small><input type="text" id="l" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">Ward</small><input type="text" id="w" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">Ward Code</small><input type="text" id="wc" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">PU Code</small><input type="text" id="pc" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">Polling Unit</small><input type="text" id="loc" class="form-control" readonly></div>
                </div>
            </div>

            <div class="card p-4">
                <span class="section-label">2. Official 14-Party Scorecard — 2026 Osun Governorship</span>
                <div class="row g-2">{party_cards}</div>
            </div>

            <div class="card p-4">
                <span class="section-label">3. Accreditation & Audit</span>
                <div class="row g-3">
                    <div class="col-4"><label class="small">Registered</label><input type="number" id="rv" class="form-control" value="0"></div>
                    <div class="col-4"><label class="small">Accredited</label><input type="number" id="ta" class="form-control" oninput="calculateTotals()"></div>
                    <div class="col-4"><label class="small">Rejected</label><input type="number" id="rj" class="form-control" value="0" oninput="calculateTotals()"></div>
                    <div class="col-6"><label class="small text-success">Valid</label><input type="number" id="vv" class="form-control" readonly></div>
                    <div class="col-6"><label class="small text-primary">Total Cast</label><input type="number" id="tc" class="form-control" readonly></div>
                </div>
                <div id="auditStatus" class="mt-3 p-2 rounded text-center d-none small fw-bold"></div>
            </div>

            <!-- BUG FIX #5: GPS no longer hard-blocks — shows warning but allows submission -->
            <button class="btn btn-outline-dark w-100 mb-2" onclick="getGPS()">
                <span id="gpsLabel">📍 Fix GPS Location (Recommended)</span>
            </button>
            <div id="gpsWarning" class="alert alert-warning d-none small py-2 mb-2">
                ⚠️ GPS not captured. You can still submit, but location will not be recorded.
            </div>

            <div class="card mb-3">
                <div class="card-body p-3">
                    <span class="section-label">4. EC 8E FORM IMAGE UPLOAD</span>
                    <div class="mt-2">
                        <label class="form-label small text-muted mb-1">Attach a clear photo of the signed EC 8E form (optional)</label>
                        <input type="file" id="ec8eFile" accept="image/*" capture="environment" class="form-control form-control-sm bg-dark text-white border-secondary">
                        <div id="ec8ePreview" class="mt-2 text-center d-none">
                            <img id="ec8eImg" src="#" alt="EC8E Preview" style="max-width:100%;max-height:200px;border-radius:8px;border:1px solid #ffc107;">
                        </div>
                    </div>
                </div>
            </div>
            <div id="submitError" class="alert alert-danger d-none small py-2 mb-2 fw-bold" style="border-radius:8px;"></div>
            <button class="btn btn-success btn-lg w-100 py-3 fw-bold" onclick="reviewSubmission()">UPLOAD PU RESULT</button>
        </div>
    </div>

        </div><!-- /formArea -->
    </div>
    <!-- Confirmation Modal -->
    <div class="modal fade" id="confirmModal" tabindex="-1">
      <div class="modal-dialog modal-dialog-centered">
        <div class="modal-content" style="background:#1a1a1a;color:#fff;border:1px solid #ffc107;">
          <div class="modal-header" style="border-bottom:1px solid #333;">
            <h6 class="modal-title text-warning fw-bold">⚠️ CONFIRM RESULT SUBMISSION</h6>
          </div>
          <div class="modal-body">
            <div id="confirmPUInfo" class="mb-3 p-2 rounded" style="background:#111;font-size:0.8rem;"></div>
            <table class="table table-sm table-dark table-bordered mb-2" style="font-size:0.8rem;">
              <thead><tr><th>Party</th><th class="text-end">Votes</th></tr></thead>
              <tbody id="confirmPartyRows"></tbody>
            </table>
            <div id="confirmAuditRows" class="p-2 rounded" style="background:#111;font-size:0.8rem;"></div>
          </div>
          <div class="modal-footer" style="border-top:1px solid #333;">
            <button type="button" class="btn btn-outline-secondary btn-sm" data-bs-dismiss="modal">← EDIT</button>
            <button type="button" class="btn btn-success btn-sm fw-bold" onclick="confirmAndSubmit()">✅ CONFIRM & SUBMIT</button>
          </div>
        </div>
      </div>
    </div>
    <div class="container pb-2" style="max-width:850px;">
        <div style="text-align:center;padding:18px 0 10px;border-top:1px solid rgba(255,255,255,0.12);margin-top:10px;">
            <img src="/static/logos/popson-logo.png" style="height:28px;object-fit:contain;vertical-align:middle;margin-right:8px;opacity:0.85;">
            <span style="color:rgba(255,255,255,0.55);font-size:0.75rem;vertical-align:middle;">Powered by <strong style="color:#fff;">Popson Geospatial Services</strong></span>
        </div>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>

    <script>
        let lat = null, lon = null, officerId, officerLga = '', submitToken = null, pendingPayload = null;
        let _resendTimer = null;

        // ── Step 1: request OTP ──────────────────────────────────────────────
        async function requestOTP() {{
            const rawId = document.getElementById('oid').value.trim().toUpperCase();
            const lga   = document.getElementById('lgaSelect').value;
            if (!lga) {{
                const errEl = document.getElementById('step1Error');
                errEl.innerText = 'Please select your LGA first.';
                errEl.classList.remove('d-none');
                return;
            }}
            if (!rawId) return;
            const btn  = document.getElementById('step1Btn');
            const errEl = document.getElementById('step1Error');
            btn.disabled = true; btn.innerText = 'Sending OTP...';
            errEl.classList.add('d-none');
            try {{
                const res = await fetch('/api/request-otp', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{ officer_id: rawId, lg: lga }})
                }});
                const out = await res.json();
                if (!res.ok) {{
                    errEl.innerText = out.detail || 'Error. Try again.';
                    errEl.classList.remove('d-none');
                    btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                    return;
                }}
                if (out.status === 'no_phone') {{
                    errEl.innerText = out.message;
                    errEl.classList.remove('d-none');
                    btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                    return;
                }}
                // OTP sent — move to step 2
                officerId = rawId;
                officerLga = lga;
                document.getElementById('phoneHintDisplay').innerText = out.phone_hint;
                document.getElementById('step1Area').classList.add('d-none');
                document.getElementById('step2Area').classList.remove('d-none');
                document.getElementById('otpInput').focus();
                startResendCountdown(60);
            }} catch(e) {{
                errEl.innerText = 'Server error. Try again.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
            }}
        }}

        // ── Step 2: verify OTP ───────────────────────────────────────────────
        async function verifyOTP() {{
            const otp   = document.getElementById('otpInput').value.trim();
            if (!otp || otp.length !== 6) return;
            const btn   = document.getElementById('step2Btn');
            const errEl = document.getElementById('step2Error');
            btn.disabled = true; btn.innerText = 'Verifying...';
            errEl.classList.add('d-none');
            try {{
                const res = await fetch('/api/verify-otp', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{ officer_id: officerId, lg: officerLga, otp }})
                }});
                const out = await res.json();
                if (!res.ok) {{
                    errEl.innerText = out.detail || 'Incorrect OTP.';
                    errEl.classList.remove('d-none');
                    btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
                    return;
                }}
                // ✅ Auth complete — store token, fill PU fields, show form
                submitToken = out.token;
                document.getElementById('s').value   = (out.state    || 'osun').toUpperCase();
                document.getElementById('l').value   = (out.lg       || '').toUpperCase();
                document.getElementById('w').value   = (out.ward     || '').toUpperCase();
                document.getElementById('wc').value  = out.ward_code || '';
                document.getElementById('pc').value  = out.pu_code   || '';
                document.getElementById('loc').value = (out.location || '').toUpperCase();
                document.getElementById('step2Area').classList.add('d-none');
                document.getElementById('formArea').classList.remove('d-none');
                if (_resendTimer) clearInterval(_resendTimer);
            }} catch(e) {{
                errEl.innerText = 'Server error. Try again.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
            }}
        }}

        // ── Resend OTP ───────────────────────────────────────────────────────
        async function resendOTP() {{
            document.getElementById('resendBtn').disabled = true;
            document.getElementById('otpInput').value = '';
            document.getElementById('step2Error').classList.add('d-none');
            // Re-use requestOTP logic — put ID back and call it
            document.getElementById('step2Area').classList.add('d-none');
            document.getElementById('step1Area').classList.remove('d-none');
            document.getElementById('step1Btn').disabled = false;
            document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
            await requestOTP();
        }}

        function backToStep1() {{
            if (_resendTimer) clearInterval(_resendTimer);
            document.getElementById('step2Area').classList.add('d-none');
            document.getElementById('step1Area').classList.remove('d-none');
            document.getElementById('step1Btn').disabled = false;
            document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
            document.getElementById('step1Error').classList.add('d-none');
            document.getElementById('otpInput').value = '';
            officerLga = '';
        }}

        function startResendCountdown(seconds) {{
            const btn = document.getElementById('resendBtn');
            const cd  = document.getElementById('resendCountdown');
            btn.disabled = true;
            cd.classList.remove('d-none');
            let remaining = seconds;
            cd.innerText = `Resend available in ${{remaining}}s`;
            _resendTimer = setInterval(() => {{
                remaining--;
                if (remaining <= 0) {{
                    clearInterval(_resendTimer);
                    btn.disabled = false;
                    cd.classList.add('d-none');
                }} else {{
                    cd.innerText = `Resend available in ${{remaining}}s`;
                }}
            }}, 1000);
        }}

        // PU fields are now auto-filled from verify-otp — no manual dropdowns needed
        function calculateTotals() {{
            let valid = 0;
            document.querySelectorAll('.party-v').forEach(i => valid += parseInt(i.value || 0));
            const rej = parseInt(document.getElementById('rj').value || 0);
            const acc = parseInt(document.getElementById('ta').value || 0);
            const cast = valid + rej;
            document.getElementById('vv').value = valid;
            document.getElementById('tc').value = cast;
            const msg = document.getElementById('auditStatus');
            if (acc > 0 && cast > acc) {{
                msg.innerHTML = "⚠️ ERROR: Over-voting!";
                msg.className = "mt-3 p-2 bg-danger text-white rounded text-center small fw-bold d-block";
            }} else if (cast > 0 && cast === acc) {{
                msg.innerHTML = "✅ AUDIT BALANCED";
                msg.className = "mt-3 p-2 bg-success text-white rounded text-center small fw-bold d-block";
            }} else {{ msg.className = "d-none"; }}
        }}

        // BUG FIX #5: GPS is now optional — shows label + warning, does not block submission
        function getGPS() {{
            navigator.geolocation.getCurrentPosition(
                pos => {{
                    lat = pos.coords.latitude;
                    lon = pos.coords.longitude;
                    document.getElementById('gpsLabel').innerText = `✅ GPS Fixed (${{lat.toFixed(4)}}, ${{lon.toFixed(4)}})`;
                    document.getElementById('gpsWarning').classList.add('d-none');
                }},
                err => {{
                    document.getElementById('gpsWarning').classList.remove('d-none');
                }}
            );
        }}

        // BUG FIX #2: ec8eFile listener wrapped in DOMContentLoaded
        document.addEventListener('DOMContentLoaded', function() {{
            const ec8eInput = document.getElementById('ec8eFile');
            if (ec8eInput) {{
                ec8eInput.addEventListener('change', function() {{
                    const file = this.files[0];
                    if (file) {{
                        const reader = new FileReader();
                        reader.onload = e => {{
                            document.getElementById('ec8eImg').src = e.target.result;
                            document.getElementById('ec8ePreview').classList.remove('d-none');
                        }};
                        reader.readAsDataURL(file);
                    }}
                }});
            }}
        }});

        async function reviewSubmission() {{
            // BUG FIX #5: GPS warning shown but not blocking
            if(!lat) {{
                document.getElementById('gpsWarning').classList.remove('d-none');
            }}
            const v = {{}};
            document.querySelectorAll('.party-v').forEach(i => v[i.dataset.p] = parseInt(i.value || 0));
            pendingPayload = {{
                officer_id: officerId, submit_token: submitToken,
                state: document.getElementById('s').value, lg: document.getElementById('l').value,
                ward: document.getElementById('w').value, ward_code: document.getElementById('wc').value,
                pu_code: document.getElementById('pc').value, location: document.getElementById('loc').value,
                reg_voters: parseInt(document.getElementById('rv').value || 0), total_accredited: parseInt(document.getElementById('ta').value || 0),
                valid_votes: parseInt(document.getElementById('vv').value || 0), rejected_votes: parseInt(document.getElementById('rj').value || 0),
                total_cast: parseInt(document.getElementById('tc').value || 0), lat, lon, votes: v
            }};
            document.getElementById('confirmPUInfo').innerHTML =
                `<b>📍 PU:</b> ${{pendingPayload.location}}<br>` +
                `<b>🗳 Ward:</b> ${{pendingPayload.ward}} &nbsp;|&nbsp; <b>LGA:</b> ${{pendingPayload.lg}}<br>` +
                `<b>🔑 PU Code:</b> ${{pendingPayload.pu_code}} &nbsp;|&nbsp; <b>Officer:</b> ${{pendingPayload.officer_id}}` +
                (lat ? `<br><b>📡 GPS:</b> ${{lat.toFixed(4)}}, ${{lon.toFixed(4)}}` : `<br><span class="text-warning">⚠️ No GPS captured</span>`);
            const tbody = document.getElementById('confirmPartyRows');
            tbody.innerHTML = '';
            Object.entries(v).forEach(([p, score]) => {{
                const tr = document.createElement('tr');
                tr.innerHTML = `<td>${{p}}</td><td class="text-end fw-bold ${{score > 0 ? 'text-warning' : 'text-secondary'}}">${{score}}</td>`;
                tbody.appendChild(tr);
            }});
            document.getElementById('confirmAuditRows').innerHTML =
                `<b>Registered:</b> ${{pendingPayload.reg_voters}} &nbsp;|&nbsp; ` +
                `<b>Accredited:</b> ${{pendingPayload.total_accredited}} &nbsp;|&nbsp; ` +
                `<b>Valid:</b> ${{pendingPayload.valid_votes}} &nbsp;|&nbsp; ` +
                `<b>Rejected:</b> ${{pendingPayload.rejected_votes}} &nbsp;|&nbsp; ` +
                `<b>Total Cast:</b> ${{pendingPayload.total_cast}}`;
            new bootstrap.Modal(document.getElementById('confirmModal')).show();
        }}

        async function confirmAndSubmit() {{
            const btn = document.querySelector('#confirmModal .btn-success');
            btn.disabled = true; btn.innerText = 'Submitting...';
            const fd = new FormData();
            fd.append('data', JSON.stringify(pendingPayload));
            const ec8eInput = document.getElementById('ec8eFile');
            if (ec8eInput && ec8eInput.files[0]) {{ fd.append('ec8e_image', ec8eInput.files[0]); }}
            const res = await fetch('/submit', {{ method: 'POST', body: fd }});
            const out = await res.json();
            bootstrap.Modal.getInstance(document.getElementById('confirmModal')).hide();
            alert(out.message);
            if(out.status === 'success') location.reload();
            else {{ btn.disabled = false; btn.innerText = '✅ CONFIRM & SUBMIT'; }}
        }}
    </script>
</body>
</html>
"""



# ── Incident Report OTP (no LGA required — officer ID lookup only) ─────────────
@app.post("/api/request-incident-otp")
async def request_incident_otp(request: Request):
    """
    Same as /api/request-otp but without requiring LGA.
    Looks up officer in Supabase first (by officer_id only), then SQLite (without LGA filter).
    Used by the Incident Report form so officers don't need to select their LGA.
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    import re as _re_inc
    officer_id = _re_inc.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)

    parts = officer_id.split("-", 1)
    if len(parts) != 2:
        raise HTTPException(status_code=400, detail="Invalid officer ID format. Expected: WARDCODE-PUCODE")
    ward_code, pu_code = parts[0].strip(), parts[1].strip()

    # Lockout check (no LGA in key for incidents)
    otp_key = f"incident|{officer_id}"
    entry = _OTP_STORE.get(otp_key, {})
    locked_until = entry.get("locked_until", 0)
    if time.time() < locked_until:
        remaining = int(locked_until - time.time())
        raise HTTPException(
            status_code=429,
            detail=f"Too many failed attempts. Try again in {remaining // 60}m {remaining % 60}s."
        )

    phone = None
    pu_data = {}

    # ── 1. Try Supabase (no LGA filter) ──────────────────────────────────────
    sb_officer = _get_supabase_officer(officer_id, "")  # empty lg = no filter
    if sb_officer:
        phone = _clean_phone(sb_officer.get("phone", "") or "")
        lg = sb_officer.get("lga", "")
        try:
            with get_db() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT ward, lg, location, pu_code, ward_code, state
                           FROM polling_units
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun'
                           LIMIT 1""",
                        (ward_code, pu_code)
                    )
                    row = cur.fetchone()
                    if row:
                        pu_data = {
                            "state":     row["state"] or "osun",
                            "ward":      row["ward"],
                            "lg":        row["lg"],
                            "location":  row["location"],
                            "pu_code":   row["pu_code"],
                            "ward_code": row["ward_code"],
                        }
        except Exception:
            pass
        if not pu_data:
            pu_data = {
                "state":     "osun",
                "ward":      sb_officer.get("ward", ""),
                "lg":        lg,
                "location":  sb_officer.get("polling_unit", ""),
                "pu_code":   pu_code,
                "ward_code": ward_code,
            }
    else:
        # ── 2. Fallback: SQLite polling_units (no LGA filter) ─────────────────
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT ward, lg, location, pu_code, ward_code, state, officer_phone
                       FROM polling_units
                       WHERE ward_code = ? AND pu_code = ?
                       AND LOWER(state) = 'osun'
                       LIMIT 1""",
                    (ward_code, pu_code)
                )
                row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Officer ID not found. Access Denied.")

        phone = _clean_phone(row["officer_phone"] or "") if row["officer_phone"] else None
        pu_data = {
            "state":     row["state"] or "osun",
            "ward":      row["ward"],
            "lg":        row["lg"],
            "location":  row["location"],
            "pu_code":   row["pu_code"],
            "ward_code": row["ward_code"],
        }

    if not phone or len(phone) < 10:
        return {
            "status":  "no_phone",
            "message": "No phone number registered for this officer ID. Contact your supervisor."
        }

    # Generate & store OTP
    otp = _generate_otp()
    _OTP_STORE[otp_key] = {
        "otp":          otp,
        "expiry":       time.time() + _OTP_TTL,
        "phone_hint":   _mask_phone(phone),
        "phone":        phone,
        "used":         False,
        "attempts":     0,
        "locked_until": 0,
        "pu_data":      pu_data,
    }

    # Send via Twilio WhatsApp
    account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
    auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
    from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")

    if not account_sid or not auth_token:
        dev_mode = os.environ.get("OTP_DEV_MODE", "").lower() in ("1", "true", "yes")
        if dev_mode:
            logger.warning("Twilio not configured — OTP not sent (dev mode)")
            logger.info(f"[DEV OTP] {officer_id}: {otp}")
        else:
            raise HTTPException(status_code=503, detail="OTP service is not configured. Contact your administrator.")
    else:
        try:
            from twilio.rest import Client as _TC2
            client = _TC2(account_sid, auth_token)
            msg = (
                f"🚨 *ACCORD INCIDENT REPORTING*\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"Your One-Time Password:\n\n"
                f"*{otp}*\n\n"
                f"Valid for *5 minutes*. Do NOT share this code.\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"_Powered by Popson Geospatial Services_"
            )
            client.messages.create(from_=f"whatsapp:{from_number}", to=f"whatsapp:{phone}", body=msg)
            logger.info(f"✅ Incident OTP sent to {_mask_phone(phone)} for officer {officer_id}")
        except Exception as e:
            logger.error(f"Twilio OTP send failed for {officer_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to send OTP: {str(e)}")

    return {"status": "sent", "phone_hint": _mask_phone(phone)}


@app.post("/api/verify-incident-otp")
async def verify_incident_otp(request: Request):
    """
    Verify OTP for incident reporting. Returns officer/PU data on success.
    No submit token required — incident form uses its own payload submission.
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    import re as _re_inc2
    officer_id = _re_inc2.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)
    given_otp  = str(body.get("otp", "")).strip()[:6]

    otp_key = f"incident|{officer_id}"
    entry   = _OTP_STORE.get(otp_key)

    if not entry:
        raise HTTPException(status_code=400, detail="No OTP requested for this officer. Start again.")

    if time.time() < entry.get("locked_until", 0):
        remaining = int(entry["locked_until"] - time.time())
        raise HTTPException(status_code=429, detail=f"Account locked. Try again in {remaining // 60}m {remaining % 60}s.")

    if entry.get("used"):
        raise HTTPException(status_code=400, detail="OTP already used. Request a new one.")

    if time.time() > entry["expiry"]:
        _OTP_STORE.pop(otp_key, None)
        raise HTTPException(status_code=400, detail="OTP expired. Request a new one.")

    if not secrets.compare_digest(given_otp, entry["otp"]):
        entry["attempts"] = entry.get("attempts", 0) + 1
        remaining_tries = _OTP_MAX_TRIES - entry["attempts"]
        if entry["attempts"] >= _OTP_MAX_TRIES:
            entry["locked_until"] = time.time() + _OTP_LOCKOUT
            entry["used"] = True
            raise HTTPException(status_code=429, detail=f"Too many wrong attempts. Account locked for {_OTP_LOCKOUT // 60} minutes.")
        raise HTTPException(status_code=401, detail=f"Incorrect OTP. {remaining_tries} attempt{'s' if remaining_tries != 1 else ''} remaining.")

    # ✅ Correct OTP
    entry["used"] = True
    pu_data = entry["pu_data"]
    return {
        "status":    "ok",
        "officer_id": officer_id,
        "state":     pu_data["state"],
        "ward":      pu_data["ward"],
        "lg":        pu_data["lg"],
        "location":  pu_data["location"],
        "pu_code":   pu_data["pu_code"],
        "ward_code": pu_data["ward_code"],
    }

# ── INCIDENT REPORT FORM ──────────────────────────────────────────────────────
@app.get("/report", response_class=HTMLResponse)
async def report_page():
    return HTMLResponse(content=REPORT_HTML)

REPORT_HTML = """
<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <meta name="mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="theme-color" content="#cc0000">
    <title>ACCORD — INCIDENT REPORT</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background: linear-gradient(rgba(0,0,0,0.7), rgba(0,0,0,0.7)), url('/static/bg.png'); background-size: cover; background-position: center; background-attachment: fixed; min-height: 100vh; margin: 0; }
        .navbar { background: rgba(180,0,0,0.92) !important; backdrop-filter: blur(10px); color: white; border-bottom: 4px solid #ff6600; }
        .card { background: rgba(255,255,255,0.96) !important; border-radius: 12px; border: none; box-shadow: 0 10px 30px rgba(0,0,0,0.4) !important; margin-bottom: 20px; color: #222; }
        .section-label { font-size: 0.75rem; font-weight: bold; color: #cc0000; text-transform: uppercase; border-left: 3px solid #ff6600; padding-left: 10px; margin-bottom: 15px; display: block; }
        .severity-btn { border: 2px solid #ddd; border-radius: 8px; padding: 10px; cursor: pointer; text-align: center; transition: all 0.2s; background: #fff; }
        .severity-btn.active-low { border-color: #28a745; background: #d4edda; }
        .severity-btn.active-medium { border-color: #ffc107; background: #fff3cd; }
        .severity-btn.active-critical { border-color: #dc3545; background: #f8d7da; }
        #step1Area, #step2Area { margin-top: 100px; }
        @media (max-width: 768px) {
            body { background-attachment: scroll; }
            .navbar h5 { font-size: 0.75rem; }
            #step1Area, #step2Area { margin-top: 50px; }
            .form-control, .form-select { min-height: 48px !important; font-size: 1rem !important; border-radius: 10px !important; }
            .btn-lg { min-height: 56px !important; font-size: 1.05rem !important; border-radius: 14px !important; }
        }
    </style>
</head>
<body>
<nav class="navbar py-2 mb-4 text-center"><h5>⚠️ ACCORD INCIDENT REPORTING SYSTEM</h5></nav>
<div class="container pb-5" style="max-width: 750px;">

    <!-- Step 1: Officer ID only (no LGA needed) -->
    <div id="step1Area" class="card p-5 text-center mx-auto" style="max-width: 420px;">
        <div style="width:56px;height:56px;border-radius:50%;background:rgba(180,0,0,0.1);border:1px solid rgba(180,0,0,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">🚨</div>
        <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ff6600;background:rgba(255,102,0,0.1);border:1px solid rgba(255,102,0,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 1 of 2 — Officer Verification</div>
        <h6 class="fw-bold mb-1 text-danger">🔐 Officer Verification</h6>
        <p class="small text-muted mb-3">Enter your Officer ID to receive an OTP on WhatsApp</p>
        <input type="text" id="oid" class="form-control mb-3 text-center fw-bold" placeholder="e.g. 10-001" autocomplete="off" style="letter-spacing:0.08em;" onkeydown="if(event.key==='Enter')requestOTP()">
        <div id="step1Error" class="alert alert-danger d-none small py-2 mb-2"></div>
        <button class="btn btn-danger w-100 fw-bold" id="step1Btn" onclick="requestOTP()">Send OTP to My WhatsApp →</button>
        <div class="mt-3">
            <a href="/vote" class="small text-muted">← Back to Vote Submission</a>
        </div>
    </div>

    <!-- Step 2: OTP entry -->
    <div id="step2Area" class="card p-5 text-center mx-auto d-none" style="max-width: 420px;">
        <div style="width:56px;height:56px;border-radius:50%;background:rgba(180,0,0,0.1);border:1px solid rgba(180,0,0,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">📱</div>
        <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ff6600;background:rgba(255,102,0,0.1);border:1px solid rgba(255,102,0,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 2 of 2 — OTP Verification</div>
        <h6 class="fw-bold mb-1">Check Your WhatsApp</h6>
        <p class="small text-muted mb-1">A 6-digit code was sent to</p>
        <p class="fw-bold mb-3" id="phoneHintDisplay" style="color:#cc0000;font-size:1rem;letter-spacing:0.1em;">+234***0000</p>
        <input type="text" id="otpInput" class="form-control mb-2 text-center fw-bold" placeholder="000000" maxlength="6" inputmode="numeric" autocomplete="one-time-code" style="font-size:1.4rem;letter-spacing:0.3em;" onkeydown="if(event.key==='Enter')verifyOTP()">
        <div id="step2Error" class="alert alert-danger d-none small py-2 mb-2"></div>
        <button class="btn btn-danger w-100 fw-bold mb-2" id="step2Btn" onclick="verifyOTP()">Verify OTP &amp; Unlock Form →</button>
        <div class="d-flex justify-content-between align-items-center">
            <button class="btn btn-link btn-sm text-muted p-0" onclick="backToStep1()">← Change ID</button>
            <button class="btn btn-link btn-sm p-0" id="resendBtn" onclick="resendOTP()" style="color:#cc0000;">Resend OTP</button>
        </div>
        <div id="resendCountdown" class="small text-muted mt-1 d-none"></div>
    </div>

    <!-- Incident Form (shown after OTP verified) -->
    <div id="formArea" class="d-none">

        <div class="card p-4">
            <span class="section-label">1. Officer &amp; Location</span>
            <div class="row g-2">
                <div class="col-6"><small class="text-muted">Officer ID</small><input type="text" id="disp_officer" class="form-control" readonly></div>
                <div class="col-6"><small class="text-muted">PU Code</small><input type="text" id="disp_pu" class="form-control" readonly></div>
                <div class="col-12"><small class="text-muted">Polling Unit</small><input type="text" id="disp_loc" class="form-control" readonly></div>
                <div class="col-6"><small class="text-muted">Ward</small><input type="text" id="disp_ward" class="form-control" readonly></div>
                <div class="col-6"><small class="text-muted">LGA</small><input type="text" id="disp_lg" class="form-control" readonly></div>
            </div>
        </div>

        <div class="card p-4">
            <span class="section-label">2. Incident Type</span>
            <select id="incident_type" class="form-select">
                <option value="">-- Select Incident Type --</option>
                <option value="Violence / security threat">⚔️ Violence / Security Threat</option>
                <option value="Ballot box snatching">🗳️ Ballot Box Snatching</option>
                <option value="INEC official misconduct">🏛️ INEC Official Misconduct</option>
                <option value="Voter intimidation">😰 Voter Intimidation</option>
                <option value="Equipment failure (BVAS etc)">🖥️ Equipment Failure (BVAS etc)</option>
                <option value="Other irregularities">⚠️ Other Irregularities</option>
            </select>
        </div>

        <div class="card p-4">
            <span class="section-label">3. Severity Level</span>
            <div class="row g-2">
                <div class="col-4">
                    <div class="severity-btn" id="btn-low" onclick="setSeverity('Low')">
                        <div style="font-size:1.5rem;">🟢</div>
                        <div class="fw-bold small">LOW</div>
                        <div style="font-size:0.65rem; color:#666;">Minor issue</div>
                    </div>
                </div>
                <div class="col-4">
                    <div class="severity-btn" id="btn-medium" onclick="setSeverity('Medium')">
                        <div style="font-size:1.5rem;">🟡</div>
                        <div class="fw-bold small">MEDIUM</div>
                        <div style="font-size:0.65rem; color:#666;">Needs attention</div>
                    </div>
                </div>
                <div class="col-4">
                    <div class="severity-btn" id="btn-critical" onclick="setSeverity('Critical')">
                        <div style="font-size:1.5rem;">🔴</div>
                        <div class="fw-bold small">CRITICAL</div>
                        <div style="font-size:0.65rem; color:#666;">Urgent response</div>
                    </div>
                </div>
            </div>
            <input type="hidden" id="severity_val" value="">
        </div>

        <div class="card p-4">
            <span class="section-label">4. Description</span>
            <textarea id="description" class="form-control" rows="4" placeholder="Describe what happened in detail — who, what, when, how many people involved..."></textarea>
        </div>

        <div class="card p-4">
            <span class="section-label">5. Evidence (Photo / Video)</span>
            <label class="form-label small text-muted">Attach photo or video evidence (optional but strongly recommended)</label>
            <input type="file" id="evidenceFile" accept="image/*,video/*" capture="environment" class="form-control bg-dark text-white border-secondary">
            <div id="evidencePreview" class="mt-2 text-center d-none">
                <img id="evidenceImg" src="#" alt="Preview" style="max-width:100%;max-height:200px;border-radius:8px;border:2px solid #ff6600;">
            </div>
        </div>

        <button class="btn btn-outline-dark w-100 mb-2" onclick="getGPS()">
            <span id="gpsLabel">📍 Fix GPS Location (Recommended)</span>
        </button>
        <div id="gpsWarning" class="alert alert-warning d-none small py-2 mb-2">
            ⚠️ GPS not captured. You can still submit.
        </div>

        <div id="submitError" class="alert alert-danger d-none small py-2 mb-2 fw-bold"></div>
        <button class="btn btn-danger btn-lg w-100 py-3 fw-bold" onclick="submitIncident()">🚨 SUBMIT INCIDENT REPORT</button>
        <div class="mt-3 text-center">
            <a href="/vote" class="small text-white">← Back to Vote Submission</a>
        </div>
    </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script>
    let lat = null, lon = null, officerData = {}, _resendTimer = null;

    // ── Step 1: request OTP ──────────────────────────────────────────────────
    async function requestOTP() {
        const rawId = document.getElementById('oid').value.trim().toUpperCase();
        if (!rawId) return;
        const btn   = document.getElementById('step1Btn');
        const errEl = document.getElementById('step1Error');
        btn.disabled = true; btn.innerText = 'Sending OTP...';
        errEl.classList.add('d-none');
        try {
            const res = await fetch('/api/request-incident-otp', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ officer_id: rawId })
            });
            const out = await res.json();
            if (!res.ok) {
                errEl.innerText = out.detail || 'Error. Try again.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                return;
            }
            if (out.status === 'no_phone') {
                errEl.innerText = out.message;
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                return;
            }
            // OTP sent — move to step 2
            officerData.officer_id = rawId;
            document.getElementById('phoneHintDisplay').innerText = out.phone_hint;
            document.getElementById('step1Area').classList.add('d-none');
            document.getElementById('step2Area').classList.remove('d-none');
            document.getElementById('otpInput').focus();
            startResendCountdown(60);
        } catch(e) {
            errEl.innerText = 'Server error. Try again.';
            errEl.classList.remove('d-none');
            btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
        }
    }

    // ── Step 2: verify OTP ───────────────────────────────────────────────────
    async function verifyOTP() {
        const otp   = document.getElementById('otpInput').value.trim();
        if (!otp || otp.length !== 6) return;
        const btn   = document.getElementById('step2Btn');
        const errEl = document.getElementById('step2Error');
        btn.disabled = true; btn.innerText = 'Verifying...';
        errEl.classList.add('d-none');
        try {
            const res = await fetch('/api/verify-incident-otp', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ officer_id: officerData.officer_id, otp })
            });
            const out = await res.json();
            if (!res.ok) {
                errEl.innerText = out.detail || 'Incorrect OTP.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
                return;
            }
            // ✅ Auth done — fill officer/PU data and show incident form
            officerData = {
                officer_id: out.officer_id,
                pu_code:    out.pu_code,
                ward:       out.ward,
                ward_code:  out.ward_code,
                lg:         out.lg,
                state:      out.state,
                location:   out.location
            };
            document.getElementById('disp_officer').value = out.officer_id;
            document.getElementById('disp_pu').value      = out.pu_code  || '';
            document.getElementById('disp_loc').value     = out.location || '';
            document.getElementById('disp_ward').value    = out.ward     || '';
            document.getElementById('disp_lg').value      = out.lg       || '';
            document.getElementById('step2Area').classList.add('d-none');
            document.getElementById('formArea').classList.remove('d-none');
            if (_resendTimer) clearInterval(_resendTimer);
        } catch(e) {
            errEl.innerText = 'Server error. Try again.';
            errEl.classList.remove('d-none');
            btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
        }
    }

    // ── Resend OTP ───────────────────────────────────────────────────────────
    async function resendOTP() {
        document.getElementById('resendBtn').disabled = true;
        document.getElementById('otpInput').value = '';
        document.getElementById('step2Error').classList.add('d-none');
        document.getElementById('step2Area').classList.add('d-none');
        document.getElementById('step1Area').classList.remove('d-none');
        document.getElementById('step1Btn').disabled = false;
        document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
        await requestOTP();
    }

    function backToStep1() {
        if (_resendTimer) clearInterval(_resendTimer);
        document.getElementById('step2Area').classList.add('d-none');
        document.getElementById('step1Area').classList.remove('d-none');
        document.getElementById('step1Btn').disabled = false;
        document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
        document.getElementById('step1Error').classList.add('d-none');
        document.getElementById('otpInput').value = '';
    }

    function startResendCountdown(seconds) {
        const btn = document.getElementById('resendBtn');
        const cd  = document.getElementById('resendCountdown');
        btn.disabled = true;
        cd.classList.remove('d-none');
        let remaining = seconds;
        cd.innerText = `Resend available in ${remaining}s`;
        _resendTimer = setInterval(() => {
            remaining--;
            if (remaining <= 0) {
                clearInterval(_resendTimer);
                btn.disabled = false;
                cd.classList.add('d-none');
            } else {
                cd.innerText = `Resend available in ${remaining}s`;
            }
        }, 1000);
    }

    function setSeverity(level) {
        document.getElementById('severity_val').value = level;
        ['low','medium','critical'].forEach(l => {
            const btn = document.getElementById('btn-' + l);
            btn.className = 'severity-btn' + (l === level.toLowerCase() ? ' active-' + l : '');
        });
    }

    function getGPS() {
        navigator.geolocation.getCurrentPosition(
            pos => {
                lat = pos.coords.latitude;
                lon = pos.coords.longitude;
                document.getElementById('gpsLabel').innerText = `✅ GPS Fixed (${lat.toFixed(4)}, ${lon.toFixed(4)})`;
                document.getElementById('gpsWarning').classList.add('d-none');
            },
            () => { document.getElementById('gpsWarning').classList.remove('d-none'); }
        );
    }

    document.addEventListener('DOMContentLoaded', function() {
        const ef = document.getElementById('evidenceFile');
        if (ef) {
            ef.addEventListener('change', function() {
                const file = this.files[0];
                if (file && file.type.startsWith('image/')) {
                    const reader = new FileReader();
                    reader.onload = e => {
                        document.getElementById('evidenceImg').src = e.target.result;
                        document.getElementById('evidencePreview').classList.remove('d-none');
                    };
                    reader.readAsDataURL(file);
                }
            });
        }
    });

    async function submitIncident() {
        const errEl = document.getElementById('submitError');
        errEl.classList.add('d-none');

        const incident_type = document.getElementById('incident_type').value;
        const severity = document.getElementById('severity_val').value;
        const description = document.getElementById('description').value.trim();

        if (!incident_type) { errEl.innerText = 'Please select an incident type.'; errEl.classList.remove('d-none'); return; }
        if (!severity) { errEl.innerText = 'Please select a severity level.'; errEl.classList.remove('d-none'); return; }
        if (!description) { errEl.innerText = 'Please describe the incident.'; errEl.classList.remove('d-none'); return; }

        const btn = document.querySelector('button.btn-danger.btn-lg');
        btn.disabled = true; btn.innerText = 'Submitting...';

        const payload = {
            officer_id: officerData.officer_id,
            pu_code: officerData.pu_code,
            ward: officerData.ward,
            ward_code: officerData.ward_code || '',
            lg: officerData.lg,
            state: officerData.state || 'osun',
            location: officerData.location,
            incident_type, severity, description, lat, lon
        };

        const fd = new FormData();
        fd.append('data', JSON.stringify(payload));
        const ef = document.getElementById('evidenceFile');
        if (ef && ef.files[0]) fd.append('evidence', ef.files[0]);

        try {
            const res = await fetch('/submit-incident', { method: 'POST', body: fd });
            const out = await res.json();
            alert(out.message);
            if (out.status === 'success') {
                document.getElementById('incident_type').value = '';
                document.getElementById('severity_val').value = '';
                document.getElementById('description').value = '';
                document.getElementById('evidencePreview').classList.add('d-none');
                ['low','medium','critical'].forEach(l => { document.getElementById('btn-'+l).className = 'severity-btn'; });
                btn.disabled = false; btn.innerText = '🚨 SUBMIT INCIDENT REPORT';
            } else {
                btn.disabled = false; btn.innerText = '🚨 SUBMIT INCIDENT REPORT';
            }
        } catch (e) {
            errEl.innerText = 'Server error. Try again.';
            errEl.classList.remove('d-none');
            btn.disabled = false; btn.innerText = '🚨 SUBMIT INCIDENT REPORT';
        }
    }
</script>
</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────

# --- DASHBOARD PAGE ---
# ── INCIDENT DASHBOARD ────────────────────────────────────────────────────────
@app.get("/incident-dashboard", response_class=HTMLResponse)
async def incident_dashboard_page(request: Request):
    token = request.cookies.get("ds_session")
    if not _is_valid_token(token):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/dashboard", status_code=302)
    return HTMLResponse(content=INCIDENT_DASHBOARD_HTML)

INCIDENT_DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ACCORD — Incident Command Dashboard</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css">
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <style>
        :root { --red: #cc0000; --orange: #ff6600; --gold: #ffc107; --dark: #0a0a0a; --panel: #141414; }
        body { background: var(--dark); color: #fff; font-family: 'Segoe UI', sans-serif; margin: 0; overflow-y: auto; }

        .navbar-custom { background: #1a0000; border-bottom: 3px solid var(--red); padding: 10px 20px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px; }
        .brand-title { color: var(--red); font-weight: 900; font-size: 1.1rem; letter-spacing: 1px; }

        .kpi-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; padding: 10px; }
        .kpi-card { background: var(--panel); border-radius: 10px; padding: 15px; text-align: center; border: 1px solid #222; }
        .kpi-val { font-size: 2rem; font-weight: 900; display: block; }
        .kpi-label { font-size: 0.65rem; color: #aaa; text-transform: uppercase; margin-top: 4px; }
        .kpi-critical { border-top: 3px solid #ff0000; }
        .kpi-medium   { border-top: 3px solid #ffc107; }
        .kpi-low      { border-top: 3px solid #00cc44; }
        .kpi-total    { border-top: 3px solid var(--orange); }

        .main-grid { display: grid; grid-template-columns: 340px 1fr; gap: 10px; padding: 0 10px 10px; min-height: 0; }
        .side-panel { background: var(--panel); border-radius: 12px; border: 1px solid #222; display: flex; flex-direction: column; overflow: hidden; max-height: calc(100vh - 200px); }
        .panel-header { background: #1c0000; padding: 10px 15px; font-size: 0.75rem; font-weight: bold; color: var(--red); border-bottom: 1px solid #330000; text-transform: uppercase; display: flex; align-items: center; justify-content: space-between; }

        .filter-row { padding: 8px; display: flex; gap: 6px; flex-wrap: wrap; background: #0f0f0f; border-bottom: 1px solid #1a1a1a; }
        .filter-row select { flex: 1; min-width: 90px; background: #1a1a1a; color: #fff; border: 1px solid #333; border-radius: 6px; font-size: 0.72rem; padding: 4px 6px; }

        .feed-container { flex: 1; overflow-y: auto; padding: 8px; }
        .incident-card { background: #1e1e1e; border-radius: 8px; padding: 10px 12px; margin-bottom: 8px; cursor: pointer; transition: background 0.15s; }
        .incident-card:hover { background: #2a2a2a; }
        .incident-card.critical { border-left: 4px solid #ff0000; }
        .incident-card.medium   { border-left: 4px solid #ffc107; }
        .incident-card.low      { border-left: 4px solid #00cc44; }

        .sev-badge { display: inline-block; font-size: 0.6rem; font-weight: bold; padding: 2px 7px; border-radius: 10px; margin-left: 6px; }
        .sev-critical { background: #ff0000; color: #fff; }
        .sev-medium   { background: #ffc107; color: #000; }
        .sev-low      { background: #00cc44; color: #000; }

        #map { height: 380px; border-radius: 12px; background: #111; margin-bottom: 10px; }

        .evidence-panel { background: var(--panel); border-radius: 12px; border: 1px solid #222; padding: 12px; margin-bottom: 10px; min-height: 120px; }
        .evidence-panel-title { font-size: 0.7rem; color: var(--orange); font-weight: bold; text-transform: uppercase; margin-bottom: 8px; border-bottom: 1px solid #2a2a2a; padding-bottom: 6px; }

        .detail-panel { background: var(--panel); border-radius: 12px; border: 1px solid #222; padding: 14px; }
        .detail-row { display: flex; justify-content: space-between; padding: 5px 0; border-bottom: 1px solid #1a1a1a; font-size: 0.75rem; }
        .detail-row:last-child { border-bottom: none; }
        .detail-label { color: #666; }
        .detail-val { color: #fff; font-weight: bold; text-align: right; max-width: 60%; }

        @media (max-width: 768px) {
            .kpi-row { grid-template-columns: repeat(2, 1fr); }
            .main-grid { grid-template-columns: 1fr; }
            .side-panel { max-height: 400px; }
        }
    </style>
</head>
<body>

<nav class="navbar-custom">
    <div>
        <div class="brand-title">🚨 ACCORD INCIDENT COMMAND — OSUN 2026</div>
        <div style="font-size:0.65rem;color:#666;margin-top:2px;">Security & Command Team Dashboard · Auto-refresh every 30s</div>
    </div>
    <div style="display:flex;gap:8px;align-items:center;">
        <span id="last-refresh" style="font-size:0.65rem;color:#555;"></span>
        <button class="btn btn-sm btn-outline-danger py-1 px-3" style="font-size:11px;" onclick="loadIncidents()">
            <i class="bi bi-arrow-clockwise"></i> REFRESH
        </button>
        <a href="/dashboard" class="btn btn-sm btn-outline-warning py-1 px-3" style="font-size:11px;">
            📊 Vote Dashboard
        </a>
    </div>
</nav>

<!-- KPI Row -->
<div class="kpi-row">
    <div class="kpi-card kpi-total">
        <span id="kpi-total" class="kpi-val" style="color:var(--orange);">0</span>
        <div class="kpi-label">Total Incidents</div>
    </div>
    <div class="kpi-card kpi-critical">
        <span id="kpi-critical" class="kpi-val" style="color:#ff4444;">0</span>
        <div class="kpi-label">🔴 Critical</div>
    </div>
    <div class="kpi-card kpi-medium">
        <span id="kpi-medium" class="kpi-val" style="color:#ffc107;">0</span>
        <div class="kpi-label">🟡 Medium</div>
    </div>
    <div class="kpi-card kpi-low">
        <span id="kpi-low" class="kpi-val" style="color:#00cc44;">0</span>
        <div class="kpi-label">🟢 Low</div>
    </div>
</div>

<div class="main-grid">

    <!-- Left: Incident Feed -->
    <div class="side-panel">
        <div class="panel-header">
            LIVE INCIDENT FEED
            <span id="feed-count" class="badge bg-danger ms-1">0</span>
        </div>
        <div class="filter-row">
            <select id="f-severity" onchange="applyFilters()">
                <option value="">All Severity</option>
                <option value="Critical">🔴 Critical</option>
                <option value="Medium">🟡 Medium</option>
                <option value="Low">🟢 Low</option>
            </select>
            <select id="f-type" onchange="applyFilters()">
                <option value="">All Types</option>
                <option value="Violence / security threat">Violence</option>
                <option value="Ballot box snatching">Ballot Snatching</option>
                <option value="INEC official misconduct">INEC Misconduct</option>
                <option value="Voter intimidation">Voter Intimidation</option>
                <option value="Equipment failure (BVAS etc)">Equipment Failure</option>
                <option value="Other irregularities">Other</option>
            </select>
            <select id="f-lga" onchange="applyFilters()">
                <option value="">All LGAs</option>
            </select>
        </div>
        <div class="feed-container" id="incidentFeed"></div>
    </div>

    <!-- Right: Map + Detail -->
    <div>
        <div id="map"></div>

        <div class="evidence-panel" id="evidencePanel">
            <div class="evidence-panel-title">📷 Evidence Viewer — Click an incident to view</div>
            <div id="evidenceContent" style="text-align:center;color:#444;font-size:0.75rem;padding:20px 0;">
                No incident selected
            </div>
        </div>

        <div class="detail-panel" id="detailPanel">
            <div style="font-size:0.7rem;color:var(--orange);font-weight:bold;text-transform:uppercase;margin-bottom:10px;border-bottom:1px solid #2a2a2a;padding-bottom:6px;">
                📋 Incident Detail
            </div>
            <div id="detailContent" style="color:#555;font-size:0.75rem;">Select an incident from the feed</div>
        </div>
    </div>
</div>

<script>
    let map, allIncidents = [], markers = [];

    const SEV_COLOR = { Critical: '#ff0000', Medium: '#ffc107', Low: '#00cc44' };
    const TYPE_ICON = {
        'Violence / security threat': '⚔️',
        'Ballot box snatching': '🗳️',
        'INEC official misconduct': '🏛️',
        'Voter intimidation': '😰',
        'Equipment failure (BVAS etc)': '🖥️',
        'Other irregularities': '⚠️'
    };

    function initMap() {
        map = L.map('map', { zoomControl: true }).setView([7.56, 4.52], 9);
        L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png').addTo(map);
    }

    async function loadIncidents() {
        try {
            const res = await fetch('/api/incidents', { credentials: 'include' });
            allIncidents = await res.json();
            document.getElementById('last-refresh').textContent = 'Last refresh: ' + new Date().toLocaleTimeString('en-NG');
            populateLGAFilter();
            applyFilters();
        } catch(e) { console.error('Load error', e); }
    }

    function populateLGAFilter() {
        const lgas = [...new Set(allIncidents.map(i => i.lg).filter(Boolean))].sort();
        const sel = document.getElementById('f-lga');
        const cur = sel.value;
        sel.innerHTML = '<option value="">All LGAs</option>';
        lgas.forEach(l => sel.add(new Option(l.toUpperCase(), l)));
        sel.value = cur;
    }

    function applyFilters() {
        const sev = document.getElementById('f-severity').value;
        const typ = document.getElementById('f-type').value;
        const lga = document.getElementById('f-lga').value;

        let filtered = allIncidents;
        if (sev) filtered = filtered.filter(i => i.severity === sev);
        if (typ) filtered = filtered.filter(i => i.incident_type === typ);
        if (lga) filtered = filtered.filter(i => i.lg === lga);

        updateKPIs(filtered);
        renderFeed(filtered);
        renderMap(filtered);
    }

    function updateKPIs(data) {
        document.getElementById('kpi-total').textContent = data.length;
        document.getElementById('kpi-critical').textContent = data.filter(i => i.severity === 'Critical').length;
        document.getElementById('kpi-medium').textContent = data.filter(i => i.severity === 'Medium').length;
        document.getElementById('kpi-low').textContent = data.filter(i => i.severity === 'Low').length;
        document.getElementById('feed-count').textContent = data.length;
    }

    function renderFeed(data) {
        const feed = document.getElementById('incidentFeed');
        if (!data.length) {
            feed.innerHTML = '<div style="color:#555;font-size:0.75rem;padding:20px;text-align:center;">No incidents found</div>';
            return;
        }
        feed.innerHTML = data.map((inc, idx) => {
            const sev = (inc.severity || 'low').toLowerCase();
            const icon = TYPE_ICON[inc.incident_type] || '⚠️';
            const ts = inc.timestamp ? new Date(inc.timestamp).toLocaleString('en-NG', {day:'2-digit',month:'short',hour:'2-digit',minute:'2-digit'}) : '--';
            return `<div class="incident-card ${sev}" onclick="selectIncident(${idx})">
                <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px;">
                    <span style="font-size:0.8rem;font-weight:bold;">${icon} ${inc.incident_type || 'Unknown'}</span>
                    <span class="sev-badge sev-${sev}">${inc.severity || ''}</span>
                </div>
                <div style="font-size:0.72rem;color:#aaa;">${(inc.location || '').toUpperCase()} · ${(inc.lg || '').toUpperCase()}</div>
                <div style="font-size:0.68rem;color:#666;margin-top:3px;">👤 ${inc.officer_id || ''} · ${ts}</div>
                ${inc.description ? `<div style="font-size:0.7rem;color:#ccc;margin-top:5px;border-top:1px solid #2a2a2a;padding-top:5px;">${inc.description.substring(0,80)}${inc.description.length > 80 ? '...' : ''}</div>` : ''}
            </div>`;
        }).join('');
    }

    function renderMap(data) {
        markers.forEach(m => map.removeLayer(m));
        markers = [];
        data.forEach((inc, idx) => {
            if (!inc.lat || !inc.lon) return;
            const color = SEV_COLOR[inc.severity] || '#aaa';
            const m = L.circleMarker([inc.lat, inc.lon], {
                radius: inc.severity === 'Critical' ? 10 : inc.severity === 'Medium' ? 7 : 5,
                color: color, fillColor: color, fillOpacity: 0.85, weight: 2
            }).addTo(map);
            m.bindPopup(`<b>${inc.incident_type}</b><br>${inc.location}<br><span style="color:${color};font-weight:bold;">${inc.severity}</span>`);
            m.on('click', () => selectIncident(idx));
            markers.push(m);
        });
    }

    function selectIncident(idx) {
        const filtered = getFiltered();
        const inc = filtered[idx];
        if (!inc) return;

        // Pan map
        if (inc.lat && inc.lon) map.setView([inc.lat, inc.lon], 14);

        // Evidence
        const evEl = document.getElementById('evidenceContent');
        if (inc.evidence_url) {
            evEl.innerHTML = `<img src="${inc.evidence_url}" style="max-width:100%;max-height:220px;border-radius:8px;border:2px solid var(--orange);cursor:zoom-in;" onclick="window.open('${inc.evidence_url}','_blank')"><div style="font-size:0.65rem;color:#555;margin-top:4px;">Click to open full size</div>`;
        } else {
            evEl.innerHTML = '<div style="color:#444;font-size:0.75rem;padding:20px 0;">No evidence uploaded for this incident</div>';
        }

        // Detail
        const sev = (inc.severity || '').toLowerCase();
        const sevColor = SEV_COLOR[inc.severity] || '#aaa';
        const ts = inc.timestamp ? new Date(inc.timestamp).toLocaleString('en-NG') : '--';
        document.getElementById('detailContent').innerHTML = `
            <div class="detail-row"><span class="detail-label">Type</span><span class="detail-val">${TYPE_ICON[inc.incident_type] || ''} ${inc.incident_type || '--'}</span></div>
            <div class="detail-row"><span class="detail-label">Severity</span><span class="detail-val"><span class="sev-badge sev-${sev}">${inc.severity || '--'}</span></span></div>
            <div class="detail-row"><span class="detail-label">Officer</span><span class="detail-val">${inc.officer_id || '--'}</span></div>
            <div class="detail-row"><span class="detail-label">Polling Unit</span><span class="detail-val">${(inc.location || '--').toUpperCase()}</span></div>
            <div class="detail-row"><span class="detail-label">PU Code</span><span class="detail-val">${inc.pu_code || '--'}</span></div>
            <div class="detail-row"><span class="detail-label">Ward</span><span class="detail-val">${(inc.ward || '--').toUpperCase()}</span></div>
            <div class="detail-row"><span class="detail-label">LGA</span><span class="detail-val">${(inc.lg || '--').toUpperCase()}</span></div>
            <div class="detail-row"><span class="detail-label">GPS</span><span class="detail-val">${inc.lat ? inc.lat.toFixed(4) + ', ' + inc.lon.toFixed(4) : 'Not captured'}</span></div>
            <div class="detail-row"><span class="detail-label">Reported</span><span class="detail-val">${ts}</span></div>
            ${inc.description ? `<div style="margin-top:10px;padding:8px;background:#1a1a1a;border-radius:6px;font-size:0.72rem;color:#ccc;border-left:3px solid ${sevColor};">${inc.description}</div>` : ''}
        `;
    }

    function getFiltered() {
        const sev = document.getElementById('f-severity').value;
        const typ = document.getElementById('f-type').value;
        const lga = document.getElementById('f-lga').value;
        let filtered = allIncidents;
        if (sev) filtered = filtered.filter(i => i.severity === sev);
        if (typ) filtered = filtered.filter(i => i.incident_type === typ);
        if (lga) filtered = filtered.filter(i => i.lg === lga);
        return filtered;
    }

    document.addEventListener('DOMContentLoaded', () => {
        initMap();
        loadIncidents();
        setInterval(loadIncidents, 30000);
    });
</script>
</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────



# ── INSIGHT API ENDPOINTS ─────────────────────────────────────────────────────

@app.get("/api/lga_completion")
async def lga_completion(request: Request):
    _require_dashboard(request)
    TOTAL_PUS_PER_LGA = {
        "osogbo": 218,"olorunda": 210,"egbedore": 72,"ede north": 88,"ede south": 50,
        "ejigbo": 120,"ife central": 148,"ife east": 118,"ife north": 74,"ife south": 84,
        "ifedayo": 50,"ifelodun": 136,"ila": 90,"ilesa east": 98,"ilesa west": 100,
        "irepodun": 130,"irewole": 100,"isokan": 72,"iwo": 200,"obokun": 88,
        "odo-otin": 110,"ola-oluwa": 64,"oriade": 100,"orolu": 96,
        "atakumosa east": 76,"atakumosa west": 68,"ayedaade": 112,"ayedire": 68,
        "boluwaduro": 56,"boripe": 90
    }
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT LOWER(lg) as lga, COUNT(*) as submitted FROM field_submissions GROUP BY LOWER(lg)")
        rows = cur.fetchall(); cur.close(); conn.close()
        result = []
        for r in rows:
            lga = r["lga"] or ""
            submitted = r["submitted"]
            total = TOTAL_PUS_PER_LGA.get(lga.lower(), 100)
            pct = round((submitted / total) * 100, 1)
            result.append({"lga": lga.upper(), "submitted": submitted, "total": total, "pct": pct})
        result.sort(key=lambda x: x["pct"], reverse=True)
        return result
    except Exception as e:
        return []

@app.get("/api/swing_pus")
async def swing_pus(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT * FROM field_submissions")
        rows = cur.fetchall(); cur.close(); conn.close()
        PARTIES_LIST = ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"]
        swing = []
        for r in rows:
            try:
                v = json.loads(r.get("edited_votes_json") or r["votes_json"] or "{}") if isinstance(r.get("edited_votes_json") or r.get("votes_json"), str) else (r.get("edited_votes_json") or r.get("votes_json") or {})
            except Exception:
                v = {}
            accord = v.get("ACCORD", 0)
            rivals = {p: v.get(p, 0) for p in PARTIES_LIST if p != "ACCORD"}
            if not rivals: continue
            top_rival = max(rivals, key=rivals.get)
            margin = accord - rivals[top_rival]
            if abs(margin) <= 15:
                swing.append({"pu_name": r["location"], "lga": r["lg"], "ward": r["ward"],
                               "accord": accord, "rival": top_rival,
                               "rival_votes": rivals[top_rival], "margin": margin})
        swing.sort(key=lambda x: abs(x["margin"]))
        return swing
    except Exception as e:
        return []

@app.get("/api/integrity_flags")
async def integrity_flags(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT * FROM field_submissions")
        rows = cur.fetchall(); cur.close(); conn.close()
        flags = []
        for r in rows:
            issues = []
            total_cast = r.get("total_cast") or 0
            accredited = r.get("total_accredited") or 0
            ec8e = r.get("ec8e_image")
            if accredited > 0 and total_cast > accredited:
                issues.append("Overvoting: votes exceed accredited")
            if not ec8e:
                issues.append("No EC8E image uploaded")
            if total_cast == 0:
                issues.append("Zero total votes recorded")
            if issues:
                flags.append({"pu_name": r["location"], "lga": r["lg"], "ward": r["ward"],
                               "issues": issues,
                               "severity": "high" if any("Overvoting" in i for i in issues) else "medium"})
        return flags
    except Exception as e:
        return []

@app.get("/api/collation_timeline")
async def collation_timeline(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT location, lg, timestamp FROM field_submissions ORDER BY timestamp ASC")
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"pu_name": r["location"], "lga": r["lg"], "timestamp": str(r["timestamp"])} for r in rows]
    except Exception as e:
        return []

@app.get("/api/agent_leaderboard")
async def agent_leaderboard(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("""SELECT officer_id, COUNT(*) as submissions, MAX(timestamp) as last_submission
                       FROM field_submissions GROUP BY officer_id ORDER BY submissions DESC LIMIT 20""")
        rows = cur.fetchall(); cur.close(); conn.close()
        return [dict(r) for r in rows]
    except Exception as e:
        return []

# ─────────────────────────────────────────────────────────────────────────────
# ── INCIDENT ENDPOINTS ────────────────────────────────────────────────────────

@app.get("/api/incidents")
async def get_incidents(request: Request):
    _require_dashboard(request)
    try:
        with get_pg() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM incidents ORDER BY timestamp DESC")
                rows = cur.fetchall()
                return [dict(r) for r in rows]
    except Exception as e:
        return []

@app.post("/submit-incident")
async def submit_incident(
    data: str = Form(...),
    evidence: UploadFile = File(None)
):
    try:
        payload = json.loads(data)
        evidence_url = None

        if evidence and evidence.filename:
            safe_pu = str(payload.get("pu_code", "unk")).replace("/", "_").replace(" ", "_")
            public_id = f"incidents/{safe_pu}_{uuid.uuid4().hex[:8]}"
            try:
                img_bytes = await _safe_read_image(evidence)
                upload_result = cloudinary.uploader.upload(
                    img_bytes,
                    public_id=public_id,
                    resource_type="auto",
                    overwrite=True
                )
                evidence_url = upload_result["secure_url"]
                logger.info(f"Evidence uploaded to Cloudinary: {evidence_url}")
            except Exception as cloud_err:
                logger.error(f"Cloudinary upload failed for incident: {cloud_err}")

        with get_pg() as conn:
            with conn.cursor() as cur:
                cur.execute("""INSERT INTO incidents (
                    officer_id, pu_code, ward, ward_code, lg, state, location,
                    incident_type, severity, description, evidence_url,
                    lat, lon, timestamp, status
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
                    payload.get("officer_id"),
                    payload.get("pu_code"),
                    payload.get("ward"),
                    payload.get("ward_code"),
                    payload.get("lg"),
                    payload.get("state"),
                    payload.get("location"),
                    payload.get("incident_type"),
                    payload.get("severity"),
                    payload.get("description"),
                    evidence_url,
                    payload.get("lat"),
                    payload.get("lon"),
                    datetime.now().isoformat(),
                    "open"
                ))
            conn.commit()

        alert_payload = {**payload, "timestamp": datetime.now().strftime("%d %b %Y %H:%M")}
        threading.Thread(target=send_incident_alert, args=(alert_payload,)).start()

        return {"status": "success", "message": "Incident reported successfully"}
    except Exception as e:
        logger.error(f"Incident submission error: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}

# ─────────────────────────────────────────────────────────────────────────────
# ── Dashboard login page ──────────────────────────────────────────────────────
DASHBOARD_LOGIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ACCORD — Dashboard Access</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'Inter', sans-serif;
            background: #020c06;
            color: #fff;
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            overflow: hidden;
        }
        body::before {
            content: '';
            position: fixed;
            inset: 0;
            background:
                radial-gradient(ellipse 120% 80% at 50% -10%, rgba(0,135,81,0.18) 0%, transparent 60%),
                linear-gradient(160deg, #020c06 0%, #041508 40%, #020c06 100%);
            z-index: 0;
        }
        .grid {
            position: fixed; inset: 0; z-index: 0;
            background-image:
                linear-gradient(rgba(0,135,81,0.05) 1px, transparent 1px),
                linear-gradient(90deg, rgba(0,135,81,0.05) 1px, transparent 1px);
            background-size: 60px 60px;
            pointer-events: none;
        }
        .card {
            position: relative; z-index: 1;
            background: rgba(255,255,255,0.04);
            border: 1px solid rgba(0,135,81,0.25);
            border-radius: 20px;
            padding: 48px 40px;
            width: 100%;
            max-width: 420px;
            text-align: center;
            box-shadow: 0 0 60px rgba(0,135,81,0.1);
        }
        .lock-icon {
            width: 64px; height: 64px;
            border-radius: 50%;
            background: rgba(0,135,81,0.12);
            border: 1px solid rgba(0,135,81,0.3);
            display: flex; align-items: center; justify-content: center;
            font-size: 1.6rem;
            margin: 0 auto 24px;
            box-shadow: 0 0 30px rgba(0,135,81,0.2);
        }
        .tag {
            display: inline-block;
            font-size: 0.6rem; font-weight: 700;
            letter-spacing: 0.18em; text-transform: uppercase;
            color: #ffc107;
            background: rgba(255,193,7,0.1);
            border: 1px solid rgba(255,193,7,0.2);
            border-radius: 20px;
            padding: 3px 12px;
            margin-bottom: 14px;
        }
        h1 { font-size: 1.4rem; font-weight: 900; margin-bottom: 8px; }
        p { font-size: 0.8rem; color: rgba(255,255,255,0.4); margin-bottom: 32px; line-height: 1.5; }
        .input-wrap { position: relative; margin-bottom: 12px; }
        input[type="password"] {
            width: 100%;
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(0,135,81,0.3);
            border-radius: 12px;
            color: #fff;
            font-size: 1rem;
            padding: 14px 48px 14px 18px;
            outline: none;
            transition: border-color 0.2s;
            letter-spacing: 0.12em;
        }
        input[type="password"]:focus { border-color: rgba(0,135,81,0.7); }
        .toggle-eye {
            position: absolute; right: 14px; top: 50%;
            transform: translateY(-50%);
            background: none; border: none; color: rgba(255,255,255,0.3);
            cursor: pointer; font-size: 1rem; padding: 0;
        }
        .toggle-eye:hover { color: rgba(255,255,255,0.7); }
        button.submit {
            width: 100%;
            background: linear-gradient(135deg, #008751, #00b368);
            border: none; border-radius: 12px;
            color: #fff; font-size: 0.95rem; font-weight: 700;
            padding: 14px;
            cursor: pointer;
            transition: opacity 0.2s, transform 0.1s;
            margin-bottom: 16px;
        }
        button.submit:hover { opacity: 0.9; }
        button.submit:active { transform: scale(0.98); }
        button.submit:disabled { opacity: 0.5; cursor: not-allowed; }
        .error {
            background: rgba(220,53,69,0.15);
            border: 1px solid rgba(220,53,69,0.3);
            border-radius: 8px;
            color: #ff6b6b;
            font-size: 0.78rem;
            padding: 10px 14px;
            margin-bottom: 12px;
            display: none;
        }
        .back { font-size: 0.72rem; color: rgba(255,255,255,0.25); text-decoration: none; }
        .back:hover { color: rgba(255,255,255,0.5); }
    </style>
</head>
<body>
<div class="grid"></div>
<div class="card">
    <div class="lock-icon">🔐</div>
    <div class="tag">Restricted Access</div>
    <h1>Dashboard Access</h1>
    <p>This dashboard is restricted to authorised command team members only. Enter your access key to continue.</p>
    <div class="error" id="err"></div>
    <div class="input-wrap">
        <input type="password" id="key" placeholder="Enter access key" autocomplete="off" onkeydown="if(event.key==='Enter')verify()">
        <button class="toggle-eye" onclick="toggleEye()" id="eyeBtn" title="Show/hide">👁</button>
    </div>
    <button class="submit" id="btn" onclick="verify()">Unlock Dashboard →</button>
    <a href="/" class="back">← Back to home</a>
</div>
<script>
    function toggleEye() {
        const inp = document.getElementById('key');
        inp.type = inp.type === 'password' ? 'text' : 'password';
    }
    async function verify() {
        const key = document.getElementById('key').value.trim();
        const btn = document.getElementById('btn');
        const err = document.getElementById('err');
        if (!key) return;
        btn.disabled = true; btn.textContent = 'Verifying...';
        err.style.display = 'none';
        try {
            const res = await fetch('/api/verify-dashboard', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ key })
            });
            if (res.ok) {
                window.location.href = '/dashboard';
            } else {
                err.textContent = 'Invalid access key. Contact your system administrator.';
                err.style.display = 'block';
                btn.disabled = false; btn.textContent = 'Unlock Dashboard →';
            }
        } catch(e) {
            err.textContent = 'Connection error. Try again.';
            err.style.display = 'block';
            btn.disabled = false; btn.textContent = 'Unlock Dashboard →';
        }
    }
</script>
</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────



# ── Demo: clear all submissions ────────────────────────────────────────────────
@app.post("/api/admin/clear-submissions")
async def clear_submissions(request: Request):
    auth = request.headers.get("Authorization", "")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(),
        _DASHBOARD_KEY_HASH
    )):
        raise HTTPException(status_code=403, detail="Not authorised")
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as c FROM field_submissions")
            row   = cur.fetchone()
            count = row["c"] if row else 0
            cur.execute("TRUNCATE TABLE field_submissions RESTART IDENTITY")
    return {"status": "ok", "message": f"Cleared {count} submissions", "deleted": count}


# ── Demo: create submission directly (no OTP) ───────────────────────────────────
@app.post("/api/admin/create-submission")
async def create_submission(request: Request):
    auth = request.headers.get("Authorization", "")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(),
        _DASHBOARD_KEY_HASH
    )):
        raise HTTPException(status_code=403, detail="Not authorised")
    body = await request.json()
    now  = datetime.now().isoformat()
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO field_submissions
                (officer_id, state, lg, ward, ward_code, pu_code, location,
                 reg_voters, total_accredited, valid_votes, rejected_votes,
                 total_cast, lat, lon, timestamp, votes_json)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (pu_code) DO UPDATE SET
                    officer_id = EXCLUDED.officer_id,
                    state = EXCLUDED.state,
                    lg = EXCLUDED.lg,
                    ward = EXCLUDED.ward,
                    ward_code = EXCLUDED.ward_code,
                    location = EXCLUDED.location,
                    reg_voters = EXCLUDED.reg_voters,
                    total_accredited = EXCLUDED.total_accredited,
                    valid_votes = EXCLUDED.valid_votes,
                    rejected_votes = EXCLUDED.rejected_votes,
                    total_cast = EXCLUDED.total_cast,
                    lat = EXCLUDED.lat,
                    lon = EXCLUDED.lon,
                    timestamp = EXCLUDED.timestamp,
                    votes_json = EXCLUDED.votes_json
            """, (
                body.get("officer_id", "DEMO"),
                (body.get("state", "osun")).lower(),
                body.get("lg", ""),
                body.get("ward", ""),
                body.get("ward_code", ""),
                body.get("pu_code", ""),
                body.get("location", ""),
                body.get("reg_voters", 0),
                body.get("total_accredited", 0),
                body.get("valid_votes", 0),
                body.get("rejected_votes", 0),
                body.get("total_cast", 0),
                body.get("lat", 0.0),
                body.get("lon", 0.0),
                now,
                json.dumps(body.get("votes", {})),
            ))
    return {"status": "ok", "message": "Submission created"}


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(request: Request):
    from fastapi.responses import HTMLResponse as _HR
    token = request.cookies.get("ds_session")
    if not _is_valid_token(token):
        return _HR(content=DASHBOARD_LOGIN_HTML)
    return _HR(content=DASHBOARD_HTML)

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>ACCORD — Situation Room · Osun 2026</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css" rel="stylesheet">
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;900&display=swap" rel="stylesheet">
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <script src="https://unpkg.com/leaflet.heat@0.2.0/dist/leaflet-heat.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.0.0"></script>
    <style>
    :root {
        --gold:   #F5A623;
        --apc:    #1A56DB;
        --adc:    #00875A;
        --pdp:    #E02424;
        --bg:     #0D0F14;
        --s1:     #161B26;
        --s2:     #1E2433;
        --s3:     #252D3D;
        --border: rgba(255,255,255,0.07);
        --text:   #F0F2F5;
        --muted:  #6B7280;
        --nav-h:  62px;
        --kpi-h:  82px;
    }
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}
    html,body{height:100%;background:var(--bg);color:var(--text);
        font-family:'Inter','Segoe UI',sans-serif;overflow:hidden;}

    /* ── NAVBAR ── */
    .navbar-custom{
        height:var(--nav-h);
        background:linear-gradient(180deg,rgba(22,27,38,0.98) 0%,rgba(13,15,20,0.98) 100%);
        border-bottom:1px solid rgba(245,166,35,0.25);
        backdrop-filter:blur(16px);
        display:flex;align-items:center;justify-content:space-between;
        padding:0 20px;gap:12px;flex-shrink:0;position:relative;z-index:10;
    }
    .brand{display:flex;align-items:center;gap:10px;}
    .live-dot{width:8px;height:8px;background:#22c55e;border-radius:50%;flex-shrink:0;
        box-shadow:0 0 0 0 rgba(34,197,94,0.5);animation:pulse-dot 1.5s infinite;}
    @keyframes pulse-dot{0%{box-shadow:0 0 0 0 rgba(34,197,94,0.5);}
        70%{box-shadow:0 0 0 8px rgba(34,197,94,0);}
        100%{box-shadow:0 0 0 0 rgba(34,197,94,0);}}
    .nav-divider{width:1px;height:32px;background:var(--border);flex-shrink:0;}
    .brand-name{font-weight:900;font-size:1rem;color:var(--gold);letter-spacing:1.5px;}
    .brand-tag{font-size:0.55rem;color:var(--muted);letter-spacing:2px;text-transform:uppercase;margin-top:1px;}

    .party-pills{display:flex;gap:8px;}
    .pill{background:var(--s2);border:1px solid var(--border);border-radius:10px;
        padding:5px 12px 5px 8px;display:flex;align-items:center;gap:8px;
        min-width:115px;transition:border-color 0.2s;}
    .pill-accord{border-bottom:2px solid var(--gold);}
    .pill-apc   {border-bottom:2px solid var(--apc);}
    .pill-adc   {border-bottom:2px solid var(--adc);}
    .pill-logo{width:28px;height:28px;border-radius:6px;background:var(--s3);
        display:flex;align-items:center;justify-content:center;overflow:hidden;flex-shrink:0;}
    .pill-logo img{width:100%;height:100%;object-fit:contain;padding:2px;}
    .pill-info label{display:block;font-size:0.52rem;color:var(--muted);
        text-transform:uppercase;letter-spacing:0.5px;}
    .pill-info span{font-size:1rem;font-weight:800;line-height:1.1;}
    .pill-accord .pill-info span{color:var(--gold);}
    .pill-apc    .pill-info span{color:#60a5fa;}
    .pill-adc    .pill-info span{color:#34d399;}

    .nav-right{display:flex;align-items:center;gap:12px;flex-shrink:0;}
    .pu-badge{background:var(--s2);border:1px solid var(--border);border-radius:8px;
        padding:5px 12px;font-size:0.7rem;color:var(--muted);
        display:flex;align-items:center;gap:6px;}
    .pu-badge b{color:var(--text);font-weight:700;}
    .pu-badge i{color:#a855f7;font-size:0.75rem;}
    .live-clock{font-size:0.88rem;font-weight:700;color:var(--gold);
        font-variant-numeric:tabular-nums;letter-spacing:1px;}
    .logout-btn{background:none;border:1px solid var(--border);color:var(--muted);
        border-radius:8px;padding:5px 10px;font-size:0.7rem;cursor:pointer;
        transition:all 0.2s;font-family:inherit;}
    .logout-btn:hover{border-color:var(--pdp);color:var(--pdp);}

    /* ── KPI STRIP ── */
    .kpi-strip{display:grid;grid-template-columns:repeat(6,1fr);
        gap:8px;padding:8px 16px;flex-shrink:0;}
    .kpi{background:var(--s1);border:1px solid var(--border);border-radius:10px;
        padding:12px 14px;position:relative;overflow:hidden;cursor:default;}
    .kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;border-radius:2px 2px 0 0;}
    .kpi-accord::before{background:linear-gradient(90deg,var(--gold),#ffd666);}
    .kpi-margin::before{background:linear-gradient(90deg,var(--apc),#60a5fa);}
    .kpi-pus::before   {background:linear-gradient(90deg,#7c3aed,#a855f7);}
    .kpi-accred::before{background:linear-gradient(90deg,#0891b2,#22d3ee);}
    .kpi-inc::before   {background:linear-gradient(90deg,var(--pdp),#f87171);}
    .kpi-turn::before  {background:linear-gradient(90deg,var(--adc),#34d399);}
    .kpi-icon{font-size:1rem;margin-bottom:4px;display:block;opacity:0.9;}
    .kpi-accord .kpi-icon{color:var(--gold);}
    .kpi-margin .kpi-icon{color:#60a5fa;}
    .kpi-pus    .kpi-icon{color:#c084fc;}
    .kpi-accred .kpi-icon{color:#22d3ee;}
    .kpi-inc    .kpi-icon{color:#f87171;}
    .kpi-turn   .kpi-icon{color:#34d399;}
    .kpi-val{font-size:1.25rem;font-weight:900;display:block;line-height:1.1;
        transition:all 0.4s;letter-spacing:-0.5px;}
    .kpi-label{font-size:0.58rem;color:#9CA3AF;text-transform:uppercase;
        letter-spacing:0.5px;margin-top:4px;display:block;font-weight:500;}

    /* ── MAIN GRID ── */
    .main-content{
        display:grid;grid-template-columns:290px 1fr 280px;
        gap:8px;padding:0 8px 8px;
        height:calc(100vh - var(--nav-h) - var(--kpi-h));overflow:hidden;
    }
    .side-panel{background:var(--s1);border:1px solid var(--border);border-radius:12px;
        display:flex;flex-direction:column;overflow:hidden;min-height:0;}
    .panel-header{padding:9px 14px;font-size:0.63rem;font-weight:700;
        color:#9CA3AF;text-transform:uppercase;letter-spacing:1px;
        border-bottom:1px solid var(--border);flex-shrink:0;
        display:flex;align-items:center;justify-content:space-between;}
    .ph-badge{background:var(--s2);border:1px solid var(--border);
        border-radius:20px;padding:1px 8px;font-size:0.58rem;color:var(--gold);}

    /* Filter bar */
    .filter-row{padding:7px;display:flex;gap:5px;border-bottom:1px solid var(--border);flex-shrink:0;}
    .filter-row input,.filter-row select{flex:1;background:var(--s2);color:var(--text);
        border:1px solid var(--border);border-radius:8px;font-size:0.68rem;
        padding:5px 8px;font-family:inherit;outline:none;}
    .filter-row input:focus,.filter-row select:focus{border-color:rgba(245,166,35,0.4);}
    .filter-row input::placeholder{color:var(--muted);}

    /* PU Feed */
    .feed-container{flex:1;overflow-y:auto;padding:8px;min-height:0;}
    .feed-container::-webkit-scrollbar{width:3px;}
    .feed-container::-webkit-scrollbar-thumb{background:var(--s3);border-radius:3px;}

    .pu-card{background:var(--s2);border:1px solid var(--border);border-radius:10px;
        padding:9px 11px;margin-bottom:6px;cursor:pointer;
        transition:background 0.15s,border-color 0.15s,transform 0.1s;}
    .pu-card:hover{background:var(--s3);border-color:rgba(245,166,35,0.2);transform:translateX(2px);}
    .pu-card-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:3px;}
    .pu-card-name{font-size:0.72rem;font-weight:600;flex:1;min-width:0;
        overflow:hidden;text-overflow:ellipsis;white-space:nowrap;margin-right:6px;}
    .winner-badge{font-size:0.52rem;font-weight:700;padding:2px 7px;border-radius:20px;flex-shrink:0;}
    .wb-ACCORD{background:rgba(245,166,35,0.12);color:var(--gold);border:1px solid rgba(245,166,35,0.2);}
    .wb-APC   {background:rgba(26,86,219,0.12);color:#60a5fa;border:1px solid rgba(26,86,219,0.2);}
    .wb-ADC   {background:rgba(0,135,90,0.12);color:#34d399;border:1px solid rgba(0,135,90,0.2);}
    .wb-PDP   {background:rgba(224,36,36,0.12);color:#f87171;border:1px solid rgba(224,36,36,0.2);}
    .wb-other {background:rgba(107,114,128,0.12);color:var(--muted);border:1px solid var(--border);}
    .pu-card-meta{font-size:0.6rem;color:var(--muted);margin-bottom:5px;}
    .score-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:3px;}
    .score-item{background:var(--s1);border-radius:6px;padding:4px 5px;text-align:center;border:1px solid transparent;}
    .si-ACCORD{border-color:rgba(245,166,35,0.15);}
    .si-APC   {border-color:rgba(26,86,219,0.15);}
    .si-ADC   {border-color:rgba(0,135,90,0.15);}
    .score-label{font-size:0.5rem;color:var(--muted);display:block;margin-bottom:1px;}
    .score-val{font-size:0.72rem;font-weight:700;}
    .sv-ACCORD{color:var(--gold);}
    .sv-APC   {color:#60a5fa;}
    .sv-ADC   {color:#34d399;}
    .sv-PDP   {color:#f87171;}
    .sv-other {color:var(--muted);}
    .ec8e-badge{font-size:0.55rem;font-weight:700;background:rgba(245,166,35,0.15);
        color:var(--gold);padding:1px 5px;border-radius:4px;margin-left:4px;}

    /* Centre column */
    .centre-col{display:flex;flex-direction:column;gap:8px;min-height:0;overflow:hidden;}
    #map{height:45%;border-radius:12px;background:#111;
        border:1px solid rgba(245,166,35,0.15);flex-shrink:0;z-index:1;}
    .leaflet-heatmap-layer{opacity:0.55!important;}
    .leaflet-tile-pane{z-index:2;opacity:1!important;}
    .leaflet-overlay-pane{z-index:3;opacity:0.65!important;}
    .leaflet-marker-pane{z-index:4;}
    .leaflet-tooltip-pane{z-index:5;}
    .leaflet-popup-pane{z-index:6;}
    .chart-row{display:grid;grid-template-columns:1fr 1fr;gap:8px;flex:1;min-height:120px;}
    .chart-box{background:var(--s1);border:1px solid var(--border);border-radius:12px;
        padding:12px;display:flex;flex-direction:column;min-height:0;overflow:hidden;}
    .chart-title{font-size:0.62rem;font-weight:700;color:#9CA3AF;
        text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;flex-shrink:0;}
    .chart-box canvas{display:block;width:100%!important;flex:1;min-height:0;}

    /* Right panel */
    .margin-card{margin:10px;background:var(--s2);border:1px solid rgba(245,166,35,0.2);
        border-radius:10px;padding:12px;text-align:center;flex-shrink:0;}
    .margin-val{font-size:1.8rem;font-weight:900;color:var(--gold);display:block;line-height:1;}
    .margin-label{font-size:0.58rem;color:var(--muted);margin-top:3px;}
    .margin-rival{font-size:0.63rem;color:#34d399;margin-top:4px;
        display:flex;align-items:center;justify-content:center;gap:3px;}

    .proj-card{margin:0 10px 8px;background:var(--s2);
        border:1px solid rgba(0,135,90,0.2);border-radius:10px;
        padding:9px 12px;flex-shrink:0;display:flex;align-items:center;justify-content:space-between;}
    .proj-val{font-size:1.25rem;font-weight:900;color:#34d399;}
    .proj-label{font-size:0.57rem;color:var(--muted);}

    .lga-list{flex:1;overflow-y:auto;padding:8px;min-height:0;}
    .lga-list::-webkit-scrollbar{width:3px;}
    .lga-list::-webkit-scrollbar-thumb{background:var(--s3);}
    .lga-row{display:flex;align-items:center;gap:7px;margin-bottom:6px;}
    .lga-name{font-size:0.64rem;width:80px;flex-shrink:0;color:var(--text);
        white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
    .lga-bar-bg{flex:1;height:5px;background:var(--s3);border-radius:3px;overflow:hidden;}
    .lga-bar-fill{height:100%;border-radius:3px;
        background:linear-gradient(90deg,var(--adc),var(--gold));
        transition:width 0.6s ease;}
    .lga-pct{font-size:0.58rem;color:var(--muted);width:26px;text-align:right;flex-shrink:0;}

    .ai-box{margin:0 10px 10px;background:#000;border:1px solid rgba(34,197,94,0.15);
        border-radius:10px;padding:10px;font-family:'Courier New',monospace;
        font-size:0.62rem;color:#22c55e;flex:1;overflow-y:auto;min-height:0;
        line-height:1.7;white-space:pre-wrap;}

    /* EC8E viewer */
    #ec8eViewerPanel{background:var(--s2);border-top:1px solid var(--border);
        padding:8px;flex-shrink:0;display:none;}
    #ec8eViewerPanel.active{display:block;}

    /* Overlays */
    .ov-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.85);
        z-index:1000;align-items:center;justify-content:center;}
    .ov-overlay.active{display:flex;}
    .ov-inner{background:var(--s1);border:1px solid var(--border);border-radius:16px;
        padding:24px;max-width:600px;width:90%;max-height:80vh;overflow-y:auto;position:relative;}
    .ov-close{position:absolute;top:12px;right:14px;background:none;border:none;
        color:var(--muted);font-size:1.2rem;cursor:pointer;}
    .ov-close:hover{color:var(--text);}
    </style>
</head>
<body>

<!-- NAVBAR -->
<nav class="navbar-custom">
    <div class="brand">
        <div class="live-dot"></div>
        <div class="nav-divider"></div>
        <div>
            <div class="brand-name">ACCORD · OSUN 2026</div>
            <div class="brand-tag">Situation Room — Live Command</div>
        </div>
    </div>

    <div class="party-pills">
        <div class="pill pill-accord">
            <div class="pill-logo">
                <img src="/static/logos/ACCORD.png" alt="ACCORD"
                     onerror="this.style.display='none'">
            </div>
            <div class="pill-info">
                <label>ACCORD</label>
                <span id="nav-ACCORD">0</span>
            </div>
        </div>
        <div class="pill pill-apc">
            <div class="pill-logo">
                <img src="/static/logos/APC.png" alt="APC"
                     onerror="this.style.display='none'">
            </div>
            <div class="pill-info">
                <label>APC</label>
                <span id="nav-APC">0</span>
            </div>
        </div>
        <div class="pill pill-adc">
            <div class="pill-logo">
                <img src="/static/logos/ADC.png" alt="ADC"
                     onerror="this.style.display='none'">
            </div>
            <div class="pill-info">
                <label>ADC</label>
                <span id="nav-ADC">0</span>
            </div>
        </div>
        <!-- Hidden spans for other parties (needed by existing JS) -->
        <span id="nav-PDP" style="display:none">0</span>
        <span id="nav-LP"  style="display:none">0</span>
        <span id="nav-NNPP" style="display:none">0</span>
        <span id="nav-APGA" style="display:none">0</span>
        <span id="nav-PRP"  style="display:none">0</span>
        <span id="nav-SDP"  style="display:none">0</span>
        <span id="nav-AA"   style="display:none">0</span>
        <span id="nav-AAC"  style="display:none">0</span>
        <span id="nav-ADP"  style="display:none">0</span>
        <span id="nav-APM"  style="display:none">0</span>
        <span id="nav-APP"  style="display:none">0</span>
        <span id="nav-BP"   style="display:none">0</span>
        <span id="nav-YPP"  style="display:none">0</span>
        <span id="nav-ZLP"  style="display:none">0</span>
    </div>

    <div class="nav-right">
        <div class="pu-badge">
            <i class="bi bi-geo-alt-fill"></i>
            <b id="pu-count">0</b>&nbsp;PUs Reported
        </div>
        <div class="live-clock" id="liveClock">--:--:--</div>
        <button class="logout-btn" onclick="logoutDash()">
            <i class="bi bi-box-arrow-right"></i> Logout
        </button>
    </div>
</nav>

<!-- KPI STRIP -->
<div class="kpi-strip">
    <div class="kpi kpi-accord">
        <i class="bi bi-check-circle-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-accord-val">0</span>
        <span class="kpi-label">ACCORD Total</span>
    </div>
    <div class="kpi kpi-margin">
        <i class="bi bi-graph-up-arrow kpi-icon"></i>
        <span class="kpi-val" id="kpi-margin-val">0</span>
        <span class="kpi-label">Lead Margin</span>
    </div>
    <div class="kpi kpi-pus">
        <i class="bi bi-geo-alt-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-pus-val">0</span>
        <span class="kpi-label">PUs Reported</span>
    </div>
    <div class="kpi kpi-accred">
        <i class="bi bi-people-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-accred-val">0</span>
        <span class="kpi-label">Accredited</span>
    </div>
    <div class="kpi kpi-inc">
        <i class="bi bi-exclamation-triangle-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-inc-val">0</span>
        <span class="kpi-label">Incidents</span>
    </div>
    <div class="kpi kpi-turn">
        <i class="bi bi-percent kpi-icon"></i>
        <span class="kpi-val" id="kpi-turn-val">0%</span>
        <span class="kpi-label">Avg Turnout</span>
    </div>
</div>

<!-- MAIN GRID -->
<div class="main-content">

    <!-- LEFT: PU FEED -->
    <div class="side-panel">
        <div class="panel-header">
            PU Results Feed
            <span class="ph-badge" id="feed-count">0</span>
        </div>
        <div class="filter-row">
            <input type="text" id="puSearch" placeholder="🔍  Search PU..." oninput="applyFilters()">
        </div>
        <div class="filter-row" style="border-top:none;padding-top:0;">
            <select id="fState" onchange="loadFilters()">
                <option value="">All States</option>
                <option value="osun" selected>Osun</option>
            </select>
            <select id="fLGA" onchange="updateWards()">
                <option value="">All LGAs</option>
            </select>
            <select id="fWard" onchange="applyFilters()">
                <option value="">All Wards</option>
            </select>
        </div>
        <div class="feed-container" id="feedList"></div>
        <div id="ec8eViewerPanel">
            <div id="ec8eContent" style="font-size:0.7rem;color:var(--muted);">
                Click a PU card to view EC8E form image
            </div>
        </div>
    </div>

    <!-- CENTRE: MAP + CHARTS -->
    <div class="centre-col">
        <div id="map"></div>
        <div class="chart-row">
            <div class="chart-box">
                <div class="chart-title">🗳 Vote Share %</div>
                <canvas id="pieChart"></canvas>
            </div>
            <div class="chart-box">
                <div class="chart-title">📊 Party Vote Count</div>
                <canvas id="barChart"></canvas>
            </div>
        </div>
    </div>

    <!-- RIGHT: ANALYTICS -->
    <div class="side-panel">
        <div class="panel-header">Analytics & Projection</div>

        <div class="margin-card">
            <span class="margin-val" id="marginVal">0</span>
            <div class="margin-label" id="marginLead">ACCORD LEAD</div>
            <div class="margin-rival">
                <i class="bi bi-arrow-up-short"></i>
                <span id="projectionNote">Awaiting data...</span>
            </div>
        </div>

        <div class="proj-card">
            <div>
                <div class="proj-val" id="projectionVal">—</div>
                <div class="proj-label">Projected ACCORD Total</div>
            </div>
            <i class="bi bi-lightning-fill" style="color:var(--gold);font-size:1.1rem;"></i>
        </div>

        <div class="panel-header" style="margin-top:2px;">
            LGA Completion
            <span class="ph-badge" id="lga-count">0 LGAs</span>
        </div>
        <div class="lga-list" id="lgaCompletionList"></div>

        <div class="panel-header">AI Insight</div>
        <div class="ai-box" id="ai_box">Awaiting data...</div>
    </div>
</div>

<!-- OVERLAYS (kept from original) -->
<div id="ov-lga" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-lga')">✕</button>
        <h5 style="color:var(--gold);margin-bottom:12px;">LGA COMPLETION</h5>
        <div id="ov-lga-inner" style="max-height:60vh;overflow-y:auto;"></div>
    </div>
</div>
<div id="ov-swing" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-swing')">✕</button>
        <h5 style="color:#ff4444;margin-bottom:12px;">SWING POLLING UNITS</h5>
        <div id="ov-swing-inner" style="max-height:60vh;overflow-y:auto;"></div>
    </div>
</div>
<div id="ov-flags" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-flags')">✕</button>
        <h5 style="color:#ff6600;margin-bottom:12px;">RESULT INTEGRITY FLAGS</h5>
        <div id="ov-flags-inner" style="max-height:60vh;overflow-y:auto;"></div>
    </div>
</div>
<div id="ov-proj" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-proj')">✕</button>
        <h5 style="color:#00ff88;margin-bottom:12px;">PROJECTED TALLY + AGENTS</h5>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;">
            <div>
                <div style="font-size:0.75rem;color:#aaa;margin-bottom:8px;">PROJECTED ACCORD TOTAL</div>
                <div id="ov-projVal" style="font-size:3rem;font-weight:900;color:#00ff88;">--</div>
                <div id="ov-projNote" style="font-size:0.75rem;color:#555;margin-top:4px;"></div>
            </div>
            <div>
                <div style="font-size:0.75rem;color:#ffc107;font-weight:bold;margin-bottom:8px;">AGENT LEADERBOARD</div>
                <div id="ov-agentList"></div>
            </div>
        </div>
    </div>
</div>
<div id="ov-timeline" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-timeline')">✕</button>
        <h5 style="color:var(--gold);margin-bottom:12px;">COLLATION TIMELINE</h5>
        <div style="position:relative;height:320px;"><canvas id="ov-timelineChart"></canvas></div>
        <div id="ov-timeline-list" style="max-height:200px;overflow-y:auto;margin-top:12px;"></div>
    </div>
</div>

<script>
    const PARTIES = ['ACCORD','AA','AAC','ADC','ADP','APGA','APC','APM','APP','BP','NNPP','PRP','YPP','ZLP'];
    const BIG3    = ['ACCORD','APC','ADC'];
    const COLORS  = { ACCORD:'#F5A623', APC:'#1A56DB', ADC:'#00875A', PDP:'#E02424' };

    let map, markers = [], heatLayer = null, allIncidents = [];
    let globalData = [], globalTotals = {};
    let pie = null, bar = null;

    // ── Live clock ──
    function tickClock() {
        const n = new Date();
        const el = document.getElementById('liveClock');
        if (el) el.textContent = n.toLocaleTimeString('en-NG',
            {hour:'2-digit',minute:'2-digit',second:'2-digit'});
    }
    setInterval(tickClock, 1000); tickClock();

    // ── Map init ──
    function init() {
        // Locked to Osun State
                map = L.map('map', {zoomControl:false}).setView([7.56, 4.52], 9);
        L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png')
         .addTo(map);

        refreshData();
        loadInsights();
    }

    // ── Data refresh ──
    async function refreshData() {
        try {
            const res = await fetch(window.location.origin + '/submissions', {credentials:'include'});
            globalData = await res.json();
            globalData = globalData.map(x => ({
                ...x,
                state: (x.state||'').toLowerCase(),
                lga:   (x.lga||'').toLowerCase(),
                ward:  (x.ward||'').toLowerCase()
            }));
            applyFilters();
        } catch(e) { console.error('Data refresh error', e); }
    }

    // ── Filters ──
    async function loadFilters() {
        try {
            const s = document.getElementById('fState').value;
            if (!s) return;
            const res = await fetch('/api/lgas/' + s, {credentials:'include'});
            const lgas = await res.json();
            const sel = document.getElementById('fLGA');
            sel.innerHTML = '<option value="">All LGAs</option>';
            lgas.forEach(l => {
                const o = document.createElement('option');
                o.value = l; o.textContent = l.charAt(0).toUpperCase() + l.slice(1);
                sel.appendChild(o);
            });
            applyFilters();
        } catch(e) {}
    }

    async function updateWards() {
        const s = document.getElementById('fState').value;
        const l = document.getElementById('fLGA').value;
        const sel = document.getElementById('fWard');
        sel.innerHTML = '<option value="">All Wards</option>';
        if (s && l) {
            try {
                const res = await fetch(`/api/wards/${s}/${l}`, {credentials:'include'});
                const wards = await res.json();
                wards.forEach(w => {
                    const o = document.createElement('option');
                    o.value = w.name || w; o.textContent = w.name || w;
                    sel.appendChild(o);
                });
            } catch(e) {}
        }
        applyFilters();
    }

    function applyFilters() {
        const s = document.getElementById('fState').value.toLowerCase();
        const l = document.getElementById('fLGA').value.toLowerCase();
        const w = document.getElementById('fWard').value.toLowerCase();
        let filtered = globalData;
        if(s) filtered = filtered.filter(x => x.state === s);
        if(l) filtered = filtered.filter(x => x.lga   === l);
        if(w) filtered = filtered.filter(x => x.ward  === w);
        updateUI(filtered);
    }

    // ── Main UI update — wires ALL metrics ──
    function updateUI(data) {
        const t = {};
        PARTIES.forEach(p => t[p] = 0);
        let totalAccred = 0, totalReg = 0;

        const list = document.getElementById('feedList');
        list.innerHTML = '';
        markers.forEach(m => map.removeLayer(m));
        markers = [];
        const heatPoints = [];

        const searchTerm = (document.getElementById('puSearch').value || '').toLowerCase();

        data.forEach(d => {
            PARTIES.forEach(p => { t[p] += (d['votes_party_'+p] || 0); });
            totalAccred += (d.total_accredited || 0);
            totalReg    += (d.reg_voters       || 0);

            if (searchTerm && !(d.pu_name||'').toLowerCase().includes(searchTerm)) return;

            // Determine winner
            const partyVotes = BIG3.map(p => ({p, v: d['votes_party_'+p]||0}));
            partyVotes.sort((a,b) => b.v - a.v);
            const winner = partyVotes[0].p;
            const wbClass = 'wb-' + (COLORS[winner] ? winner : 'other');

            const ec8eBadge = d.ec8e_image
                ? `<span class="ec8e-badge">📷 EC8E</span>` : '';

            const card = document.createElement('div');
            card.className = 'pu-card';
            card.innerHTML = `
                <div class="pu-card-head">
                    <span class="pu-card-name">${d.pu_name||'—'}${ec8eBadge}</span>
                    <span class="winner-badge ${wbClass}">${winner} ✓</span>
                </div>
                <div class="pu-card-meta">${(d.lga||'').toUpperCase()} · ${d.ward||''}</div>
                <div class="score-grid">
                    <div class="score-item si-ACCORD">
                        <span class="score-label">ACCORD</span>
                        <span class="score-val sv-ACCORD">${(d.votes_party_ACCORD||0).toLocaleString()}</span>
                    </div>
                    <div class="score-item si-APC">
                        <span class="score-label">APC</span>
                        <span class="score-val sv-APC">${(d.votes_party_APC||0).toLocaleString()}</span>
                    </div>
                    <div class="score-item si-ADC">
                        <span class="score-label">ADC</span>
                        <span class="score-val sv-ADC">${(d.votes_party_ADC||0).toLocaleString()}</span>
                    </div>
                    <div class="score-item">
                        <span class="score-label">PDP</span>
                        <span class="score-val sv-PDP">${(d.votes_party_PDP||0).toLocaleString()}</span>
                    </div>
                </div>`;
            card.onclick = () => {
                if(d.latitude) map.setView([d.latitude, d.longitude], 14);
                showEc8e(d.ec8e_image, d.pu_name);
            };
            list.appendChild(card);

            // Map marker — coloured by winner
            if (d.latitude) {
                const col = COLORS[winner] || '#ffc107';
                const m = L.circleMarker([d.latitude, d.longitude], {
                    radius:7, color:col, fillColor:col, fillOpacity:0.85, weight:2
                }).addTo(map);
                m.bindPopup(`<b>${d.pu_name}</b><br>
                    ACCORD: <b style="color:#F5A623">${d.votes_party_ACCORD||0}</b>
                    &nbsp; APC: ${d.votes_party_APC||0}
                    &nbsp; ADC: <span style="color:#34d399">${d.votes_party_ADC||0}</span>`);
                markers.push(m);
                heatPoints.push([d.latitude, d.longitude,
                    Math.min(1, (d.votes_party_ACCORD||0) / 400)]);
            }
        });

        // Heatmap — subtle overlay, tiles stay visible
        if (heatLayer) { map.removeLayer(heatLayer); heatLayer = null; }
        if (typeof L.heatLayer !== 'undefined' && heatPoints.length > 0) {
            heatLayer = L.heatLayer(heatPoints, {
                radius:35, blur:25, maxZoom:14, minOpacity:0.4,
                gradient:{
                    0.0:'#00000000',
                    0.3:'#006400',
                    0.5:'#ffa500',
                    0.7:'#ff4500',
                    1.0:'#ffffff'
                }
            }).addTo(map);
            // Bring markers on top of heatmap
            markers.forEach(m => { try{ m.bringToFront(); }catch(e){} });
        }

        // ── Update ALL nav party spans ──
        PARTIES.forEach(p => {
            const el = document.getElementById('nav-'+p);
            if(el) el.innerText = (t[p]||0).toLocaleString();
        });

        // ── Update KPI cards ──
        const rivals = {};
        PARTIES.filter(p => p !== 'ACCORD').forEach(p => rivals[p] = t[p]);
        const topRival = Object.keys(rivals).reduce((a,b) => rivals[a]>rivals[b]?a:b, 'APC');
        const margin   = t.ACCORD - rivals[topRival];
        const turnout  = totalReg > 0 ? ((totalAccred/totalReg)*100).toFixed(1) : '0.0';

        const set = (id, val) => { const el=document.getElementById(id); if(el) el.innerText=val; };
        set('kpi-accord-val', (t.ACCORD||0).toLocaleString());
        set('kpi-margin-val', (margin>=0?'+':'')+margin.toLocaleString());
        set('kpi-pus-val',    data.length.toLocaleString());
        set('kpi-accred-val', totalAccred.toLocaleString());
        set('kpi-turn-val',   turnout+'%');
        set('pu-count',       data.length.toLocaleString());
        set('feed-count',     data.length.toLocaleString());

        // Margin card
        set('marginVal',  Math.abs(margin).toLocaleString());
        set('marginLead', margin>=0 ? `ACCORD LEAD OVER ${topRival}` : `TRAILING ${topRival}`);
        const mEl = document.getElementById('marginVal');
        if(mEl) mEl.style.color = margin>=0 ? '#F5A623' : '#f87171';

        globalTotals = {...t};
        updateProjection(globalTotals, data.length);
        updateCharts(t);
        runAI(t);
        loadInsights();
    }

    // ── Charts ──
    function updateCharts(t) {
        const labels = BIG3;
        const vals   = labels.map(p => t[p]||0);
        const colors = ['#F5A623','#1A56DB','#00875A'];
        const total  = vals.reduce((a,b)=>a+b,0);

        if(pie) pie.destroy();
        pie = new Chart(document.getElementById('pieChart'), {
            type:'doughnut',
            data:{labels, datasets:[{data:vals, backgroundColor:colors,
                borderWidth:2, borderColor:'#0D0F14', hoverOffset:6}]},
            options:{
                responsive:true, maintainAspectRatio:false,
                plugins:{
                    legend:{
                        position:'bottom',
                        labels:{color:'#9CA3AF',font:{size:9},
                            boxWidth:10,padding:10,usePointStyle:true}
                    },
                    datalabels:{
                        color:'#fff',
                        font:{weight:'900',size:11},
                        textShadowBlur:4,
                        textShadowColor:'rgba(0,0,0,0.8)',
                        formatter:(val,ctx)=>{
                            if(total===0||val===0) return '';
                            const pct=((val/total)*100).toFixed(1);
                            return pct>5 ? pct+'%' : '';
                        }
                    },
                    tooltip:{callbacks:{
                        label:ctx=>`${ctx.label}: ${ctx.parsed.toLocaleString()} votes (${total>0?((ctx.parsed/total)*100).toFixed(1):0}%)`
                    }}
                },
                cutout:'60%'
            }
        });

        if(bar) bar.destroy();
        bar = new Chart(document.getElementById('barChart'), {
            type:'bar',
            data:{labels, datasets:[{data:vals, backgroundColor:colors,
                borderRadius:4, borderSkipped:false}]},
            options:{
                responsive:true, maintainAspectRatio:false,
                indexAxis:'y',
                plugins:{
                    legend:{display:false},
                    datalabels:{color:'#fff',anchor:'end',align:'right',
                        font:{weight:'bold',size:9},
                        formatter:(val)=>val>0?val.toLocaleString():''}
                },
                scales:{
                    x:{beginAtZero:true, grid:{color:'rgba(255,255,255,0.04)'},
                       ticks:{color:'#6B7280',font:{size:9}}, border:{display:false}},
                    y:{grid:{display:false}, ticks:{color:'#F0F2F5',font:{size:10,weight:'600'}},
                       border:{display:false}}
                }
            }
        });
    }

    // ── Projection ──
    function updateProjection(totals, puCount) {
        const TOTAL_PUS = 500;
        if (puCount < 1) return;
        const proj = Math.round((totals.ACCORD / puCount) * TOTAL_PUS);
        const set = (id,v) => { const el=document.getElementById(id); if(el) el.innerText=v; };
        set('projectionVal',  '~' + proj.toLocaleString());
        set('ov-projVal',     '~' + proj.toLocaleString());
        set('projectionNote', `Based on ${puCount} of ~${TOTAL_PUS} PUs`);
        set('ov-projNote',    `Based on ${puCount} of ~${TOTAL_PUS} PUs`);
    }

    // ── AI Insight ──
    async function runAI(totals) {
        try {
            const s = document.getElementById('fState').value.toLowerCase();
            const l = document.getElementById('fLGA').value.toLowerCase();
            const payload = Object.assign({}, totals, {lg:l||'ALL', state:s||'Osun'});
            const res = await fetch(window.location.origin + '/api/ai_interpret', {
                method:'POST', headers:{'Content-Type':'application/json'},
                credentials:'include', body:JSON.stringify(payload)
            });
            const out = await res.json();
            const el = document.getElementById('ai_box');
            if(el) el.innerText = out.analysis || 'No insight available.';
        } catch(e) {}
    }

    // ── Insights (LGA completion etc) ──
    async function loadInsights() {
        loadLGACompletion();
    }

    async function loadLGACompletion() {
        try {
            const res = await fetch('/api/lga_completion', {credentials:'include'});
            const data = await res.json();
            const el = document.getElementById('lgaCompletionList');
            if(!el) return;
            el.innerHTML = '';
            const lgas = data.slice(0,15);
            lgas.forEach(row => {
                const pct = row.pct !== undefined ? row.pct
                    : Math.round((row.submitted||row.reported||0) / Math.max(row.total,1) * 100);
                const div = document.createElement('div');
                div.className = 'lga-row';
                div.innerHTML = `
                    <span class="lga-name">${row.lga||row.name||''}</span>
                    <div class="lga-bar-bg">
                        <div class="lga-bar-fill" style="width:${Math.min(pct,100)}%"></div>
                    </div>
                    <span class="lga-pct">${pct}%</span>`;
                el.appendChild(div);
            });
            const countEl = document.getElementById('lga-count');
            if(countEl) countEl.textContent =
                `${lgas.filter(r=>(r.submitted||r.reported||0)>0).length} / 30`;
        } catch(e) {}
    }

    // ── Incidents KPI ──
    async function loadIncidents() {
        try {
            const res = await fetch('/api/incidents', {credentials:'include'});
            const data = await res.json();
            const el = document.getElementById('kpi-inc-val');
            if(el) el.innerText = data.length;
        } catch(e) {}
    }

    // ── EC8E viewer ──
    function showEc8e(url, name) {
        const panel = document.getElementById('ec8eViewerPanel');
        const content = document.getElementById('ec8eContent');
        if(!panel || !content) return;
        if(url) {
            panel.classList.add('active');
            content.innerHTML = `<div style="font-size:0.65rem;color:var(--muted);margin-bottom:6px;">${name}</div>
                <img src="${url}" style="max-width:100%;max-height:180px;border-radius:8px;
                    border:1px solid var(--border);cursor:zoom-in;"
                    onclick="window.open('${url}','_blank')">`;
        } else {
            panel.classList.remove('active');
            content.innerHTML = 'No EC8E image for this PU.';
        }
    }

    // ── Overlays ──
    function openOverlay(id) {
        const el = document.getElementById(id);
        if(el) el.classList.add('active');
    }
    function closeOverlay(id) {
        const el = document.getElementById(id);
        if(el) el.classList.remove('active');
    }

    // ── Logout ──
    async function logoutDash() {
        await fetch('/api/logout-dashboard', {method:'POST'});
        window.location.href = '/dashboard';
    }

    // ── Boot ──
    document.addEventListener('DOMContentLoaded', init);
    setInterval(refreshData,   2000);
    setInterval(loadIncidents, 2000);
    setInterval(loadInsights,  2000);
</script>
</body>
</html>
"""
